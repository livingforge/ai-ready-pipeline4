"""配る見本の資料を作る。**中身が読める形（このコード）が正本である。**

書き出したものは `examples/from-excel/資料/` と `examples/from-documents/資料/`
に**コミットしてある** ―― `arp4` を試す人が Python を動かさずに開いて確かめ
られるようにするためで、`tests/` の検体（61 本・11MB あり、変更のたびに全部
差し替わる）とは扱いが違う。**数と変更頻度が違うから、置き方も違う。**

コミットする以上、**再生成して差分が出ないこと**が要る（`reproducible`）――
毎回差分が出るものは、差分が出ても誰も見なくなる。

**新しい形式の見本は検体から作る。** ここで 2 本目の生成器を書くと、
テストが見ているもの（`tests/dataset/`）と配っているものが別々に腐る ――
見本が壊れていても、どのテストも落ちない状態になる。
"""

import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook

_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT / "build"))
import reproducible  # noqa: E402


def build(directory: Path) -> list[Path]:
    """Excel だけの最小サンプル（`examples/from-excel/資料/`）。

    `tests/test_example.py` がここを起点に**通しで**（parse → 整理 → freeze →
    build → publish）走るので、**足すときは通しの重さも増える**ことに注意する。
    """
    directory.mkdir(parents=True, exist_ok=True)
    made = [_design(directory / "基本設計書.xlsx"),
            _operation(directory / "運用設計.xlsx")]
    return made


#: 文書形式の見本に選んだ検体（`tests/dataset/*.yml` の ``置き場``）。
#: **1 形式に 1 冊**である ―― 見本は「開いて何が読めるかを確かめる」ためのもので、
#: 網羅は検体の仕事だからである（増やすと、どれを開けばいいか分からなくなる）。
DOCUMENTS = {
    "資料/O/新販売管理システム方式提案（第2.1版）.pptx": "方式提案.pptx",
    "資料/P/受注登録機能仕様書（第1.2版）.docx": "受注登録機能仕様書.docx",
    "資料/Q/受注管理システム検収仕様書（第1.0版）.pdf": "検収仕様書.pdf",
    "資料/N/得意先マスタ移行.csv": "得意先マスタ移行.csv",
}


def build_documents(directory: Path) -> list[Path]:
    """Excel 以外の見本（`examples/from-documents/資料/`）。

    **検体そのものを 4 冊だけ書き出して置く。** 見本用の生成器を別に書かない
    のは、書けば必ず**テストが見ているものと配っているものが割れる**からである
    ―― `tests/dataset/*.yml` の ``なぜ`` に、その 1 冊が何を突くかまで書いてある。
    """
    sys.path.insert(0, str(_ROOT / "tests"))
    import dataset                                   # noqa: PLC0415

    directory.mkdir(parents=True, exist_ok=True)
    made: list[Path] = []
    with tempfile.TemporaryDirectory() as work:
        for path in dataset.build(Path(work), only=DOCUMENTS):
            where = directory / DOCUMENTS[
                path.relative_to(Path(work)).as_posix()]
            shutil.copyfile(path, where)
            made.append(where)
    return sorted(made)


def _design(path: Path) -> Path:
    book = Workbook()
    sheet = book.active
    sheet.title = "受注テーブル"
    sheet["B2"] = "受注テーブル（T_ORDER）定義書  改訂 2.1"
    rows = [["論理名", "物理名", "型", "桁", "PK", "必須", "備考"],
            ["受注番号", "ORDER_NO", "文字列", 10, "○", "○", "採番ルールは別紙"],
            ["受注日", "ORDER_DATE", "日付", None, None, "○", None],
            ["顧客コード", "CUSTOMER_CD", "文字列", 8, None, "○", None]]
    for r, row in enumerate(rows, start=5):
        for c, value in enumerate(row, start=2):
            if value is not None:
                sheet.cell(row=r, column=c, value=value)

    cover = book.create_sheet("表紙")
    cover["A1"] = "基本設計書"
    cover["A2"] = "2026-08-01 版"

    reproducible.stamp(book)
    book.save(path)
    return reproducible.freeze(reproducible.restamp(path))


def _operation(path: Path) -> Path:
    book = Workbook()
    sheet = book.active
    sheet.title = "運用方針"
    sheet["A1"] = "受注データの保持期間"
    sheet["A2"] = "受注データは 13 か月でアーカイブする"
    reproducible.stamp(book)
    book.save(path)
    return reproducible.freeze(reproducible.restamp(path))


if __name__ == "__main__":
    here = Path(__file__).parent
    for made in (build(here / "from-excel" / "資料")
                 + build_documents(here / "from-documents" / "資料")):
        print(made)
