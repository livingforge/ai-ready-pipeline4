"""日本の SIer が書く Excel 設計書を組み立てる道具立て。

テスト資材の見た目と癖を本物に寄せるためのヘルパで、パイプライン本体
（``arp``）からは参照しない。ここが再現するのは次の 4 つ:

1. **表紙 → 改訂履歴 → 本体シート**の 3 点セット（どの成果物にも付く）
2. **格子の表**（見出し行に色、全セルに罫線、区分列は縦結合で「同上」）
3. **方眼紙シート**（列幅を詰めた升目。構成図・フロー図はこの上に描く）
4. **オートシェイプ＋コネクタ**で描いた図（openpyxl は図形を書けないので、
   保存後の xlsx（zip）に drawing 部を直接注入する）

4 は ``arp.ingest.extractors.xlsx`` が読む側の実装と対になっている。
読み側は ``xl/drawings/drawingN.xml`` の ``<xdr:sp>``（ノード）と
``<xdr:cxnSp>``（コネクタ）を見るので、書き側もその形で出す。

**列幅はシート単位の資源**である。1 シートに列数・列幅の違う表を積むと後の表が
先の表の幅を上書きして先の表が崩れるので、``_set_width`` が食い違いを
``LayoutError`` で弾く。同じシートに積む表は列幅を揃えるか、章ごとに
シートを分ける（資材は後者を基本にしてある —— 1 シート = 1 章）。
"""

from __future__ import annotations

import posixpath
import re
import shutil
import sys
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from xml.etree import ElementTree as ET

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "build"))
import reproducible  # noqa: E402
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── 体裁 ──────────────────────────────────────────────────────────
FONT = "ＭＳ Ｐゴシック"

F_TITLE = Font(name=FONT, size=18, bold=True)
F_SUBTITLE = Font(name=FONT, size=12)
F_HEAD = Font(name=FONT, size=9, bold=True)
F_BODY = Font(name=FONT, size=9)
F_NOTE = Font(name=FONT, size=9, color="808080")
F_SECTION = Font(name=FONT, size=11, bold=True)

FILL_HEAD = PatternFill("solid", fgColor="D9E1F2")
FILL_HEAD2 = PatternFill("solid", fgColor="FFF2CC")
FILL_LABEL = PatternFill("solid", fgColor="F2F2F2")

_thin = Side(style="thin", color="808080")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

AL_L = Alignment(horizontal="left", vertical="top", wrap_text=True)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_CV = Alignment(horizontal="left", vertical="center", wrap_text=True)


class LayoutError(Exception):
    """1 つの列に食い違う幅を指定した（＝同じシートのどちらかの表が崩れる）。"""


def _set_width(ws: Worksheet, col: int, width: float) -> None:
    """列幅を決める。同じシートの同じ列に違う幅を指定したら弾く。

    列幅はシート単位の資源なので、1 シートに列数・列幅の違う表を積むと、
    後から書いた表の幅が先に書いた表の幅を上書きし、**先の表が崩れる**。
    Excel で開くまで気づけない類の崩れなので、生成時に落としてしまう。

    直し方は 2 つ。**同じシートに積む表は列幅を揃える**（狭いほうの表は
    広いほうの列幅の一部を使う）か、**章ごとにシートを分ける**。
    """
    letter = get_column_letter(col)
    declared = getattr(ws, "_declared_widths", None)
    if declared is None:
        declared = {}
        ws._declared_widths = declared
    if declared.get(letter, width) != width:
        raise LayoutError(
            f"[{ws.title}] {letter} 列の幅が食い違う（{declared[letter]} → {width}）。"
            "同じシートに積む表は列幅を揃えるか、章ごとにシートを分けること。"
        )
    declared[letter] = width
    ws.column_dimensions[letter].width = width


def new_book() -> Workbook:
    """既定シートを消した空のブック。"""
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def add_sheet(wb: Workbook, title: str, *, grid: bool = False) -> Worksheet:
    """シートを足す。``grid=True`` で方眼紙（列幅 2.5・行高 15）にする。"""
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False
    if grid:
        for c in range(1, 81):
            _set_width(ws, c, 2.5)
        for r in range(1, 121):
            ws.row_dimensions[r].height = 15
    return ws


def set_print(ws: Worksheet, *, landscape: bool = True, doc_name: str = "") -> None:
    """A4・横・幅 1 ページに収める。フッタに文書名とページ番号。"""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.left.text = doc_name
    ws.oddFooter.right.text = "&P / &N"


# ── 表紙・改訂履歴 ────────────────────────────────────────────────
@dataclass
class DocInfo:
    """1 つの成果物の表紙に載る情報。"""

    doc_name: str                 # 文書名（例: 基本設計書（システム方式））
    subsystem: str = "全体"       # サブシステム
    version: str = "1.0"
    date: str = "2025/06/30"
    author: str = "ファブリカム株式会社"   # = VENDOR（下で定義）
    revisions: list[tuple[str, str, str, str]] = field(default_factory=list)
    # (版数, 改訂日, 改訂内容, 改訂者)


SYSTEM_NAME = "新販売管理システム"
PROJECT_NAME = "新販売管理システム構築プロジェクト"
#: 発注者・受託者・銀行はいずれも「例示用に用意された実在しない会社名」を使う
#: （Contoso / Fabrikam / Woodgrove Bank の日本語表記）。実在の企業名を
#: 架空の設計書に書かないための決めごとで、名前を変えるときも同じ性質のものを選ぶ。
CLIENT = "コントソ食品株式会社"
CLIENT_SHORT = "コントソ食品"      # 議事録の出席者欄など、社名を略して書くところ
VENDOR = "ファブリカム株式会社"
BANK = "ウッドグローブ銀行"


def cover(wb: Workbook, info: DocInfo) -> Worksheet:
    """表紙シート。SIer 標準の「文書名・システム名・版数・承認欄」。"""
    ws = add_sheet(wb, "表紙")
    for i, width in enumerate((3, 14, 20, 20, 20, 20, 3), start=1):
        _set_width(ws, i, width)

    ws.merge_cells("B3:F3")
    ws["B3"] = CLIENT + " 御中"
    ws["B3"].font = F_SUBTITLE
    ws["B3"].alignment = AL_C

    ws.merge_cells("B6:F6")
    ws["B6"] = PROJECT_NAME
    ws["B6"].font = Font(name=FONT, size=14, bold=True)
    ws["B6"].alignment = AL_C

    ws.merge_cells("B8:F9")
    ws["B8"] = info.doc_name
    ws["B8"].font = F_TITLE
    ws["B8"].alignment = AL_C
    ws.row_dimensions[8].height = 30
    ws.row_dimensions[9].height = 30

    ws.merge_cells("B11:F11")
    ws["B11"] = f"（{info.subsystem}）"
    ws["B11"].font = F_SUBTITLE
    ws["B11"].alignment = AL_C

    meta = [
        ("版数", f"第 {info.version} 版"),
        ("作成年月日", info.date),
        ("作成者", info.author),
    ]
    row = 14
    for label, value in meta:
        ws.cell(row=row, column=3, value=label)
        ws.cell(row=row, column=3).font = F_BODY
        ws.cell(row=row, column=3).fill = FILL_LABEL
        ws.cell(row=row, column=3).border = BORDER
        ws.cell(row=row, column=3).alignment = AL_CV
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.cell(row=row, column=4, value=value)
        ws.cell(row=row, column=4).font = F_BODY
        ws.cell(row=row, column=4).alignment = AL_CV
        for c in range(4, 7):
            ws.cell(row=row, column=c).border = BORDER
        ws.row_dimensions[row].height = 20
        row += 1

    # 承認欄（押印欄）。日本の設計書に必ず付く。
    top = row + 2
    ws.cell(row=top, column=3, value="承認").font = F_BODY
    for i, label in enumerate(("承認", "確認", "作成")):
        col = 4 + i
        ws.cell(row=top, column=col, value=label)
        ws.cell(row=top, column=col).font = F_HEAD
        ws.cell(row=top, column=col).fill = FILL_HEAD
        ws.cell(row=top, column=col).alignment = AL_C
        ws.cell(row=top, column=col).border = BORDER
        ws.cell(row=top + 1, column=col).border = BORDER
    ws.cell(row=top, column=3).fill = FILL_LABEL
    ws.cell(row=top, column=3).border = BORDER
    ws.cell(row=top, column=3).alignment = AL_C
    ws.cell(row=top + 1, column=3, value="").border = BORDER
    ws.row_dimensions[top + 1].height = 44

    set_print(ws, landscape=False, doc_name=info.doc_name)
    return ws


def revisions(wb: Workbook, info: DocInfo) -> Worksheet:
    """改訂履歴シート。版を重ねた形跡（＝下流との食い違いの遠因）を残す。"""
    ws = add_sheet(wb, "改訂履歴")
    heading(ws, 2, 2, "改訂履歴")
    rows = info.revisions or [("1.0", info.date, "初版作成", "山田")]
    table(
        ws,
        top=4,
        left=2,
        header=["版数", "改訂日", "改訂内容", "改訂者"],
        rows=[list(r) for r in rows],
        widths=[8, 12, 70, 10],
    )
    set_print(ws, doc_name=info.doc_name)
    return ws


# ── 見出し・表 ────────────────────────────────────────────────────
def heading(ws: Worksheet, row: int, col: int, text: str) -> int:
    """章見出し。次に書ける行を返す。"""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = F_SECTION
    cell.alignment = AL_L
    return row + 2


def note(ws: Worksheet, row: int, col: int, text: str) -> int:
    """注記（1 セルのテキスト。抽出側では TextElement になる）。"""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = F_NOTE
    cell.alignment = AL_L
    return row + 2


def table(
    ws: Worksheet,
    *,
    top: int,
    left: int,
    header: list[str],
    rows: list[list[str]],
    widths: list[int] | None = None,
    merge_cols: tuple[int, ...] = (),
    header_fill: PatternFill = FILL_HEAD,
    center_cols: tuple[int, ...] = (),
    groups: list[tuple[str, int]] | None = None,
    caption: str = "",
) -> int:
    """格子の表を書く。次に書ける行（1 行あけた後）を返す。

    ``merge_cols`` に 0 始まりの列番号を渡すと、同じ値が続く区間を**縦結合**する
    （日本の設計書で区分列に使う「同上」の書き方）。抽出側はこの縦結合を各行へ
    展開して読むので、結合の有無で表の意味が変わらないことの検査にもなる。

    ``groups`` を渡すと見出しが 2 段になる。``[("項目", 3), ("", 1)]`` のように
    (見出し, 列数) を並べ、見出しが空の列は下段の見出しと**縦に結合**する。
    横結合・面結合は抽出側では左上セルに値が残るだけなので、2 段見出しは
    「結合を展開しないほうが正しい」側の検体になる。

    ``caption`` を渡すと表の直上に横結合した表題の帯を置く。

    ``widths`` はシート単位で共有される資源なので、同じシートに複数の表を積む
    ときは**列幅を揃える**こと（狭いほうの表は広いほうの列幅の先頭を使う）。
    揃えられない組み合わせは章ごとにシートを分ける。食い違いは ``_set_width``
    が ``LayoutError`` で弾く。
    """
    if widths:
        for i, w in enumerate(widths):
            _set_width(ws, left + i, w)

    if caption:
        cell = ws.cell(row=top, column=left, value=caption)
        cell.font = F_HEAD
        cell.fill = FILL_LABEL
        cell.alignment = AL_CV
        ws.merge_cells(
            start_row=top, start_column=left, end_row=top, end_column=left + len(header) - 1
        )
        for c in range(len(header)):
            ws.cell(row=top, column=left + c).border = BORDER
        top += 1

    head_row = top
    if groups:
        col = left
        for label, span in groups:
            if label:
                cell = ws.cell(row=top, column=col, value=label)
                cell.font = F_HEAD
                cell.fill = header_fill
                cell.alignment = AL_C
                if span > 1:
                    ws.merge_cells(
                        start_row=top, start_column=col, end_row=top, end_column=col + span - 1
                    )
            for i in range(span):
                ws.cell(row=top, column=col + i).border = BORDER
                ws.cell(row=top, column=col + i).fill = header_fill
            col += span
        ws.row_dimensions[top].height = 22
        head_row = top + 1

    for i, name in enumerate(header):
        cell = ws.cell(row=head_row, column=left + i, value=name)
        cell.font = F_HEAD
        cell.fill = header_fill
        cell.alignment = AL_C
        cell.border = BORDER
    ws.row_dimensions[head_row].height = 24

    if groups:
        # 見出しが空の列は上段と下段を縦に結合する（1 列だけの見出し）。
        col = left
        for label, span in groups:
            if not label and span == 1:
                ws.merge_cells(
                    start_row=top, start_column=col, end_row=head_row, end_column=col
                )
                ws.cell(row=top, column=col, value=header[col - left])
                ws.cell(row=top, column=col).font = F_HEAD
                ws.cell(row=top, column=col).fill = header_fill
                ws.cell(row=top, column=col).alignment = AL_C
            col += span

    for r, values in enumerate(rows):
        for c, value in enumerate(values):
            cell = ws.cell(row=head_row + 1 + r, column=left + c, value=value)
            cell.font = F_BODY
            cell.border = BORDER
            cell.alignment = AL_C if c in center_cols else AL_L

    for c in merge_cols:
        start = 0
        for r in range(1, len(rows) + 1):
            same = r < len(rows) and rows[r][c] == rows[start][c]
            if not same:
                if r - start > 1:
                    ws.merge_cells(
                        start_row=head_row + 1 + start,
                        start_column=left + c,
                        end_row=head_row + r,
                        end_column=left + c,
                    )
                    ws.cell(row=head_row + 1 + start, column=left + c).alignment = AL_C
                start = r
    return head_row + len(rows) + 3


def kv_table(
    ws: Worksheet,
    *,
    top: int,
    left: int,
    pairs: list[tuple[str, str]],
    width: int = 70,
    label_width: int = 18,
) -> int:
    """ラベル・値の 2 列表（画面仕様の「概要」欄などに使う縦持ちの表）。"""
    _set_width(ws, left, label_width)
    _set_width(ws, left + 1, width)
    for i, (label, value) in enumerate(pairs):
        lc = ws.cell(row=top + i, column=left, value=label)
        lc.font = F_HEAD
        lc.fill = FILL_LABEL
        lc.border = BORDER
        lc.alignment = AL_CV
        vc = ws.cell(row=top + i, column=left + 1, value=value)
        vc.font = F_BODY
        vc.border = BORDER
        vc.alignment = AL_L
    return top + len(pairs) + 2


def kv_group_table(
    ws: Worksheet,
    *,
    top: int,
    left: int,
    groups: list[tuple[str, list[tuple[str, str]]]],
    width: int = 88,
    group_width: int = 14,
    label_width: int = 20,
) -> int:
    """3 列の縦持ち表（大分類・項目・内容）。大分類は**縦結合**する。

    「画面概要」「処理概要」のような縦長の定義を、区分でくくって書く形。

    同じシートに他の表を積むときは、``group_width`` / ``label_width`` /
    ``width`` をその表の列幅に合わせる（``table`` の注記を参照）。
    """
    _set_width(ws, left, group_width)
    _set_width(ws, left + 1, label_width)
    _set_width(ws, left + 2, width)

    row = top
    for group, pairs in groups:
        start = row
        for label, value in pairs:
            gc = ws.cell(row=row, column=left, value=group if row == start else None)
            gc.font = F_HEAD
            gc.fill = FILL_HEAD
            gc.border = BORDER
            gc.alignment = AL_C
            lc = ws.cell(row=row, column=left + 1, value=label)
            lc.font = F_HEAD
            lc.fill = FILL_LABEL
            lc.border = BORDER
            lc.alignment = AL_CV
            vc = ws.cell(row=row, column=left + 2, value=value)
            vc.font = F_BODY
            vc.border = BORDER
            vc.alignment = AL_L
            row += 1
        if row - start > 1:
            ws.merge_cells(
                start_row=start, start_column=left, end_row=row - 1, end_column=left
            )
    return row + 2


# ── 画面レイアウト図・帳票様式イメージ ────────────────────────────
#: 枠の種類ごとの (塗り, フォント, 寄せ)。
_BOX_STYLE = {
    "window": (None, F_SECTION, AL_L),
    "title": (FILL_HEAD, F_HEAD, AL_C),
    "label": (FILL_LABEL, F_BODY, AL_C),
    "input": (PatternFill("solid", fgColor="FFFFFF"), F_BODY, AL_CV),
    "display": (PatternFill("solid", fgColor="F2F2F2"), F_BODY, AL_CV),
    "button": (FILL_HEAD2, F_HEAD, AL_C),
    "area": (None, F_BODY, AL_L),
    "plain": (None, F_BODY, AL_L),
}


@dataclass
class Box:
    """レイアウト図の 1 つの枠。位置・大きさはセル単位（方眼紙前提）。"""

    row: int
    col: int
    w: int
    h: int = 1
    text: str = ""
    kind: str = "input"


def layout(ws: Worksheet, boxes: list[Box]) -> None:
    """方眼紙のシートに、セル結合と罫線で画面／帳票の様式を描く。

    日本の設計書で画面レイアウト・帳票レイアウトを表すときの定番の書き方で、
    図形ではなく**結合セルの塊**として現れる。抽出側では 1 つの大きな表領域に
    なり、値は各結合範囲の左上セルにだけ入る。
    """
    for box in boxes:
        fill, font, align = _BOX_STYLE.get(box.kind, _BOX_STYLE["input"])
        r0, c0 = box.row, box.col
        r1, c1 = box.row + box.h - 1, box.col + box.w - 1
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                cell = ws.cell(row=r, column=c)
                if box.kind != "plain":
                    cell.border = BORDER
                if fill is not None:
                    cell.fill = fill
        if r1 > r0 or c1 > c0:
            ws.merge_cells(start_row=r0, start_column=c0, end_row=r1, end_column=c1)
        cell = ws.cell(row=r0, column=c0, value=box.text or None)
        cell.font = font
        cell.alignment = align


# ── 図形（オートシェイプ＋コネクタ）────────────────────────────────
@dataclass
class Node:
    """図のノード 1 つ。位置はセル単位（方眼紙シート前提）。"""

    key: str          # コネクタから参照する識別子
    text: str
    col: int          # 0 始まり
    row: int          # 0 始まり
    w: int = 10       # 幅（セル数）
    h: int = 3        # 高さ（セル数）
    shape: str = "roundRect"
    fill: str = "DEEBF7"
    line: str = "2E75B6"


@dataclass
class Edge:
    """ノード間の接続。``label`` はコネクタの図形名に入れる。

    ``dash`` は線種（DrawingML の ``a:prstDash/@val``）。**既定は空で、
    そのときは ``a:prstDash`` を出さない** ―― 出さないのが実線なので、
    既存の呼び出し側の drawing XML は 1 バイトも変わらない。

    体制図で「実線＝指揮命令／破線＝委託」のように**線種で描き分ける**のは
    日本の設計書では普通の書き方で、凡例がそう宣言している以上、実物の
    ``a:prstDash`` も出ていなければ資料と中身が食い違う。

    使える値は DrawingML の既定値（``dash`` ``sysDash`` ``dashDot``
    ``lgDash`` ``dot`` など）。
    """

    src: str
    dst: str
    label: str = ""
    dash: str = ""


@dataclass
class Diagram:
    """1 シートに描く図。"""

    sheet: str
    nodes: list[Node]
    edges: list[Edge] = field(default_factory=list)


# 方眼紙 1 セル = 20px（列幅 2.5・行高 15）を EMU に換算した値。
_EMU_PER_CELL = 20 * 9525

_NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_PR = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_DRAWING = _NS_R + "/drawing"
_CT_DRAWING = "application/vnd.openxmlformats-officedocument.drawing+xml"


def save(wb: Workbook, path: Path, diagrams: list[Diagram] | None = None) -> Path:
    """ブックを保存し、図があれば drawing 部を注入する。

    **書き出したものは git に入る**（`examples/*/資料/`）ので、保存時刻と zip の
    エントリ日時を固定する ―― 揃えないと、中身が 1 文字も変わっていなくても
    生成器を回すたびに全冊が差分に出て、**見本が古いかどうかを誰も判定
    できなくなる**（`build/reproducible.py`）。
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    reproducible.stamp(wb)
    wb.save(str(path))
    reproducible.restamp(path)
    if diagrams:
        _inject_diagrams(path, diagrams)
    return reproducible.freeze(path)


def _inject_diagrams(path: Path, diagrams: list[Diagram]) -> None:
    """保存済み xlsx（zip）へ図形を書き込む。

    openpyxl はオートシェイプ・コネクタを書けないため、
    ``[Content_Types].xml`` / シートの ``.rels`` / ワークシート XML を
    直接書き換えて ``xl/drawings/drawingN.xml`` を足す。
    """
    with zipfile.ZipFile(path) as zf:
        parts = {name: zf.read(name) for name in zf.namelist()}

    sheet_parts = _sheet_parts(parts)
    new_parts: dict[str, bytes] = {}
    used = sum(1 for n in parts if n.startswith("xl/drawings/drawing"))

    for diagram in diagrams:
        sheet_part = sheet_parts.get(diagram.sheet)
        if sheet_part is None:
            raise KeyError(f"シートが見つかりません: {diagram.sheet}")
        used += 1
        drawing_part = f"xl/drawings/drawing{used}.xml"
        new_parts[drawing_part] = _drawing_xml(diagram).encode("utf-8")

        rid = _add_sheet_rel(parts, new_parts, sheet_part, drawing_part)
        parts[sheet_part] = _add_drawing_ref(parts[sheet_part], rid)
        parts["[Content_Types].xml"] = _add_content_type(
            parts["[Content_Types].xml"], drawing_part
        )

    parts.update(new_parts)
    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zf:
        for name, data in parts.items():
            zf.writestr(name, data)
    shutil.move(str(tmp), str(path))


def _sheet_parts(parts: dict[str, bytes]) -> dict[str, str]:
    """{シート名: ワークシート部のパス}。workbook.xml と its rels から引く。"""
    rels = _rels_map(parts["xl/_rels/workbook.xml.rels"])
    wb = ET.fromstring(parts["xl/workbook.xml"])
    out: dict[str, str] = {}
    sheets = wb.find(f"{{{_NS_MAIN}}}sheets")
    for sheet in sheets or []:
        name = sheet.get("name")
        target = rels.get(sheet.get(f"{{{_NS_R}}}id", ""))
        if name and target:
            out[name] = posixpath.normpath(posixpath.join("xl", target)).lstrip("/")
    return out


def _rels_map(data: bytes) -> dict[str, str]:
    root = ET.fromstring(data)
    return {
        rel.get("Id", ""): rel.get("Target", "")
        for rel in root.findall(f"{{{_NS_PR}}}Relationship")
    }


def _add_sheet_rel(
    parts: dict[str, bytes],
    new_parts: dict[str, bytes],
    sheet_part: str,
    drawing_part: str,
) -> str:
    """ワークシートの .rels に drawing 関係を足し、割り当てた rId を返す。"""
    rels_part = posixpath.join(
        posixpath.dirname(sheet_part), "_rels", posixpath.basename(sheet_part) + ".rels"
    )
    existing = parts.get(rels_part) or new_parts.get(rels_part)
    if existing is None:
        body = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<Relationships xmlns="{_NS_PR}"></Relationships>'
        existing = body.encode("utf-8")
    text = existing.decode("utf-8")
    rid = f"rId{len(_rels_map(existing)) + 1}"
    target = "../drawings/" + posixpath.basename(drawing_part)
    rel = f'<Relationship Id="{rid}" Type="{_REL_DRAWING}" Target="{target}"/>'
    text = text.replace("</Relationships>", rel + "</Relationships>")
    new_parts[rels_part] = text.encode("utf-8")
    return rid


def _add_drawing_ref(data: bytes, rid: str) -> bytes:
    """ワークシート XML の末尾に ``<drawing r:id=.../>`` を足す。

    CT_Worksheet の要素順では drawing は末尾側（pageSetup / headerFooter の後）
    なので ``</worksheet>`` の直前に置けばよい。
    """
    text = data.decode("utf-8")
    if f'xmlns:r="{_NS_R}"' not in text:
        text = text.replace("<worksheet ", f'<worksheet xmlns:r="{_NS_R}" ', 1)
    return text.replace(
        "</worksheet>", f'<drawing r:id="{rid}"/></worksheet>'
    ).encode("utf-8")


def _add_content_type(data: bytes, drawing_part: str) -> bytes:
    text = data.decode("utf-8")
    override = f'<Override PartName="/{drawing_part}" ContentType="{_CT_DRAWING}"/>'
    if override in text:
        return data
    return text.replace("</Types>", override + "</Types>").encode("utf-8")


def _crossings(diagram: Diagram) -> list[str]:
    """**関係の無い箱を突っ切る線**を数える。

    通り道は接続点どうしを結んだ矩形にしか無く、この道具は障害物を避けない
    ―― 途中に箱があると線がその上を走る。接続は ``stCxn`` / ``endCxn`` に
    id で書いてあるのでパースは正しく通り、**実物の Excel で開いた人にだけ
    別の繋がりに見える**（銀行 → 入金消込 の線が会計システムの箱を突っ切り、
    「入金消込 → 会計システム → 銀行」の直列に読めた）。潰すのは箱の置き方
    でしかないので、ここでは数えて申告するだけにする。
    """
    by_key = {node.key: node for node in diagram.nodes}
    found = []
    for edge in diagram.edges:
        c0, r0, c1, r1 = _route(by_key[edge.src], by_key[edge.dst])[:4]
        x0, x1 = min(c0, c1), max(c0, c1)
        y0, y1 = min(r0, r1), max(r0, r1)
        for node in diagram.nodes:
            if node.key in (edge.src, edge.dst):
                continue
            if x0 < node.col + node.w and node.col < x1 and \
               y0 < node.row + node.h and node.row < y1:
                head = node.text.splitlines()[0]
                found.append(f"{diagram.sheet}: {edge.src} → {edge.dst} の線が"
                             f"『{head}』を突っ切る")
    return found


def _drawing_xml(diagram: Diagram) -> str:
    """図をまるごと drawing XML にする（ノード → コネクタの順）。"""
    by_key: dict[str, tuple[Node, int]] = {}
    body: list[str] = []
    shape_id = 1

    for warning in _crossings(diagram):
        print(f"  警告 {warning}")

    for node in diagram.nodes:
        shape_id += 1
        by_key[node.key] = (node, shape_id)
        body.append(_shape_xml(node, shape_id))

    for edge in diagram.edges:
        if edge.src not in by_key or edge.dst not in by_key:
            raise KeyError(f"接続先が図に無い: {edge.src} → {edge.dst}")
        shape_id += 1
        body.append(_connector_xml(edge, by_key, shape_id))

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<xdr:wsDr xmlns:xdr="{_NS_XDR}" xmlns:a="{_NS_A}" xmlns:r="{_NS_R}">'
        + "".join(body)
        + "</xdr:wsDr>"
    )


def _anchor(c0: int, r0: int, c1: int, r1: int) -> tuple[str, str]:
    """twoCellAnchor の from / to。"""
    frm = (
        f"<xdr:from><xdr:col>{c0}</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{r0}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
    )
    to = (
        f"<xdr:to><xdr:col>{c1}</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{r1}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
    )
    return frm, to


def _shape_xml(node: Node, shape_id: int) -> str:
    frm, to = _anchor(node.col, node.row, node.col + node.w, node.row + node.h)
    paragraphs = "".join(
        '<a:p><a:pPr algn="ctr"/><a:r>'
        f'<a:rPr lang="ja-JP" sz="900" b="0"><a:solidFill><a:srgbClr val="1F1F1F"/></a:solidFill></a:rPr>'
        f"<a:t>{escape(line)}</a:t></a:r></a:p>"
        for line in node.text.splitlines()
    )
    return (
        "<xdr:twoCellAnchor>"
        + frm
        + to
        + '<xdr:sp macro="" textlink="">'
        + "<xdr:nvSpPr>"
        + f'<xdr:cNvPr id="{shape_id}" name="{escape(node.key)}"/>'
        + "<xdr:cNvSpPr/>"
        + "</xdr:nvSpPr>"
        + "<xdr:spPr>"
        + "<a:xfrm>"
        + f'<a:off x="{node.col * _EMU_PER_CELL}" y="{node.row * _EMU_PER_CELL}"/>'
        + f'<a:ext cx="{node.w * _EMU_PER_CELL}" cy="{node.h * _EMU_PER_CELL}"/>'
        + "</a:xfrm>"
        + f'<a:prstGeom prst="{node.shape}"><a:avLst/></a:prstGeom>'
        + f'<a:solidFill><a:srgbClr val="{node.fill}"/></a:solidFill>'
        + f'<a:ln w="9525"><a:solidFill><a:srgbClr val="{node.line}"/></a:solidFill></a:ln>'
        + "</xdr:spPr>"
        + "<xdr:txBody>"
        + '<a:bodyPr vertOverflow="clip" horzOverflow="clip" wrap="square" anchor="ctr"/>'
        + "<a:lstStyle/>"
        + paragraphs
        + "</xdr:txBody>"
        + "</xdr:sp><xdr:clientData/></xdr:twoCellAnchor>"
    )


#: prstGeom の接続点 index（上 / 左 / 下 / 右）。
_SITE = {"top": 0, "left": 1, "bottom": 2, "right": 3}


def _route(src: Node, dst: Node) -> tuple[int, int, int, int, str, str]:
    """線の両端（``c0,r0`` → ``c1,r1``）と、出入りする辺を決める。

    位置関係で決める（右へ流れるなら右→左、下へ流れるなら下→上）。**端は辺の
    真ん中**であって番地の格子ではない ―― 角から角へ斜めに走る線は、接続が
    正しくても図としては読めない。
    """
    dx = (dst.col + dst.w / 2) - (src.col + src.w / 2)
    dy = (dst.row + dst.h / 2) - (src.row + src.h / 2)
    if abs(dy) >= abs(dx):
        s_site, d_site = ("bottom", "top") if dy > 0 else ("top", "bottom")
        c0 = src.col + src.w // 2
        r0 = src.row + src.h if dy > 0 else src.row
        c1 = dst.col + dst.w // 2
        r1 = dst.row if dy > 0 else dst.row + dst.h
    else:
        s_site, d_site = ("right", "left") if dx > 0 else ("left", "right")
        c0 = src.col + src.w if dx > 0 else src.col
        r0 = src.row + src.h // 2
        c1 = dst.col if dx > 0 else dst.col + dst.w
        r1 = dst.row + dst.h // 2
    return c0, r0, c1, r1, s_site, d_site


def _connector_xml(
    edge: Edge, by_key: dict[str, tuple[Node, int]], shape_id: int
) -> str:
    src, src_id = by_key[edge.src]
    dst, dst_id = by_key[edge.dst]
    c0, r0, c1, r1, s_site, d_site = _route(src, dst)

    # 矩形は左上から右下へしか書けないので、右から左へ・下から上へ引いた線は
    # 反転で表す。**矢羽根（tailEnd）の付く端が入れ替わる**ので、これを落とすと
    # 接続（stCxn / endCxn）は正しいのに、Excel で開いた人にだけ逆向きの矢印が
    # 見える ―― パースは id を読むので通り、テストも全部通る。
    flips = "".join(f' flip{axis}="1"' for axis, on in zip("HV", (c0 > c1, r0 > r1)) if on)
    frm, to = _anchor(min(c0, c1), min(r0, r1), max(c0, c1), max(r0, r1))
    name = edge.label or f"コネクタ {shape_id}"
    return (
        "<xdr:twoCellAnchor>"
        + frm
        + to
        + '<xdr:cxnSp macro="">'
        + "<xdr:nvCxnSpPr>"
        + f'<xdr:cNvPr id="{shape_id}" name="{escape(name)}"/>'
        + "<xdr:cNvCxnSpPr>"
        + f'<a:stCxn id="{src_id}" idx="{_SITE[s_site]}"/>'
        + f'<a:endCxn id="{dst_id}" idx="{_SITE[d_site]}"/>'
        + "</xdr:cNvCxnSpPr>"
        + "</xdr:nvCxnSpPr>"
        + "<xdr:spPr>"
        + f"<a:xfrm{flips}>"
        + f'<a:off x="{min(c0, c1) * _EMU_PER_CELL}" y="{min(r0, r1) * _EMU_PER_CELL}"/>'
        + f'<a:ext cx="{abs(c1 - c0) * _EMU_PER_CELL}" cy="{abs(r1 - r0) * _EMU_PER_CELL}"/>'
        + "</a:xfrm>"
        + '<a:prstGeom prst="bentConnector3"><a:avLst/></a:prstGeom>'
        # a:ln の子は順番が決まっている（塗り → prstDash → 端の形）。
        # dash が空のときは prstDash を出さない＝実線で、既存の図と同じ XML。
        + '<a:ln w="12700"><a:solidFill><a:srgbClr val="404040"/></a:solidFill>'
        + (f'<a:prstDash val="{escape(edge.dash)}"/>' if edge.dash else "")
        + '<a:tailEnd type="triangle"/></a:ln>'
        + "</xdr:spPr>"
        + "</xdr:cxnSp><xdr:clientData/></xdr:twoCellAnchor>"
    )


def build(
    out_dir: Path,
    rel_path: str,
    info: DocInfo,
    body: "callable",
    diagrams: list[Diagram] | None = None,
) -> Path:
    """表紙・改訂履歴を付けたブックを組み、``body(wb, info)`` に本体を書かせる。"""
    wb = new_book()
    cover(wb, info)
    revisions(wb, info)
    body(wb, info)
    path = out_dir / rel_path
    return save(wb, path, diagrams)


_SAFE = re.compile(r"[\\/:*?\"<>|]")


def safe_name(text: str) -> str:
    return _SAFE.sub("_", text)
