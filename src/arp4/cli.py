"""``arp4`` ― LLM が整理し、機械が管理する。

::

    arp4 init            置き場の骨組みを作る
    arp4 model           使ってよい語彙（種別・関係）を出す ―― **整理層が最初に読む**
    arp4 schema          書いてよい形（整理結果の契約）を出す
    arp4 lint <パス>…    整理結果を 1 ファイル単位で検査する（freeze の部分集合）
    arp4 parse <パス>…   ① 既存資産 → パース結果（.arp/rounds/<ラウンド>/parsed/）
    arp4 render <パス>…  機械が読めなかった範囲を絵にする（図形のシート）
    arp4 declare <型>…   同じ構成のシートを一括で対象外宣言する（表紙・改訂履歴）
    arp4 draft           コードのパース結果 → 整理結果の骨格（②の機械分。文章は TODO で空く）
    arp4 freeze          ② 凍結ゲート（未整理 0 / 語彙外 0 / concept 実在 / 関係の組み合わせ）
    arp4 build           ③ 整理結果 → 正本
    arp4 check           正本の機械検証（構造・参照整合性・多重度・出典）
    arp4 publish         設計書を生成する（--audience stakeholder で PM・顧客向け）
    arp4 auto <パス>…    parse → draft → freeze → build → number → check → publish
                         を 1 コマンドで（止まるのは矛盾と「整理層の手番」だけ）

    arp4 lock / conform  標準パック準拠（CI 用）

**整理はコマンドではない。** ``parsed/`` を読んで ``organized/`` を書くのは
エージェントの仕事である。

**arp4 が作るものは全部 ``.arp/`` の中に入る。** 配布先の直下には何も置かない
（``rounds/`` も ``sources/`` も一般名詞で、相手の持ち物と衝突する）。

``--root`` を省くと **cwd から上方探索**するので、``.arp/`` の奥にいても動く。
ただし**エージェントに渡す手順では必ず ``--root`` を明示する** ―― 省くと cwd 次第で
置き場が動き、一時フォルダで作業した結果が誰にも見えない場所に出る。
"""

from __future__ import annotations

import argparse
import os
import subprocess
import sys
from pathlib import Path
from typing import Any

from arp4 import audit as audit_module
from arp4 import build as build_module
from arp4 import concepts as concepts_module
from arp4 import conform as conform_module
from arp4 import audience as audience_module
from arp4 import auto as auto_module
from arp4 import decisions as decisions_module
from arp4 import derived as derived_module
from arp4 import digest as digest_module
from arp4 import draft as draft_module
from arp4 import fix as fix_module
from arp4 import freeze as freeze_module
from arp4 import gate as gate_module
from arp4 import mdio
from arp4 import metamodel as mm
from arp4 import organized as organized_module
from arp4 import pack as pack_module
from arp4 import parse as parse_module
from arp4 import paths as paths_module
from arp4 import publish as publish_module
from arp4 import render as render_module
from arp4 import report as report_module
from arp4 import sequence as sequence_module
from arp4 import shape as shape_module
from arp4 import spec as spec_module
from arp4 import trace as trace_module
from arp4.finding import Finding, counts, order
from arp4.paths import Round
from arp4.validate import validate
from arp4.yamlio import YamlError


#: ``freeze`` が既定で並べる作業キューの件数（残りは ``--list``）。
_QUEUE_HEAD = 20

#: 全件の置き場（``.arp/out/`` の下 ―― 生成物なので ``.gitignore`` 済み）。
#: **上書きしていく。** 実行ごとに増やすと、読み手が最新を選ぶ手間を負う。
_FINDINGS = "findings.json"


def _relative(path: Path, root: Path) -> str:
    """根からの相対。外に出ていたらそのまま出す（表示だけの話で落とさない）。"""
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return str(path)


def _load(args: argparse.Namespace) -> tuple[spec_module.Spec, list[Finding]]:
    return spec_module.load(paths_module.resolve(args.root))


def _exit_code(findings: list[Finding], strict: bool) -> int:
    """**終了コードは書式で変わらない。** 人が読んでも機械が読んでも同じ結論。

    3 値である ―― error あり = 1 / warn のみ（``--strict``）= 2 / clean = 0。
    ``--strict`` で warn も 1 にしていたころは、**CI から「壊れている」と
    「気になる点が残っている」が区別できなかった** ―― error 0 なのに exit 1 に
    なるので、呼ぶ側は結局出力を文字列で読み直すことになる。warn を止めない
    運用（``--strict`` 無し）の意味は変えない。
    """
    tally = counts(findings)
    if tally["error"]:
        return 1
    if strict and tally["warn"]:
        return 2
    return 0


def _machine(args: argparse.Namespace, command: str, findings: list[Finding],
             metrics: dict[str, Any] | None = None) -> bool:
    """機械可読で出したなら ``True``。

    **出したなら人向けの出力は 1 行も混ぜない。** 混ぜると標準出力が JSON として
    読めなくなり、呼ぶ側が「JSON でない行を捨てる」処理を書くことになる ――
    そこまでやるなら書式を約束していないのと同じである。
    """
    fmt = getattr(args, "format", "text")
    if fmt == "text":
        return False
    print(report_module.render(command, findings, fmt, metrics))
    return True


def _report(findings: list[Finding], strict: bool, *, args: Any = None,
            command: str = "", paths: paths_module.Paths | None = None,
            metrics: dict[str, Any] | None = None) -> int:
    """人向けの一覧。**畳むのは表示だけで、件数は畳まない**（→ :mod:`arp4.digest`）。

    ``paths`` を渡すと**全件をファイルへ置く**（→ :func:`arp4.report.write`）。
    標準出力は読み手の文脈に必ず載るが、ファイルは要るときだけ載る。
    """
    for line in digest_module.lines(
            findings,
            show_known=getattr(args, "show_known", False),
            codes=getattr(args, "code", None) or (),
            summary=getattr(args, "summary", False)):
        print(line)

    if paths is not None and findings:
        full = report_module.write(paths.out / _FINDINGS, command or "check",
                                   findings, metrics)
        print(f"\n全件 {len(findings)} 件: "
              f"{_relative(full, paths.root)}（UTF-8 の JSON）" if full
              else "\n全件は --format json で出せます")
    tally = counts(findings)
    print(f"\nerror {tally['error']} / warn {tally['warn']}")
    return _exit_code(findings, strict)


def _round(paths: paths_module.Paths, name: str | None) -> Round | None:
    """対象のラウンド。既定は**いちばん新しいもの**。"""
    if name:
        return paths.round(name)
    return paths.latest_round()


# ── サブコマンド ────────────────────────────────────────────────
def _init(args: argparse.Namespace) -> int:
    paths = paths_module.create(args.root or Path.cwd())
    print(f"{paths.root} を用意しました")
    for path in (paths.rounds_dir, paths.metamodel,
                 paths.concepts, paths.items, paths.relations, paths.out):
        print(f"  {path.relative_to(paths.root)}")
    for finding in _ignored_wholesale(paths):
        print(finding.render())
    # **資料の置き場は作らない。** 既存資産は既にこの中にあるので、その場所を
    # そのまま指す ―― 集める先を作ると、資料が原本と写しの 2 か所に増える。
    print(f"次にやること: 既存資産の場所を指す"
          f"。arp4 parse --root {paths.root} <パス>…")
    return 0


def _ignored_wholesale(paths: paths_module.Paths) -> list[Finding]:
    """``.arp/`` が丸ごと ``.gitignore`` されていないか（``I001``。warn）。

    **危険を文章で警告しておきながら、検知手段が無かった。** 手順書は「``.arp/`` を
    丸ごと ``.gitignore`` に足さない ―― 無視した瞬間に区別が消えるが、**エラーには
    ならない**」と正しく書いている。``arp4 init`` は ``.arp/.gitignore``（``out/``
    だけを無視する）を書くが、**親の ``.gitignore`` は見ていない**ので、上位の
    ``.arp/`` に負けたことは誰も言わない。

    消えるのは「機械が最初に出したもの」と「人が直したもの」の区別である ――
    パース結果は編集してよい設計で、その区別は**初回コミットだけが持っている**
    （別に原本を持っていない）。実測（sales-corpus・201 ファイル）で、
    パース結果も整理結果も ``git status`` に 1 行も出ない状態のまま通しが
    終わっていた。

    **止めはしない。** git を使わない配布先はあるし、無視したいという判断も
    ありうる ―― 言わないことだけが問題である。
    """
    try:
        done = subprocess.run(
            ["git", "check-ignore", "-q", str(paths.arp)],
            cwd=paths.root, capture_output=True, timeout=10)
    except (OSError, subprocess.SubprocessError):
        return []                     # git が無い・リポジトリでない
    if done.returncode != 0:
        return []
    return [Finding(
        "warn", "I001", "",
        f"{paths.arp.name}/ が丸ごと .gitignore されています。パース結果の"
        "「機械が最初に出したもの」と「人が直したもの」の区別は初回コミットだけが"
        "持っています（別に原本はありません）。"
        f"無視するなら {paths.arp.name}/out/ だけにしてください")]


def _target_round(paths: paths_module.Paths,
                  args: argparse.Namespace) -> tuple[Round, str]:
    """``parse`` の宛先。**既定は「いま開いているラウンドの続き」。**

    以前は今日の日付でラウンドを起こしていたので、**同じ資料を別の日に処理し直すと
    別ラウンドになり、整理をやり直す羽目になった**。ラウンドは時刻ではなく
    「どの資料の版を扱っているか」の単位なので、日付では切らない。
    """
    if args.round:
        return paths.round(args.round), "指定されたラウンド"
    if getattr(args, "new_round", False):
        return paths.new_round(), "新しいラウンド（--new-round）"
    open_round = paths.open_round()
    if open_round is not None:
        return open_round, "作業中のラウンドの続き"
    latest = paths.latest_round()
    if latest is not None:
        return paths.new_round(), f"新しいラウンド（{latest.name} は凍結済み）"
    return paths.new_round(), "最初のラウンド"


def _targets(args: argparse.Namespace, paths: paths_module.Paths
             ) -> tuple[list[Path], Path]:
    """資料のパスと、**フォルダ構造を写す基準**（``--base``）。

    既定の基準は**渡されたパスの共通の親**である。``arp4 parse src/ docs/`` なら
    親はどちらもプロジェクト根なので、パース結果は ``src/…`` ``docs/…`` と
    元の木をそのまま写す ―― 基準を渡したパス自身にすると ``src`` が落ちて、
    別のフォルダの同名ファイルが同じ場所へ出る。

    別ドライブに散っていると共通の親が取れない。そのときは根を基準にするので、
    根の外の資料は :func:`arp4.parse.relative_path` がファイル名だけに畳む
    （**平らになるが消えはしない**）。
    """
    sources = [Path(s).resolve() for s in args.source]
    if args.base:
        return sources, Path(args.base).resolve()
    try:
        return sources, Path(os.path.commonpath([p.parent for p in sources]))
    except ValueError:                              # 別ドライブに散っている
        return sources, paths.root


def _no_sources(command: str) -> int:
    """**「資料が 0 冊」を成功にしない**の入口版。次の一手まで言う。

    ここを既定値で埋めない ―― 埋めていたころの既定は ``sources/`` で、
    **打ち間違えても 0 冊で正常終了**していた。
    """
    print("資料のパスを指定してください", file=sys.stderr)
    print("  既存資産は動かしません。いまある場所をそのまま指します:",
          file=sys.stderr)
    print(f"    arp4 {command} --root <プロジェクト> src/ ddl/ docs/",
          file=sys.stderr)
    print("  （sources/ という置き場は作らなくなりました）", file=sys.stderr)
    return 2


def _parse(args: argparse.Namespace) -> int:
    if not args.source:
        return _no_sources("parse")
    paths = paths_module.resolve(args.root)
    round_, why = _target_round(paths, args)
    sources, base = _targets(args, paths)

    print(f"ラウンド {round_.name} ― {why}")
    targets, findings = parse_module.plan(round_, sources, base,
                                          exclude=args.exclude)
    for finding in order(findings):
        print(finding.render())
    if not targets:
        print("読めた資料がありません")
        return 1 if any(f.level == "error" for f in findings) else 0

    chosen, skipped = _confirm(targets, args)
    if args.dry_run:
        print(f"\n--dry-run のため書き込みませんでした（{len(chosen)} ファイル）")
        return 0

    round_.open(reason=why)
    written, failed = parse_module.write(chosen)
    # **撮った版を残す。** 書けたものだけを渡す（上書きを見送ったパース結果は
    # 古い原本の写しのままなので、指紋を更新するとずれが見えなくなる）。
    parse_module.record(round_, chosen, written)
    for finding in order(failed):
        # **書けなかったものを黙らない。** 読めた資料が書けずに消えるのは、
        # 読めなかったものが消えるのと同じである。
        print(finding.render())
    print(f"\n{len(written)} ファイルを {round_.parsed} へ書きました"
          f"（ラウンド {round_.name}）")
    for path in written[:20]:
        print(f"  {path.relative_to(round_.parsed).as_posix()}")
    if len(written) > 20:
        print(f"  …ほか {len(written) - 20} ファイル")
    if skipped:
        # **黙って飛ばさない。** 編集済みを守ったことは必ず見えるようにする。
        print(f"\n編集済みのため上書きしませんでした: {len(skipped)} ファイル")
        for target in skipped[:10]:
            print(f"  {target.path.relative_to(round_.parsed).as_posix()}")
        print("  上書きするなら --yes、資料が更新されたなら新しいラウンドを起こす")
    print("次にやること: エージェントが parsed/ を読んで organized/ を書く → arp4 freeze")
    return 1 if failed else 0


def _render(args: argparse.Namespace) -> int:
    """機械が読めなかった範囲を絵にする。

    **既定は図形のあるシートだけ**である。全シートを撮ると 1 冊で分単位かかるうえ、
    テキストで足りているシートの絵は受け取り側の予算を食うだけで何も足さない
    （「何を撮らなかったか」は必ず出す ―― 黙って絞ると撮ったつもりになる）。
    """
    if not args.source and not args.base:
        # **`--pending` でも要る。** 宣言が指すのはパース結果の相対パスなので、
        # 元の 1 冊へ戻るには parse のときと**同じ基準**が要る。既定で埋めると、
        # 基準がずれたまま「元シートに辿り着けませんでした」だけが並ぶ。
        return _no_sources("render")
    paths = paths_module.resolve(args.root)
    round_ = (paths.round(args.round) if args.round
              else paths.latest_round() or paths.new_round())
    sources, base = _targets(args, paths)
    out_root = Path(args.out).resolve() if args.out else round_.images

    if args.range and not args.sheet:
        # 範囲はシートごとの座標である。**どのシートの B2:L20 かを機械は決めない。**
        print("--range には --sheet が要ります（範囲はシートごとの座標です）",
              file=sys.stderr)
        return 1
    if args.pending and (args.sheet or args.range or args.all):
        print("--pending は宣言から対象を決めます"
              "（--sheet / --range / --all とは併せられません）", file=sys.stderr)
        return 1

    if args.pending:
        plans, skipped, note, unopenable = _pending_plan(round_, base, args)
        if not plans:
            print(f"ラウンド {round_.name} ― {note or '未読取 の宣言がありません'}")
            _report_unopenable(unopenable, base)
            return 1 if unopenable else 0
    else:
        books = sorted(p for p in parse_module.expand(sources)
                       if p.suffix.lower() == ".xlsx")
        if not books:
            print("撮れる資料がありません（.xlsx だけが対象です）")
            return 1
        try:
            plans, skipped, unopenable = _render_plan(books, args)
        except ValueError as exc:                   # --range の書式違い
            print(exc, file=sys.stderr)
            return 1
        note = f"{len(books)} 冊"

    total = sum(len(tiles) for _, _, tiles in plans)
    # **計画と結果を同じ単位で名乗らない。** 下の「N 枚を…へ書きました」と語が
    # 揃っていると、`1 枚` の直後に `0 枚を書きました` が出たときにどちらが実績か
    # 読み取れない ―― Excel が開けずに 1 枚も撮れないことは普通にある（openpyxl が
    # 生成したブックなど）。ここは**予定**だと語で言い切る。
    print(f"ラウンド {round_.name} ― {note} / 撮る予定 {len(plans)} シート"
          f" / {total} 枚")
    if skipped:
        print(f"  撮らないシート {skipped} 枚"
              + ("（元シートに辿り着けませんでした）" if args.pending
                 else "（図形なし。全部撮るなら --all）"))
    _report_unopenable(unopenable, base)

    for book, sheet, tiles in plans:
        where = parse_module.relative_path(book, base).as_posix()
        print(f"  {where} / {sheet}: {len(tiles)} 枚 "
              + " ".join(t.range for t in tiles))
    if args.dry_run:
        print("\n--dry-run のため撮りませんでした")
        return 1 if unopenable else 0

    jobs = []
    for book, sheet, tiles in plans:
        where = parse_module.relative_path(book, base)
        parsed, anchor = _drawing_anchor(round_, where, sheet)
        jobs.append(render_module.Job(
            xlsx=book, sheet=sheet, tiles=tuple(tiles),
            out_dir=out_root / where, title_rows=args.title_rows,
            parsed=parsed, anchor=anchor,
            # 拡大図は範囲を名前に持つ（全体図と別ファイルになり、貼っても消さない）。
            stem=_zoom_stem(sheet, args.range)))
    try:
        results = render_module.render_all(jobs, target_px=args.target_px)
    except render_module.RenderUnavailableError as exc:
        # 環境が無いのは**1 冊の問題ではない**ので、ここは止まってよい。
        print(f"\n撮れません: {exc}", file=sys.stderr)
        return 1

    print()
    written = blanks = 0
    failed: list[render_module.Result] = []
    for result in results:
        written += len(result.shots)
        blanks += len(result.blank)
        if result.error:
            failed.append(result)
        for shot in result.shots:
            print(f"  {shot.path.relative_to(out_root).as_posix()} "
                  f"{shot.size[0]}x{shot.size[1]}")

    attached = [r for r in results if r.attached]
    print(f"\n実際に書いたのは {written} 枚（{out_root}）")
    if total and not written:
        # **「予定 N 枚」と「0 枚」だけを並べて終わらない。** 撮れなかったのが
        # 全部なのか一部なのかは、終了コード（1）からは区別できない ―― この行が
        # 無いと「1 枚と出た直後に 0 枚」に見え、読み違えの余地が残る。
        print("  1 枚も撮れていません。下の理由を見てください"
              "（Excel が開けないブックは 30 冊に 1 冊はあります）")
    if attached:
        # **撮っただけでは使われない。** 整理層が読むのは parsed/ なので、
        # そこから辿れるようにして初めて「読める」ことになる。
        print(f"  パース結果 {len(attached)} ファイルへ絵を貼りました"
              f"（{round_.parsed} の図形アンカー）")
    # **「貼り先が無い」と「貼る内容が変わらなかった」は別**である。後者を警告に
    # すると、2 回目以降の撮り直しが毎回「何か足りない」と言い出す。
    missing = [r for r in results
               if r.shots and (r.job.parsed is None or not r.job.anchor)]
    if missing:
        # **原因を言い分ける。** パース結果が 1 本も無いなら `parse` がまだで、
        # あるのに当たらないならシートが読まれていない（非表示・名前の衝突・
        # 同名のブックが 2 冊）。前者の案内を後者に出すと、**既に回した
        # `arp4 parse` をもう一度回して同じところへ戻る。**
        parsed_any = round_.parsed.is_dir() and any(round_.parsed.rglob(f"*{mdio.EXT}"))
        print(f"  貼り先が無かった {len(missing)} 件"
              + ("（このラウンドにパース結果がありません。"
                 "先に arp4 parse を回してください）" if not parsed_any else
                 f"（{round_.parsed} に相方のシートが見つかりません。"
                 "パース結果のブック名・シート名と突き合わせてください）"))
    if blanks:
        # **黙って飛ばさない。** 白紙は「資料が空」ではなく「見積もりが行き過ぎた」。
        print(f"  白紙だった区画 {blanks} 個（図形の広がりを大きく見積もりました）")
    for result in failed:
        print(f"  撮れなかった: {parse_module.relative_path(result.job.xlsx, base).as_posix()}"
              f" / {result.job.sheet}: {result.error}", file=sys.stderr)
    print("次にやること: エージェントが絵を読んで parsed/ を直す → arp4 freeze")
    return 1 if failed or unopenable else 0


def _open_book(book: Path, unopenable: list[tuple[Path, str]]) -> Any | None:
    """ブックを開く。開けなければ**理由を控えて None**（例外にしない）。

    **1 冊の失敗で全部を落とさない。** 撮影側（:func:`render.render_all`）は
    最初からそうしてあり、`parse` も飛ばして続けるのに、**計画側にだけ守りが
    無かった** ―― zip として壊れた 1 冊（途中で切れた添付・拡張子だけ付け替えた
    もの）を渡すと `BadZipFile` が素通しになり、**残りの 13 冊が 1 枚も撮れない。**
    30 冊に 1 冊はこうなる（→ `tests/dataset/規模.yml`）。
    """
    from openpyxl import load_workbook

    try:
        return load_workbook(book, data_only=True)
    except Exception as exc:                        # noqa: BLE001 壊れた資料で止めない
        unopenable.append((book, f"{type(exc).__name__}: {exc}"))
        return None


def _report_unopenable(unopenable: list[tuple[Path, str]], base: Path) -> None:
    """開けなかった資料を**数えて名指しで**出す。

    「撮らないシート」に混ぜない ―― **図形が無いから撮らない**のと
    **読めなかったから撮れない**のとでは、次のラウンドですることが正反対である。
    """
    if not unopenable:
        return
    print(f"  開けなかった資料 {len(unopenable)} 冊（ほかは撮ります）",
          file=sys.stderr)
    for book, reason in unopenable:
        print(f"    {parse_module.relative_path(book, base).as_posix()}"
              f": {reason}", file=sys.stderr)


def _pending_plan(round_: Round, base: Path, args: argparse.Namespace
                  ) -> tuple[list[tuple[Path, str, list]], int, str,
                             list[tuple[Path, str]]]:
    """``未読取`` と宣言されたシートだけを撮る計画。

    **宣言はアンカーに付くが、撮るのはシート全体**である。「業務フローは図形で
    描かれており本体が出ていない」という宣言は表題のセルに付くのが普通で、
    その番地（`B2`）を撮っても何も写らない ―― 宣言が言っているのは
    **このシートが読めていない**ことである。寄りたければ `--range` を使う。
    """
    result, _ = organized_module.load(round_)
    wanted: dict[str, str] = {}                     # パース結果 → 宣言のアンカー
    for entry in result.out_of_scope:
        if entry.unreadable:
            wanted.setdefault(entry.file, entry.anchor)
    if not wanted:
        return [], 0, "", []

    plans: list[tuple[Path, str, list]] = []
    unopenable: list[tuple[Path, str]] = []
    skipped = 0
    for relative, anchor in sorted(wanted.items()):
        path = round_.parsed / f"{relative}{mdio.EXT}"
        origin = (parse_module.sheet_origin(mdio.read(path).source)
                  if path.is_file() else None)
        book = base / origin[0] if origin else None
        if book is None or not book.is_file():
            # **黙って飛ばさない。** 元シートに戻れないことは分かるようにする。
            skipped += 1
            continue
        workbook = _open_book(book, unopenable)
        if workbook is None:                        # 一巡目に読めた資料が壊れた
            continue
        try:
            if origin[1] not in workbook.sheetnames:
                skipped += 1
                continue
            boxes = render_module.drawing_boxes(book).get(origin[1], [])
            tiles = render_module.plan_worksheet(
                workbook[origin[1]], render_module.extent_of(boxes),
                boxes=boxes, max_px=args.target_px, wide_px=args.wide_px)
        finally:
            workbook.close()
        if tiles:
            plans.append((book, origin[1], tiles))
        else:
            skipped += 1
    return plans, skipped, f"未読取 の宣言 {len(wanted)} 件", unopenable


def _zoom_stem(sheet: str, cell_range: str | None) -> str:
    """拡大図のファイル名。**範囲を名前に持たせる**（全体図を潰さない）。"""
    if not cell_range:
        return ""
    return parse_module.safe_name(f"{sheet}-{cell_range.replace(':', '_')}")


def _drawing_anchor(round_: Round, where: Path, sheet: str
                    ) -> tuple[Path | None, str]:
    """絵を貼る先（パース結果と図形のアンカー）。**まだ parse していなければ無し。**

    アンカー名を組み立てず**パース結果から探す**のは、``s{n}-g1`` の ``n`` が
    ブック内のシート番号（非表示シートを飛ばした後の番号）だからである。
    ここで数え直すと、パース側の数え方を変えた瞬間に静かにずれる。

    **``parse`` と違うパスの渡し方をしても貼れるようにする。** 置き場は「渡した
    パスの共通の親」から決まるので、``parse`` に ``資料/`` を渡して ``render`` に
    1 冊だけ渡すと ``where`` が食い違い、絵は撮れているのに貼り先が 1 つも
    見つからない ―― しかも終了コードは 0 で、出るのは「先に arp4 parse を
    回すと」という**原因と違う案内**だった（実測で、`parse` は回してあった）。
    貼られていないことに気づかない読み手は、絵があるのに `未読取` を宣言する。

    直せる範囲で直す ―― **ブックのフォルダ名とシート名で 1 冊だけに定まるなら
    それを使う。** 同名のブックが 2 冊あるときは決められないので、そのときだけ
    「貼り先が無い」として扱う（当てずっぽうで別の資料へ貼らない）。
    """
    name = f"{parse_module.safe_name(sheet)}{mdio.EXT}"
    path = round_.parsed / where / name
    if not path.is_file():
        book = where.name
        found = [p for p in round_.parsed.rglob(name) if p.parent.name == book]
        if len(found) != 1:
            return None, ""
        path = found[0]
    for anchor in mdio.read(path).anchors:
        if anchor.id.endswith("-g1"):
            return path, anchor.id
    return path, ""


def _render_plan(books: list[Path], args: argparse.Namespace
                 ) -> tuple[list[tuple[Path, str, list]], int,
                            list[tuple[Path, str]]]:
    """撮る対象を決める。**図形の有無は機械が数えられる**ので自己申告に頼らない。"""
    plans: list[tuple[Path, str, list]] = []
    unopenable: list[tuple[Path, str]] = []
    skipped = 0
    for book in books:
        # 図形のあるシートは**描画のアンカー**に出る（別に数え直さない）。
        # 位置は広がり（撮る範囲）と切り口（割ってよい場所）の両方に要る。
        drawings = render_module.drawing_boxes(book)
        workbook = _open_book(book, unopenable)
        if workbook is None:                        # **残りの資料は撮る**
            continue
        try:
            for sheet in workbook.worksheets:
                if sheet.sheet_state != "visible":
                    continue
                if args.sheet and sheet.title not in args.sheet:
                    continue
                boxes = drawings.get(sheet.title, [])
                if not args.all and not args.sheet and not boxes:
                    skipped += 1
                    continue
                if args.range:
                    # **人が範囲を決めたら、機械は広げない**（切り口だけは避ける）。
                    tiles = render_module.plan_range(
                        sheet, args.range, boxes=boxes,
                        max_px=args.target_px, wide_px=args.wide_px)
                else:
                    tiles = render_module.plan_worksheet(
                        sheet, render_module.extent_of(boxes), boxes=boxes,
                        max_px=args.target_px, wide_px=args.wide_px)
                if tiles:
                    plans.append((book, sheet.title, tiles))
        finally:
            workbook.close()
    return plans, skipped, unopenable


def _confirm(targets: list[parse_module.Target],
             args: argparse.Namespace) -> tuple[list[parse_module.Target],
                                                list[parse_module.Target]]:
    """上書きの確認。**未編集のものは黙って上書きしてよい。**

    全部を同じ重さで聞くと、確認そのものが読み飛ばされる。対話できない場面
    （エージェント・CI）では**編集済みを守って報告する**ほうへ倒す。
    """
    chosen: list[parse_module.Target] = []
    skipped: list[parse_module.Target] = []
    interactive = sys.stdin.isatty() and not args.yes
    always = args.yes

    for target in targets:
        if not target.needs_confirm:
            chosen.append(target)
            continue
        if always:
            chosen.append(target)
            continue
        if not interactive:
            skipped.append(target)
            continue
        answer = input(f"{target.path} は編集されています。上書きしますか？"
                       " [y/N/a(すべて)] ").strip().lower()
        if answer == "a":
            always = True
            chosen.append(target)
        elif answer == "y":
            chosen.append(target)
        else:
            skipped.append(target)
    return chosen, skipped


def _declare(args: argparse.Namespace) -> int:
    """同じ構成のシートを一括で対象外にする。**理由は人が与える。**"""
    paths = paths_module.resolve(args.root)
    round_ = _round(paths, args.round)
    if round_ is None or not round_.parsed.is_dir():
        print("ラウンドがありません（arp4 parse を先に実行してください）",
              file=sys.stderr)
        return 1
    if round_.is_frozen():
        print(f"ラウンド {round_.name} は凍結済みです（整理結果は編集しません）",
              file=sys.stderr)
        return 1

    plans, findings = organized_module.plan_declare(
        round_, args.pattern, args.reason, args.kind)
    for finding in order(findings):
        print(finding.render())
    if not plans:
        print("宣言する未整理のアンカーはありません"
              "（パターンが当たっていないか、既に整理済みです）")
        return 0

    total = sum(len(p.anchors) for p in plans)
    print(f"\n{len(plans)} ファイル / アンカー {total} 件を "
          f"{args.kind} として宣言します（理由: {args.reason}）")
    # **一括で仕様の外へ出す操作なのに、全部を確かめる手が無かった。** 既定の
    # 先頭 20 件は打ち間違いの確認には足りる（実測 54 ファイル / 189 アンカーの
    # うち 34 ファイルが「…ほか」に畳まれた）が、**パターンが余計なシートに
    # 当たっていても畳まれた側は見えない** ―― 落とした側は誰にも見えないので、
    # `freeze` の未整理からも消える。
    shown = plans if args.list else plans[:20]
    for plan in shown:
        mark = "追記" if plan.existed else "新規"
        print(f"  [{mark}] {plan.file}  {len(plan.anchors)} 件")
    if len(shown) < len(plans):
        print(f"  …ほか {len(plans) - len(shown)} ファイル（全部出すには --list）")

    if args.dry_run:
        print("\n--dry-run のため書き込みませんでした")
        return 0
    written = organized_module.write_declarations(plans)
    print(f"\n{len(written)} ファイルを書きました")
    if any(p.existed for p in plans):
        print("注意: 追記したファイルのコメントは失われます（差分で確認してください）")
    print("次にやること: arp4 freeze --dry-run で残作業を見る")
    return 0


def _schema(args: argparse.Namespace) -> int:
    """**書いてよい形**を出す。``arp4 model`（語彙）と対になる。

    語彙は「使ってよい種別・関係・属性」、形は「どの欄が要るか・どこに何を書くか」
    である。**持ち主が違う** ―― 語彙はプロジェクトが決め、形は arp4 が決める。

    畳んだ結果ではなく原文を出す。**なぜそう決めたかはコメントにしか無い**ので、
    値だけ出すと「何を書けばよいか」は分かっても「なぜ 3 つ揃えるのか」が落ちる。
    """
    print(shape_module.text(args.name))
    return 0


def _lint_targets(round_: Round,
                  given: list[str]) -> tuple[list[Path] | None, list[str]]:
    """引数のパスを整理結果の実ファイルに解決する。

    **当たらなかったものは必ず言う。** 黙って落とすと「lint が通った」と
    「そもそも検査していない」が同じ顔になる ―― 打ち間違い 1 つで、直したはずの
    ファイルが検査されないまま緑になる。

    **ディレクトリも受ける。** 手順書は「1 ファイル書くたびに打て」と言うが、
    分担で配る単位は 1 ブック（＝1 フォルダ）である ―― 実測（11 ロット）で、
    ディレクトリを渡して ``exit 2`` になったと**8 ロット中 6 つ**が報告した。
    ファイルを 1 本ずつ並べる回避策は、日本の設計書のシート名に**空白が普通に
    入る**（`受注入力 項目.yml`）ためシェルで壊れやすく、**壊れても「lint は
    通った」と同じ顔で終わる**（当たらなかったものを言う上の規律が、シェルの
    分割には効かない）。
    """
    if not given:
        return None, []
    targets: list[Path] = []
    missing: list[str] = []
    organized_dir = round_.organized.resolve()
    for one in given:
        candidates = [Path(one), round_.organized / one]
        found = next((c for c in candidates
                      if c.is_file() or c.is_dir()), None)
        if found is None:
            missing.append(one)
            continue
        found = found.resolve()
        if found.is_dir():
            # **中に整理結果が 1 本も無いディレクトリは「当たらなかった」側**に
            # する。0 件を黙って通すと、フォルダ名を打ち間違えたときに
            # 「レコード 0・error 0」で緑になる。
            if found != organized_dir and organized_dir not in found.parents:
                missing.append(one)
                continue
            # **`load` が読む対象と揃える**（凍結マニフェストだけ外す）。
            # **予約名も入れる** ―― `yaml_files` で展開していたので
            # `arp4 lint <organized>` は `_concepts.yml` を一度も読まず、`G002` も
            # `G021` も黙って出なかった。書き換えてよい対象（`--fix`）とは別物で、
            # そちらは `_apply_fixes` が改めて `yaml_files` で絞る。
            inside = [p.resolve()
                      for p in organized_module.lintable(round_)
                      if found == p.resolve() or found in p.resolve().parents]
            if not inside:
                missing.append(one)
                continue
            targets += inside
            continue
        if organized_dir not in found.parents:
            missing.append(one)
            continue
        targets.append(found)
    # 同じファイルを 2 度渡されても 1 度だけ検査する（ディレクトリとその中の
    # ファイルを両方渡す打ち方は普通にある）。
    return list(dict.fromkeys(targets)), missing


def _apply_fixes(round_: Round, model: mm.Metamodel,
                 targets: list[Path] | None
                 ) -> tuple[list[fix_module.Fix], list[Finding]]:
    """対象のファイルを直す。**検算を通ったものだけが書かれる。**

    **予約名は直さない。** :func:`arp4.fix.repair` が知っているのは ``records:``
    の形だけで、``_concepts.yml`` / ``_metamodel-add.yml`` は別の形である ――
    検査の対象（:func:`arp4.organized.lintable`）には入るが、書き換えの対象では
    ない。
    """
    paths = [p for p in (targets if targets is not None
                         else organized_module.yaml_files(round_))
             if p.stem not in organized_module.SPECIAL]
    fixed: list[fix_module.Fix] = []
    refused: list[Finding] = []
    for path in paths:
        location = path.relative_to(round_.root).as_posix()
        applied, trouble = fix_module.repair(path, model, location)
        fixed += applied
        refused += trouble
    return fixed, refused


def _lint(args: argparse.Namespace) -> int:
    """**書いている最中に回す検査。** 1 ファイルだけで決まるものだけを見る。"""
    paths = paths_module.resolve(args.root)
    round_ = _round(paths, args.round)
    if round_ is None or not round_.organized.is_dir():
        print("整理結果がありません（parsed/ を読んで organized/ を書いてください）",
              file=sys.stderr)
        return 1

    targets, missing = _lint_targets(round_, args.path)
    if missing:
        for one in missing:
            print(f"整理結果として見つかりません: {one}"
                  f"（{round_.organized} の下のファイルかフォルダを指してください）",
                  file=sys.stderr)
        return 2

    model, model_findings = mm.load(paths.metamodel)
    known, concept_findings = concepts_module.load(paths)

    fixed: list[fix_module.Fix] = []
    if args.fix:
        # **凍結後の整理結果は編集しない。** 直すと `G009` でハッシュが合わなく
        # なり、「正本側で直す」という決まりを機械が破ることになる。
        if round_.is_frozen():
            print(f"ラウンド {round_.name} は凍結済みです"
                  "（直すなら正本側か、新しいラウンドを起こしてください）",
                  file=sys.stderr)
            return 2
        fixed, refused = _apply_fixes(round_, model, targets)
        model_findings += refused

    report = freeze_module.lint(round_, model, known, only=targets)
    findings = model_findings + concept_findings + report.findings

    metrics = dict(report.metrics)
    if fixed:
        metrics["fixed"] = [f.as_dict() for f in fixed]
    if _machine(args, "lint", findings, metrics):
        return _exit_code(findings, args.strict)

    for one in fixed:
        print(one.render())
    if fixed:
        print(f"{len(fixed)} 件を直しました（残りは下に出ます）\n")

    # **読んだ予約名を必ず言う。** `_concepts.yml` はレコードではないので
    # `files`・`records`・`out_of_scope` のどれにも入らない ―― それだけを渡すと
    # 「0 ファイル / 0 レコード / 0 対象外 / error 0 / warn 0」になり、検査した
    # 結果が白なのか、そもそも読まれなかったのかが**打った人から区別できない。**
    print(f"ラウンド {round_.name} ― 整理結果 {metrics['files']} ファイル / "
          f"レコード {metrics['records']}"
          + (f"（うち参照だけ {metrics['references']}）"
             if metrics.get("references") else "")
          + f" / 対象外 {metrics['out_of_scope']}"
          # **書いた宣言が読まれたことを数で言う。** `G020` は lint では出ない
          # （関係は別のファイルにありうる）ので、宣言が効いたかを確かめる手が
          # 件数しか無い ―― 0 なら書いた場所か綴りが違う（名前の誤りは G031）。
          + (f" / known_gaps の宣言 {metrics['known_gaps']}"
             if metrics.get("known_gaps") else "")
          + (f" / {'・'.join(metrics['proposals'])} を検査しました"
             if metrics.get("proposals") else ""))
    code = _report(findings, args.strict, args=args, command="lint",
                   paths=paths, metrics=metrics)
    # **lint が通ったことは凍結できることを意味しない。** 未整理（G001）も
    # concept の実在（G003）も、1 ファイルだけでは決まらないので見ていない。
    print("横断の検査（未整理・concept 実在・参照だけのレコード）は "
          "arp4 freeze --dry-run で見てください")
    return code


def _draft(args: argparse.Namespace) -> int:
    """コード由来のパース結果から整理結果の骨格を機械生成する（②の機械分）。

    **意味の判断はしない** ―― 規則（organize.md のコード整理の規約）を実行する
    だけで、statement / description は ``<TODO 抽出元 …>`` のまま空ける。埋める
    のは整理層（LLM か人）で、残りは ``arp4 freeze --dry-run`` が G026 で数える。
    """
    paths = paths_module.resolve(args.root)
    round_ = _round(paths, args.round)
    if round_ is None or not round_.parsed.is_dir():
        print("ラウンドがありません（arp4 parse を先に実行してください）",
              file=sys.stderr)
        return 1
    if round_.is_frozen():
        print(f"ラウンド {round_.name} は凍結済みです（整理結果は編集しません）",
              file=sys.stderr)
        return 1

    result = draft_module.plan(round_)
    print(f"ラウンド {round_.name} ― コード {len(result.drafted)} ファイル / "
          f"シート {result.sheets}（シートは従来どおり整理層が書く）")
    if result.skipped:
        # **書いたものは上書きしない。** 文章を埋めたあとの再実行で潰さないため。
        print(f"  整理結果が既にあるため飛ばした {len(result.skipped)} ファイル"
              "（作り直すなら該当の organized/.yml を消してから）")
    for drafted in result.drafted[:20]:
        print(f"  {drafted.file}: レコード {drafted.records} / "
              f"文章化スロット {drafted.todo}")
    if len(result.drafted) > 20:
        print(f"  …ほか {len(result.drafted) - 20} ファイル")

    if args.dry_run:
        print(f"\n--dry-run のため書き込みませんでした"
              f"（文章化スロット {result.todo} 件）")
        return 0
    written = draft_module.write(round_, result)
    print(f"\n{len(written)} ファイルを {round_.organized} へ書きました")
    print(f"  機械が下した判断 {len(result.decisions)} 件 → "
          f"{decisions_module.path_of(round_).relative_to(paths.root)}")
    if result.todo:
        print(f"\n文章化スロット {result.todo} 件（これが整理層の作業キュー）")
        print("  organized/.yml の <TODO 抽出元 …> を、抽出元を読んで埋める")
        print("  残りは arp4 freeze --dry-run が G026 で数える")
    print("次にやること: 文章化 → arp4 lint → arp4 freeze")
    return 0


def _keys_of(prefixes: list[str] | None) -> list[str]:
    """``--path`` の値を照合できる形に揃える（区切りと前後の ``/`` だけ）。"""
    return [p.replace("\\", "/").strip("/") for p in (prefixes or []) if p]


def _only_under(findings: list[Finding],
                prefixes: list[str] | None) -> tuple[list[Finding], list[Finding]]:
    """指摘を担当ぶんに絞る。**隠したほうも返す**（黙って消さないため）。

    絞り込みは表示だけの話で、**判定は絞らない** ―― ゲートの条件はファイルを
    またぐので、自分のロットだけを見て「凍結できる」と言えることは無い。

    **件数ではなく指摘そのものを返す。** 「隠した N 件」だけを持ち帰っていた
    ころは、隠したぶんの error と warn が分けられず、末尾が必ず「凍結できません
    （上の error を潰してください）」で終わった ―― **その error は画面に 1 件も
    出ていない**ので、担当は自分の失敗と読み違える（実測で 3 人が独立に報告）。
    """
    if not prefixes:
        return findings, []
    keys = _keys_of(prefixes)
    kept: list[Finding] = []
    hidden: list[Finding] = []
    for finding in findings:
        where = (finding.file or "").replace("\\", "/")
        (kept if any(key in where for key in keys) else hidden).append(finding)
    return kept, hidden


def _unmatched(round_: Round, prefixes: list[str] | None) -> list[str]:
    """どのファイルにも当たらなかった ``--path``。**0 件と区別する。**

    当たらないパスを黙って通すと、打ち間違い 1 つが「担当ぶんは全部きれい」に
    見える ―― 絞り込みは指摘を消す仕組みなので、**消えたのか無かったのかが
    打った人から区別できない。** ``lint`` は当たらないパスで ``exit 2`` になる
    （→ :func:`_lint_targets`）ので、それに揃える。

    照合の相手は**指摘が名指ししうるファイル**の一覧である（整理結果と
    パース結果）。指摘の側で照合すると、error も warn も無いきれいな担当ぶんが
    「当たらなかった」になる。
    """
    keys = _keys_of(prefixes)
    if not keys:
        return []
    files = [p.relative_to(round_.root).as_posix()
             for p in organized_module.lintable(round_) + mdio.scan(round_.parsed)]
    return [key for key in keys if not any(key in one for one in files)]


def _tally(findings: list[Finding]) -> str:
    """``error N / warn M``。**0 も書く**（無いことを言うのが仕事である）。"""
    counted = counts(findings)
    return f"error {counted['error']} 件 / warn {counted['warn']} 件"


def _freeze_blocked(args: argparse.Namespace, shown: list[Finding],
                    hidden: list[Finding]) -> str:
    """凍結できなかったときの結び 1 行。**「上の error」が上に無い場合を分ける。**

    ``--path`` で絞ると、自分の担当が error 0 でも末尾は必ず「凍結できません
    （上の error を潰してください）」で終わっていた ―― その error は隠したぶんで
    **画面には 1 件も出ていない**ので、「上の」を探しても無い。実測で 3 人が
    独立に「自分の担当が落ちていると読んだ」と報告した。
    """
    mine = counts(shown)["error"]
    theirs = counts(hidden)["error"]
    if not getattr(args, "path", None) or mine:
        return ("凍結できません（上の error を潰してください）"
                + (f"。ほかに --path で隠した error が {theirs} 件あります"
                   if theirs else ""))
    return (f"あなたの担当（--path {'、'.join(args.path)}）は凍結の条件を"
            f"満たしています。ラウンド全体では他の担当ぶんに error が "
            f"{theirs} 件残っているので、まだ凍結できません"
            "（その error は --path で隠しています。全部見るには --path を"
            "外してください）")


def _freeze(args: argparse.Namespace) -> int:
    spec, findings = _load(args)
    paths = spec.paths
    round_ = _round(paths, args.round)
    if round_ is None or not round_.dir.is_dir():
        print("ラウンドがありません（arp4 parse を先に実行してください）",
              file=sys.stderr)
        return 1

    # **当たらない --path は 0 件と区別する。** 打ち間違いを黙って通すと、
    # 「担当ぶんは全部きれい」に見える（`lint` の規律に揃える）。ゲートを
    # 回す前に止めるのは、200 ファイル読んでから打ち直させないためである。
    astray = _unmatched(round_, getattr(args, "path", None))
    if astray:
        for one in astray:
            print(f"--path がどのファイルにも当たりません: {one}"
                  f"（{round_.organized} / {round_.parsed} の下の"
                  "部分パスを指してください）", file=sys.stderr)
        return 2

    known, concept_findings = concepts_module.load(paths)
    report = freeze_module.gate(round_, spec.metamodel, known)
    findings = findings + concept_findings + report.findings

    # **G001 で埋めない。** 整理の途中は未整理が数百件あるのが正常なので、素直に
    # 全部並べると G002 / G008 / G012 が最後の 1 画面より上へ流れて完全に埋もれる。
    # 集計を先に、そのほかの指摘を次に、作業キュー（G001 と、draft の文章化残り
    # G026 ―― どちらも「これから書くもの」の一覧である）は最後に置く。
    # **分担のときは自分のロットだけを見たい。** ゲートはラウンド全体を見て
    # 判定する（そこは変えない ―― 食い違いはファイルをまたぐ）が、**出す指摘は
    # 絞れる。** 実測（11 ロットの分担）で、1 人あたり数百行の他人の指摘を
    # `grep` で除けてから自分の 1 行を探していた。しかもパース結果のファイル名に
    # 空白が入るので、素直に grep すると件数が化ける。
    shown_all, hidden = _only_under(findings, getattr(args, "path", None))
    queue = [f for f in shown_all if f.code in ("G001", "G026")]
    others = [f for f in shown_all if f.code not in ("G001", "G026")]

    metrics = report.metrics
    if getattr(args, "path", None):
        # **絞ったぶんの内訳は機械にも渡す。** 人向けの行だけに書くと、CI が
        # 「自分の担当は通ったが他人のぶんで止まっている」を判定できない。
        metrics["path"] = list(args.path)
        metrics["path_findings"] = counts(shown_all)
        metrics["hidden_findings"] = counts(hidden)
    # **書式は出し方だけを変える。** 凍結するかどうかは書式で変わらない ――
    # 変わると「JSON で見たときだけ凍っていない」という差が生まれ、CI と手元で
    # 結論がずれる。人向けの行を 1 つも混ぜないために、以後は `quiet` で括る。
    quiet = _machine(args, "freeze", findings, metrics)
    if not quiet:
        print(f"ラウンド {round_.name} ― パース {metrics['parsed_files']} ファイル / "
              f"アンカー {metrics['anchors']}")
        print(f"  レコード {metrics['records']}"
              + (f"（うち参照だけ {metrics['references']}）"
                 if metrics.get("references") else "")
              + f" / 対象外 {metrics['out_of_scope']} / 未整理 {metrics['unclaimed']}")
        if metrics.get("unreadable"):
            # 「資料に無い」と「機械が読めていない」は別物。後者は拾い直す対象。
            print(f"  うち未読取 {metrics['unreadable']} 件"
                  "（機械が読めていないだけ。次のラウンドで拾い直す対象）")
        if metrics.get("known_gaps"):
            # **宣言で降ろしたぶんを黙らない。** `known_gaps` は指摘を消す仕組み
            # なので、件数を出さないと「直したから減った」と「宣言で降ろした」が
            # 画面から区別できない ―― 黙って消えるのがいちばん悪い形である。
            print(f"  known_gaps の宣言 {metrics['known_gaps']} 件"
                  f"（うち G020 を {metrics['known_gaps_silenced']} 件"
                  "「調べたうえで相手がいない」として出していません。"
                  "build が正本へ引き継ぎ、check では W032 として理由つきで"
                  "出続けます）")
        if others:
            tally: dict[str, int] = {}
            for finding in others:
                tally[finding.code] = tally.get(finding.code, 0) + 1
            print("  そのほかの指摘: "
                  + " / ".join(f"{code} {n} 件" for code, n in sorted(tally.items())))
            print()
            for finding in order(others):
                print(finding.render())

        if queue:
            shown = queue if args.list else order(queue)[:_QUEUE_HEAD]
            print(f"\n残作業 {len(queue)} 件"
                  "（未整理のアンカーと文章化の残り。これが作業キュー）")
            for finding in order(shown):
                # **どのファイルの話かを落とさない。** 作業キューは 300 行並ぶ
                # ことがあり、アンカー（`s1-t1`）だけでは資料をまたいで同じ
                # 番地が並ぶ。
                print(f"  {finding.where}: {finding.message}")
            if len(shown) < len(queue):
                print(f"  …ほか {len(queue) - len(shown)} 件（全部出すには --list）")

        if getattr(args, "path", None):
            # **絞ったことを黙らない。** しかも「隠した N 件」だけでは足りない
            # ―― 隠したぶんに error があっても凍結は止まる（判定は絞っていない）
            # ので、**この path のぶんと隠したぶんを分けて**出さないと、担当は
            # 「error 0 なのに凍結できません」を自分の失敗と読む。
            print(f"\n--path {'、'.join(args.path)} のぶん: {_tally(shown_all)}")
            print(f"--path で隠したぶん（ほかの担当）: {_tally(hidden)}"
                  "（判定はラウンド全体で行っています。外すと全部出ます）")

    if any(f.level == "error" for f in findings):
        if not quiet:
            # **stdout を先に流し切る。** 混ぜると、この結び 1 行が error の
            # 一覧の途中に割り込んで出る（実測で報告された）。
            sys.stdout.flush()
            print(f"\n{_freeze_blocked(args, shown_all, hidden)}", file=sys.stderr)
        return 1
    if args.dry_run:
        if not quiet:
            print("\n--dry-run のため凍結しませんでした")
        return 0

    manifest = freeze_module.apply(round_, report)
    if not quiet:
        print(f"\n凍結しました: {round_.frozen}"
              f"（{len(manifest['files'])} ファイル / {manifest['frozen_at']}）")
        print("次にやること: arp4 build で正本を組み立てる")
    return 0


def _build(args: argparse.Namespace) -> int:
    spec, findings = _load(args)
    paths = spec.paths
    round_ = _round(paths, args.round)
    if round_ is None or not round_.organized.is_dir():
        print("整理結果がありません（arp4 parse → 整理 → arp4 freeze）",
              file=sys.stderr)
        return 1
    if not round_.is_frozen() and not args.force:
        print(f"ラウンド {round_.name} は凍結されていません"
              "（arp4 freeze を先に。承知のうえなら --force）", file=sys.stderr)
        return 1

    result, load_findings = organized_module.load(round_)
    known, concept_findings = concepts_module.load(paths)
    added, proposal_findings = concepts_module.apply_proposal(
        known, result.concepts, round_.name)

    plan = build_module.plan(spec, result, known, round_.name)
    issues = build_module.issues(spec, result, known, plan)
    findings = (findings + load_findings + concept_findings
                + proposal_findings + plan.findings + issues.findings)
    for finding in order(findings):
        print(finding.render())

    metrics = plan.metrics
    print(f"\nラウンド {round_.name} ― レコード {len(result.records)} 件 / "
          f"concept {metrics['concepts']}（新規 {len(added)}）")
    print(f"  アイテム 新規 {metrics['created']} / 更新 {metrics['updated']} / "
          f"関係 {metrics['relations']}（更新 {metrics['relation_updates']}）")
    if metrics["protected"]:
        # 黙って守ると「なぜ資料の記述が反映されないのか」が見えない。
        print(f"  overridden で守った属性 {metrics['protected']} 件")
    if issues.items:
        print(f"  矛盾から起こした課題 {len(issues)} 件 / 争点への disputes "
              f"{len(issues.relations)} 本（どちらが正しいかは決めていません）")

    if plan.empty and not issues.items and not added:
        print("組み立てるものはありません（正本は最新です）")
        return 0
    if args.dry_run:
        print("\n--dry-run のため書き込みませんでした")
        return 0

    items, relations = build_module.apply(spec, plan)
    applied_items, applied_relations = build_module.apply_issues(spec, issues)
    written = spec_module.save_in_place(spec, items | applied_items,
                                        relations | applied_relations)
    concepts_module.save(paths, known)
    # **build も判断を残す。** 採らなかった値（B024）・決められなかった向き（B026）・
    # 矛盾からの自動起票は、これまで端末に流れて消えるだけだった ―― Excel だけの
    # 資産は draft を通らないので decisions.yml がそもそも作られず、
    # `out/決定記録.md` が生成されなかった（実測・sales-corpus r001）。
    # **`replace` にする**（`append` ではない）―― build は同じ整理結果から同じ判断を
    # 出すので、打ち直すたびに積むと件数が判断の数を言わなくなる。
    logged = plan.logged + issues.logged
    if logged or decisions_module.path_of(round_).is_file():
        decisions_module.replace(round_, "build", logged)
    print(f"\n{len(written)} ファイルを更新しました（status: review で登録）")
    if logged:
        print(f"  機械が下した判断 {len(logged)} 件 → "
              f"{decisions_module.path_of(round_).relative_to(paths.root)}"
              "（out/決定記録.md に出ます）")
    print(f"  concept 台帳: {paths.concepts.relative_to(paths.root)}")
    print("次にやること: arp4 check で必須属性の欠落（E010）を確認し、人が埋める")
    return 0


def _check(args: argparse.Namespace) -> int:
    spec, findings = _load(args)
    known, concept_findings = concepts_module.load(spec.paths)
    findings = (findings + concept_findings
                # **逆向きの検査。** 台帳 → 正本の写像が切れても、順方向
                # （G003）だけでは誰も気づかない（→ E029）。
                + concepts_module.check(known, set(spec.by_id))
                + validate(spec) + sequence_module.missing_format(spec)
                + sequence_module.nonconforming(spec)
                + trace_module.check(spec, spec.paths)
                # **段の境界を前倒しする。** 「資料は届いているのに設計書が空」を
                # publish の出力を人が目で見るまで誰も言わなかった（→ W034）。
                + publish_module.pending(spec)
                # 文書定義そのものの検査。解決できない列は**エラーにならず全行空の
                # 列**になり、「全行が空だったので省略」と畳まれて誰も気づかない
                # （→ E040 / W043）。
                + publish_module.lint(spec)
                # **組み上がった設計書のほうも見る**（P0xx）。lint が見ているのは
                # 文書定義（レシピ）で、出来上がった表（料理）は誰も見ていなかった
                # ―― 「列見出しは日本語」と書いてあるのに `7.1 business` が出荷され、
                # 「母集合を並べない」と書いてあるのに 80 行が全行 `―` で出た。
                + audit_module.audit(spec)
                # 解釈層（derived）の根拠の実在。**幻覚の最頻形は存在しない
                # 根拠**なので、正本の検証と同じ場所で機械が全件潰す。
                + derived_module.check(spec))

    rounds = spec.paths.rounds()
    frozen = [r for r in rounds if r.is_frozen()]
    metrics = {"rounds": len(rounds), "frozen_rounds": len(frozen),
               "items": len(spec.items), "relations": len(spec.relations)}
    if _machine(args, "check", findings, metrics):
        return _exit_code(findings, args.strict)

    print(f"ラウンド {len(rounds)}（凍結済み {len(frozen)}）")
    print(f"アイテム {len(spec.items)} 件 / 関係 {len(spec.relations)} 件")
    return _report(findings, args.strict, args=args, command="check",
                   paths=spec.paths, metrics=metrics)


def _number(args: argparse.Namespace) -> int:
    spec, findings = _load(args)
    blocking = [f for f in findings if f.level == "error"]
    if blocking:
        for finding in order(blocking)[:10]:
            print(finding.render(), file=sys.stderr)
        return 1

    assignments, seq_findings = sequence_module.assign(
        spec, renumber=args.renumber, fix_format=args.fix_format)
    # **自分が生成した値を検査する。** 採番の欠陥は書き込むまで無症状で、
    # 気づくのは次の工程の E012 ―― そのときには正本に入っている。
    broken = sequence_module.collisions(spec, assignments)
    reported = seq_findings + broken + sequence_module.missing_format(spec)
    stray = sequence_module.nonconforming(spec)
    # --fix-format のときは下の一覧が「何をどう直したか」を言う。同じものを
    # error としても並べると、打ち手が 2 つに割れて読めなくなる。
    if not args.fix_format:
        reported += stray
    for finding in order(reported):
        print(finding.render())
    if stray and not args.fix_format:
        print(f"書式外の表示 ID が {len(stray)} 件あります。"
              "元資料の番号なら overridden に理由を書き、"
              "そうでなければ arp4 number --fix-format で振り直してください。")
    if not assignments:
        print("採番が必要なアイテムはありません。")
        return 0
    if args.verbose:
        for assignment in assignments:
            print(f"  {assignment.render()}")
    else:
        # **既定はサマリ。** 全件出すと 890 件・15,000 字超が stdout に流れ、
        # 呼んだ側（エージェント）の予算を食うだけで 1 行も検算されない（r001
        # 実測）。種別ごとの件数と範囲なら、番号の抜け・体系違いは読み取れる。
        for line in _number_summary(assignments):
            print(line)
        print("  （1 件ずつ見るなら --verbose）")
    if broken:
        # サマリを出したあとで止める ―― どの体系が壊れたかは上の行にしか出ない。
        print(f"\n採番が重複した表示 ID を {len(broken)} 件作りました。"
              "表示 ID は一意でなければならないので書き込みません"
              "（メタモデルの sequence.format と by を確認してください）。")
        return 1
    if args.dry_run:
        print(f"\n--dry-run のため書き込みませんでした（{len(assignments)} 件）")
        return 0

    changed = sequence_module.apply(spec, assignments)
    written = spec_module.save_in_place(spec, changed)
    print(f"\n{len(assignments)} 件を採番し、{len(written)} ファイルを更新しました")
    print("注意: 書き戻したファイルのコメントは失われます（差分で確認してください）")
    return 0


def _number_summary(assignments: list[sequence_module.Assignment]) -> list[str]:
    """採番の要約 ―― 種別ごとに**件数と範囲**を 1 行で言う。

    範囲は表示 ID の並び（:func:`arp4.sequence.sort_key`）で最小と最大を出す。
    接頭辞・グループ（``FR-`` と ``NFR-``）が混ざる種別は、値の頭で分ける ――
    ``FR-001 〜 NFR-012`` という範囲は 2 つの体系をまたいでいて読めない。
    """
    buckets: dict[tuple[str, str, str], list[str]] = {}
    for assignment in assignments:
        head = assignment.value.rsplit("-", 1)[0] if "-" in assignment.value else ""
        key = (assignment.type_name, assignment.attribute, head)
        buckets.setdefault(key, []).append(assignment.value)

    lines: list[str] = []
    for (type_name, attribute, _head), values in sorted(buckets.items()):
        ordered = sorted(values, key=sequence_module.sort_key)
        span = (ordered[0] if len(ordered) == 1
                else f"{ordered[0]} 〜 {ordered[-1]}")
        line = f"  {type_name}.{attribute}: {span}（{len(ordered)} 件）"
        # **重複は言う。歯抜けは言わない。** この行は表示 ID の頭（`TC`）で
        # 束ねるので、件数を数えない限り衝突が見えない ―― 実測で 12 件の重複が
        # `TC-0001 〜 TC-0010（12 件）` の 1 行に畳まれ、`check` まで誰も
        # 気づかなかった。歯抜けのほうは、埋めるのは空いているアイテムだけと
        # 決めた（規律 1）以上、1 回の採番が飛び飛びになるのは正常な穴埋めである。
        duplicated = len(ordered) - len(set(ordered))
        if duplicated:
            line += f" ← 表示 ID が {duplicated} 件重複しています"
        lines.append(line)
    return lines


def _publish(args: argparse.Namespace) -> int:
    spec, findings = _load(args)
    if args.list:
        for definition in publish_module.catalog(spec):
            phase = str(definition.get("phase") or "―")
            print(f"  {definition['name']:<24} {phase:<8}"
                  f" {definition.get('title', '')}")
        return 0

    # 文書定義の誤り（E040）は**生成前に**止める ―― 解決できない列は全行空の列に
    # なり、生成物からは「資料に無い」と見分けが付かない。直す場所は正本ではなく
    # パックの documents/*.yml である。
    findings = findings + validate(spec) + publish_module.lint(spec)
    stakeholder = args.audience == "stakeholder"
    derived_data = None
    if stakeholder:
        # 解釈層は stakeholder 文書の材料なので、根拠の実在をここでも見る ――
        # 指せない basis のまま生成すると、根拠の無い要約が顧客に渡る。
        derived_data, derived_findings = derived_module.load(spec.paths)
        findings = (findings + derived_findings
                    + derived_module.check(spec, derived_data))
    blocking = [f for f in findings if f.level == "error"]
    if blocking and not args.force:
        for finding in order(blocking)[:10]:
            print(finding.render(), file=sys.stderr)
        print(f"error が {len(blocking)} 件あります。"
              "設計書ではなく正本を直して再実行してください（承知のうえなら --force）。",
              file=sys.stderr)
        return 1

    # **組み上がった設計書のほうも、出す前に見る**（P1xx）。ここが `check` にしか
    # 無かったあいだ、帯・`_gate.json`・穴の 1 枚は **publish が知らない件数**を
    # 刻んでいた ―― 実測（r001）で `check` は warn 143 件、生成物は 124 件と
    # 名乗り、差の 19 件がまるごと表の形の指摘だった。手順書は「`0_この設計書の
    # 穴.md` に出る」と書いていたので、**読み手は出ていないものを探しに行く。**
    #
    # 母集合は生成する束と揃える（`--document` / `--full` をそのまま渡す）。
    # error があるまま `--force` で来たときは組み立てない ―― 文書定義が壊れて
    # いれば（`E040`）ここは組めず、**止めるべき場所で例外が出る**。帯は既に
    # `--force` を刻むので、その場合の指摘は `arp4 check` の側で読む。
    if not blocking and not stakeholder:
        findings = findings + audit_module.audit(
            spec, full=args.full, names=args.document or None)

    # **通った条件を生成物に残す。** --force の cost が「端末に 10 行流れる」だけ
    # だと、次のラウンドでも同じ判断が繰り返される（→ arp4.gate）。
    gate = gate_module.summarize(findings, forced=bool(blocking))

    out = Path(args.out).resolve() if args.out else spec.paths.out
    if stakeholder:
        written = audience_module.publish_stakeholder(spec, derived_data, out)
        # 読者別の生成でも記録は残す ―― **顧客向けのほうが痕跡が要る。**
        #
        # **書き先は `out/stakeholder/` である（`out/` 直下ではない）。** 直下へ
        # 書くと developer の記録を上書きする ―― stakeholder は P1xx 監査を回さ
        # ない（母集合が developer の文書定義に紐づくため。上記）ので、上書きは
        # **表の形の指摘だけが消えた件数**に化ける。実測（sales-corpus r001）で
        # developer の帯と穴の 1 枚が warn 132（P106 4・P107 4 を含む）と刻んだ
        # 直後に stakeholder が `_gate.json` を 124 で上書きし、**手順書が
        # 「3 つが同じ件数になった」と書いている状態が嘘になっていた。**
        written.append(gate_module.record(out / audience_module.DIR, gate))
    else:
        # 決定記録の付録（4-2）。機械が下した判断の全件 ―― 事後拒否権の入口。
        # **目次より先に書く。** 目次は `out/` にある付録を実際に見て並べるので
        # （→ :func:`publish._index`）、あとに書くと初回だけ載らない。
        appendix = audience_module.decision_report(spec, out)
        written = publish_module.publish(spec, out, names=args.document or None,
                                         flat=args.flat, full=args.full,
                                         gate=gate, findings=findings)
        if appendix is not None:
            written.append(appendix)
    if not written:
        print("生成できる文書がありません（arp4 publish --list で確認）",
              file=sys.stderr)
        return 1
    # **「N ファイル」の N を設計書の数として読ませない。** 束には `_gate.json`
    # （通った条件の記録）が混じっており、`check` はこのあと同じ場所へ
    # `findings.json` を置く ―― 実測（sales-corpus r001）で、報告するエージェントが
    # この 32 を「設計書 32 ファイル」と書き写し、`out/` の実数は 33、設計書は
    # 31（md 16 + html 15）だった。**数えたものが何かを言う。**
    docs = [p for p in written if p.suffix != ".json"]
    aside = len(written) - len(docs)
    tail = f"（うち設計書 {len(docs)} ／ 記録 {aside}）" if aside else ""
    print(f"生成しました: {len(written)} ファイル{tail} → {out}")
    for path in written:
        print(f"  {path.relative_to(out).as_posix()}")
    print("次にやること: 内容が違うなら out/ ではなく正本を直して再生成する")
    return 0


#: ``arp4 auto`` が「整理層（LLM）の持ち場が残っている」と言うときの終了コード。
#: 0（完了）とも 1（機械が判断できない矛盾・error）とも違う ―― 呼んだ側は
#: 文章化・整理を済ませて**同じコマンドをもう一度打てばよい**。
_AUTO_PENDING = 3


def _step_args(args: argparse.Namespace, **overrides: Any) -> argparse.Namespace:
    """auto の 1 段に渡す引数。**各コマンドの既定と同じ値**を明示して埋める。"""
    base = {"root": args.root, "strict": False, "format": "text",
            "round": None, "dry_run": False}
    base.update(overrides)
    return argparse.Namespace(**base)


def _auto(args: argparse.Namespace) -> int:
    """parse → draft → 文章化キュー → freeze → build → number → check → publish を
    1 コマンドで回す（Phase 4-1）。

    **止まるのは 2 つだけ**である ―― 機械が判断できない矛盾・error（exit 1）と、
    整理層の持ち場が残っているとき（exit 3 ―― 文章化スロットか、コードでない
    資料の整理）。承認のための停止は無い。判断は全件 ``decisions.yml`` に残り、
    publish の「決定記録」が事後拒否権の入口になる。
    """
    if not args.source:
        return _no_sources("auto")

    try:
        paths_module.resolve(args.root)
    except paths_module.ArpNotFound:
        # 骨組みが無いだけなら作る ―― auto は「コード一式を渡すと承認なしで
        # ドキュメントが出る」の入口なので、init を別打ちさせない。
        created = paths_module.create(args.root or Path.cwd())
        print(f"{created.root / paths_module.ARP_DIR} を用意しました（arp4 init 相当）")

    print("=== ① parse ===")
    code = _parse(_step_args(args, source=args.source, base=args.base,
                             exclude=args.exclude, new_round=args.new_round,
                             yes=True))
    if code != 0:
        return code

    print("\n=== ② draft（コード整理の機械分）===")
    code = _draft(_step_args(args))
    if code != 0:
        return code

    print("\n=== ③ freeze ===")
    spec, findings = _load(args)
    paths = spec.paths
    round_ = paths.latest_round()
    known, concept_findings = concepts_module.load(paths)
    report = freeze_module.gate(round_, spec.metamodel, known)
    gate_findings = findings + concept_findings + report.findings
    queue = [f for f in gate_findings if f.code in ("G001", "G026")]
    hard = [f for f in gate_findings
            if f.level == "error" and f.code not in ("G001", "G026")]
    if hard:
        for finding in order(hard)[:20]:
            print(finding.render(), file=sys.stderr)
        print(f"\n機械では判断できない error が {len(hard)} 件あります"
              "（凍結しませんでした）", file=sys.stderr)
        return 1
    if queue:
        # **整理層の持ち場。** ここは停止ではなく手番の交代である。
        print(f"整理層の作業が {len(queue)} 件残っています（作業キュー）:")
        for finding in order(queue)[:_QUEUE_HEAD]:
            print(f"  {finding.where}: {finding.message}")
        if len(queue) > _QUEUE_HEAD:
            print(f"  …ほか {len(queue) - _QUEUE_HEAD} 件"
                  "（arp4 freeze --dry-run --list で全部出る）")
        print("\n文章化（<TODO …> を埋める）と資料の整理を済ませて、"
              "もう一度 arp4 auto を打ってください")
        return _AUTO_PENDING
    if not round_.is_frozen():
        freeze_module.apply(round_, report)
        print(f"凍結しました: {round_.frozen}")
    else:
        print(f"ラウンド {round_.name} は凍結済みです（そのまま進めます）")

    print("\n=== ④ build ===")
    code = _build(_step_args(args, force=False))
    if code != 0:
        return code

    # 4-4: 「相手が資料に無い」と機械が判定できる W031 を自動宣言する。
    spec, _ = _load(args)
    changed, logged = auto_module.declare_gaps(spec, round_.name)
    if changed:
        spec_module.save_in_place(spec, changed)
        decisions_module.append(round_, logged)
        print(f"\nknown_gaps を {len(changed)} 件自動宣言しました"
              "（相手になれる種別が正本に 1 件も無いもの。決定記録に残ります）")

    print("\n=== ⑤ number ===")
    code = _number(_step_args(args, renumber=False, fix_format=False,
                              verbose=False))
    if code != 0:
        return code

    print("\n=== ⑥ check ===")
    spec, findings = _load(args)
    known, concept_findings = concepts_module.load(paths)
    check_findings = (findings + concept_findings
                      + concepts_module.check(known, set(spec.by_id))
                      + validate(spec) + sequence_module.missing_format(spec)
                      + sequence_module.nonconforming(spec)
                      + trace_module.check(spec, spec.paths)
                      + publish_module.pending(spec)
                      + publish_module.lint(spec)
                      + derived_module.check(spec))
    errors = [f for f in check_findings if f.level == "error"]
    for finding in order(check_findings)[:30]:
        print(finding.render())
    tally = counts(check_findings)
    print(f"error {tally['error']} / warn {tally['warn']}")
    if errors:
        print("\ncheck の error は機械では埋められません（正本を直して"
              "もう一度 arp4 auto）", file=sys.stderr)
        return 1

    # 4-3: 自動昇格。**既定は現行どおり review 止まり** ―― policy.yml で
    # auto_approve を選んだプロジェクトだけが有効になる。
    if auto_module.policy(spec).get("auto_approve"):
        promoted, logged = auto_module.promote(spec, round_.name)
        if promoted:
            spec_module.save_in_place(spec, promoted)
            decisions_module.append(round_, logged)
            print(f"\ncheck error 0 のため {len(promoted)} 件を approved へ"
                  "昇格しました（policy.yml の auto_approve）")

    print("\n=== ⑦ publish ===")
    for audience in audience_module.AUDIENCES:
        code = _publish(_step_args(args, document=[], out=None, list=False,
                                   flat=False, full=False, force=False,
                                   audience=audience))
        if code != 0:
            return code
    print("\narp4 auto: 人の介入なしで設計書一式まで到達しました"
          "（判断は decisions.yml と out/決定記録.md にあります）")
    return 0


def _correctable(definition: dict) -> bool:
    """向きを機械が直せる関係か。

    ``from`` と ``to`` に共通の種別があると、その種別どうしの関係は**両向きとも
    宣言に合う**ので直しようがない（``refines`` の ``same_type_only`` が典型）。
    共通が無ければ、逆向きで書かれていても一意に戻せる。
    """
    origin = set(definition.get("from") or [])
    target = set(definition.get("to") or [])
    if not origin or not target:          # 宣言が無い＝何でもよい
        return False
    return not (origin & target)


def _attribute_hint(name: str, attr: dict, groups: str = "") -> str:
    """属性 1 つの説明。**enum の値まで出す** ―― 整理層が選ぶのはここからである。"""
    kind = str(attr.get("kind") or "?")
    if kind == "enum":
        values = "/".join(str(v) for v in (attr.get("values") or []))
        kind = f"enum: {values}" + ("…" if attr.get("extensible") else "")
    marks = [flag for flag, on in (("複数値", attr.get("multi")),
                                   ("一意", attr.get("unique"))) if on]
    if attr.get("pattern"):
        marks.append(f"書式 {attr['pattern']}")
    if groups:
        marks.append(f"※設計書の節になる: {groups}")
    return f"{name}（{'、'.join([kind] + marks)}）"


def _grouping(chain: list) -> dict[str, dict[str, str]]:
    """**どの属性が設計書の章立てになるか**を、文書定義から機械的に集める。

    整理層が最初に読むのは ``arp4 model`` である。そこで ``category`` が
    ``owner`` や ``description`` と同じ「任意の文字列」にしか見えないと、
    **これが目次を決めると知らないまま**整理結果が出来上がる ―― 実測では
    17 分類のうち 15 がモジュール名と 1 対 1 になり、要件定義書の目次が
    ソースのファイル一覧になった（→ ``docs/decisions.md`` 決定 19）。

    宣言を新しく足さない。``documents/*.yml`` の ``group_by`` が**既にその宣言**
    なので、見せていなかっただけである。
    """
    grouped: dict[str, dict[str, list[str]]] = {}
    for definition in pack_module.documents(list(chain)):
        title = str(definition.get("title") or definition.get("name") or "")
        for section in definition.get("sections") or []:
            attribute = str(section.get("group_by") or "")
            type_name = str(section.get("type") or "")
            # 関係の章は ``group_by: from``（属性ではない）なので数えない。
            if not attribute or not type_name or str(
                    section.get("kind") or "items") != "items":
                continue
            where = grouped.setdefault(type_name, {}).setdefault(attribute, [])
            where.append(f"{title}「{section.get('heading')}」")
    return {t: {a: "・".join(v) for a, v in attrs.items()}
            for t, attrs in grouped.items()}


def _attribute_lines(attributes: dict, verbose: bool, indent: str,
                     groups: dict[str, str] | None = None) -> list[str]:
    """必須・任意の属性名。

    **任意属性の名前をどこにも出していなかった** ので、``B021``（その種別に無い
    属性名）を踏んだ人はパッケージのソースにある ``metamodel.yml`` を直接読むしか
    なかった。``arp4 model`` を唯一の入口だと言っている以上、ここに出す。
    """
    groups = groups or {}
    required = {k: v for k, v in attributes.items()
                if isinstance(v, dict) and v.get("required")}
    optional = {k: v for k, v in attributes.items() if k not in required}

    lines: list[str] = []
    for label, group in (("必須", required), ("任意", optional)):
        if not group:
            continue
        if verbose:
            shown = [_attribute_hint(k, v or {}, groups.get(k, ""))
                     for k, v in group.items()]
        else:
            shown = [k + ("※" if k in groups else "") for k in group]
        lines.append(f"{indent}{label}: " + " ".join(shown))
    return lines


def _sequence_hint(definition: dict) -> str:
    """``arp4 number`` がこの種別に何を振るか。**無いなら無いと言う。**

    「機械が振らない」と「表示 ID を持たない」は別である。資料から取った ID を
    整理層が書く種別（採番を外した ``message`` 等）を「名前で参照する」と言うと、
    **書くべき属性を書かないまま整理結果が出来上がる。**
    """
    sequence = definition.get("sequence") or {}
    attribute = sequence.get("attribute")
    if not attribute:
        manual = sequence_module.display_attribute(definition)
        if manual:
            return f"なし（機械は振らない。{manual} は資料から取って整理層が書く）"
        return "なし（表示 ID を持たない。設計書からは名前で参照する）"
    formats = sequence.get("format")
    if isinstance(formats, dict):
        shown = "、".join(f"{k}: {v}" for k, v in formats.items())
    else:
        shown = str(formats or "書式なし")
    by = f"（{sequence['by']} ごと）" if sequence.get("by") else ""
    return f"{attribute} = {shown}{by}"


def _model(args: argparse.Namespace) -> int:
    """**整理層が最初に読むもの。** 語彙を知らずに書いた type は凍結で弾かれる。"""
    try:
        spec, findings = _load(args)
        model = spec.metamodel
        chain = list(model.chain)
        origin = f"{spec.paths.metamodel}（extends: {model.extends or 'なし'}）"
    except (FileNotFoundError, paths_module.ArpNotFound):
        raw = mm.load_pack(args.pack)
        model, findings = mm.resolve(raw)
        # 正本がまだ無いときも文書定義は読める（パックそのものを鎖にする）。
        chain, chain_findings = pack_module.resolve_chain(args.pack)
        findings += chain_findings
        origin = f"pack {args.pack}"

    grouping = _grouping(chain)

    print(f"{origin} ― version {model.version}")
    print("\n[整理結果に書く type]（種別 → 正本のアイテム種別）")
    for fact_type in sorted(model.fact_types):
        item_type, fixed = model.fact_types[fact_type]
        extra = f"  {fixed}" if fixed else ""
        print(f"  {fact_type:<16} → {item_type}{extra}")

    # **整理結果から起こせない種別**は、写像（fact_types）を持たないもの。
    # 「正本にはあるのに整理から書けない」ことが model の出力から読めないと、
    # 議事録の決定事項・テスト結果をどこへ寄せるかで必ず迷う。
    orphans = [name for name, definition in model.item_types.items()
               if not definition.get("fact_types")]
    if orphans:
        print("\n[整理結果からは書けない種別]（正本側で人が起こす。整理では寄せない）")
        for name in orphans:
            print(f"  {name:<20} {model.item_types[name].get('label', '')}")

    for layer in model.layers:
        names = model.types_in_layer(layer)
        if not names:
            continue
        print(f"\n[{layer}] {len(names)} 種別")
        for name in names:
            definition = model.item_types[name]
            attributes = definition.get("attributes") or {}
            print(f"  {name:<20} {definition.get('label', ''):<12}"
                  f" 属性 {len(attributes)}")
            for line in _attribute_lines(attributes, args.attributes, "    ",
                                         grouping.get(name)):
                print(line)
            # 表示 ID が振られるかどうかは、整理の時点で知りたい（この種別を
            # あとから ID で参照できるか、が決まる）。
            print(f"    採番 {_sequence_hint(definition)}")

    print(f"\n[関係] {len(model.relation_types)} 種")
    for name, definition in model.relation_types.items():
        origin_types = "|".join(definition.get("from") or ["*"])
        target = "|".join(definition.get("to") or ["*"])
        attributes = definition.get("attributes") or {}
        # from と to に共通の種別があると、**向きは機械が直せない**（B026）。
        # 整理層が「どちらでもよい」と読むと、トレースがそっくり落ちる。
        fixed = "" if _correctable(definition) else "  ※向きは書いたまま入る"
        print(f"  {name:<14} {origin_types} → {target}"
              + (" +属性" if attributes and not args.attributes else "") + fixed)
        if attributes and args.attributes:
            for line in _attribute_lines(attributes, True, "    "):
                print(line)

    for finding in order(findings):
        print(finding.render(), file=sys.stderr)
    return 1 if counts(findings)["error"] else 0


def _lock(args: argparse.Namespace) -> int:
    spec, findings = _load(args)
    chain = list(spec.metamodel.chain)
    blocking = [f for f in findings if f.level == "error"]
    if blocking:
        for finding in order(blocking):
            print(finding.render(), file=sys.stderr)
        print("チェーンを解決できないため lock を更新しませんでした。", file=sys.stderr)
        return 1
    if not chain:
        print("extends が宣言されていないので lock は不要です。")
        return 0
    path = pack_module.write_lock(spec.paths.spec, chain)
    print(f"pack.lock を更新しました: {path}")
    print("  チェーン: " + " → ".join(f"{p.name}@{p.version}" for p in chain))
    return 0


def _conform(args: argparse.Namespace) -> int:
    spec, findings = _load(args)
    findings = findings + conform_module.conform(
        spec, frozen=args.frozen, baseline=args.baseline)
    chain = list(spec.metamodel.chain)
    if _machine(args, "conform", findings,
                {"chain": [f"{p.name}@{p.version}" for p in chain]}):
        return _exit_code(findings, args.strict)

    if chain:
        print("チェーン: " + " → ".join(f"{p.name}@{p.version}" for p in chain))
    else:
        print("extends が宣言されていません（準拠検証の対象外）")
    return _report(findings, args.strict, args=args, command="conform",
                   paths=spec.paths)


# ── 入口 ────────────────────────────────────────────────────────
def _resilient_output() -> None:
    """**コンソールの文字コードで落ちない。**

    日本語 Windows の既定コンソールは cp932 で、そこに無い文字を出そうとすると
    ``UnicodeEncodeError`` で**処理そのものが落ちる**。出力の都合で仕事が止まっては
    いけないので、出せない文字は置き換える。**資料の文字は選べない** ―― 半角の
    ``¥``・絵文字・外字は客先の設計書のシート名に普通に入っている。

    置き換え方は ``backslashreplace`` である。``replace``（``?``）にしていた頃は、
    **出せなかった文字と、資料に元から書いてある ``?`` が見分けられなかった**
    ―― しかも ``?`` は Windows のファイル名に使えない文字なので、
    ``受?一覧.md`` と出たものを探しても**そんな名前のファイルは無い**。
    ``受\\xa5一覧.md`` なら符号位置が残るので、どの文字だったかを読み直せる。
    「資料に無い」と「機械が出せていない」を混ぜないのは、パースの側で
    やっていることと同じである。
    """
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(                        # type: ignore[union-attr]
                errors="backslashreplace")
        except (AttributeError, OSError, ValueError):
            pass


def main(argv: list[str] | None = None) -> int:
    _resilient_output()
    parser = argparse.ArgumentParser(prog="arp4", description=__doc__)
    sub = parser.add_subparsers(dest="command", required=True)

    def add(name: str, help_text: str, handler, strict: bool = True,
            machine: bool = False, digest: bool = False):
        command = sub.add_parser(name, help=help_text)
        command.add_argument("--root", metavar="<パス>",
                             help="プロジェクト根（既定: cwd から .arp/ を上方探索"
                                  "。**エージェントは必ず明示する**）")
        if strict:
            command.add_argument("--strict", action="store_true",
                                 help="warn も失敗として扱う")
        if machine:
            # **検査するコマンドにだけ付ける。** 何かを作るコマンド（parse /
            # build / publish）の出力は「何をしたか」であって指摘ではないので、
            # 同じ器に入れると読み手が中身を推測することになる。
            command.add_argument("--format", choices=report_module.FORMATS,
                                 default="text",
                                 help="指摘の出し方（json / sarif は機械向け）")
        if digest:
            # **段階的に開く。** 既定は種別ごとに先頭
            # `digest.CAP` 件（→ :mod:`arp4.digest`）。全体像を見てから 1 種類を
            # 開くのが読み方で、絞り込みは**表示だけ**に効く（判定は絞らない ――
            # 絞った結果で終了コードが変わると、CI と手元で結論がずれる）。
            command.add_argument("--summary", action="store_true",
                                 help="code ごとの内訳だけを出す（1 回目の呼び出し用）")
            command.add_argument("--code", action="append", metavar="<コード>",
                                 help="その code だけを切らずに出す"
                                      "（前方一致。--code W0 で系ごと。複数可）")
        command.set_defaults(handler=handler, strict=False, format="text",
                             summary=False, code=None)
        return command

    add("init", "置き場の骨組みを作る", _init, strict=False)

    parse = add("parse", "既存資産をパース結果へ（①）", _parse, strict=False)
    parse.add_argument("source", nargs="*",
                       help="資料のパス（必須。既存資産のいまある場所を指す）")
    parse.add_argument("--round", metavar="<名前>",
                       help="ラウンド名（既定: 作業中のラウンド。無ければ新規）")
    parse.add_argument("--new-round", action="store_true",
                       help="資料が更新されたので新しいラウンドを起こす")
    parse.add_argument("--base", metavar="<パス>",
                       help="フォルダ構造を写す基準（既定: 資料のパスの共通の親）")
    parse.add_argument("--exclude", action="append", metavar="<glob>",
                       help="資料として読まないパターン（パスにも名前にも当たる。"
                            "テストの期待値・フィクスチャを除くのに使う。複数可）")
    parse.add_argument("--yes", action="store_true",
                       help="編集済みのパース結果も上書きする")
    parse.add_argument("--dry-run", action="store_true",
                       help="書き込まずに件数だけ出す")

    render = add("render", "機械が読めなかった範囲を絵にする", _render, strict=False)
    render.add_argument("source", nargs="*",
                        help="資料のパス（必須。parse に渡したものと同じ）")
    render.add_argument("--sheet", action="append", metavar="<名前>",
                        help="シートを名指しする（既定: 図形のあるシートだけ）")
    render.add_argument("--pending", action="store_true",
                        help="`未読取` と宣言されたシートだけを撮る（二巡目）"
                             "。宣言そのものが撮り直しの依頼になる")
    render.add_argument("--range", metavar="<A1レンジ>",
                        help="範囲を名指しして撮り直す（B2:L20）。拡大して読むためのもので、"
                             "線の始点・終点は全体図では確定しきれない。"
                             "--sheet と併せて使い、絵は差し替えず追加される")
    render.add_argument("--all", action="store_true",
                        help="図形の無いシートも撮る（遅い。1 冊で分単位）")
    render.add_argument("--round", metavar="<名前>", help="対象のラウンド（既定: 最新）")
    render.add_argument("--base", metavar="<パス>",
                        help="フォルダ構造を写す基準（既定: 資料のパスの共通の親）")
    render.add_argument("--out", metavar="<パス>",
                        help="書き出し先（既定: .arp/rounds/<ラウンド>/images/）")
    render.add_argument("--target-px", type=int, default=render_module.TARGET_PX,
                        metavar="<px>", help="1 枚の長辺（既定: "
                        f"{render_module.TARGET_PX}）")
    render.add_argument("--wide-px", type=int, default=render_module.WIDE_PX,
                        metavar="<px>", help="横も割る幅（既定: "
                        f"{render_module.WIDE_PX}）")
    render.add_argument("--title-rows", type=int, default=0, metavar="<行数>",
                        help="全部の絵に載せる見出し行（長い表を割るとき。"
                             "どこまでが見出しかは機械が当てない）")
    render.add_argument("--dry-run", action="store_true",
                        help="撮らずに枚数と範囲だけ出す（Excel を起動しない）")

    declare = add("declare", "同じ構成のシートを一括で対象外宣言する",
                  _declare, strict=False)
    declare.add_argument("pattern", nargs="+",
                         help="パース結果の名前（表紙 / 改訂履歴 / '*/レイアウト' 等）")
    declare.add_argument("--reason", required=True, metavar="<理由>",
                         help="対象外にする理由（必須。機械が埋めない）")
    declare.add_argument("--kind", default=organized_module.SCOPE_DEFAULT,
                         choices=list(organized_module.SCOPE_KINDS),
                         help="対象外＝資料に仕様が無い / 未読取＝機械が読めていない")
    declare.add_argument("--round", metavar="<名前>", help="対象のラウンド（既定: 最新）")
    declare.add_argument("--dry-run", action="store_true",
                         help="書き込まずに対象だけ出す")
    declare.add_argument("--list", action="store_true",
                         help="当たったファイルを全部出す（既定は先頭 20 件）")

    schema = add("schema", "書いてよい形を出す（arp4 model と対になる）",
                 _schema, strict=False)
    schema.add_argument("name", nargs="?", default="organized",
                        choices=shape_module.names(),
                        help="スキーマの名前（既定: organized。整理結果の形。"
                             "ほかに concepts / metamodel-add）")

    lint = add("lint", "整理結果を 1 ファイル単位で検査する（freeze の部分集合）",
               _lint, machine=True, digest=True)
    lint.add_argument("path", nargs="*",
                      help="整理結果のパス（ファイルでもフォルダでも可。"
                           "既定: そのラウンドの全部）")
    lint.add_argument("--round", metavar="<名前>", help="対象のラウンド（既定: 最新）")
    lint.add_argument("--fix", action="store_true",
                      help="機械的に確実なものだけ直す（attrs の外の属性を中へ）")

    draft = add("draft", "コードのパース結果から整理結果の骨格を機械生成する",
                _draft, strict=False)
    draft.add_argument("--round", metavar="<名前>", help="対象のラウンド（既定: 最新）")
    draft.add_argument("--dry-run", action="store_true",
                       help="書き込まずに件数だけ出す")

    freeze = add("freeze", "整理結果を凍結する（②）", _freeze, strict=False,
                 machine=True)
    freeze.add_argument("--round", metavar="<名前>", help="対象のラウンド（既定: 最新）")
    freeze.add_argument("--dry-run", action="store_true",
                        help="ゲートだけ見る（凍結しない）")
    freeze.add_argument("--list", action="store_true",
                        help=f"未整理のアンカーを全部出す（既定は先頭 {_QUEUE_HEAD} 件）")
    freeze.add_argument("--path", action="append", metavar="<部分パス>",
                        help="指摘を担当ぶんに絞る（分担用。判定は絞りません。"
                             "この path のぶんと隠したぶんを分けて集計します。"
                             "どのファイルにも当たらなければ exit 2）")

    build = add("build", "整理結果を正本へ（③）", _build, strict=False)
    build.add_argument("--round", metavar="<名前>", help="対象のラウンド（既定: 最新）")
    build.add_argument("--dry-run", action="store_true", help="書き込まずに案だけ出す")
    build.add_argument("--force", action="store_true",
                       help="凍結していないラウンドでも組み立てる")

    check = add("check", "正本を検証する", _check, machine=True, digest=True)
    check.add_argument("--show-known", action="store_true",
                       help=f"known_gaps で承知済みの欠落（{digest_module.KNOWN}）"
                            "も 1 件ずつ出す（既定は件数だけ）")

    number = add("number", "表示 ID を採番する", _number, strict=False)
    number.add_argument("--dry-run", action="store_true", help="書き込まずに案だけ出す")
    number.add_argument("--verbose", action="store_true",
                        help="採番を 1 件ずつ出す（既定は種別ごとの件数と範囲）")
    number.add_argument("--renumber", action="store_true",
                        help="既存の番号も振り直す（参照が壊れる。移行用）")
    number.add_argument("--fix-format", action="store_true",
                        help="採番の書式に合わない表示 ID だけ振り直す"
                             "（overridden は触らない。移行用）")

    publish = add("publish", "設計書を生成する", _publish, strict=False)
    publish.add_argument("document", nargs="*", help="生成する文書名（既定は全部）")
    publish.add_argument("--audience", choices=list(audience_module.AUDIENCES),
                         default="developer",
                         help="読者（developer: 12 種＋決定記録 / "
                              "stakeholder: PM・顧客向けの概要・機能一覧・"
                              "用語集・テスト状況・構成図・処理フロー図）")
    publish.add_argument("--out", metavar="<パス>", help="出力先（既定 .arp/out）")
    publish.add_argument("--list", action="store_true", help="生成できる文書を一覧する")
    publish.add_argument("--flat", action="store_true",
                         help="工程で分けず out/ 直下へ並べる")
    publish.add_argument("--full", action="store_true",
                         help="マトリクスの空行・空列も省略せずに出す")
    publish.add_argument("--force", action="store_true",
                         help="error があっても生成する（非推奨）")

    auto = add("auto", "parse → draft → freeze → build → number → check → "
               "publish を 1 コマンドで回す", _auto, strict=False)
    auto.add_argument("source", nargs="*",
                      help="資料のパス（必須。arp4 parse と同じ）")
    auto.add_argument("--base", metavar="<パス>",
                      help="フォルダ構造を写す基準（既定: 資料のパスの共通の親）")
    auto.add_argument("--exclude", action="append", metavar="<glob>",
                      help="資料として読まないパターン（arp4 parse と同じ）")
    auto.add_argument("--new-round", action="store_true",
                      help="資料が更新されたので新しいラウンドを起こす")

    add("lock", "pack.lock を更新する", _lock, strict=False)

    conform = add("conform", "標準パックへの準拠を検証する", _conform,
                  machine=True, digest=True)
    conform.add_argument("--frozen", action="store_true",
                         help="pack.lock との不一致を error にする（CI 用）")
    conform.add_argument("--baseline", action="store_true",
                         help="ベースライン前提（全アイテムが approved）も検査する")

    model = add("model", "使ってよい語彙を出す", _model, strict=False)
    model.add_argument("--pack", default="jp-sier-std",
                       help="プロジェクト外で実行するときの既定パック")
    model.add_argument("--attributes", "-a", action="store_true",
                       help="属性の kind・enum の値・書式まで出す（関係の属性も）")

    args = parser.parse_args(argv)
    try:
        return int(args.handler(args))
    except (FileNotFoundError, YamlError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
