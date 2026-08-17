"""① 既存資産 → パース結果。**意味を判断しない**のが唯一の規律である。

Excel とコードで役割が違う。

**Excel** ―― 機械がやるのは 4 つだけで、どれも意味の判断ではない。

1. **非空セルの位置と値をそのまま出す**（書式だけのセル・削除済みデータの残骸で
   使用範囲が膨らむので、完全な空行・空列は落とす）
2. **縦に広がる結合セルを下の行へ展開する** ―― 表の分類列で「同上」を表す日本の
   慣習。openpyxl は左上にしか値を返さない。**判断ではなく忠実性の回復**である
   （幅は問わない ―― 区分と小区分をまとめて括った面結合も画面では全行に掛かって
   見えている。広げないのは**横方向**だけで、1 行だけの横結合は表題である）
3. **画面に見えている表記へ直す**（日付・時刻・真偽・パーセント）―― ``str()`` は
   ``2026-08-02 00:00:00`` や ``0.153`` を返すが、資料にそう書いてあったことは
   一度も無い。**2 と同じ忠実性の回復**であって、値の解釈ではない
4. **塊ごとに区切って出す** ―― トークンを減らすための**提示上の区切り**であって
   意味ではない。区切りを間違えても読み直せるよう、**セル番地を必ず併記する**。
   併記した番地が使えるように、**塊の中の空行・空列は詰めない**（詰めると
   ``at=B3:F8`` と書いてある表の列が 4 本しかなくなり、3 本目が D 列か E 列かを
   読み手が決められない）。**すかすかな塊は表にしない** ―― 工程表は枠だけ
   大きくて 9 割が空欄になるので、番地付きの箇条書きへ回す（:data:`_SPARSE`）

3 が持っていた「表の型を当てる」（``tables.yml``）は無い。それは意味の判断なので
整理層の仕事である。

**読めなかったものは数えて申告する** ―― 計算結果の保存されていない数式セル
（``data_only`` が None を返す）、エラー値（``#REF!``）、非表示の行・列、非表示の
シート、貼り付け画像。どれも黙っていると**空欄に見える**か**値に見える**ので、
整理層は「資料に無い」と読む。「資料に無い」と「機械が読めていない」は、次の
ラウンドで拾い直すかどうかが正反対になる。

**セルの値ではないところに乗っているものも取る** ―― セルのコメント（メモ・
スレッドコメント）とハイパーリンクの行き先と外部ブック参照は**表に出てこない**。
1 つ目には決定の理由と積み残しが溜まり（本文より新しいことがある）、あとの 2 つは
**どの資料がまだ手元に無いか**の一覧になる。

**取り消し線だけは書式から取る**（:func:`_struck`）。罫線・塗り・太字と違って、
取り消し線は**その文字が消されていることを表す**ので、値だけ抜き出すと画面に
見えているものが消える ―― 「廃止かどうか」は決めず、番地を申告する。

**コード** ―― 骨格は AST で取る。クラス・メソッド・シグネチャは構文木から曖昧さ
ゼロで取れるので、ここに LLM を使っても精度は上がらない。**意図の層（業務ルール）は
出さない** ―― コードは元ファイルがテキストなので、整理層が原本を直接読める。

**図形（オートシェイプ）はテキストと接続を取る。** 業務フロー・ER 図・状態遷移図は
図形で描かれていることが多い。箱の中の文字（``xl/drawings/*.xml`` の ``<a:t>``）は
専用のアンカー（``s4-g1``）へ、線がどの図形とどの図形を結んでいるかは別のアンカー
（``s4-c1``）へ出す。

**図は ``xdr:sp`` だけではない。** SmartArt・グラフ・グラフシートは図形を 1 つも
持たず、数えていなかった間は**セル 0 個・図形 0 個としてファイルごと消えていた**
―― 資料に無いのでも読めなかったのでもなく、シートが存在したことすら伝わらない。
何であるかは ``a:graphicData`` の uri に書いてあるので分類は転記であり、3 つを
別々に数えるのは**次にやることが 3 つとも違う**からである（グラフは元データの
シートを読む、SmartArt は箱の文字が取れる、埋め込みは元ファイルを足す）。

**貼り付け画像は中の文字も読む**（:mod:`arp4.ocr` ―― Windows OCR）。実体を
``images/`` へ出すだけだった頃は、「文字しか無い平坦な画像」（表・画面・帳票を
そのまま撮ったもの）まで整理層が 1 枚ずつ開く必要があった ―― 字であるなら
機械が読める。**読んだ字は別のアンカー**（``s4-o1``）へ出し、セルの値とも
代替テキストとも混ぜない ―― OCR は必ず読み違えるので、出自が同じ枠に入ると
整理層は「資料にそう書いてある」と読む。これも意味の判断ではない（engine が
出した字の転記である）。**使えない環境では黙って劣化させず、理由を出す**
（``P016`` と ``o1`` の両方 ―― 空の ``o1`` は「画像に字が無かった」に見える）。

**接続を取るのは意味の判断ではない。** 以前ここには「『A の次が B』は矢印が持って
おり、座標から復元するしかない（＝意味の判断になる）ので取らない」と書いてあった
が、**前提が事実と違った。** 接続子は繋がっている図形の id を ``a:stCxn`` /
``a:endCxn`` に、矢羽根を ``a:headEnd`` / ``a:tailEnd`` に**明示的に持っている**。
読むのは座標からの復元ではなく書いてあることの転記である（検証コーパスでは
接続子 160 本すべてが両端の id を持っていた）。取らずにいた間、図に描かれた流れは
1 件も仕様にならなかった ―― 14 シートが `未読取` のまま終わり、正本の業務フローは
0 件だった。

**取れないのは配置と、繋がっていない線である。** 枠で括られたゾーン・段組み・注記の
位置は座標にしか無く、線を目分量で置いただけの図（``stCxn`` を持たない接続子）も
どこからどこへの線か決められない。**取れなかったぶんは数えて申告する**
（:attr:`arp4.mdio.Doc.notes`）―― 「資料に情報が無い」と「機械が読めていない」は、
次のラウンドで拾い直すかどうかが正反対になる。目で見るしかないものは
:mod:`arp4.render` が絵にする。
"""

from __future__ import annotations

import ast
import datetime as dt
import hashlib
import os
import re
import subprocess
import warnings
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field, replace as _replace
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable, Iterable, NamedTuple

from arp4 import mdio, ocr, yamlio
from arp4.finding import Finding
from arp4.paths import Round

#: 読める拡張子。増やすときはここと :func:`_parse_one` を対で足す。
SUPPORTED = (".xlsx", ".xlsm", ".pptx", ".pptm", ".docx", ".docm", ".pdf",
             ".csv", ".tsv",
             ".py", ".md", ".txt", ".sql", ".ddl", ".java")

#: Excel として開くもの。``.xlsm`` は**マクロが付いているだけで中身は同じ**である。
#: 実案件の設計書はマクロ付きで配られることが珍しくないので、これを弾いていたぶん
#: **読めるものを「読めない」と言っていた**。
_EXCEL = (".xlsx", ".xlsm")

#: PowerPoint として開くもの。``.pptm`` は ``.xlsm`` と同じ理由で読む
#: （マクロが付いているだけで、スライドの中身は変わらない）。
_SLIDES = (".pptx", ".pptm")

#: Word として開くもの。``.docm`` は ``.xlsm`` と同じ理由で読む。
_WORDS = (".docx", ".docm")

#: 区切りで割るテキスト。**長らく「Excel で開き直してください」で終わっていた**
#: ―― 助言そのものは正しいが、移行データの一覧は数百本の CSV で配られるので、
#: 1 本ずつ開き直す作業を誰も引き受けず、**数百本が 1 件も仕様にならなかった**。
_CSV = (".csv", ".tsv")

#: 素のテキストとして読むもの。**設計の正本が Markdown で置かれている**現場は
#: もう珍しくない（このリポジトリ自身の正本が `docs/` と `surface/` である）。
#: 弾いていたあいだ、arp4 は**自分の仕様を読めなかった**。
_TEXT = (".md", ".txt")

#: DDL。**テーブル定義書は現場の 1 冊でいちばん先に要求される**のに、語彙
#: （`entity` / `data-item` / `index`）も様式も揃っていて中身だけが空だった。
#: DDL は AST と同じで曖昧でないので、当てにいくところが 1 つも無い。
_SQL = (".sql", ".ddl")

#: 開けないが、**何をすれば読めるかは言える**形式。「読めません」で終わると、
#: 資料が 1 冊落ちたまま誰も拾い直さない。
_ADVICE = {
    ".xls": "旧形式（BIFF）です。Excel で開いて .xlsx として保存し直してください",
    ".xlsb": "バイナリ形式です。Excel で開いて .xlsx として保存し直してください",
    # **旧 OLE 形式。** `.xls` と同じで、中身は zip ではないので開けない。
    ".ppt": "旧形式です。PowerPoint で開いて .pptx として保存し直してください",
    ".doc": "旧形式です。Word で開いて .docx として保存し直してください",
    # **圧縮したまま置かれた資料。** 客先からの受け渡しはこの形が普通で、
    # 展開し忘れると**中の 30 冊がまるごと 1 行の申告になる。**
    ".zip": "圧縮ファイルです。展開してから sources/ に置いてください"
            "（中の資料はまだ 1 冊も読んでいません）",
    ".7z": "圧縮ファイルです。展開してから sources/ に置いてください"
           "（中の資料はまだ 1 冊も読んでいません）",
    ".lzh": "圧縮ファイルです。展開してから sources/ に置いてください"
            "（中の資料はまだ 1 冊も読んでいません）",
}

#: 開けなかったときに、**先頭の数バイトから言えること**。例外の文言
#: （``File is not a zip file``）は「壊れている」としか言っておらず、**次に何を
#: すればいいかが 1 つも書いていない** ―― 拡張子を付け替えただけの PDF と、
#: パスワードで保護されたブックと、途中で切れた添付では、やることが 3 つとも違う。
#:
#: 読むのは先頭のバイト列そのものなので、これは推測ではなく転記である。
_MAGIC: tuple[tuple[bytes, str], ...] = (
    (b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1",
     "中身は OLE 複合ドキュメントです。旧形式（.xls / .doc）か、"
     "パスワードで保護されたブックのどちらかです。Excel で開き、"
     "保護を外して .xlsx として保存し直してください"),
    (b"%PDF", "中身は PDF です（拡張子だけが .xlsx になっています）。"
              "PDF は読めるので、拡張子を .pdf に直してから"
              "もう一度 arp4 parse に渡してください"),
    (b"\x89PNG", "中身は PNG 画像です（拡張子だけが .xlsx になっています）"),
    (b"{\\rt", "中身は RTF です（拡張子だけが .xlsx になっています）"),
)

#: 表とみなす最小の広がり。1x1 はタイトルや注記なのでテキストへ回す。
_MIN_TABLE = 2

#: 連結とみなす距離（空白 1 行／列までは同じ塊）。
_GAP = 2

#: 表の体裁で出すのをやめる**すかすか具合**。枠の面積が実セル数のこの倍数を
#: 超えたら（＝ 8 分の 7 以上が空欄なら）番地付きの箇条書きへ回す。
#:
#: **塊の中の空きを詰めない**という約束（:func:`_grid`）は、たかだか空行・空列
#: 1 本ぶんの膨らみを想定していた。工程表（ガントチャート）はその想定の外にある
#: ―― 斜めに並んだ ``■`` は空白 1 マスずつで繋がるので 1 つの塊になり、
#: **400 行 × 400 列の枠に値が 800 個**という表が出る。中身は 800 セルなのに
#: Markdown は 485KB で、その 99% は ``|  |`` である。トークンを減らすための
#: 区切り（規律の 4）が、逆にいちばん大きなファイルを作っていた。
#:
#: 箇条書きへ回しても**失うものは無い** ―― どのセルにも番地が付くので、
#: 表の格子から位置を数えるより精確である（区切りの前提そのもの）。
_SPARSE = 8

#: 上のすかすか判定を掛ける最小の面積。小さい枠は空欄が多くても表のままにする
#: ―― 3 行 3 列に値が 3 つは**書きかけの表**であって工程表ではない。
_SPARSE_AREA = 100


@dataclass(frozen=True)
class Media:
    """ブックから取り出した画像 1 枚。**中身は入っていたそのまま**（変換しない）。

    機械が読めないものを「読めません」と申告するだけで終わっていたころ、
    貼り付け画像に描かれた業務フローは**誰の手にも渡らなかった** ―― 申告の宛先
    （整理層）が画像を開けるのに、開ける場所へ出していなかったからである。

    名前は 2 か所に出す。**パース結果の中**（``s<番号>-i1`` の Markdown リンク）と
    **``sources.yml``** で、前者は読むため、後者は撮った版を突き合わせるためにある。

    **中の文字は :mod:`arp4.ocr` が読む。** 絵は絵のまま渡すが、絵の中の字まで
    渡さないでいると、文字しか無い画像（表・画面・帳票を撮ったもの）まで
    整理層が 1 枚ずつ開くことになる。
    """

    #: 書き出す名前（``<シート>-p1.jpeg``）。:mod:`arp4.render` の ``-1.png`` と
    #: 混ざらないよう ``p`` を挟む ―― 隣り合って置かれるので、**撮った絵**と
    #: **貼ってあった画像**が同じ名前の並びになると出自が分からなくなる。
    name: str
    body: bytes
    #: この画像から**機械が読んだ字**（``None`` は読みにいかなかったとき）。
    #: 画像そのものとは別に持つ ―― バイト列は資料の写しだが、こちらは
    #: **機械の読み**であって資料ではない。
    reading: ocr.Reading | None = None


@dataclass
class Target:
    """書き出し 1 件。**まだ書いていない。**"""

    path: Path                     # 書き出し先
    doc: mdio.Doc
    exists: bool = False
    dirty: bool = False            # git 上で編集されているか
    #: 元にした原本。**1 冊が何枚ものパース結果になる**ので、書き出し先からは
    #: 逆に辿れない（``受注.xlsx`` → ``受注/一覧.md`` ほか 8 枚）。指紋を残すのは
    #: 原本 1 件につき 1 つなので、ここで持って :func:`record` へ渡す。
    origin: Path | None = None
    #: このパース結果の隣へ出す画像 ``(書き出し先, 中身)``。
    images: list[tuple[Path, bytes]] = field(default_factory=list)

    @property
    def needs_confirm(self) -> bool:
        """**未編集のものは黙って上書きしてよい。** 全部を同じ重さで聞くと、
        確認そのものが読み飛ばされる。"""
        return self.exists and self.dirty


# ── 入口 ────────────────────────────────────────────────────────
def plan(round_: Round, sources: Iterable[Path], base: Path,
         exclude: Iterable[str] | None = None, use_ocr: bool = True
         ) -> tuple[list[Target], list[Finding]]:
    """資料を読んで**書き出し計画**を返す。書き込みは :func:`write`。

    分けてあるのは、上書き確認を CLI（人と話せる層）に任せるためである。

    ``exclude`` は**資料でないもの**を名指しで外すためにある ―― ``tests/`` を
    丸ごと渡すと、パース結果の正解（``dataset/正解/*.md``）やフィクスチャを
    資料として拾う。除外は黙って行わない（何件飛ばしたかを ``P014`` で言う。
    1 件にも当たらないパターンは打ち間違いの疑いなので、それも言う）。

    ``use_ocr`` を落とすと**貼り付け画像の中の字を読まない**（既定は読む）。
    切れるようにしてあるのは 2 つの理由からで、どちらも「読まないほうがよい」
    ではない ―― 資料が数百冊あって画像だけで何分も掛かるとき、そして
    **環境の違うマシンで同じパース結果を出したい**ときである（読める字は
    入っている言語パックで変わる）。落としたことは ``o1`` に必ず書く。
    """
    targets: list[Target] = []
    findings: list[Finding] = []
    dirty = _dirty_paths(round_.parsed)
    missed: list[tuple[Path, str]] = []

    expanded = sorted(expand(sources, missed))
    expanded, excluded = _excluded(expanded, base, list(exclude or []))
    findings += excluded

    for path in expanded:
        if path.suffix.lower() not in SUPPORTED:
            advice = _ADVICE.get(path.suffix.lower())
            findings.append(Finding(
                "warn", "P001", path.name,
                f"読めない形式です: {path.suffix}"
                + (f"（{advice}）" if advice else
                   f"（読めるのは {'、'.join(SUPPORTED)} です）")))
            continue
        try:
            docs, said, media = _parse_one(path, base, use_ocr)
        except Exception as exc:                       # 壊れたファイルで止めない
            findings.append(Finding("error", "P010", path.name,
                                    f"読み込みに失敗しました: {exc}"
                                    f"{_why(path, exc)}"))
            continue
        findings += said
        for relative, doc in docs:
            out = round_.parsed / relative
            # **画像はパース結果の隣**（``images/`` の同じ枝）へ置く。並びを
            # 揃えてあるので、どの md のどのアンカーの画像かを人が突き合わせ
            # ずに済む（:attr:`arp4.paths.Round.images`）。
            where = round_.images / relative.parent
            targets.append(Target(path=out, doc=doc, exists=out.is_file(),
                                  dirty=out.is_file() and _edited(out, dirty),
                                  origin=path,
                                  images=[(where / one.name, one.body)
                                          for one in media.get(relative, [])]))

    findings += _missed_note(missed)
    findings += _ocr_note(targets, use_ocr)
    targets, clashes = _unique(targets, dirty)
    findings += [Finding("warn", "P002", was.name,
                         f"書き出し先が重なったので名前を変えました → {now.name}"
                         "（シート名の記号を落とすと別のシートが同じファイル名に"
                         "なる。そのまま上書きするとシートが 1 枚消える）")
                 for was, now in clashes]
    return targets, findings


def _excluded(paths: list[Path], base: Path, patterns: list[str]
              ) -> tuple[list[Path], list[Finding]]:
    """``--exclude`` を適用する。**黙って飛ばさない**（件数を必ず言う）。

    当て先は :func:`relative_path`（パース結果の置き場と同じ相対）とファイル名の
    両方である ―― ``正解/*.md`` でも ``*.golden.md`` でも書ける。
    """
    if not patterns:
        return paths, []
    from fnmatch import fnmatch

    kept: list[Path] = []
    hits: dict[str, int] = {pattern: 0 for pattern in patterns}
    for path in paths:
        relative = relative_path(path, base).as_posix()
        matched = [p for p in patterns if fnmatch(relative, p) or fnmatch(path.name, p)]
        if matched:
            hits[matched[0]] += 1
        else:
            kept.append(path)

    findings = []
    for pattern in patterns:
        if hits[pattern]:
            findings.append(Finding(
                "warn", "P014", pattern,
                f"--exclude により {hits[pattern]} ファイルを資料から外しました"))
        else:
            findings.append(Finding(
                "warn", "P014", pattern,
                "--exclude が 1 ファイルにも当たりませんでした"
                "（打ち間違いなら、外したつもりの資料が読まれています）"))
    return kept, findings


def _unique(targets: list[Target], dirty: set[Path] | None
            ) -> tuple[list[Target], list[tuple[Path, Path]]]:
    """書き出し先を一意にする。**衝突を黙って上書きしない。**

    シート名からファイル名にできない文字を落とす（:func:`safe_name`）ので、
    ``受注/一覧`` と ``受注:一覧`` は同じ ``受注_一覧.md`` になる。先に出たほうは
    **動かさない** ―― 既に整理結果がそのパスを指しているかもしれないので、
    付け替えていいのはあとから来たほうだけである。
    """
    taken: set[Path] = set()
    fixed: list[Target] = []
    clashes: list[tuple[Path, Path]] = []

    for target in targets:
        if target.path not in taken:
            taken.add(target.path)
            fixed.append(target)
            continue
        renamed, index = target.path, 2
        while renamed in taken:
            renamed = target.path.with_name(
                f"{target.path.stem}~{index}{target.path.suffix}")
            index += 1
        taken.add(renamed)
        clashes.append((target.path, renamed))
        fixed.append(_replace(target, path=renamed, exists=renamed.is_file(),
                              dirty=renamed.is_file() and _edited(renamed, dirty),
                              images=_restem(target, renamed.stem)))
    return fixed, clashes


def _restem(target: Target, stem: str) -> list[tuple[Path, bytes]]:
    """パース結果を改名したら、**画像も同じだけ改名する。**

    画像の名前はシート名から作る（``<シート>-p1.jpeg``）ので、``受注/一覧`` と
    ``受注:一覧`` は md と同じく画像でもぶつかる ―― md だけ ``~2`` を付けて
    画像を放っておくと、**片方のシートの画像がもう片方に上書きされる**。
    md は名前を変えたと申告するので気づけるが、画像は黙って消える。
    """
    was = target.path.stem
    return [(path.with_name(f"{stem}{path.name[len(was):]}"), body)
            for path, body in target.images]


def write(targets: Iterable[Target]) -> tuple[list[Path], list[Finding]]:
    """計画どおりに書く。**確認は済んでいる前提。**

    **1 本書けなくても残りを書く。** :func:`plan` は「1 冊が壊れていて 29 冊が
    落ちる」を避けるために例外を握っているのに、書き出しのほうは 1 本目で
    止まっていた ―― 途中まで書いたパース結果だけが残り、**どこまで書けたのかは
    誰にも分からない**。読めた資料が書けずに消えるのは、読めなかったものが
    消えるのと同じである。
    """
    written: list[Path] = []
    findings: list[Finding] = []
    for target in targets:
        try:
            written.append(mdio.write(target.path, target.doc))
        except OSError as exc:
            findings.append(Finding("error", "P011", target.path.name,
                                    f"書き出しに失敗しました: {exc}"
                                    f"{_write_advice(target.path)}"))
            continue
        # **画像が書けなくてもパース結果は残す。** 出せなかったぶんは名前を挙げて
        # 申告する ―― 黙ると、パース結果の中のリンクだけが**実体の無い先**を
        # 指したまま残り、読み手は「開けない画像」を自分の環境のせいだと読む。
        for image, body in target.images:
            try:
                image.parent.mkdir(parents=True, exist_ok=True)
                image.write_bytes(body)
            except OSError as exc:
                findings.append(Finding("warn", "P015", image.name,
                                        f"画像の書き出しに失敗しました: {exc}"
                                        f"{_write_advice(image)}"))
    return written, findings


# ── 原本の指紋 ──────────────────────────────────────────────────
def fingerprint(path: Path) -> str:
    """原本 1 件の指紋。**バイト列で取る。**

    :func:`arp4.freeze.digest` は文字列を取るが、原本は Excel（zip）のことが
    あるので decode を挟めない。桁を揃えてあるので読み手は同じものとして扱える。
    """
    return hashlib.sha256(path.read_bytes()).hexdigest()[:12]


def record(round_: Round, targets: Iterable[Target],
           written: Iterable[Path] | None = None) -> Path | None:
    """**撮った時点の原本**を ``sources.yml`` に残す。

    残す相手は**書けたパース結果の原本だけ**である。編集済みで上書きを見送った
    ものまで記録すると、ディスクの上には古いパース結果が残っているのに指紋だけが
    新しくなり、**ずれが消えたように見える** ―― 守ったつもりの編集が、いちばん
    黙って腐る形になる。

    既にある記録には**重ねる**（消さない）。``arp4 parse`` は資料の一部だけを
    撮り直す使い方をするので、置き換えにすると前に撮った 29 冊の指紋が消える。
    """
    keep = None if written is None else {Path(p) for p in written}
    entries: dict[str, dict[str, Any]] = {}
    for target in targets:
        if target.origin is None:
            continue
        if keep is not None and target.path not in keep:
            continue
        key = _under_root(round_, target.origin)
        try:
            entry = entries.setdefault(
                key, {"digest": fingerprint(target.origin),
                      "parsed": [], "images": []})
        except OSError:
            continue                       # 撮った直後に消えた。次の照合で言う
        name = target.path.relative_to(round_.parsed).as_posix()
        if name not in entry["parsed"]:
            entry["parsed"].append(name)
        # **取り出した画像も名前で残す。** パース結果と同じ理由である ―― 原本を
        # 撮り直したときに、前の版から出ていた画像がどれかが分からないと、
        # **消えた画像と、まだ出していない画像の区別が付かない。**
        for image, _ in target.images:
            where = image.relative_to(round_.images).as_posix()
            if where not in entry["images"]:
                entry["images"].append(where)
    if not entries:
        return None

    previous = yamlio.load(round_.prints) if round_.prints.is_file() else None
    files = dict((previous or {}).get("files") or {}) if isinstance(previous, dict) else {}
    for key, entry in entries.items():
        entry["parsed"] = sorted(entry["parsed"])
        # **画像の無い資料に空の欄を出さない。** 30 冊のうち画像があるのは
        # 数冊で、残り全部に `images: []` が並ぶと「出せなかった」と読める。
        if entry["images"]:
            entry["images"] = sorted(entry["images"])
        else:
            entry.pop("images")
        files[key] = entry
    yamlio.dump(round_.prints, {"files": dict(sorted(files.items()))})
    return round_.prints


def drifted(round_: Round) -> list[Finding]:
    """撮った時点の原本と、いま手元にある原本の差。

    **「原本を変えるな」ではない。** 資料はラウンドのあいだに当然改訂されるし、
    それを咎める立場に arp4 は無い。言えるのは「**この出典が撮られたときの原本と
    いまの原本は違う**」という事実だけで、それは次のラウンドを起こすかどうかの
    判断材料そのものである。

    :func:`arp4.freeze.verify` と**向きが対称**である ―― あちらは凍結後に
    下流（整理結果）が動いていないかを見る。上流が動いたときに何も言わないと、
    パース結果は古いまま、整理層は渡されたものを正しく整理し、**正本は静かに
    古い版の写しになる**。落ちるところが 1 つも無いので、気付くのは生成された
    設計書を人が読み直したときである。

    段は warn である。``build`` は落ちない（門の約束「通れば build は原理的に
    失敗しない」は守られている）ので、``G005`` / ``G018`` と同じ扱いにする。
    止めたいなら ``--strict``。
    """
    if not round_.prints.is_file():
        return []                          # 指紋を残す前に撮ったラウンド
    manifest = yamlio.load(round_.prints) or {}
    files = (manifest.get("files") or {}) if isinstance(manifest, dict) else {}

    findings: list[Finding] = []
    for key in sorted(files):
        entry = files[key] if isinstance(files[key], dict) else {}
        origin = round_.root / key
        # **次に何を撮り直すかまで言う。** 1 冊が何枚ものパース結果になるので、
        # 原本の名前だけでは「どれが古いのか」が読み手に分からない。
        stale = ("、".join(entry.get("parsed") or [])) or "(不明)"
        if not origin.is_file():
            findings.append(Finding(
                "warn", "G019", round_.name,
                f"撮った原本が見つかりません: {key}"
                f"（このラウンドの {stale} は確かめられません）", file=key))
            continue
        try:
            now = fingerprint(origin)
        except OSError as exc:
            findings.append(Finding("warn", "G019", round_.name,
                                    f"原本を読めません: {key}（{exc}）", file=key))
            continue
        if now != entry.get("digest"):
            findings.append(Finding(
                "warn", "G019", round_.name,
                f"撮ったあとで原本が変わっています: {key}"
                f"（{stale} は古い版の写しです。新しいラウンドを起こして"
                "撮り直すか、この版のままでよいなら arp4 parse で撮り直す）",
                file=key))
    return findings


def _under_root(round_: Round, path: Path) -> str:
    """プロジェクト根からの相対。**照合のときに開く先**なので 1 つに揃える。

    根の外にある資料（``--base`` で外を指した・別ドライブ）は絶対パスのまま
    残す ―― 相対にできないものを無理に相対にすると、照合が別のファイルを開く。
    """
    try:
        return path.resolve().relative_to(round_.root).as_posix()
    except ValueError:
        return path.resolve().as_posix()


#: Windows のパス長。**超えると `FileNotFoundError` になる**（「無い」と言われる
#: ので、権限やディスクを疑って時間を溶かす）。深いフォルダ ＋ 長いブック名 ＋
#: 長いシート名で普通に届く ―― 資料の側は 1 つも間違っていない。
_MAX_PATH = 260


def _write_advice(path: Path) -> str:
    if len(str(path)) >= _MAX_PATH - 20:
        return (f"（パスが {len(str(path))} 文字あります。Windows は 260 文字で"
                "頭打ちになるので、プロジェクトを浅い場所へ移すか、"
                "長いパスを有効にしてください）")
    return "（フォルダの権限・同名のフォルダ・ディスクの空きを見てください）"


#: **資料置き場ではないもの。** 客先からもらったフォルダをそのまま ``sources/``
#: へ置くと、`.git` だけで数万ファイルある ―― 1 つずつ「読めない形式です」
#: （``P001``）と言えば、**本当に読めなかった 1 冊がその山に埋もれる**。
#: 隠しフォルダを資料として置く書き方は無いので、名前で切ってよい。
_NOT_SOURCES = {"node_modules", "__pycache__", "$RECYCLE.BIN"}


#: :func:`expand` が「開けなかった」と言うときの内訳。**次にやることが 3 つとも
#: 違う**ので、同じ文句にまとめない ―― 同期を待つ／パスを直す／権限を直す。
_NO_TARGET = "先が無い"
_NO_PATH = "名前ごと無い"
_CANNOT_LIST = "一覧できない"


def expand(sources: Iterable[Path],
           missed: list[tuple[Path, str]] | None = None) -> list[Path]:
    """フォルダを資料の一覧へ。**一時ファイル（``~$``）は資料ではない。**

    ``missed`` を渡すと、**名前はあるのに開けなかったもの**をそこへ入れる
    （:func:`_missed_note`）。ここは長く ``is_file()`` でも ``is_dir()`` でも
    ないものを黙って落としていた ―― どちらも偽になるのは、**リンクの先が
    消えている**ときと、OneDrive の「オンデマンド」が実体を持っていないとき、
    そして**打ったパスが間違っている**ときである。落とすと 0 冊で正常終了する
    ので、資料を渡したつもりの人には**中身が空だった**ようにしか見えない。
    """
    found: list[Path] = []
    for source in sources:
        if source.is_dir():
            found += _walk(source, missed)
        elif source.is_file():
            found.append(source)
        elif missed is not None:
            missed.append((source, _NO_TARGET if os.path.lexists(source)
                           else _NO_PATH))
    return found


#: 開けなかったときに言うこと。**3 つとも次にやることが違う**ので分けてある。
_MISSED: dict[str, str] = {
    _NO_TARGET:
        "リンクはありますが先が見つかりません。切れたショートカット"
        "（シンボリックリンク）か、クラウド同期が実体をまだ落としてきていない"
        "ファイルです。同期を待つか、実体をもらい直してください。",
    _NO_PATH:
        "そのパスにはありません。打ち間違いか、渡したつもりのフォルダが"
        "空です。読んだ資料が 0 冊でも parse は正常終了するので、ここで言わないと"
        "中身が空だったように見えます。",
    _CANNOT_LIST:
        "フォルダの中を一覧できません（権限・ロック・同期待ち）。"
        "この下にある資料は 1 冊も読んでいません。フォルダの権限を"
        "確かめるか、開いているプログラムを閉じてから回し直してください。",
}


def _missed_note(missed: list[tuple[Path, str]]) -> list[Finding]:
    """**名前はあるのに開けなかったもの。** 黙ると 0 冊で正常終了する。

    3 つに分けるのは、**次にやることが 3 つとも違う**からである ―― 同期を待つ、
    パスを直す、権限を直す。どれであるかはファイルシステムに書いてあるので、
    これは転記である（開けない資料の先頭バイトを読むのと同じ理屈）。
    """
    return [Finding("warn", "P007", path.name, _MISSED[kind])
            for path, kind in sorted(missed)]


def _ocr_note(targets: list[Target], use_ocr: bool) -> list[Finding]:
    """**画像の中の字を誰も読まなかったこと**を、走らせた人にも言う。

    パース結果の ``o1`` には理由が入るが、それを読むのは整理層である ――
    足りないのは**言語パック 1 つ**というような、いま手元で直せる話が、
    直せる人の目に触れないまま終わる。

    **1 度だけ出す。** 環境の話なので、資料 30 冊ぶん同じ行を並べても
    分かることは増えない（ブック単位の申告と役割が違う）。
    """
    if not use_ocr or not any(target.images for target in targets):
        return []
    why = ocr.trouble()
    if not why:
        return []
    return [Finding("warn", "P016", "Windows OCR",
                    f"貼り付け画像の中の文字を読めませんでした（{why}）。"
                    "画像の実体は出してあるので、整理層が開けば読めます ―― "
                    "ただし平坦な画像（表・画面・帳票を撮ったもの）まで"
                    "1 枚ずつ開くことになります。Windows で言語パックを"
                    "入れると次のラウンドから読めます（設定 > 時刻と言語 > "
                    "言語と地域 > 言語のオプション > 基本的なタイピング）。")]


def _walk(root: Path, missed: list[Path] | None = None) -> list[Path]:
    """フォルダの下のファイル。**同じフォルダを二度歩かない。**

    ``rglob`` は接合点（ジャンクション）を素直に辿るので、自分の親を指す
    リンクが 1 つあると**同じ資料が何度も出てくる** ―― 実測では 1 つの
    ``a.txt`` が 64 本になった（止まったのは Windows のパス長 260 文字で
    切れたからで、リンクを見分けたからではない）。同じ資料が何度もパースされ、
    書き出し先が重なって ``P002`` の山になる。

    リンクそのものは飛ばさない。**別のところを指すリンクの先には資料がある**
    ので、まだ歩いていない実体なら歩く ―― 落とすのは「もう歩いた実体」だけ
    なので、ここで消える資料は無い。
    """
    found: list[Path] = []
    seen = {Path(os.path.realpath(root))}
    stack = [root]
    while stack:
        where = stack.pop()
        try:
            entries = sorted(where.iterdir())
        except OSError:
            # **一覧できないフォルダ。** ここには長く「既知の穴」とだけ書いて
            # あった ―― 権限の付け替え漏れ、まだ落ちてきていない同期フォルダ、
            # 使用中のロックでこうなる。中の資料は 1 冊残らず消えるのに、
            # **どのフォルダが消えたのかを 1 行も言っていなかった。**
            if missed is not None:
                missed.append((where, _CANNOT_LIST))
            continue
        for entry in entries:
            if not _is_source(entry.relative_to(root)):
                continue
            if entry.is_file():
                found.append(entry)
            elif entry.is_dir():
                real = Path(os.path.realpath(entry))
                if real in seen:
                    continue                       # 歩いた実体（＝ 循環）
                seen.add(real)
                stack.append(entry)
            elif missed is not None:
                # **一覧には出るのに開けない。** リンクの先が消えている・
                # クラウド同期が実体を持っていない ―― 黙って落とすと、
                # 資料が 1 冊減ったことが誰にも伝わらない。
                missed.append((entry, _NO_TARGET))
    return found


def _is_source(relative: Path) -> bool:
    """**機械が置いたものを資料として数えない。** 判定は元の置き場からの相対で
    行う ―― 絶対パスで見ると、プロジェクトが `.venv` の下にあるだけで
    資料が 1 冊も見えなくなる。"""
    return not any(part.startswith("~$") or part.startswith(".")
                   or part in _NOT_SOURCES for part in relative.parts)


def relative_path(path: Path, base: Path) -> Path:
    """元のフォルダ構造をそのまま写す。**一意性はパスで担保する。**"""
    try:
        return path.resolve().relative_to(base.resolve())
    except ValueError:
        return Path(path.name)


def _edited(path: Path, dirty: set[Path] | None) -> bool:
    """git 上で編集されているか。**``None`` は「git が使えない」**の意味である。

    そのときは全部を「編集あり」に倒す ―― 分からないまま黙って上書きするより、
    確認が 1 回多いほうがましである。
    """
    return True if dirty is None else path.resolve() in dirty


def _dirty_paths(root: Path) -> set[Path] | None:
    """``root`` の下で git が「変わっている」と言うパス。**1 度で全部聞く。**

    以前は 1 ファイルにつき ``git status`` を 1 回起動していた。1 件 23ms は
    どうということのない数だが、**シート 1 枚がファイル 1 本**なので 30 冊
    201 シートの再実行では 201 プロセス ―― 実測 4.6 秒で、パース本体
    （1.3 秒）より長い。まとめて 1 回なら 0.02 秒である。

    ここは「遅いから通さない」が始まる場所である（使用範囲を端から端まで
    回して 1 シート 7.5 秒かかっていたのと同じ形）―― **性能の問題が
    網羅性の問題になる。**

    ``None`` を返すのは git が使えないときだけで、**空集合とは意味が違う**
    （空集合は「聞けたが 1 件も編集されていない」）。
    """
    if not root.exists():
        return set()                               # 最初のラウンド（まだ何も無い）
    try:
        top = subprocess.run(["git", "rev-parse", "--show-toplevel"],
                             capture_output=True, text=True, cwd=root, timeout=10)
        if top.returncode != 0:
            return None
        # ``-z`` は NUL 区切りで**引用も退避もしない** ―― 日本語のファイル名が
        # 普通に並ぶこの置き場では、既定の引用形式だと自前で戻す羽目になる。
        # ``--untracked-files=all`` はフォルダにまとめさせないため（まとめられると
        # 書いたばかりのファイルが 1 件も出てこない）。
        status = subprocess.run(
            ["git", "status", "--porcelain=v1", "-z", "--no-renames",
             "--untracked-files=all", "--", str(root)],
            capture_output=True, cwd=root, timeout=60)
    except (OSError, subprocess.SubprocessError):
        return None
    if status.returncode != 0:
        return None
    try:
        text = status.stdout.decode("utf-8")
    except UnicodeDecodeError:
        return None                                # 読めないなら安全側（全部確認）
    base = Path(top.stdout.strip())
    return {(base / entry[3:]).resolve() for entry in text.split("\0") if entry}


def _why(path: Path, exc: Exception) -> str:
    """読めなかった理由の続き。**形式ごとに別のことを言う。**

    ここは長く :func:`_unopenable` を無条件に足していた。あそこに書いてあるのは
    **Excel の話だけ**なので、構文の通らないソース（``print "…"`` ―― 棚卸しの
    対象は 10 年もののコードなので普通に出てくる）に
    ``（zip として開けません ―― 添付が途中で切れたか、別形式に拡張子だけを
    付け替えた資料です）`` と言っていた ―― **メールを探しに行かせて、そこに
    資料は無い。** `veryHidden` に「再表示してください」と言ったのと同じ失敗で、
    案内どおりにやって届かない申告は、申告しないのと同じである。

    しかもソースの側は**壊れていない** ―― 原本はテキストなので整理層が直接
    読める。取れていないのは骨格（クラス・メソッドの一覧）だけである。
    """
    suffix = path.suffix.lower()
    if suffix in _EXCEL:
        return _unopenable(path)
    if suffix == ".py":
        return _unskeletal(exc)
    if suffix == ".pdf":
        return _unreadable_pdf(exc)
    return ""


def _unreadable_pdf(exc: Exception) -> str:
    """**PDF が開けない理由は 2 つで、やることが正反対である。**

    パスワードが掛かっているなら外してもらい直す（資料は壊れていない）。
    途中で切れているなら受け渡しをやり直す ―― どちらか分からないまま
    「読み込みに失敗しました」だけを出すと、その 1 冊は拾い直されない。
    """
    said = str(exc).lower()
    if "password" in said or "encrypt" in said:
        return ("（パスワードで保護されています。資料は壊れていません ―― "
                "保護を外して保存し直したものをもらい直してください）")
    return ("（PDF として開けません。受け渡しの途中で切れたか、"
            "拡張子だけを付け替えた資料です。先頭が `%PDF` で"
            "始まっているかを確かめてください）")


def _unskeletal(exc: Exception) -> str:
    """**ソースは読めているが、骨格が取れない。** 次にやることが Excel と違う。

    Excel は開けなければ中身が 1 つも見えないので「保存し直す・もらい直す」に
    なるが、コードは**元ファイルがテキストである** ―― 骨格を出せなくても、
    整理層は原本をそのまま読める（:mod:`arp4.parse` の冒頭に書いてあるとおり、
    意図の層はもともとそうやって読ませている）。資料を取り直す話にしないこと
    のほうが大事である。
    """
    if isinstance(exc, SyntaxError):
        return ("（Python として構文が通りません。Python 2 のコード"
                "（`print \"…\"`）がこの形になります。資料は壊れていません。"
                "原本はテキストなので整理層が直接読めます。"
                "取れていないのは骨格（クラス・メソッドの一覧）だけです）")
    if isinstance(exc, UnicodeDecodeError):
        return ("（UTF-8 でも cp932 でも読めません。EUC-JP・UTF-16 で"
                "保存されたソースがこうなります。先頭に "
                "`# -*- coding: euc-jp -*-` を足すか、UTF-8 で保存し直して"
                "ください）")
    return ""


def _unopenable(path: Path) -> str:
    """**開けなかった理由は先頭の数バイトに書いてある。**

    ここを黙っていた頃の申告は ``読み込みに失敗しました: File is not a zip file``
    だけだった ―― 事実ではあるが、**その 1 冊を拾い直す手が 1 つも書いていない**。
    「読めません」で終わる申告は、資料が 1 冊落ちたまま誰も拾い直さない。

    読むのは magic bytes の転記であって、中身の判断ではない。
    """
    try:
        head = path.open("rb").read(8)
    except PermissionError:
        # **中身を見る前に開けない。** 客先から付いてきた権限、開いたままの
        # Excel、まだ落ちてきていない同期ファイルがこの形になる ―― ここを
        # 黙ると `[Errno 13] Permission denied` だけが出て、資料が壊れて
        # いるのだと思われる（**資料の側は 1 つも間違っていない**）。
        return ("（ファイルを開く権限がありません。開いたままの"
                "プログラムを閉じるか、フォルダの権限を確かめてください。"
                "資料そのものは壊れていません）")
    except OSError:
        return ""
    if not head:
        return "（0 バイトです。コピーか添付の取り出しが途中で終わっています）"
    for magic, why in _MAGIC:
        if head.startswith(magic):
            return f"（{why}）"
    if not head.startswith(b"PK"):
        return ("（zip として開けません。添付が途中で切れたか、"
                "別形式に拡張子だけを付け替えた資料です）")
    try:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
    except (OSError, zipfile.BadZipFile):
        return ("（zip の途中で切れています。メールの添付を"
                "取り出し直してください）")
    if "[Content_Types].xml" not in names:
        return ("（zip ですが Excel ブックの中身がありません。"
                ".ods / .docx を .xlsx と名付けた資料がこうなります。"
                "Excel で開いて .xlsx として保存し直してください）")
    return ""


def _parse_one(path: Path, base: Path, use_ocr: bool = True
               ) -> tuple[list[tuple[Path, mdio.Doc]], list[Finding],
                          dict[Path, list[Media]]]:
    """**画像を持ちうるのは Excel だけ**なので、ほかの形式は空の第 3 要素を返す。"""
    relative = relative_path(path, base)
    if path.suffix.lower() in _EXCEL:
        # 空なら `_excel` が P009 を出す
        return _excel(path, relative, use_ocr)
    if path.suffix.lower() in _SLIDES:
        # **読み込むのは使うときだけ**（:mod:`arp4.pptx` はこちらを import する）。
        from arp4 import pptx as pptx_module

        return pptx_module.read(path, relative, use_ocr)
    if path.suffix.lower() == ".pdf":
        from arp4 import pdf as pdf_module

        return pdf_module.read(path, relative, use_ocr)
    if path.suffix.lower() in _WORDS:
        from arp4 import docx as docx_module

        return docx_module.read(path, relative, use_ocr)
    if path.suffix.lower() in _CSV:
        made, said = _csv(path, relative)
        made, empty = _empty_note(made, path)
        return made, said + empty, {}
    if path.suffix.lower() in _TEXT:
        return (*_empty_note(_markdown(path, relative), path), {})
    if path.suffix.lower() in _SQL:
        return (*_empty_note(_ddl(path, relative), path), {})
    if path.suffix.lower() == ".java":
        return (*_empty_note(_java(path, relative), path), {})
    return (*_empty_note(_python(path, relative), path), {})


def _empty_note(made: list[tuple[Path, mdio.Doc]], path: Path
                ) -> tuple[list[tuple[Path, mdio.Doc]], list[Finding]]:
    """**1 行も出なかったら言う。** 数えて申告するのはブックだけではない。

    Excel は `P009` で申告していたのに、コード・DDL・Markdown は黙って `[]` を
    返していた ―― `sources/` に 23 本入れて `parsed/` が 22 本でも、**差の 1 本が
    どれかを言うものが無い**。読めなかったものほど静かに消えるのは、arp4 が
    いちばん避けたい壊れ方である。

    警告に留めるのは、**空で正しいことがある**ため（`py.typed` のような目印の
    ファイル、見出しの無いメモ）。
    """
    if made:
        return made, []
    return made, [Finding("warn", "P009", path.name,
                          "パース結果が 1 本も出ませんでした（この 1 本は "
                          "`parsed/` に 1 行も残りません）。中身があるはずなら、"
                          "この形式から機械が拾える塊が無かったということです。"
                          "原本を開いて確かめてください。")]


# ── Excel ───────────────────────────────────────────────────────
#: 救出モードで読んだシートに必ず載せる申告。**取れなかったものを名指しする** ――
#: とくに結合の展開が効かないので、区分の列が 2 行目以降だけ空欄になる（画面では
#: 全行に掛かって見えている）。黙ると整理層には「区分の無い行」に見え、
#: **資料に無いのか機械が読めていないのかが混ざる**。
_RESCUE_NOTE = (
    "このブックは壊れたシートがあったため救出モードで読みました。"
    "このシートの値は取れていますが、縦結合の展開・非表示行と列の申告・"
    "セルのコメント・リンクの行き先・印刷したときだけ見えるもの（ヘッダ・"
    "フッタ）は取れていません。とくに縦結合が広がらないので、"
    "区分の列は 2 行目以降が空欄になります（画面では全行に掛かって"
    "見えているものです）。ここでの空欄は「資料に無い」とは読めません。"
    "Excel でも開けないことがあります（実物で確かめたところ、修復でも"
    "戻りませんでした）。確実なのは元の資料をもらい直すことです。")


def _open(path: Path) -> tuple[Any, bool, list[str]]:
    """ブックを開く。**1 枚が壊れていても残りを読む。**

    シート 1 枚の XML が壊れていると ``load_workbook`` は 1 冊まるごと投げる
    ―― 20 枚のうち 19 枚は読めるのに、`P010` で 1 冊ごと落ちていた。
    「1 冊が壊れていて 29 冊が落ちる」を避けているのと同じことが、**ブックの
    中では守られていなかった。**

    救出は ``read_only`` で読み直す。openpyxl はこの形だとシートを 1 枚ずつ
    遅延で読むので、**壊れた 1 枚だけが落ちる** ―― 代わりに結合・行列の寸法・
    コメント・リンク・印刷設定を持たない worksheet になるので、
    :data:`_RESCUE_NOTE` で何が取れていないかを必ず言う。

    **既定では使わない。** 救出モードを常用すると、取れているはずのものが
    静かに取れなくなる ―― 落ちたときにだけ、落ちたと言ったうえで使う。
    """
    from openpyxl import load_workbook          # 依存は使うときだけ読む

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        try:
            # **開けなかったときにファイルを掴んだままにしない。**
            # ``load_workbook`` は自分で開くので、シート 1 枚の XML で投げると
            # handle が閉じられずに残り、回収されるのは GC のときになる ――
            # そのとき出る ``unclosed file`` は、**次に読んだブックの言い分**
            # （`P013`）として出る。読めているブックに身に覚えのない申告が付き、
            # しかも**資料を 1 冊足すだけで付く先が変わる**ので追えない。
            with path.open("rb") as stream:
                return load_workbook(stream, data_only=True), False, _said(caught)
        except Exception:
            pass
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        return load_workbook(path, data_only=True, read_only=True), True, \
            _said(caught)


#: **図形は自前で読んでいる**ので、openpyxl の「Shapes and drawings will be
#: lost」は事実に反する。出したままにすると、いちばん気にしてほしい申告
#: （:func:`_shape_note`）が「どうせ失われる」で読み飛ばされる。
_MUTED = "DrawingML"


def _said(caught: list[Any]) -> list[str]:
    """openpyxl が言ったこと。**握り潰さず申告へ回す。**

    ここは長く素通りだった ―― 警告は stderr へ 1 行流れて消え、`findings` にも
    パース結果にも残らない。いちばん効くのは**日付として範囲外のシリアル**で、
    openpyxl はそのセルを ``#VALUE!`` にして警告する ―― こちらはそれを
    「エラー値（計算できていない）」として申告するので、**資料は壊れていないのに
    壊れていると言う**ことになっていた（申告のほうが嘘になる形である）。
    """
    return sorted({str(one.message) for one in caught
                   if _MUTED not in str(one.message)})


def _excel(path: Path, relative: Path, use_ocr: bool = True
           ) -> tuple[list[tuple[Path, mdio.Doc]], list[Finding],
                      dict[Path, list[Media]]]:
    """**シート 1 枚 = ファイル 1 本。** ブックのファイル名がディレクトリになる。"""
    from openpyxl.worksheet.worksheet import Worksheet

    workbook, rescued, said = _open(path)
    trouble: list[str] = []
    related = _related_many(path, _WANTED, trouble)
    drawings = _drawings(related)
    extras = book_extras(path)
    threads = threaded_comments(
        {title: [part.body for part in parts]
         for title, parts in related["threaded"].items()}, extras.people)
    facts = sheet_facts(path)
    # **画像の実体はブック 1 冊ぶんで 1 つ**である（同じ画像を 2 枚のシートが
    # 使い回すことがある）。パート名で引けるようにしておいて、シートごとに
    # 名前を付け直して出す。
    bodies = {part.name: part.body
              for parts in related["image"].values() for part in parts}
    # **中の字はブック 1 冊ぶんまとめて読む。** シートごとに読みにいくと、
    # 20 枚のブックで engine を 20 回起こすことになる（:func:`arp4.ocr.read`
    # は同じ実体を 1 度しか読まないので、貼り回された社章も 1 回で済む）。
    readings = _readings(drawings, bodies) if use_ocr else None
    made: list[tuple[Path, mdio.Doc]] = []
    media: dict[Path, list[Media]] = {}
    skipped: list[tuple[str, str]] = []
    broken: list[str] = []

    # **シートは並び順で数える**（`workbook.worksheets` ではない）。グラフシートを
    # 飛ばして数えると、同じブックを次のラウンドで読んだときにアンカーの番号が
    # ずれる ―― 整理結果が指す先が黙って別のシートになる。
    for index, title in enumerate(workbook.sheetnames, start=1):
        sheet = workbook[title]
        if sheet.sheet_state != "visible":
            skipped.append((title, sheet.sheet_state))
            continue
        if not rescued and not isinstance(sheet, Worksheet):
            # **グラフシート**（セルを持たないシート）。値は元データのシートに
            # あるが、**このシートが存在したことは伝える** ―― 黙って落とすと、
            # 「資料に無い」と「機械が読めていない」のどちらでもない
            # 「機械が見なかった」が 1 枚ぶん混ざる。
            drawing = drawings.get(title, Drawing())
            shots, chunk, said_ocr = _pictures(index, title, relative, drawing,
                                               bodies, readings)
            out, doc = _chartsheet(path, relative, index, title, drawing)
            if chunk is not None:
                doc.chunks.append(chunk)
                if said_ocr is not None:
                    doc.chunks.append(said_ocr)
                media[out] = shots
            made.append((out, doc))
            continue
        try:
            sheet_cells = _cells(sheet)
        except Exception:
            # **1 枚が読めなくても残りを読む。** 壊れているのがシート 1 枚の
            # ときに 1 冊まるごと落とすのは、1 冊が壊れているときに 30 冊を
            # 落とすのと同じである ―― 名前を挙げて申告し、次へ進む。
            broken.append(title)
            continue
        _merge_threads(sheet_cells, threads.get(sheet.title, []))
        _merge_struck(sheet_cells, facts.struck.get(sheet.title, {}))
        cells = sheet_cells.values
        drawing = drawings.get(sheet.title, Drawing())
        blanks = facts.blanks.get(sheet.title, [])
        setup = _print_setup(sheet)
        # **数式だけ・コメントだけのシートを丸ごと落とさない。** 計算結果が
        # 保存されていないと非空セルが 0 になり、シートが存在したことすら
        # 整理層に伝わらなくなる。
        #
        # **印刷設定は数に入れない。** ヘッダ・フッタはブック全体へまとめて
        # 掛けるものなので、それだけで 1 本出すと**作業用の白紙がぜんぶ
        # パース結果になる** ―― 資料が増えたのではなく、器が増えるだけである。
        if not (cells or drawing.total or blanks or drawing.unreadable
                or sheet_cells.comments or sheet_cells.links):
            continue

        doc = mdio.Doc(title=f"{path.name} / {sheet.title}",
                       source=sheet_source(relative, sheet.title))
        if rescued:
            # **申告はシートの中に置く。** ブック単位の findings はコンソールに
            # 1 度出るだけで、パース結果を読むときには手元に無い ―― 救出モードで
            # 空欄になったものを「資料に無い」と読ませないために、値の隣に置く。
            doc.notes.append(_RESCUE_NOTE)
        if drawing.total or drawing.unreadable:
            doc.notes.append(_shape_note(index, drawing, readings))
        if blanks:
            doc.notes.append(_formula_note(index, blanks))
        if sheet_cells.errors:
            doc.notes.append(_error_note(index, sheet_cells.errors))
        if sheet_cells.comments:
            doc.notes.append(_comment_note(index, sheet_cells.comments))
        if sheet_cells.links:
            doc.notes.append(_link_note(index, sheet_cells.links))
        if setup:
            doc.notes.append(_print_note(index, setup))
        if sheet_cells.hidden:
            doc.notes.append(_hidden_note(sheet_cells))
        if sheet_cells.struck:
            doc.notes.append(_struck_note(index, sheet_cells))
        if drawing.series:
            doc.notes.append(_chart_note(index, drawing))
        tables = texts = 0
        for region in _regions(cells):
            frame = _frame(cells, region)
            if not frame.height:
                continue
            wide = frame.height >= _MIN_TABLE and frame.width >= _MIN_TABLE
            if wide and not frame.sparse:
                tables += 1
                doc.chunks.append(mdio.Chunk(
                    anchor=f"s{index}-t{tables}", at=frame.at,
                    heading=f"表 {frame.at}", rows=_grid(cells, frame)))
                if frame.height >= _BIG or len(frame.addresses) >= _BIG:
                    doc.notes.append(_big_note(index, f"t{tables}", frame))
            else:
                texts += 1
                shape = (f"（{frame.height} 行 × {frame.width} 列に"
                         f" {len(frame.addresses)} セル）" if wide else "")
                doc.chunks.append(mdio.Chunk(
                    anchor=f"s{index}-x{texts}", at=frame.at,
                    heading=f"セル {frame.at}{shape}", cells=frame.addresses))
                if wide:
                    doc.notes.append(_sparse_note(index, texts, frame))
                elif len(frame.addresses) >= _BIG:
                    # **1 列に数千行**は「すかすか」に当たらない（枠に隙間が
                    # 無い）ので、上の申告では拾えない ―― 縦に長い一覧は
                    # 実物にごく普通にある。
                    doc.notes.append(_big_note(index, f"x{texts}", frame))

        if blanks:
            # **アンカーを出すのは、`未読取` を宣言する先が要る**からである
            # （図形と同じ理屈）。表の中では空欄にしか見えないので、ここに
            # 番地を並べておかないと「読めていない」と言う場所が無い。
            listed = _listed([(ref, "計算結果が保存されていません（空欄ではありません）")
                              for ref in blanks])
            doc.chunks.append(mdio.Chunk(
                anchor=f"s{index}-f1", at=f"数式 {len(blanks)} 個",
                heading="数式（計算結果が保存されていない）", cells=listed))

        if sheet_cells.errors:
            # **エラー値は空欄ではない。** 表にはそのまま出ている（画面にもそう
            # 見えている）が、**その欄の仕様は資料から読み取れていない** ――
            # 宣言する先として番地を並べる。
            doc.chunks.append(mdio.Chunk(
                anchor=f"s{index}-e1", at=f"エラー値 {len(sheet_cells.errors)} 個",
                heading="エラー値（計算できていない）",
                cells=_listed(sheet_cells.errors)))

        if sheet_cells.comments:
            # **コメントは表に出てこない。** セルの上に浮いているので、表として
            # 出すと番地が持てない。番地付きの箇条書きにする。
            doc.chunks.append(mdio.Chunk(
                anchor=f"s{index}-m1", at=f"コメント {len(sheet_cells.comments)} 件",
                heading="セルのコメント（メモ）", cells=sheet_cells.comments))

        if sheet_cells.links:
            doc.chunks.append(mdio.Chunk(
                anchor=f"s{index}-l1", at=f"リンク {len(sheet_cells.links)} 件",
                heading="リンク（表には表示文字列しか出ていない）",
                cells=sheet_cells.links))

        if setup:
            # **紙にしか出ないもの。** 文書番号・版・機密区分がフッタにしか
            # 無いのは日本の設計書ではごく普通で、表だけを読むと**その 1 冊が
            # 何の文書か**が落ちる。
            doc.chunks.append(mdio.Chunk(
                anchor=f"s{index}-p1", at=f"印刷設定 {len(setup)} 件",
                heading="印刷したときだけ見えるもの（ヘッダ・フッタ・繰り返す見出し）",
                cells=_listed(setup)))

        if sheet_cells.struck:
            # **画面では消してあるのに、表では生きた行と同じに見える。**
            # 番地を並べるのは、`廃止` と読むか読まないかを整理層が決められる
            # ようにするためである（機械は決めない）。
            doc.chunks.append(mdio.Chunk(
                anchor=f"s{index}-d1",
                at=f"取り消し線 {len(sheet_cells.struck)} セル",
                heading="取り消し線の掛かったセル（画面では消してある）",
                cells=_listed(sheet_cells.struck)))

        if drawing.series:
            # **グラフは別のアンカー**にする。図形の文字（`g1`）と混ぜると、
            # 「シートに描いてある文字」と「別のシートを指す案内」が同じ出典に
            # なる ―― 後者は読みに行く先であって、資料の中身ではない。
            doc.chunks.append(mdio.Chunk(
                anchor=f"s{index}-k1", at=f"グラフ {drawing.charts} 個",
                heading="グラフ（タイトル・系列・参照範囲）",
                rows=[["グラフ", "系列", "分類", "値"],
                      *[list(one) for one in drawing.series]]))

        if drawing.alts:
            # **代替テキストは画像の中身ではない**が、何の画像かは分かる。
            # **図形と群にも書ける** ―― 構成図に 1 文だけ添えられた
            # 「現用／待機の 2 系統。DR は別紙 5」は箱の文字にも表にも無く、
            # 画像からしか取っていなかった頃はそこだけが丸ごと落ちていた。
            doc.chunks.append(mdio.Chunk(
                anchor=f"s{index}-a1", at=f"代替テキスト {len(drawing.alts)} 件",
                heading="代替テキスト（人が書いた説明）",
                cells=drawing.alts))

        shots, pictures, said_ocr = _pictures(index, sheet.title, relative,
                                              drawing, bodies, readings)
        if pictures is not None:
            # **貼り付け画像は実体ごと出す。** ここは長く「中身は取れていません」
            # の 1 行だけだった ―― 機械が読めないのは本当だが、**整理層は読める**
            # のに、開ける場所へ出していなかった。名前を md に書いておくのは、
            # 出典として指す先が要るからである（アンカーは `i1`）。
            doc.chunks.append(pictures)
        if said_ocr is not None:
            # **読んだ字は絵とは別のアンカー**（`o1`）にする。同じ `i1` に混ぜると、
            # 「ブックに入っていた画像」と「機械がその画像から読んだ字」が同じ
            # 出典になる ―― 前者は資料の写しだが、後者は**必ず読み違えを含む**。
            doc.chunks.append(said_ocr)

        if drawing.total or drawing.unreadable:
            # **セルの塊とは別のアンカー**にする。図形は番地を持たないので、
            # 表の出典（B8:J20）と同じ枠に混ぜると出典が辿れなくなる。
            #
            # **テキストが 1 つも取れなくてもアンカーは出す。** 出さないと、
            # 図形だけのシート（＝いちばん読めていないシート）がアンカー 0 になり、
            # 未整理の一覧（G001）にも上がらず、`未読取` の宣言先も無くなる。
            # 読めていないものほど静かに消えるのでは、申告している意味がない。
            #
            # **XML として読めなかったパートも同じ扱いである。** 図形 0 個と
            # 見分けが付かないうえ、そちらは宣言先すら出ないので、いっそう静かに
            # 消える（:attr:`Drawing.unreadable`）。
            labels = drawing.labels
            # **画像しか無いシートに「絵にして読む」と書かない。** 撮り直しても
            # 絵が絵のまま出るだけで、案内どおりにやった人は空振りする ――
            # :func:`_shape_note` はそう言っているのに、ここだけが render を
            # 指したままだった。実体を出しているなら、行き先は `i1` である。
            empty = (f"テキストの入った図形はありません（{drawing.summary}）。"
                     + (f"画像の実体は `s{index}-i1` に出してあります"
                        if drawing.media else "絵にして読む → arp4 render"))
            doc.chunks.append(mdio.Chunk(
                anchor=f"s{index}-g1", at=drawing.summary,
                heading=("図形（テキスト）" if labels else "図形（テキストなし）"),
                cells=([(f"図形{i}", text) for i, text in enumerate(labels, start=1)]
                       or [("図形", empty)])))

        if drawing.links:
            # **接続はテキストとは別のアンカー**にする。箱の名前（何があるか）と
            # 線（どう繋がるか）は別の事実で、片方だけを出典にしたいことがある。
            doc.chunks.append(mdio.Chunk(
                anchor=f"s{index}-c1", at=f"接続 {len(drawing.links)} 本",
                heading="図形の接続", rows=_link_rows(drawing.links)))

        if doc.chunks or doc.notes:
            out = Path(*relative.parts) / f"{safe_name(sheet.title)}{mdio.EXT}"
            made.append((out, doc))
            if shots:
                media[out] = shots

    workbook.close()
    return made, (_skipped_note(path, skipped) + _external_note(path, extras)
                  + _macro_note(path, extras) + _properties_note(path, extras)
                  + _broken_note(path, rescued, broken, len(made))
                  + _drawing_gap_note(path, trouble)
                  + _nothing_note(path, workbook, made, skipped, broken)
                  + _warned_note(path, said)), media


def _drawing_gap_note(path: Path, trouble: list[str]) -> list[Finding]:
    """**辿れなかった関係は「図が無い」と見分けが付かない。**

    :func:`_related_many` は「読めない資料でパースそのものを止めない」ために
    例外を飲む ―― それは正しいが、**飲んだことを言っていなかった**。
    ``xl/_rels/workbook.xml.rels`` が壊れているブック（帳票ツールの出力に
    ある）では図形もグラフも SmartArt も 1 つも見えなくなり、パース結果は
    **図の無いブック**として何食わぬ顔で出てくる。

    ここは「資料に無い」と「機械が読めていない」が混ざるいちばん静かな形で、
    しかも `未読取` を宣言する先（``g1``）すら出ない。
    """
    if not trouble:
        return []
    return [Finding("warn", "P008", path.name,
                    "このブックの図（図形・グラフ・SmartArt）は 1 つも"
                    "読めていません。zip の中の関係を辿れませんでした"
                    f"（{trouble[0]}）。パース結果は図の無いブックとして"
                    "出てきますが、資料に無いのではありません。"
                    "図が仕様なら Excel で開いて確かめるか、"
                    "`arp4 render` で絵にして読んでください。")]


def _nothing_note(path: Path, workbook: Any, made: list[Any],
                  skipped: list[tuple[str, str]], broken: list[str]
                  ) -> list[Finding]:
    """**1 本も出なかったブックは、置いていないブックと区別が付かない。**

    シート 0 枚のブック（ツールが書き出したもの）・白紙だけのブック・非表示
    しか無いブックがこうなる。1 枚ずつの規律（値も図もコメントも無いシートは
    出さない ―― でないと**作業用の白紙がぜんぶパース結果になる**）は変えない
    が、それを**ブック 1 冊に掛けたときの結果**は黙れない ―― `sources/` に
    確かに置いた 1 冊が、`parsed/` から丸ごと消える。

    内訳を添えるのは、次にやることがそれで決まるからである（非表示なら
    再表示、壊れているならもらい直し、白紙なら**そのままでよい**）。
    """
    if made:
        return []
    sheets = len(getattr(workbook, "sheetnames", []) or [])
    detail = [f"シート {sheets} 枚"]
    if skipped:
        detail.append(f"うち非表示 {len(skipped)} 枚")
    if broken:
        detail.append(f"うち読めなかった {len(broken)} 枚")
    return [Finding("warn", "P009", path.name,
                    f"パース結果が 1 本も出ませんでした（{'／'.join(detail)}）。"
                    "値も図もコメントも無いシートは出さない決まりなので、"
                    "表紙だけ・白紙だけのブックならこれで正しいのですが、"
                    "`parsed/` にはこの 1 冊が 1 行も残りません。"
                    "中身があるはずなら、非表示シート・壊れたシート・"
                    "貼り付け画像だけのシートのどれかです。Excel で開いて"
                    "確かめてください。")]


def _broken_note(path: Path, rescued: bool, broken: list[str],
                 read: int) -> list[Finding]:
    """**壊れていたのがシート 1 枚でも、黙って落とさない。**

    以前はここで 1 冊まるごと `P010` になっていた ―― 20 枚のうち 19 枚は
    読めるのに、**壊れた 1 枚のせいで 1 冊が誰の目にも触れない**。ブック単位の
    事実なので findings に出す（非表示シート ``P003`` と同じ）。

    **読めた枚数を必ず添える。** 「1 枚が壊れています」だけだと、残りが読めた
    のか道連れになったのかが分からない。

    **「Excel で開いて保存し直してください」とは言わない。** 実物で確かめた
    ところ（Excel 16.0）、シートのパートが途中で切れたブックは**修復モードでも
    開けなかった** ―― 案内どおりにやって届かない申告は、`veryHidden` に
    「再表示してください」と言ったのと同じ失敗である。
    """
    if not (rescued or broken):
        return []
    where = ("シート " + "、".join(broken[:_MAX_LIST]) if broken
             else "壊れているシートがあります")
    if len(broken) > _MAX_LIST:
        where += f" ほか {len(broken) - _MAX_LIST} 枚"
    return [Finding(
        "warn", "P012", path.name,
        f"{where} が読めませんでした（残り {read} 枚は読めています）。"
        "救出モードで読んだので、縦結合の展開・非表示行と列の申告・"
        "セルのコメント・リンクの行き先・印刷したときだけ見えるものは"
        "取れていません。パース結果の先頭にも同じことが書いてあります。"
        "Excel でも開けないことがあります（修復でも戻らない壊れ方です）"
        "。中身が要るなら元の資料をもらい直してください。")]


def _warned_note(path: Path, said: list[str]) -> list[Finding]:
    """openpyxl が言ったこと。**stderr へ流して消していたものを申告へ回す。**

    いちばん効くのは**日付として範囲外のシリアル**である ―― 納期の引き算を
    日付書式のまま持っているセルがこうなり、openpyxl はそれを ``#VALUE!`` に
    して警告する。こちらはそれを「エラー値（計算できていない）」として申告して
    いたので、**資料は壊れていないのに壊れていると言っていた**（読み落としより
    悪い ―― 申告のほうが嘘になる形である）。
    """
    if not said:
        return []
    return [Finding("warn", "P013", path.name,
                    "読み込み時に openpyxl が言ったこと: "
                    + " ／ ".join(said[:_MAX_LIST])
                    + (f" ほか {len(said) - _MAX_LIST} 件"
                       if len(said) > _MAX_LIST else "")
                    + "。表の上では見分けが付きません。日付として範囲外の"
                      "シリアルはエラー値（`#VALUE!`）として出ますが、"
                      "資料が壊れているのではなく、その数が日付として"
                      "読めないだけです。")]


def _external_note(path: Path, extras: BookExtras) -> list[Finding]:
    """**このブックは別のブックを読んでいる**ことの申告。

    ハイパーリンクを取るのと同じ理屈である ―― リンクは「人が辿る先」、外部参照は
    「数式が読んでいる先」で、どちらも**まだ手元に無い資料の一覧**になる。
    違うのは**表からいっそう見えない**ことで、参照先の値はキャッシュされている
    ので、画面にはただの数が出ている（``12,000`` が別ブック由来だとは誰も
    気付かない）。その資料が集まっていなければ、その数の根拠は次のラウンドでも
    確かめられない。

    ブック単位の事実なのでシートごとの ``notes`` ではなく findings に出す
    （非表示シート ``P003`` と同じ）。
    """
    if not extras.external:
        return []
    listed = "、".join(
        where + (f"（シート: {'・'.join(sheets)}）" if sheets else "")
        for where, sheets in extras.external[:_MAX_LIST])
    return [Finding("warn", "P004", path.name,
                    f"別のブック {len(extras.external)} 冊を参照しています"
                    f"（{listed}）。参照先の値はキャッシュされているので画面には"
                    "数が出ていますが、その数の根拠はこのブックにありません。"
                    "まだ集めていないなら sources/ に足して parse を回し直して"
                    "ください。")]


def _properties_note(path: Path, extras: BookExtras) -> list[Finding]:
    """**誰がいつ触ったかは、シートにも表にも出てこない。**

    改訂履歴のシートは**人が書いた申告**で、書き忘れれば何も残らない。ブックの
    プロパティは Excel が黙って書くので、**改訂履歴に無い更新**がここにだけ
    残る ―― 課題管理のコメントが本文より新しいのと同じ形である。

    **食い違っていても機械は何も言わない。** 最終更新日時が最新版の改訂日より
    後なら「載っていない更新がある」のか「開いて保存し直しただけ」なのかは
    資料からは決まらない ―― どちらかを当てるのは意味の判断である。並べるだけ
    にして、突き合わせは整理層に渡す。

    ブック単位の事実なのでシートごとの ``notes`` ではなく findings に出す
    （非表示シート ``P003``・外部参照 ``P004`` と同じ）。
    """
    if not extras.core:
        return []
    listed = "／".join(f"{label} {value}" for label, value in extras.core)
    return [Finding("warn", "P005", path.name,
                    f"ブックのプロパティ（{listed}）。シートにも表にも"
                    "出てきません（日時は UTC）。改訂履歴は人が書いた申告"
                    "なので、そこに無い更新がここにだけ残っていることが"
                    "あります。食い違っていても、どちらが本当かは資料からは"
                    "決まらないので機械は判断していません。")]


def _macro_note(path: Path, extras: BookExtras) -> list[Finding]:
    """**マクロが入っていることを、1 行も言っていなかった。**

    ``.xlsm`` を弾くのをやめたときに書いたのは「マクロが付いているだけで中身は
    同じである」だったが、**それは表の話でしかない。** 実案件の設計書に付いて
    いるマクロは飾りではなく、採番規則・入力チェック・帳票の組み立てがそこに
    しか無いことがある ―― シートに書いてあるのは「ボタンを押す」だけである。

    ここは**申告の規律が掛かっていなかった唯一の入口**だった。数式のキャッシュ
    無し・貼り付け画像・埋め込みオブジェクトは「読めていない」と言っているのに、
    マクロだけは読めていないことすら伝えていない ―― 読み手には**マクロの無い
    ブック**と同じに見える（いちばん静かな消え方である）。

    **中身は取らない。** ``vbaProject.bin`` は OLE 複合ドキュメントで、zip の
    中には入っていても zip としては開けない ―― 取るなら別の読み手が要る。
    ブック単位の事実なので findings に出す（非表示シート ``P003``・外部参照
    ``P004``・プロパティ ``P005`` と同じ）。
    """
    if not extras.macros:
        return []
    return [Finding("warn", "P006", path.name,
                    "マクロ（VBA）が入っています。中身は取っていません"
                    "（`xl/vbaProject.bin` は zip の中にありますが、"
                    "zip としては開けません）。採番規則・入力チェック・帳票の"
                    "組み立てがマクロにしか無いことがあり、そのとき"
                    "シートに書いてあるのは「ボタンを押す」だけです。"
                    "仕様が要るなら Excel の VBE（Alt+F11）で開いて読むか、"
                    "作成者に確認してください。")]


def _merge_threads(cells: Cells, threads: list[tuple[str, str]]) -> None:
    """スレッドコメントを旧形式のコメントへ**混ぜる**（同じ ``m1`` に出す）。

    同じ番地に両方あるときは**スレッド側を採る** ―― 旧形式のほうは古い Excel
    向けのなりすまし（:data:`_LEGACY_MARK`）で、中身は同じものが 200 字の
    但し書き付きで入っているだけである。混ぜずに 2 件出すと、レビュー指摘の
    件数が倍に見える。
    """
    if not threads:
        # スレッドのパートを落として配られた資料では、なりすましだけが残る。
        # 但し書きは剥がして本文を出す（読めるものを読めない形で出さない）。
        cells.comments = [(ref, _unwrap_legacy(text)) for ref, text in cells.comments]
        return
    taken = {ref for ref, _ in threads}
    kept = [(ref, text) for ref, text in cells.comments if ref not in taken]
    cells.comments = sorted(kept + threads, key=lambda one: _sortable(one[0]))


def _merge_struck(cells: Cells, runs: dict[str, tuple[str, bool]]) -> None:
    """セルの中の**一部だけ**に掛かった取り消し線を混ぜる（同じ ``d1`` に出す）。

    セル全体の書式（:func:`_struck`）と分けずに 1 つのアンカーへ出すのは、
    読み手にとって**どちらも「画面では消してある文字」**だからである ―― 書かれ方
    （セルの書式かリッチテキストか）の違いは、資料を書いた人の操作の違いでしかない。

    **消してある文字だけは書き分ける。** 一部なら残りは生きているので、値を
    そのまま並べると「この行はまるごと廃止」に読める ―― 取り消し線を取ると
    決めた理由（画面に見えているものへ寄せる）がそこで裏返る。
    """
    if not runs:
        return
    known = {ref for ref, _ in cells.struck}
    where = {f"{column_name(column)}{row}": (row, column) for row, column in cells.values}
    extra: list[tuple[str, str]] = []
    for ref, (marked, whole) in runs.items():
        if ref in known or ref not in where:
            continue                               # 既に書式から取れている
        value = cells.values[where[ref]]
        if whole:
            extra.append((ref, value))
        else:
            cells.partly_struck += 1
            extra.append((ref, f"{value}（消してあるのは「{marked}」だけです）"))
    cells.struck = sorted(cells.struck + extra, key=lambda one: _sortable(one[0]))


def _sortable(ref: str) -> tuple[int, int, str]:
    """``B10`` を並べ替えられる形に。**文字列順だと B10 が B2 より前に来る。**"""
    found = re.match(r"^([A-Z]+)(\d+)$", ref)
    if not found:
        return (0, 0, ref)
    column = 0
    for letter in found.group(1):
        column = column * 26 + (ord(letter) - ord("A") + 1)
    return (int(found.group(2)), column, ref)


def _skipped_note(path: Path, skipped: list[tuple[str, str]]) -> list[Finding]:
    """**非表示シートは読まないが、あったことは言う。**

    行・列の非表示は読んで申告する（:func:`_hidden_note`）のに、シートは読まない
    ―― 壊れ方が違うからである。ただし**黙って落とすところまで同じにはしない。**
    ブック単位の非表示は作業用が大半だが、「旧版」「Ver1.0」を隠しただけの
    ブックは実案件にごく普通にあり、そこに現行仕様が入っていることがある。
    シートが 1 枚も出てこなければ、誰もそれを疑わない。

    **隠し方は 2 通りある。** 右クリックの「再表示」で戻せる ``hidden`` と、
    そこに出てこない ``veryHidden`` である。同じ案内をしていたぶん、後者は
    「再表示してください」と言われて**メニューに無い**ところで止まっていた ――
    案内どおりにやって届かない申告は、読まれなくなるのと同じである。
    """
    if not skipped:
        return []
    deep = [title for title, state in skipped if state == "veryHidden"]
    message = (f"非表示のシート {len(skipped)} 枚"
               f"（{'、'.join(title for title, _ in skipped)}）は読んでいません。"
               "作業用が大半ですが、旧版を隠しただけのことがあります。"
               "中身が要るなら Excel で再表示してから arp4 parse を回し直して"
               "ください。")
    if deep:
        message += (f"うち {len(deep)} 枚（{'、'.join(deep)}）は右クリックの"
                    "「再表示」には出てきません（veryHidden）。VBE の"
                    "プロパティで Visible を戻すか、作成者に確認してください。"
                    "画面から隠す手間を掛けたシートなので、作業用とは"
                    "限りません。")
    return [Finding("warn", "P003", path.name, message)]


def _chartsheet(path: Path, relative: Path, index: int, title: str,
                drawing: Drawing) -> tuple[Path, mdio.Doc]:
    """グラフシート 1 枚。**セルは無いが、グラフは読める。**

    ここは長く「グラフの中身は取れていません」の 1 行だけを出していた。
    セルが 0 個なので**それ以外に何も書いていないファイル**になり、
    シートが存在したことは伝わっても、次に何をすればいいかは伝わらない
    ―― タイトルと参照範囲を出せば、**読みに行く先が分かる。**
    """
    doc = mdio.Doc(title=f"{path.name} / {title}",
                   source=sheet_source(relative, title))
    doc.notes.append(
        f"グラフシートです（セルを持たないシート）。グラフに描かれている数"
        f"そのものはここにありません。系列が指しているのは別のシートの値です。"
        f"図そのものが仕様なら、アンカー `s{index}-g1` に "
        "out_of_scope の kind: 未読取 を宣言してください。")
    if drawing.series:
        doc.notes.append(_chart_note(index, drawing))
    doc.chunks.append(mdio.Chunk(
        anchor=f"s{index}-g1", at="グラフシート",
        heading="グラフシート（セルなし）",
        cells=[("シート", "グラフだけのシートです。セルの値は 1 つもありません")]))
    if drawing.series:
        doc.chunks.append(mdio.Chunk(
            anchor=f"s{index}-k1", at=f"グラフ {drawing.charts} 個",
            heading="グラフ（タイトル・系列・参照範囲）",
            rows=[["グラフ", "系列", "分類", "値"],
                  *[list(one) for one in drawing.series]]))
    return Path(*relative.parts) / f"{safe_name(title)}{mdio.EXT}", doc


def _readings(drawings: dict[str, Drawing], bodies: dict[str, bytes]
              ) -> dict[str, ocr.Reading]:
    """ブック 1 冊ぶんの画像を**まとめて**読む（:mod:`arp4.ocr`）。

    **貼ってある画像だけを読む。** ``bodies`` にはブックの中の画像が全部入って
    いるが、どの図にも貼られていない実体（差し替えて消し忘れた絵）は資料の
    どこにも現れないので、読んでも出す先が無い。
    """
    wanted = {part for drawing in drawings.values() for part, _ in drawing.media}
    return ocr.read({part: body for part, body in bodies.items() if part in wanted})


def _pictures(index: int, title: str, relative: Path, drawing: Drawing,
              bodies: dict[str, bytes],
              readings: dict[str, ocr.Reading] | None = None,
              prefix: str = "s"
              ) -> tuple[list[Media], mdio.Chunk | None, mdio.Chunk | None]:
    """貼り付け画像を取り出して名前を付ける。``(出す画像, 画像の塊, 読んだ字の塊)``。

    **絵は絵のまま渡し、字は字として読む。** ここがやるのは「読める場所へ置いて、
    名前を出典として指せる形にする」ことと「機械が読める字を読んでおく」ことだけで、
    中身の判断はしない（パース層の規律はそのまま）。絵として読むのは整理層である。

    **同じ実体は 1 回だけ出す。** 1 枚の画像を 2 か所に貼ったシートは ``xdr:pic``
    が 2 個になるが、実体は 1 つである ―― 枚数ぶん書き出すと、同じバイト列が
    別々の名前で並び、読む側は**違う画像だと思って両方開く。**

    **実体に辿り着けない画像は数から落とす。** リンク画像（``r:link``）は
    ブックの中に実体を持たず、指しているのは資料を作った人の手元のパスである。

    ``readings`` が ``None`` なら**読みにいかなかった**（``--no-ocr``）。
    「読んで字が無かった」と同じ形にはしない ―― 前者は誰も見ていないという
    ことで、次にやることが正反対である。
    """
    if not drawing.media:
        return [], None, None
    stem = safe_name(title)
    named: dict[str, str] = {}
    shots: list[Media] = []
    listed: list[tuple[str, str]] = []
    for part, alt in drawing.media:
        body = bodies.get(part)
        if body is None:
            continue
        if part not in named:
            named[part] = f"{stem}-p{len(named) + 1}{Path(part).suffix.lower() or '.bin'}"
            shots.append(Media(name=named[part], body=body,
                               reading=(readings or {}).get(part)))
            listed.append((named[part], alt or "（代替テキストはありません）"))
    if not shots:
        return [], None, None

    # **リンクはパース結果からの相対**にする（:mod:`arp4.render` と同じ形）。
    # ``rounds/r001/`` の中で ``parsed/`` と ``images/`` が並んでいるので、
    # エディタからもエージェントからもそのまま開ける。
    up = "../" * (len(relative.parts) + 1)
    where = f"{up}images/{relative.as_posix()}"
    return shots, mdio.Chunk(
        anchor=f"{prefix}{index}-i1", at=f"画像 {len(shots)} 枚",
        heading="画像（ブックから取り出したファイル）",
        cells=listed,
        text="\n".join(f"![{one.name}]({where}/{one.name})" for one in shots)), \
        _read_chunk(index, shots, readings is not None, prefix)


#: 読んだ字を出すときの見出し。**「機械が読んだ」と毎回書く** ―― 出典として
#: 指された整理結果を人がレビューするとき、そこに見えているのはこの見出しである。
_OCR_HEADING = "画像の中の文字（Windows OCR が読んだもの。読み違えが混ざります）"

#: 読みにいかなかったとき（``--no-ocr``）。**空にしない** ―― 空の ``o1`` は
#: 「画像に文字が無かった」に見え、それは「資料に無い」と「機械が読めていない」の
#: 取り違えそのものである。
_OCR_OFF = ("読みにいっていません（`--no-ocr` が指定されました）。"
            "画像を開いて読むのは整理層の仕事です。")

#: 読んだが字が 1 つも出なかったとき。**画像は絵である**（図・写真・網点の
#: 掛かった絵）。次にやることは「開いて見る」で、`arp4 render` ではない。
_OCR_BLANK = ("文字は見つかりませんでした（図・写真・網点の掛かった絵は"
              "こうなります）。開いて見るのは整理層の仕事です。")


def _read_chunk(index: int, shots: list[Media], attempted: bool,
                prefix: str = "s") -> mdio.Chunk | None:
    """読んだ字を ``s<番号>-o1`` へ。**必ず 1 枚ずつ、画像の名前と対で出す。**

    まとめて 1 つの塊にしないのは、読んだ字が**どの画像から出たか**を整理層が
    辿れるようにするためである（3 枚貼ってあるシートで、字が 1 枚からしか
    出ていないことは、それ自体が「残り 2 枚は絵である」という事実になる）。

    字は**インデントした塊**（4 字下げ）で出す。OCR の行には ``|`` も
    ``#`` も ``` も普通に混ざるので、表や引用に組み直すと**読めた字のほうが
    壊れる** ―― ここは原文をそのまま置く場所である。
    """
    if not shots:
        return None
    body: list[str] = []
    for one in shots:
        body.append(f"`{one.name}`")
        body.append("")
        if not attempted:
            said = _OCR_OFF
        elif one.reading is None:
            said = "読みにいっていません。"
        elif one.reading.trouble:
            said = f"読めませんでした（{one.reading.trouble}）。"
        elif not one.reading.lines:
            said = _OCR_BLANK
        else:
            said = one.reading.text
        body += [f"    {line}" for line in said.splitlines()]
        body.append("")
    return mdio.Chunk(anchor=f"{prefix}{index}-o1", at=f"画像 {len(shots)} 枚",
                      heading=_OCR_HEADING, text="\n".join(body).rstrip())


#: パース結果の ``<!-- source: … -->`` でブックとシートを繋ぐ語。
#: **書く側と読む側で 1 か所にする** ―― 表記が割れると、パース結果から元シートへ
#: 戻れなくなる（``arp4 render --pending`` はここを頼りに撮り直す）。
_SHEET_MARK = " / シート: "


def sheet_source(relative: Path, sheet: str) -> str:
    """``資料/A/基本設計書.xlsx / シート: 受注テーブル``。"""
    return f"{relative.as_posix()}{_SHEET_MARK}{sheet}"


def sheet_origin(source: str) -> tuple[Path, str] | None:
    """``source`` から ``(ブックの相対パス, シート名)`` を戻す。

    シートを持たないもの（コード）は None。**書式が変わったら両方が同時に壊れる**
    ようにしてある（:func:`sheet_source` と対）。
    """
    if _SHEET_MARK not in source:
        return None
    book, _, sheet = source.partition(_SHEET_MARK)
    if not book.strip() or not sheet.strip():
        return None
    return Path(book.strip()), sheet.strip()


# ── 図形（テキストだけ取り、取れなかったぶんを申告する） ────────
_NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
       "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
       "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
       "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
       "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
       "dgm": "http://schemas.openxmlformats.org/drawingml/2006/diagram",
       # スライドの図形（``p:sp``）。**中身は Excel とまったく同じ DrawingML**
       # で、違うのは外側の名前空間だけである（→ :data:`_SPREADSHEET`）。
       "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
       # Word の本文。**図形の入れ物としては Excel とも PowerPoint とも違う**
       # （``wps:wsp`` / ``pic:pic`` で、接続子そのものが無い）ので、
       # :func:`_shapes` は使わない ―― 使うのは申告のほうだけである。
       "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


@dataclass(frozen=True)
class Holder:
    """図形の**入れ物**。``a:`` の側は 1 つも変わらないので、ここだけが形式差である。

    Excel の描画は ``xdr:sp`` / ``xdr:cxnSp``、PowerPoint のスライドは
    ``p:sp`` / ``p:cxnSp`` と、**外側の名前空間だけが違う**。箱の中の文字
    （``a:t``）も接続の端点（``a:stCxn``）も矢羽根（``a:headEnd``）も線の見た目も
    まったく同じものなので、:func:`_shapes` を 2 本書くと**同じ規律を 2 度書いて
    片方だけ直した状態**が必ず生まれる。

    **持たせるのは「次にやること」が形式で変わるところだけ**である。図形の
    申告（:func:`_shape_note`）は「読めなかったものをどう見るか」まで書くのが
    仕事で、そこは Excel と PowerPoint で本当に違う ―― `arp4 render` は
    Excel を起こして撮るので、スライドには使えない。案内どおりにやって届かない
    申告は、申告しないのと同じである。
    """

    #: 入れ物の名前空間（``xdr`` / ``p``）。
    ns: str
    #: 申告に出す言葉（``このシートには`` / ``このスライドには``）。
    label: str
    #: アンカーの接頭辞（``s1-g1`` の ``s``）。**申告はアンカーを名指しする**
    #: ので、ここが実物と食い違うと**存在しない出典を案内する**ことになる
    #: ―― 案内どおりに開いて無いのは、申告していないのと同じである。
    prefix: str
    #: 目で見る手立て。``…読み`` ``…読んでください`` に続けられる形にしてある。
    look: str
    #: 繋がっていない線をどう見るかの 1 文。
    trace: str
    #: 貼り付け画像は撮り直しても変わらない、に当たる 1 文。
    reshoot: str
    #: グラフの値がどこにあるか。Excel は同じブックの別シート、PowerPoint は
    #: **スライドの中に埋め込まれたブック**（`ppt/embeddings/`）にある。
    chart_data: str


#: Excel の描画（``xl/drawings/*.xml``）。
_SPREADSHEET = Holder(
    ns=_NS["xdr"], label="シート", prefix="s", look="`arp4 render` で絵にして",
    trace="どこへ向かう線かは `arp4 render` で絵にすれば読めます。",
    reshoot="`arp4 render` で撮り直しても中身は変わりません。",
    chart_data="系列が指しているのは別のシートの値です")

#: PowerPoint のスライド（``ppt/slides/slide*.xml``）。**`arp4 render` は使えない**
#: ―― あれは Excel を起こして印刷する仕掛けなので、スライドは撮れない。
_SLIDE = Holder(
    ns=_NS["p"], label="スライド", prefix="s", look="PowerPoint で開いて",
    trace="どこへ向かう線かは PowerPoint で開けば読めます。",
    reshoot="実体を `images/` に出してあるので、開いて見てください。",
    chart_data="値はスライドに埋め込まれたブック（`ppt/embeddings/`）にあり、"
               "そこは取り出していません")

#: Word の本文（節 1 つぶん）。**図形の取り出しは :mod:`arp4.docx` が自前で行う**
#: ―― ここで使うのは申告の言葉だけである（:func:`_shape_note`）。
_WORD = Holder(
    ns=_NS["w"], label="節", prefix="w", look="Word で開いて",
    trace="どこへ向かう線かは Word で開けば読めます。",
    reshoot="実体を `images/` に出してあるので、開いて見てください。",
    chart_data="値は文書に埋め込まれたブック（`word/embeddings/`）にあり、"
               "そこは取り出していません")

#: 関係の種別。**種別で絞る**（拡張子で絞らない）―― シートの ``.rels`` には
#: ハイパーリンクも並んでおり、リンク先がたまたま ``.xml`` だと描画として
#: 開きに行くことになる（例外は握り潰されるので、間違えても静かに壊れる）。
_REL_DRAWING = "/drawing"
_REL_DIAGRAM = "/diagramData"
_REL_CHART = "/chart"
#: 貼り付け画像の実体（``xl/media/image1.jpeg``）。描画パートからもう 1 段たどる。
#: **``xdr:pic`` は実体を持たない** ―― 持っているのは ``a:blip`` の ``r:embed``
#: （関係の id）だけで、そこから先はこの種別を辿らないと画像に届かない。
_REL_IMAGE = "/image"
#: スレッドコメント（2018 年以降の Excel が書くコメント）と、その記入者の名簿。
#: **Microsoft の独自スキーマ**なので名前空間が ISO のものと違う。
_REL_THREADED = "/threadedComment"
_REL_PERSON = "/person"
#: 外部ブック参照（``='[外部.xlsx]シート'!A1`` の参照先）。
_REL_EXTERNAL = "/externalLink"

#: :func:`_related_many` に渡す「何を取りに行くか」。**1 か所にまとめてある**の
#: は、zip を開く回数がそのまま 30 冊ぶんの実行時間になるからである。
_WANTED = {"drawing": (_REL_DRAWING, ""),
           "diagram": (_REL_DRAWING, _REL_DIAGRAM),
           "chart": (_REL_DRAWING, _REL_CHART),
           "image": (_REL_DRAWING, _REL_IMAGE),
           "threaded": (_REL_THREADED, "")}


def drawing_parts(path: Path) -> dict[str, list[bytes]]:
    """シート名 → そのシートに載っている描画パートの XML。

    **openpyxl はオートシェイプを捨てる**ので zip を直接見る。読めない資料で
    パースそのものを止めてはいけないので、ここで起きた例外は「描画は無い」として
    飲む（読めなかったことより、資料が落ちるほうが痛い）。

    ここを公開しているのは、同じ描画を**別の目的でも読む**からである
    （:mod:`arp4.render` はシートの実際の広がりを図形のアンカーから割り出す）。
    zip と rels の辿り方が 2 か所にあると、片方だけ直した状態が必ず生まれる。
    """
    return _related(path, _REL_DRAWING)


def diagram_parts(path: Path) -> dict[str, list[bytes]]:
    """シート名 → そのシートに載っている **SmartArt のデータパート**の XML。

    描画パートからもう 1 段たどる（シート → 描画 → ``diagramData``）。
    SmartArt の箱の文字は描画パートには無く、ここにしか無い。
    """
    return _related(path, _REL_DRAWING, through=_REL_DIAGRAM)


def _related(path: Path, kind: str, through: str = "") -> dict[str, list[bytes]]:
    """シート名 → そのシートから ``kind`` で繋がっているパートの本文。

    ``through`` を渡すと**もう 1 段たどる**（シート → 描画 → SmartArt のデータ）。
    """
    return {title: [part.body for part in parts] for title, parts
            in _related_many(path, {"only": (kind, through)})["only"].items()}


@dataclass(frozen=True)
class Part:
    """zip の中のパート 1 本。**本文だけでなく名前も持つ。**

    名前が要るのは画像を出すようになってからである ―― 中身がバイト列のままでは
    **JPEG なのか PNG なのか EMF なのかが分からず**、拡張子を付けられない。
    """

    name: str
    body: bytes
    #: このパートから ``/image`` で繋がっている先 ``{rId: パート名}``。
    #: ``xdr:pic`` は ``a:blip`` の ``r:embed`` しか持たないので、**pic と実体を
    #: 突き合わせるにはこの対応表が要る** ―― 枚数だけで対応させると、代替
    #: テキストの付いている画像とそうでない画像が入れ替わる。
    links: dict[str, str] = field(default_factory=dict)


def _related_many(path: Path, wants: dict[str, tuple[str, str]],
                  trouble: list[str] | None = None,
                  parts_of: "Callable[[zipfile.ZipFile], dict[str, str]] | None" = None
                  ) -> dict[str, dict[str, list[Part]]]:
    """欲しいパートを**まとめて**取る（``種別 → {シート名: [本文]}``）。

    **zip を開くのは 1 回**である。ここは 30 冊ぶん回るところなので、取りたい
    ものが増えるたびに開き直していると、そのまま実行時間になる ―― 描画・
    SmartArt・グラフ・スレッドコメントは**どれもシートの ``.rels`` から始まる
    同じ道**を歩くので、辿った結果を使い回せば道は 1 本で足りる。

    読めない資料でパースそのものを止めてはいけないので、ここで起きた例外は
    「そういうパートは無い」として飲む（読めなかったことより、資料が落ちる
    ほうが痛い）。**ただし飲んだことは ``trouble`` に置く** ―― 黙ると
    「図の無いブック」と見分けが付かない（:func:`_drawing_gap_note`）。

    ``parts_of`` は**出発点の一覧だけ**を差し替える（既定は
    :func:`_sheet_parts`）。PowerPoint のスライドも「入れ物 → ``.rels`` →
    描画・画像」という同じ道を歩くので、歩き方まで書き直す理由が 1 つも無い
    ―― 違うのは 1 歩目をどこから始めるかだけである。
    """
    empty: dict[str, dict[str, list[Part]]] = {key: {} for key in wants}
    try:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
            walked: dict[tuple[str, str], list[str]] = {}
            linked: dict[str, dict[str, str]] = {}

            def targets(part: str, want: str) -> dict[str, str]:
                """``{rId: 解決済みパート名}``。**id ごと返す**（画像で要る）。"""
                rels = f"{_part_dir(part)}/_rels/{Path(part).name}.rels"
                if part not in names or rels not in names:
                    return {}
                return {rid: got for rid, target
                        in _rel_targets(archive, rels, want).items()
                        if (got := _resolve(_part_dir(part), target)) in names}

            def follow(part: str, want: str) -> list[str]:
                # **取りに行く先が自分自身のこともある。** Excel の図形は
                # シートから ``/drawing`` で繋がった別のパートに入っているが、
                # PowerPoint の図形は**スライドの XML そのもの**に入っている
                # ―― 道が 0 歩なだけで、そこから先（画像・グラフ）は同じである。
                if not want:
                    return [part] if part in names else []
                # **同じ道を二度歩かない。** 描画へは 3 種類（図形・SmartArt・
                # グラフ）が同じ 1 歩目で入る。
                if (part, want) in walked:
                    return walked[(part, want)]
                return walked.setdefault(
                    (part, want), list(targets(part, want).values()))

            def images_of(part: str) -> dict[str, str]:
                """描画パートから画像への対応表。**pic と実体を繋ぐ唯一の道。**"""
                if part not in linked:
                    linked[part] = targets(part, _REL_IMAGE)
                return linked[part]

            got: dict[str, dict[str, list[Part]]] = {key: {} for key in wants}
            bodies: dict[str, bytes] = {}
            for title, part in (parts_of or _sheet_parts)(archive).items():
                for key, (kind, through) in wants.items():
                    parts = follow(part, kind)
                    if through:
                        parts = [deeper for one in parts
                                 for deeper in follow(one, through)]
                    if parts:
                        got[key][title] = [
                            Part(name=p, body=bodies.setdefault(p, archive.read(p)),
                                 links=images_of(p)) for p in parts]
            return got
    except (OSError, zipfile.BadZipFile, KeyError, ET.ParseError) as exc:
        if trouble is not None:
            trouble.append(str(exc) or exc.__class__.__name__)
        return empty


@dataclass
class BookExtras:
    """ブック全体にぶら下がっているもの。**シートからはたどれない。**"""

    #: スレッドコメントの記入者 ``{personId: 表示名}``。名前は ``xl/persons/``
    #: にしか無く、コメント側は id しか持っていない。
    people: dict[str, str] = field(default_factory=dict)
    #: 参照している**別ブック** ``(パス, キャッシュされたシート名)``。
    external: list[tuple[str, list[str]]] = field(default_factory=list)
    #: ブックのプロパティ ``(項目, 値)``。**どのシートにも表にも出てこない。**
    core: list[tuple[str, str]] = field(default_factory=list)
    #: マクロ（VBA）が入っているか。**中身は取らないが、あることは言う。**
    macros: bool = False


#: マクロの入れ物。**拡張子ではなくパートの有無で見る**（``.xlsx`` に名前を
#: 付け替えただけの資料でも、入っていれば入っている）。中身は OLE 複合
#: ドキュメントなので、ここから先は zip では開けない。
_VBA_PART = "xl/vbaProject.bin"


def book_extras(path: Path) -> BookExtras:
    """``xl/workbook.xml`` からたどるものを**まとめて 1 回で**取る。"""
    found = BookExtras()
    try:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
            found.core = _properties(archive, names)
            found.macros = _VBA_PART in names
            targets = _rel_targets(archive, "xl/_rels/workbook.xml.rels")
            types = _rel_types(archive, "xl/_rels/workbook.xml.rels")
            for identity, target in targets.items():
                part = _resolve("xl", target)
                kind = types.get(identity, "")
                if part not in names:
                    continue
                if kind.endswith(_REL_PERSON):
                    found.people.update(_people(archive.read(part)))
                elif kind.endswith(_REL_EXTERNAL):
                    where = _external_path(archive, part)
                    if where:
                        found.external.append((where, _external_sheets(
                            archive.read(part))))
    except (OSError, zipfile.BadZipFile, KeyError, ET.ParseError):
        return found
    return found


#: ブックのプロパティ。**並びは固定**（誰がいつ触ったかの順に読める）。
#: ``docProps/core.xml`` と ``docProps/app.xml`` の 2 つに分かれている。
_DC = "http://purl.org/dc/elements/1.1/"
_DCTERMS = "http://purl.org/dc/terms/"
_CP = "http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
_EP = "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
_PROPERTIES = (
    ("文書の表題", f"{{{_DC}}}title"),
    ("作成者", f"{{{_DC}}}creator"),
    ("作成日時", f"{{{_DCTERMS}}}created"),
    ("最終更新者", f"{{{_CP}}}lastModifiedBy"),
    ("最終更新日時", f"{{{_DCTERMS}}}modified"),
    ("会社", f"{{{_EP}}}Company"),
    ("作成アプリ", f"{{{_EP}}}Application"),
)

#: プロパティの置き場。**関係から辿る**（``docProps/`` は慣習であって規約ではない）。
_REL_CORE = "metadata/core-properties"
_REL_APP = "extended-properties"


def _properties(archive: zipfile.ZipFile, names: set[str]) -> list[tuple[str, str]]:
    """ブックのプロパティ。**シートにも表にも出てこないのに、誰がいつ触ったかを持つ。**

    設計書の版と日付は改訂履歴のシートに書かれるが、それは**人が書いた申告**で
    ある。ブックのプロパティは Excel が黙って書くので、**改訂履歴に載っていない
    更新**がここにだけ残る ―― 課題管理のコメントが本文より新しいのと同じ形で、
    「最後に触ったのは誰か」を表からは確かめられない。

    ``作成アプリ`` を取るのは、**Excel で作られていない設計書**があるからである
    （帳票ツール・レポート出力）。計算結果の保存されていない数式・高さ 0 の行と
    同じ側の資料で、そうと分かっていれば読み手の身構え方が変わる。
    """
    found: dict[str, str] = {}
    for rels, kind in (("_rels/.rels", _REL_CORE), ("_rels/.rels", _REL_APP)):
        if rels not in names:
            continue
        for target in _rel_targets(archive, rels, kind).values():
            part = _resolve("", target)
            if part not in names:
                continue
            try:
                root = ET.fromstring(archive.read(part))
            except ET.ParseError:
                continue
            for element in root.iter():
                text = (element.text or "").strip()
                if text:
                    found.setdefault(element.tag, text)
    return [(label, found[tag]) for label, tag in _PROPERTIES if tag in found]


def _external_path(archive: zipfile.ZipFile, part: str) -> str:
    """参照先ブックの**置き場**。本文ではなく ``.rels`` の Target に書いてある。"""
    rels = f"{_part_dir(part)}/_rels/{Path(part).name}.rels"
    if rels not in set(archive.namelist()):
        return ""
    for target in _rel_targets(archive, rels).values():
        if target:
            # ``file:///C:/...`` で入っていることも相対パスのこともある。
            # **書いてあるまま出す** ―― 直すと元資料と突き合わせられなくなる。
            return target
    return ""


def _external_sheets(body: bytes) -> list[str]:
    """参照先ブックのシート名（Excel が**手元に無いとき用にキャッシュ**する）。"""
    try:
        root = ET.fromstring(body)
    except ET.ParseError:
        return []
    return [name.get("val") or "" for name
            in root.iter(f"{{{_NS['main']}}}sheetName") if name.get("val")]


def _people(body: bytes) -> dict[str, str]:
    try:
        root = ET.fromstring(body)
    except ET.ParseError:
        return {}
    return {person.get("id") or "": person.get("displayName") or ""
            for person in root.iter(f"{{{_TC}}}person")}


def _sheet_parts(archive: zipfile.ZipFile) -> dict[str, str]:
    """シート名 → パッケージ内のシート XML。**辿り方を 1 か所にする。**

    描画（:func:`drawing_parts`）とシート XML（:func:`sheet_facts`）が同じ道を歩く。
    2 か所に書くと、片方だけ直した状態が必ず生まれる。
    """
    book = ET.fromstring(archive.read("xl/workbook.xml"))
    targets = _rel_targets(archive, "xl/_rels/workbook.xml.rels")
    found: dict[str, str] = {}
    for sheet in book.findall("main:sheets/main:sheet", _NS):
        title = sheet.get("name") or ""
        relation = sheet.get(f"{{{_NS['rel']}}}id") or ""
        part = _resolve("xl", targets.get(relation, ""))
        if title and part:
            found[title] = part
    return found


@dataclass
class SheetFacts:
    """シート XML を**流し読みして**取れるもの。openpyxl からは取れない。

    2 つを 1 つの走査でまとめているのは、**zip を開く回数がそのまま 30 冊ぶんの
    実行時間になる**からである（:func:`_related_many` と同じ理屈）。シート XML は
    数十 MB になりうるので、読む回数を増やせばそのまま効く。
    """

    #: シート名 → **計算結果が保存されていない数式セル**の番地。
    blanks: dict[str, list[str]] = field(default_factory=dict)
    #: シート名 → ``{番地: (取り消し線の掛かっている文字, セル全体か)}``。
    #: **リッチテキスト**（1 つのセルの中で書式が変わる）でしか起きない。
    struck: dict[str, dict[str, tuple[str, bool]]] = field(default_factory=dict)


def sheet_facts(path: Path) -> SheetFacts:
    """シート XML を 1 度だけ流し読みして、openpyxl が返さないものを取る。

    **数式のキャッシュ無し** ―― ``data_only=True`` は値が保存されていなければ
    None を返す。数式が書いてあったことすら分からないので、**空欄とは別物**で
    ある。``data_only=False`` でもう一度開けば同じことは分かるが、それは大きな
    ブックで読み込みが 2 倍になる（実測 30 冊で効く）。

    **セルの中の一部だけに掛かった取り消し線** ―― :func:`_struck` が見ている
    ``cell.font.strike`` は**セル 1 つに 1 つ**の書式で、1 つのセルの中で書式が
    変わるもの（リッチテキスト）は表せない。``受注区分（廃止）`` の後半だけを
    消してある欄は実物にごく普通にあり、**画面では消えているのに、こちらには
    生きた文字として出ていた** ―― 取り消し線を書式から取ると決めた理由
    （その書式だけが値を偽る）が、そこでは成り立っていなかった。

    読めない資料でパースそのものを止めてはいけないので、例外は「そういうものは
    無い」として飲む（:func:`drawing_parts` と同じ）。
    """
    found = SheetFacts()
    try:
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
            runs = _struck_strings(archive, names)
            for title, part in _sheet_parts(archive).items():
                if part not in names:
                    continue
                blanks, struck = _scan_sheet(archive, part, runs)
                if blanks:
                    found.blanks[title] = blanks
                if struck:
                    found.struck[title] = struck
            return found
    except (OSError, zipfile.BadZipFile, KeyError, ET.ParseError):
        return found


def _struck_strings(archive: zipfile.ZipFile, names: set[str]
                    ) -> dict[str, tuple[str, bool]]:
    """共有文字列のうち、**取り消し線の掛かった断片を持つもの**。

    返すのは ``{索引: (消してある文字, セル全体か)}`` である。「全体か」を分けて
    持つのは、**次に読む人の読み方が変わる**からである ―― 全部消してあるなら
    セル全体の書式と同じ意味だが、一部だけなら**残りは生きている**。

    走るのは ``xl/sharedStrings.xml`` である。1 つのセルの中で書式が変わる文字列は
    Excel がここへ ``<r>``（run）に割って書く ―― openpyxl はその割れ目を
    ``rich_text=True`` でしか返さず、既定では**繋げた 1 本の文字列**にしてしまう。
    ここで読み込み方を変えないのは、**セルの値の型が変わる**（``CellRichText``）
    と、書式を見ない 30 冊ぶんの読み込みまで道連れになるからである。
    """
    part = "xl/sharedStrings.xml"
    if part not in names:
        return {}
    si = f"{{{_NS['main']}}}si"
    found: dict[str, tuple[str, bool]] = {}
    index = 0
    try:
        with archive.open(part) as stream:
            for _, element in ET.iterparse(stream, events=("end",)):
                if element.tag != si:
                    continue
                marked = _struck_runs(element)
                if marked:
                    found[str(index)] = marked
                index += 1
                element.clear()
    except (ET.ParseError, KeyError, OSError):
        return found
    return found


def _struck_runs(holder: ET.Element) -> tuple[str, bool] | None:
    """断片（``<r>``）の並びから**消してある文字**を取る。

    ``(消してある文字, セル全体か)`` を返す。断片が 1 つも無い（＝ 1 つの書式で
    書かれた）文字列は ``None`` ―― そちらはセルの書式（:func:`_struck`）で
    取れている。
    """
    run = f"{{{_NS['main']}}}r"
    text = f"{{{_NS['main']}}}t"
    strike = f"{{{_NS['main']}}}rPr/{{{_NS['main']}}}strike"
    runs = holder.findall(run)
    struck = [one for one in runs if one.find(strike) is not None]
    if not struck:
        return None
    return ("".join((one.find(text).text or "")
                    if one.find(text) is not None else ""
                    for one in struck),
            len(struck) == len(runs))


def _scan_sheet(archive: zipfile.ZipFile, part: str,
                runs: dict[str, tuple[str, bool]]
                ) -> tuple[list[str], dict[str, tuple[str, bool]]]:
    """シート 1 枚ぶん。**流し読みする**（シート XML は数十 MB になりうる）。

    数式の「値が無い」は 2 通りの書かれ方をする ―― ``<v>`` が**そもそも無い**の
    と、``<v></v>`` と**空で置かれている**のと。後者を数え落としていたぶん、
    ツールが書き出したブック（いちばん値が入っていないブック）だけが
    申告から漏れていた。

    ただし ``t="str"`` の空値は**数式が空文字を返した結果**であって読めなかった
    わけではないので数えない（``=IF(…,"","対象")`` は設計書にごく普通に出てくる）。
    """
    cell = f"{{{_NS['main']}}}c"
    formula = f"{{{_NS['main']}}}f"
    value = f"{{{_NS['main']}}}v"
    found: list[str] = []
    struck: dict[str, tuple[str, bool]] = {}
    try:
        with archive.open(part) as stream:
            for _, element in ET.iterparse(stream, events=("end",)):
                if element.tag != cell:
                    continue
                reference = element.get("r")
                if element.find(formula) is not None:
                    cached = element.find(value)
                    empty = cached is None or not (cached.text or "").strip()
                    if empty and reference and element.get("t") != "str":
                        found.append(reference)
                elif runs and element.get("t") == "s" and reference:
                    # **共有文字列を指しているセルだけ**が候補である（索引は
                    # ``<v>`` に入っている）。空振りが多いので、取り消し線の
                    # 掛かった文字列が 1 つも無いブックでは見に行かない。
                    cached = element.find(value)
                    key = (cached.text or "").strip() if cached is not None else ""
                    if key in runs:
                        struck[reference] = runs[key]
                elif element.get("t") == "inlineStr" and reference:
                    # **書き方は 2 通りある。** Excel は文字列を共有表へ集める
                    # が、ツールが書き出したブックはセルの中に直に書く
                    # （``<is>``）―― 共有表しか見ていなければ、そちらは
                    # 1 件も取れない。
                    inline = element.find(f"{{{_NS['main']}}}is")
                    marked = _struck_runs(inline) if inline is not None else None
                    if marked:
                        struck[reference] = marked
                element.clear()
    except (ET.ParseError, KeyError, OSError):
        return found, struck
    return found, struck


# ── スレッドコメント（いまの Excel が書くコメント） ─────────────
#: Microsoft の独自スキーマ。ISO の名前空間ではない。
_TC = "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"

#: 旧形式へのなりすまし。スレッドコメントを付けた Excel は、**古い Excel でも
#: 読めるように**同じ内容を旧形式のコメントとしても書く ―― その本文は
#: 「[Threaded comment]」で始まる 200 字あまりの Microsoft の但し書きで、
#: 記入者は表示名ではなく ``tc={GUID}`` である。openpyxl が返すのはこちらなので、
#: **黙って出すとレビュー指摘 1 件ごとに但し書きが 1 つ載る**（本文はその末尾に
#: 埋もれる）。読み手はそれを 3 件も見た時点でコメントを読むのをやめる。
_LEGACY_MARK = "[Threaded comment]"
_LEGACY_AUTHOR = "tc="
#: 但し書きの後ろに続く本文と返信の見出し（Excel が英語で書く）。
_LEGACY_BODY = re.compile(r"^\s*(?:Comment|Reply):\s*", re.MULTILINE)


def threaded_comments(parts: dict[str, list[bytes]], people: dict[str, str]
                      ) -> dict[str, list[tuple[str, str]]]:
    """シート名 → ``(番地, 本文)``。**スレッドは返信の順に並べる。**

    旧形式のコメント（:func:`_comment`）と分けずに同じアンカー（``m1``）へ出す
    のは、読み手にとって**どちらも「セルに付いた指摘」**だからである ―― 書かれ
    方の違いは資料の作られた年の違いでしかない。

    ここで取れるのは旧形式に無いものが 3 つある ―― **返信**（指摘への回答が
    そこにしか無い）、**記入者の表示名**（旧形式へのなりすましは GUID しか
    持たない）、**解決済みかどうか**（``done``）。最後のものが要るのは、
    片付いた指摘と生きている指摘が混ざると**整理層が全部を積み残しとして
    読む**からである。解決したことは資料に書いてあるので、転記である。
    """
    found: dict[str, list[tuple[str, str]]] = {}
    for title, bodies in parts.items():
        listed: list[tuple[str, str]] = []
        for body in bodies:
            listed += _threads(body, people)
        if listed:
            found[title] = listed
    return found


def _threads(body: bytes, people: dict[str, str]) -> list[tuple[str, str]]:
    """1 パートぶん。**親のすぐ後ろに返信を並べる**（時系列がばらけない）。"""
    try:
        root = ET.fromstring(body)
    except ET.ParseError:
        return []

    posted: list[dict[str, str]] = []
    for node in root.iter(f"{{{_TC}}}threadedComment"):
        text = node.find(f"{{{_TC}}}text")
        posted.append({
            "id": node.get("id") or "",
            "parent": node.get("parentId") or "",
            "ref": node.get("ref") or "",
            "who": people.get(node.get("personId") or "", ""),
            "when": (node.get("dT") or "")[:10],
            "done": node.get("done") or "0",
            "text": ((text.text or "").strip() if text is not None else "")})

    # **返信は 1 段とは限らず、親が居るとも限らない。**
    #
    # 親のすぐ後ろに返信を並べるのは時系列をばらけさせないためだが、``parentId``
    # の連なりを 1 段しか辿っていなかったぶん、**返信への返信が黙って落ちて
    # いた**。もっと静かに消えるのが**親を失った返信**である（起点のコメント
    # だけ削除された資料・抜粋して配られた資料でこうなる）―― 起点が無いので
    # 1 件も出ず、そのセルにコメントが付いていたことすら伝わらない。
    known = {one["id"] for one in posted}
    replies: dict[str, list[dict[str, str]]] = {}
    for one in posted:
        if one["parent"] in known:
            replies.setdefault(one["parent"], []).append(one)

    listed: list[tuple[str, str]] = []
    seen: set[str] = set()

    def walk(one: dict[str, str], depth: int, orphan: bool = False) -> None:
        if one["id"] in seen:                        # ``parentId`` の輪で回らない
            return
        seen.add(one["id"])
        head = "[解決済み] " if one["done"] == "1" else ""
        lead = "返信（起点のコメントは資料に残っていません） " if orphan else (
            "返信 " if depth else "")
        listed.append((one["ref"], head + lead + _posted(one)))
        for reply in replies.get(one["id"], []):
            walk(reply, depth + 1)

    for one in posted:
        if one["parent"] in known:
            continue
        walk(one, 0, orphan=bool(one["parent"]))
    return listed


def _posted(one: dict[str, str]) -> str:
    """``設計者（2026-07-01）: 本文``。**記入者と日付は本文の一部**である。"""
    who = one["who"] or "記入者不明"
    when = f"（{one['when']}）" if one["when"] else ""
    return f"{who}{when}: {one['text']}"


def _unwrap_legacy(text: str) -> str:
    """旧形式へのなりすましから**本文だけ**を取り出す。

    スレッドコメントのパートが読めた資料では使わない（そちらのほうが記入者も
    返信も揃っている）。ここが効くのは**パートを落として配られた資料**で、
    そのときでも但し書き 200 字より本文のほうが要る。
    """
    if _LEGACY_MARK not in text:
        return text
    pieces = [piece.strip() for piece in
              _LEGACY_BODY.split(text.split(_LEGACY_MARK, 1)[1])]
    body = [piece for piece in pieces[1:] if piece]
    return "／".join(body) if body else text


class _Link(NamedTuple):
    """接続子 1 本。**転記だけ**（座標からの復元も、凡例の解釈もしない）。"""

    source: str
    arrow: str
    target: str
    name: str = ""                                     # 線の名前（``委託``）
    dash: str = ""                                     # 線種（実線・破線…）
    color: str = ""                                    # 線色（``#404040``）
    width: str = ""                                    # 太さ（``1pt``）


@dataclass
class Drawing:
    """1 シートぶんの描画。**取れたものと取れなかったものを両方持つ。**

    数えるのは**図形そのもの**であって、アンカー（載っている枠）ではない。
    アンカーを数えていた頃は、グループ化された業務フロー（箱 10 個で
    アンカー 1 個）が「図形が 1 個あり、10 個からテキストを取り出しました」と
    **自分で矛盾したことを言っていた** ―― 申告が信用されなくなるのがいちばん困る。

    **画像を図形と混ぜない。** 貼り付けたスクリーンショットは
    「テキストの取れない図形」ではなく「そもそもテキストを持たないもの」で、
    :mod:`arp4.render` で撮り直しても中身は読めるようにならない ―― 絵は絵の
    ままである。**中の字のほうは :mod:`arp4.ocr` が読む**（``o1``）。
    """

    shapes: int = 0                                    # 図形（``xdr:sp``）
    pictures: int = 0                                  # 貼り付け画像（``xdr:pic``）
    connectors: int = 0                                # 接続子（``xdr:cxnSp``）
    charts: int = 0                                    # グラフ
    diagrams: int = 0                                  # SmartArt（図表）
    diagram_boxes: int = 0                             # SmartArt の箱
    objects: int = 0                                   # 埋め込みオブジェクト（OLE）
    labels: list[str] = field(default_factory=list)    # テキストの取れた図形
    #: 接続 ``(元, 向き, 先, 名前, 線種, 線色, 太さ)``。向きも名前も線も
    #: **書いてあるまま**で、意味は付けない ―― 「破線＝委託」と読むのは凡例を
    #: 見た整理層の仕事である。線の 3 つは分からなければ空（テーマ由来の線は
    #: 解決しない → :func:`_line`）。
    links: list[_Link] = field(default_factory=list)
    #: **どこにも繋がっていない**接続子の数（両端の id が無い ―― 目分量の線）。
    loose: int = 0
    #: **繋がってはいるが、相手が文字を持たない**接続子の数（ゾーンの囲み枠・
    #: 装飾へ落ちる線）。:attr:`loose` と分けてあるのは**次にやることが正反対**
    #: だからである ―― こちらは絵にすればどこへ向かう線か読めるが、あちらは
    #: 絵にしても読めない（資料に情報が無い）。混ぜて申告していた頃は、
    #: 繋がっている線まで「線を目分量で置いた図」と言っていた（**申告が嘘**）。
    unnamed: int = 0
    #: 図形・群・画像・図表の**代替テキスト** ``(種別, 本文)``。
    #: 名前（``Picture 1``）は Excel が自動で振るので取らない ―― 代替テキスト
    #: だけが**人が書いたもの**である。
    alts: list[tuple[str, str]] = field(default_factory=list)
    #: そのうち**画像**に書かれていた枚数。書かれていない画像は「そこに何かが
    #: 貼ってある」ことしか伝わらない ―― 中身が読めないのは同じでも、
    #: **何の画像かも分からない**ぶんだけ次にやることが変わる。
    picture_alts: int = 0
    #: 貼り付け画像の実体 ``(zip の中のパート名, 代替テキスト)``。**貼ってある
    #: 順**（``xdr:pic`` の出てくる順）で、取り出して ``images/`` へ出すのに使う。
    media: list[tuple[str, str]] = field(default_factory=list)
    #: グラフの中身 ``(グラフ名, 系列, 分類の範囲, 値の範囲)``。
    series: list[tuple[str, str, str, str]] = field(default_factory=list)
    #: **XML として読めなかった描画パートの数。** 0 個の図形と見分けが付かない
    #: ので数える ―― ここを黙ると、図の描いてあるシートが「図の無いシート」
    #: として出てくる（`未読取` を宣言する先も出ない）。
    unreadable: int = 0
    #: 図形として置かれた**表**（``a:tbl``）の中身。**Excel には出てこない**
    #: ―― あちらの表はセルの面であって図形ではない。PowerPoint の一覧は
    #: すべてこの形になるので、埋め込みオブジェクトとして数えて中身を捨てると、
    #: **スライドの表が 1 つも仕様にならない。**
    tables: list[list[list[str]]] = field(default_factory=list)

    @property
    def total(self) -> int:
        return (self.shapes + self.pictures + self.connectors
                + self.charts + self.diagrams + self.objects + len(self.tables))

    @property
    def summary(self) -> str:
        """``図形 3 個・接続子 1 本・画像 2 枚``。**数えたものだけを言う。**"""
        boxes = f"（箱 {self.diagram_boxes} 個）" if self.diagram_boxes else ""
        parts = [f"図形 {self.shapes} 個" if self.shapes else "",
                 f"表 {len(self.tables)} 枚" if self.tables else "",
                 f"接続子 {self.connectors} 本" if self.connectors else "",
                 f"SmartArt {self.diagrams} 個{boxes}" if self.diagrams else "",
                 f"グラフ {self.charts} 個" if self.charts else "",
                 f"埋め込みオブジェクト {self.objects} 個" if self.objects else "",
                 f"画像 {self.pictures} 枚" if self.pictures else ""]
        return "・".join(p for p in parts if p) or "図形 0 個"


def _drawings(related: dict[str, dict[str, list[Part]]]) -> dict[str, Drawing]:
    """シート名 → :class:`Drawing`。**パートは読み終わったものを受け取る**
    （:func:`_related_many` が zip を 1 回開いてまとめて取ってある）。"""
    found: dict[str, Drawing] = {}
    titles = (set(related["drawing"]) | set(related["diagram"])
              | set(related["chart"]))
    for title in titles:
        drawing = Drawing()
        for part in related["drawing"].get(title, []):
            _shapes(part, drawing)
        for part in related["diagram"].get(title, []):
            _diagram(part.body, drawing)
        for part in related["chart"].get(title, []):
            _chart(part.body, drawing)
        if drawing.total or drawing.series or drawing.unreadable:
            # **読めなかったパートも持ち回る。** 落とすと図形 0 個と同じになり、
            # 申告する先（:func:`_shape_note`）へ届かない。
            found[title] = drawing
    return found


#: 矢羽根の向き。``headEnd`` は線の始点側、``tailEnd`` は終点側に付く。
_ARROW = "→"
_BACK = "←"
_BOTH = "↔"
_PLAIN = "―"


def _shapes(part: Part, into: Drawing,
            holder: Holder = _SPREADSHEET) -> Drawing:
    """描画パート 1 本を読む。**テキストと接続の両方を取る。**

    ``holder`` は**入れ物の名前空間だけ**を差し替える（→ :class:`Holder`）。
    既定が Excel なのは、ここが 30 冊ぶん回る既定の道だからである ―― スライドを
    読むときは :data:`_SLIDE` を明示して渡す。

    テキストは**図形ごと**に集める（アンカーごとではない） ―― グループ化された
    業務フローはアンカー 1 個の中に箱が 10 個入っていることがあり、まとめてしまうと
    「受注入力与信判定在庫引当」と繋がって語が壊れる。段落は改行で分ける。

    **接続は推測ではない。** 接続子（``xdr:cxnSp``）は繋がっている図形の id を
    ``a:stCxn`` / ``a:endCxn`` に、矢羽根を ``a:headEnd`` / ``a:tailEnd`` に持って
    いる。これを読むのは座標からの復元ではなく**書いてあることの転記**なので、
    「意味を判断しない」という規律に反しない。id が無い接続子（線を目分量で置いた図）
    だけが取れず、その本数は :attr:`Drawing.loose` に数える。
    """
    try:
        root = ET.fromstring(part.body)
    except ET.ParseError:
        # **読めなかった描画パートは、図の無いシートと見分けが付かない。**
        into.unreadable += 1
        return into

    for picture in root.iter(f"{{{holder.ns}}}pic"):
        into.pictures += 1
        # **代替テキストは人が書いた文字である。** 貼り付け画像の中身は取れない
        # ままだが、「受注入力画面のイメージ」と書いてあれば**何の画像かは
        # 分かる** ―― 名前（``Picture 1``）は Excel が自動で振るので取らない。
        alt = _shape_alt(picture, holder)
        if alt:
            into.picture_alts += 1
            into.alts.append(("画像", alt))
        # **実体の在り処を控える。** ここを取っていなかったあいだ、貼り付け
        # 画像は「枚数」としてしか残らず、**中身は資料の中に入ったまま**だった
        # ―― 機械が読めないだけで、エージェントは開けば読める。
        embedded = _embedded(picture)
        if embedded and (where := part.links.get(embedded)):
            into.media.append((where, alt))

    for frame in root.iter(f"{{{holder.ns}}}graphicFrame"):
        # **グラフ・SmartArt・埋め込みオブジェクトは ``xdr:sp`` ではない。**
        # 数えていなかったぶん、グラフだけのシートは図形 0 個・セル 0 個になり、
        # **ファイルが 1 本も出なかった** ―― シートが存在したことすら伝わらない。
        # 何であるかは ``a:graphicData`` の uri に書いてある（転記であって判断ではない）。
        data = frame.find(f".//{{{_NS['a']}}}graphicData")
        kind = (data.get("uri") or "") if data is not None else ""
        if kind.endswith("/chart"):
            into.charts += 1
            label = "グラフ"
        elif kind.endswith("/diagram"):
            into.diagrams += 1
            label = "SmartArt"
        elif kind.endswith("/table"):
            # **PowerPoint の一覧はここに入る。** 埋め込みオブジェクトとして
            # 数えて中身を捨てていたら、スライドの表は 1 枚も仕様にならない。
            into.tables.append(_shape_table(frame))
            label = "表"
        else:
            into.objects += 1
            # **貼り込まれた Word・PDF は開けない**が、代替テキストに何の資料か
            # 書いてあれば、`sources/` に足すべきファイルの名前が分かる。
            label = "埋め込みオブジェクト"
        alt = _shape_alt(frame, holder)
        if alt:
            into.alts.append((label, alt))

    names: dict[str, str] = {}
    for shape in root.iter(f"{{{holder.ns}}}sp"):
        into.shapes += 1
        text = _shape_text(shape)
        identity = _shape_id(shape, holder)
        if identity:
            names[identity] = text
        if text:
            into.labels.append(text)
        # **図形にも代替テキストは書ける。** 画像からは取るのに図形から取って
        # いなかったぶん、箱の文字にも表にも無い 1 文（「現用／待機の 2 系統。
        # DR は別紙 5」）が、**そこにしか書かれていないまま落ちていた**。
        alt = _shape_alt(shape, holder)
        if alt:
            into.alts.append(("図形", alt))

    for group in root.iter(f"{{{holder.ns}}}grpSp"):
        # **群に書かれた代替テキストは図全体の説明である。** 構成図に 1 文だけ
        # 添えるのはこの形になる（箱ごとではなくゾーンごとに書く）。
        alt = _shape_alt(group, holder)
        if alt:
            into.alts.append(("図形の群", alt))

    for connector in root.iter(f"{{{holder.ns}}}cxnSp"):
        into.connectors += 1
        start = _endpoint(connector, "stCxn")
        end = _endpoint(connector, "endCxn")
        if start is None or end is None:
            into.loose += 1                            # どこにも繋がっていない
            continue
        source, target = names.get(start, ""), names.get(end, "")
        if not source or not target:
            # **繋がってはいるが、相手が文字を持たない**（ゾーンの囲み枠・装飾）。
            # 目分量の線と同じ数に混ぜていた頃は、繋がっている線まで
            # 「線を目分量で置いた図」と申告していた ―― 絵にすれば読めるほうを
            # 「絵にしても読めない」と言うのは、読み手に空振りをさせる。
            into.unnamed += 1
            continue
        arrow = _direction(connector)
        if arrow == _BACK:                             # 矢羽根が始点側にある
            source, target, arrow = target, source, _ARROW
        into.links.append(_Link(source, arrow, target,
                                _shape_name(connector, holder),
                                *_line(connector, holder)))
    return into


#: グラフの名前空間（DrawingML の図形とは別建て）。
_C = "http://schemas.openxmlformats.org/drawingml/2006/chart"


def _chart(body: bytes, into: Drawing) -> Drawing:
    """グラフ 1 枚から**タイトル・系列名・参照範囲**を取る。

    ここには長く「グラフの中身は取れていません（系列が指しているのは別の
    シートの値です）」と書いてあった。**言っていること自体は正しいが、
    どのシートのどこかを言っていなかった** ―― 読み手はブックを端から探す
    しかなく、実際には探されずに終わる。

    参照範囲は ``c:f`` に ``'元データ'!$B$2:$B$3`` と**そのまま書いてある**ので、
    読むのは転記である（値そのものはキャッシュを見ずに、指し先の番地を出す
    ―― そのシートは同じパース結果の中に必ずある）。タイトルと軸名は人が
    書いた文字で、図形の ``<a:t>`` を読むのと同じことをしている。
    """
    try:
        root = ET.fromstring(body)
    except ET.ParseError:
        # **読めなかった描画パートは、図の無いシートと見分けが付かない。**
        into.unreadable += 1
        return into

    heading = root.find(f"{{{_C}}}chart/{{{_C}}}title")
    title = _chart_text(heading) if heading is not None else ""
    for order, series in enumerate(root.iter(f"{{{_C}}}ser"), start=1):
        into.series.append((title or "（タイトルなし）",
                            _series_name(series, order),
                            _series_ref(series, ("cat", "xVal")),
                            _series_ref(series, ("val", "yVal"))))
    if title and not into.series:
        # 系列がキャッシュだけで書かれているグラフでも、**タイトルは人の言葉**
        # なので落とさない。
        into.series.append((title, "", "", ""))
    return into


def _chart_text(holder: ET.Element) -> str:
    """タイトル・系列名。**人が打った文字と、セルから引いた文字の 2 通りある。**

    打った文字は図形と同じ ``a:t`` に入るが、**セルから引いたものは
    ``c:strRef``**（``c:f`` に参照、``c:strCache`` に最後に見えていた文字）に
    入る。``a:t`` しか見ていなかったぶん、**表題をセルから引いたグラフだけ**が
    「（タイトルなし）」になっていた ―― 実物では引くほうが多い書き方である。

    キャッシュ（画面に出ていた文字）を先に、無ければ参照そのものを出す。
    参照だけでも「どのセルの文字か」は伝わる。
    """
    text = _shape_text(holder)
    if text:
        return text
    for tag in (f"{{{_C}}}v", f"{{{_C}}}f"):
        for element in holder.iter(tag):
            if (element.text or "").strip():
                return (element.text or "").strip()
    return ""


def _series_name(series: ET.Element, order: int) -> str:
    """系列名。**参照ではなく文字が入っていることもある**ので両方見る。"""
    holder = series.find(f"{{{_C}}}tx")
    if holder is None:
        return f"系列{order}"
    return _chart_text(holder) or f"系列{order}"


def _series_ref(series: ET.Element, tags: tuple[str, ...]) -> str:
    """系列が指している範囲（``'元データ'!$B$2:$B$3``）。

    **どのタグに入るかはグラフの種類で変わる。** 棒・折れ線・円は
    ``c:cat`` / ``c:val`` だが、**散布図とバブルは ``c:xVal`` / ``c:yVal``**
    である ―― 前者しか見ていなかったぶん、散布図だけが分類も値も空欄のまま
    「参照範囲を取り出しました」と申告していた（申告のほうが嘘になる）。
    """
    for tag in tags:
        holder = series.find(f"{{{_C}}}{tag}")
        if holder is None:
            continue
        for element in holder.iter(f"{{{_C}}}f"):
            if (element.text or "").strip():
                return (element.text or "").strip()
    return ""


#: SmartArt の中身。箱そのものと、箱をつなぐ線（``parTrans`` / ``sibTrans``）と、
#: 見た目を作るための複製（``pres``）が同じ一覧に並んでいる。**箱だけを取る**
#: ―― 複製まで拾うと同じ語が 2 回出て、整理層が箱の数を数え間違える。
_DIAGRAM_SKIP = {"doc", "parTrans", "sibTrans", "pres"}


def _diagram(body: bytes, into: Drawing) -> Drawing:
    """SmartArt のデータパートから**箱の中の文字**を取る。

    SmartArt（組織図・手順・循環）で描かれた業務フローは、図形と違って
    ``xdr:sp`` を 1 つも持たない ―― 文字は ``xl/diagrams/data*.xml`` に別建てで
    入っている。**読むのは ``<a:t>`` の転記**で、図形のときと同じことをしている。

    **繋がりは取らない。** SmartArt の親子は ``dgm:cxn`` にあるが、指しているのは
    箱ではなく内部の点（``presOf`` / ``presParOf`` など役割つきの参照）で、
    どれが「A の次が B」に当たるかはレイアウト定義まで解かないと決まらない
    ―― そこは復元であって転記ではないので、やらずに数だけ申告する。
    """
    try:
        root = ET.fromstring(body)
    except ET.ParseError:
        # **読めなかった描画パートは、図の無いシートと見分けが付かない。**
        into.unreadable += 1
        return into

    for point in root.iter(f"{{{_NS['dgm']}}}pt"):
        if (point.get("type") or "node") in _DIAGRAM_SKIP:
            continue
        into.diagram_boxes += 1
        text = _shape_text(point)
        if text:
            into.labels.append(text)
    return into


def _shape_text(shape: ET.Element) -> str:
    """箱の中の文字。**改行は 2 通りの書かれ方をする。**

    段落（``a:p``）だけを改行として扱っていたが、DrawingML には段落を割らない
    行区切り（``a:br``）がもう 1 つある ―― 読み飛ばしていたぶん、``承認待ち``
    と ``3 営業日以内`` が ``承認待ち3 営業日以内`` という 1 語に化けていた。
    **整理層はそれを 1 つの状態名として読む**（元の 2 語には戻せない）。

    **行頭の字下げは落とさない。** セルの値では残すと決めてある
    （:func:`_value` ―― 項目定義書の親子は ``受注ヘッダ`` / ``　　受注番号``
    と字下げで表す）のに、図形の段落はまとめて ``strip()`` していた。機能構成図を
    1 つのテキストボックスで書くのは日本の設計書でごく普通の書き方で、そこでは
    **親子が段落頭の全角空白にしか無い** ―― 潰すと 12 個の機能が同じ深さで並ぶ。
    行末の空白は画面に出ないので落とす（こちらもセルと同じ）。
    """
    paragraphs: list[str] = []
    for para in shape.iter(f"{{{_NS['a']}}}p"):
        pieces: list[str] = []
        for node in para.iter():
            if node.tag == f"{{{_NS['a']}}}t":
                pieces.append(node.text or "")
            elif node.tag == f"{{{_NS['a']}}}br":
                pieces.append("\n")
        line = "".join(pieces).rstrip()
        if line.strip():                    # 空の段落（行送りだけ）は落とす
            paragraphs.append(line)
    return "\n".join(paragraphs)


def _shape_table(frame: ET.Element) -> list[list[str]]:
    """``a:tbl`` を行の並びにする。**縦結合はセルの面と同じ規律で下へ広げる。**

    ``a:gridSpan`` / ``a:hMerge``（横）と ``a:rowSpan`` / ``a:vMerge``（縦）は
    Excel の結合セルとまったく同じ意味で、**openpyxl が左上にしか値を返さない
    のと同じことが XML でも起きる** ―― 続きのセルは空で入っている。

    **広げるのは縦だけである**（:func:`_cells` と同じ）。分類列の「同上」を
    回復するのは忠実性の話だが、1 行だけの横結合は表題であって、広げると
    同じ語が 5 列に並ぶ。
    """
    rows: list[list[str]] = []
    above: list[str] = []
    for row in frame.iter(f"{{{_NS['a']}}}tr"):
        line: list[str] = []
        for index, cell in enumerate(row.findall(f"{{{_NS['a']}}}tc")):
            text = _shape_text(cell)
            if not text and cell.get("vMerge") and index < len(above):
                # **縦結合の続き**（画面では上の値が全行に掛かって見えている）。
                text = above[index]
            line.append(text)
        if line:
            rows.append(line)
            above = line
    return rows


def _shape_id(shape: ET.Element, holder: Holder = _SPREADSHEET) -> str:
    element = shape.find(f".//{{{holder.ns}}}cNvPr")
    return (element.get("id") or "") if element is not None else ""


def _shape_name(shape: ET.Element, holder: Holder = _SPREADSHEET) -> str:
    """名前（``cNvPr/@name``）を**そのまま**返す。**接続子にだけ使う。**

    :func:`_shape_alt` は「名前は取らない」と決めている ―― 図形の名前は
    ``図 3`` のような自動採番で、人が書いた言葉は代替テキストのほうにあるからで
    ある。**接続子はそこが違う。** 線には代替テキストを書く習慣が無く、テキスト
    枠も持たないので、線に付けられた言葉の**唯一の在り処が名前**になる ――
    実測（kotonoha r001）の体制図では ``委託`` ``点検依頼`` ``提供``
    ``月次締め`` が名前にだけあり、**「どれが委託か」は他のどこにも書かれて
    いなかった**（線種は 7 本とも実線で、凡例が言う破線は資料に存在しない）。

    **自動で振られた名前を機械が落とさない。** ``コネクタ 10`` が混ざるのは
    整理層が見て捨てればよいノイズだが、「これは自動名だ」と機械が判定して
    落とすと**落とした判断が誰にも見えなくなる** ―― しかも判定は言語ごとの
    綴り（``コネクタ`` / ``Straight Arrow Connector``）に依るので、資料の言語が
    変わるだけで人の付けた名前を捨てる。ここは転記に徹する。
    """
    element = shape.find(f".//{{{holder.ns}}}cNvPr")
    return (element.get("name") or "") if element is not None else ""


def _shape_alt(shape: ET.Element, holder: Holder = _SPREADSHEET) -> str:
    """代替テキスト（``descr``）。**Excel は自動で入れない**ので、あれば人の言葉。

    名前（``name``）を取らないのはその裏返しである ―― ``Picture 1`` `` 図 3``
    は Excel が振った番号でしかなく、並べても行が増えるだけで何も伝わらない。
    **接続子だけは例外**で、線は代替テキストもテキスト枠も持たないぶん名前が
    言葉の唯一の在り処になる（→ :func:`_shape_name`）。
    """
    element = shape.find(f".//{{{holder.ns}}}cNvPr")
    if element is None:
        return ""
    title = (element.get("title") or "").strip()
    descr = (element.get("descr") or "").strip()
    return "／".join(p for p in (title, descr) if p)


def _embedded(picture: ET.Element) -> str:
    """``a:blip`` の ``r:embed``（画像への関係の id）。**無ければ空。**

    リンク画像（``r:link``）は取らない ―― ブックの中に実体が無く、指している先は
    **その資料を作った人の手元のパス**である。辿れないものを「取り出せなかった」
    と申告する先は :func:`_picture_note` にある。
    """
    blip = picture.find(f".//{{{_NS['a']}}}blip")
    if blip is None:
        return ""
    return blip.get(f"{{{_NS['rel']}}}embed") or ""


#: 接続の表の固定列。この後ろに :data:`_LINK_EXTRA` を**出す値があるぶんだけ**足す。
_LINK_HEAD = ("元", "向き", "先")

#: ``(属性, 見出し, 常に出すか)``。``True`` は**値が 1 つでもあれば出す**、
#: ``False`` は**行のあいだで違うときだけ出す**。
_LINK_EXTRA = (("name", "名前", True), ("dash", "線種", True),
               ("color", "線色", False), ("width", "太さ", False))


def _link_rows(links: list[_Link]) -> list[list[str]]:
    """接続の表を組む。**名前と線種は必ず出し、色と太さは違うときだけ出す。**

    線種を必ず出すのは、**「7 本とも実線でした」自体が答え**だからである ――
    凡例が「実線＝指揮命令 / 破線＝委託」と描き分けを謳っていても、図に破線が
    1 本も無いなら整理層が決められることは何も無い（資料の側に情報が無い、と
    分かる）。列が無いと、それが「線種が無い図」なのか「線種を読んでいない
    パーサ」なのか読み手から区別できない ―― 実測（kotonoha r001）ではこれが
    通し実行で唯一の「未読取」1 件になった。

    名前も同じ理由で必ず出す。**線に付いた言葉の唯一の在り処が名前**であり
    （→ :func:`_shape_name`）、その体制図では ``委託`` ``点検依頼`` がそこに
    しか無かった。``コネクタ 10`` のような自動名が混ざるが、**捨てる判断は
    整理層がする** ―― 機械が落とすと落としたことが見えなくなる。

    色と太さは**違いがあるときだけ**列にする。実測で 7 本とも ``#404040`` /
    ``1pt`` と同じで、全行同じ値の列は表を横に伸ばすだけだった（``publish`` が
    全行空の列を畳むのと同じ規律 ―― 描き分けていない属性は描き分けを伝えない）。
    """
    keep = [(field, head) for field, head, always in _LINK_EXTRA
            if (any(getattr(link, field) for link in links) if always
                else len({getattr(link, field) for link in links}) > 1)]
    return [[*_LINK_HEAD, *(head for _field, head in keep)],
            *[[link.source, link.arrow, link.target,
               *(getattr(link, field) for field, _head in keep)]
              for link in links]]


def _endpoint(connector: ET.Element, tag: str) -> str | None:
    element = connector.find(f".//{{{_NS['a']}}}{tag}")
    if element is None:
        return None
    return element.get("id") or None


def _direction(connector: ET.Element) -> str:
    """矢羽根から向きを読む。**付いていない線は無向のまま出す**（決めつけない）。"""
    head = _arrow_end(connector, "headEnd")
    tail = _arrow_end(connector, "tailEnd")
    if head and tail:
        return _BOTH
    if tail:
        return _ARROW
    if head:
        return _BACK
    return _PLAIN


def _arrow_end(connector: ET.Element, tag: str) -> bool:
    element = connector.find(f".//{{{_NS['a']}}}{tag}")
    if element is None:
        return False
    return (element.get("type") or "none") != "none"


#: ``a:prstDash/@val`` → 日本語。**訳すだけで、意味は付けない** ―― 「破線＝委託」
#: と読むのは凡例を見た整理層の仕事である。表に無い値は原文のまま出す（新しい
#: 値を勝手に「実線」へ丸めると、描き分けてあるものが同じに見える）。
_DASH = {"solid": "実線",
         "dot": "点線", "sysDot": "点線",
         "dash": "破線", "sysDash": "破線", "lgDash": "長破線",
         "dashDot": "一点鎖線", "sysDashDot": "一点鎖線",
         "lgDashDot": "長一点鎖線",
         "dashDotDot": "二点鎖線", "sysDashDotDot": "二点鎖線",
         "lgDashDotDot": "長二点鎖線"}

#: 線の太さの単位（EMU）。``a:ln/@w`` は 1pt = 12700 EMU で入っている。
_EMU_PER_POINT = 12700


def _line(connector: ET.Element,
          holder: Holder = _SPREADSHEET) -> tuple[str, str, str]:
    """接続子の線 ``(線種, 線色, 太さ)``。**分からないものは空で返す。**

    体制図・業務フロー図は「実線＝指揮命令 / 破線＝委託」のように**線種で意味を
    描き分ける**が、ここは長く ``a:stCxn`` / ``a:endCxn`` と矢羽根しか読んで
    いなかった ―― 実測（kotonoha r001）で体制図の接続 7 本はどれが委託か決め
    られず、**通し実行で唯一の「未読取」1 件**になった。

    読むのは ``cxnSp/spPr/a:ln`` に**書いてあるものだけ**である（入れ物の
    名前空間は ``holder`` が決める ―― PowerPoint も同じ形で持っている）。

    ==================  ========================================================
    ``a:prstDash``      あればその値（``dash`` → 破線）
    無い（``a:ln`` は有）  OOXML の既定どおり **実線** ―― ただし ``xdr:style``
                        （``a:lnRef`` でテーマを引く形）があるときは**言わない**
    ``a:ln`` が無い       言わない（線はテーマ側にある）
    ==================  ========================================================

    **解決しないものを断定しない。** テーマの線種を引くには ``xdr:style`` →
    ``a:lnRef`` → ``theme1.xml`` の ``lnStyleLst`` を辿る必要があり、そこまで
    やっていないので空のまま出す ―― 「実線」と埋めてしまうと、破線で描き分けて
    ある図が全部同じに見え、しかも**申告が嘘になる。**
    """
    line = connector.find(f"./{{{holder.ns}}}spPr/{{{_NS['a']}}}ln")
    if line is None:
        return "", "", ""
    dash = line.find(f"{{{_NS['a']}}}prstDash")
    if dash is not None:
        raw = dash.get("val") or ""
        style = _DASH.get(raw, raw)
    elif connector.find(f"./{{{holder.ns}}}style") is None:
        style = _DASH["solid"]
    else:
        style = ""                                     # テーマ由来 ―― 解決しない
    return style, _line_color(line), _line_width(line)


def _line_color(line: ET.Element) -> str:
    """``a:ln`` の色。**テーマ色は名前のまま**（配色を解決しない）。"""
    fill = line.find(f"{{{_NS['a']}}}solidFill")
    if fill is None:
        return ""
    srgb = fill.find(f"{{{_NS['a']}}}srgbClr")
    if srgb is not None and srgb.get("val"):
        return "#" + str(srgb.get("val")).upper()
    scheme = fill.find(f"{{{_NS['a']}}}schemeClr")
    return str(scheme.get("val") or "") if scheme is not None else ""


def _line_width(line: ET.Element) -> str:
    """``a:ln/@w``（EMU）を pt にする。**単位を付ける**（生の 12700 は読めない）。"""
    raw = line.get("w")
    if not raw:
        return ""
    try:
        points = int(raw) / _EMU_PER_POINT
    except ValueError:
        return ""
    return f"{points:g}pt"


def _shape_note(index: int, drawing: Drawing,
                readings: dict[str, ocr.Reading] | None = None,
                holder: Holder = _SPREADSHEET) -> str:
    """図形について何が取れて何が取れなかったかの申告。**黙って空を返さない。**"""
    if drawing.unreadable:
        # **XML として開けなかった描画パートは、図形 0 個と同じ形で出てくる。**
        # 数だけでも言っておかないと、図の描いてあるシートが「図の無いシート」
        # として読まれる ―― しかも `arp4 render` は撮れるので、絵にすれば読める。
        return (f"この{holder.label}の描画パート {drawing.unreadable} 本は"
                "XML として読めませんでした（ほかに"
                f" {drawing.summary}）。図が無いのではありません。"
                f"アンカー `{holder.prefix}{index}-g1` を{holder.look}読み、"
                "それでも確定できないなら out_of_scope に kind: 未読取 で"
                "宣言してください。")
    if not drawing.labels and not drawing.links:
        # **何が読めていないかで次にやることが変わる。** 文字を持ちうる図形
        # （``sp`` / 接続子 / SmartArt）が 1 つも無いシートに「業務フローが
        # 描かれていることが多い」と言うと、絵にして読んでも何も出てこない。
        drawn = drawing.shapes or drawing.connectors or drawing.diagrams
        if not drawn and drawing.series and drawing.charts == drawing.total:
            # **グラフしか無いシートで「中身は取れていません」と言わない。**
            # 隣の `k1` にタイトルも参照範囲も出ているのに読めていないと言うと、
            # 申告のほうが信用されなくなる（:func:`_chart_note` が別に言う）。
            return (f"この{holder.label}には {drawing.summary} があります"
                    f"（中身は `{holder.prefix}{index}-k1`）。セルの値は 1 つもありません。")
        if not drawn and drawing.pictures and not drawing.objects:
            # **画像しか無いシートに `arp4 render` を勧めない。** 撮り直しても
            # 絵が絵のまま出るだけで、中の文字は読めるようにならない ――
            # 案内どおりにやって届かない申告は、申告しないのと同じである
            # （表紙の会社ロゴ・スキャンした帳票見本がまさにこの形になる）。
            return (f"この{holder.label}には {drawing.summary} があります。"
                    "文字を持つ図形はありません。"
                    + _picture_note(index, drawing, readings, holder))
        return (f"この{holder.label}には {drawing.summary} があります。"
                + ("テキストも接続も取れていません（業務フロー・ER 図・"
                   "状態遷移図・画面レイアウトはここに描かれていることが多い）。"
                   if drawn else
                   "文字を持つ図形はありません。中身は取れていません。")
                + f"アンカー `{holder.prefix}{index}-g1` を{holder.look}"
                "読んでください。それでも確定できないなら out_of_scope に "
                "kind: 未読取 で宣言してください。")

    got: list[str] = []
    if drawing.labels:
        got.append(f"{len(drawing.labels)} 個からテキスト（`{holder.prefix}{index}-g1`）")
    if drawing.links:
        got.append(f"接続 {len(drawing.links)} 本（`{holder.prefix}{index}-c1`）")
    note = f"この{holder.label}には {drawing.summary} があり、{'、'.join(got)}を取り出しました。"

    if drawing.diagrams:
        # **SmartArt は箱の文字しか取れない。** 親子・順序は内部の点への参照に
        # なっており、どれが「A の次が B」かはレイアウト定義まで解かないと決まらない。
        note += ("SmartArt は箱の文字だけで、箱どうしの繋がり（親子・順序）は"
                 "取れていません。")
    if drawing.charts and not drawing.series:
        # 系列もタイトルも読めなかったグラフだけがここに来る（読めたぶんは
        # :func:`_chart_note` が別に言う）。
        note += (f"グラフ {drawing.charts} 個の中身は取れていません"
                 f"（{holder.chart_data}）。")
    if drawing.objects:
        note += (f"埋め込みオブジェクト {drawing.objects} 個の中身は取れていません"
                 "（Word・PDF などが貼り込まれている場合、元ファイルを "
                 "sources/ に足してください）。")
    if drawing.unnamed:
        # **繋がってはいるが、相手が文字を持たない。** ゾーンの囲み枠・装飾へ
        # 落ちる線で、構成図では珍しくない ―― 下の目分量の線と混ぜて数えて
        # いた頃は「線を目分量で置いた図」と申告していたが、こちらは**絵にすれば
        # どこへ向かう線か読める**（次にやることが正反対である）。
        note += (f"ただし接続子 {drawing.unnamed} 本は、繋がってはいますが"
                 "相手の図形が文字を持ちません（ゾーンの囲み枠・装飾）。"
                 f"{holder.trace}")
    if drawing.loose:
        # **繋がっていない線は取れない。** 座標から当てるのは意味の判断になる。
        note += (f"接続子 {drawing.loose} 本はどこにも繋がっていません"
                 "（線を目分量で置いた図）。両端の id が資料に無いので、"
                 "どこからどこへの線かは絵にしても決まりません。")
    if drawing.pictures:
        note += _picture_note(index, drawing, readings, holder)
    note += ("取れていないのは配置です。枠で括られたゾーン・段組み・注記の"
             f"位置が仕様なら、{holder.look}読み、それでも確定できない"
             "ときに out_of_scope に kind: 未読取 で宣言してください。")
    return note


def _picture_note(index: int, drawing: Drawing,
                  readings: dict[str, ocr.Reading] | None = None,
                  holder: Holder = _SPREADSHEET) -> str:
    """**画像は撮り直しても読めるようにならない。取り出して渡し、字は読む。**

    :mod:`arp4.render` は「人が見て読む」ための絵であって、貼り付け画像の中の
    文字を機械が読むわけではない ―― 図形と違い、画像は**絵にする以前から絵**
    である。ここで render を勧めていたぶん、スキャンした帳票見本・手書きの
    赤入れ・会社ロゴしか無いシートで読み手を空振りさせていた。

    **だから「読めません」で終わらせない。** 機械が読めないことと、**この資料が
    誰にも読まれないこと**は別である ―― 整理層（エージェント）は画像を開けるので、
    実体を ``images/`` へ出して名前を ``s<番号>-i1`` に置く。長いあいだここは
    「中身は取れていません」の 1 行だけで、貼り付け画像に描かれた業務フローは
    **ブックの中に入ったまま**だった。

    **字は機械も読む**（:mod:`arp4.ocr` ―― ``s<番号>-o1``）。ここも「絵は読めない」
    の一言で片付けていたが、貼ってある画像の多くは**表・画面・帳票をそのまま
    撮ったもの**で、絵ではなく字である。読めた枚数を言うのは、**残りが本当に絵
    である**ことをそれで示すためでもある。

    **代替テキストの無い画像は分けて数える。** 中身が読めないのは同じでも、
    書いてあれば「何の画像か」は分かる ―― 無ければそれすら分からないので、
    次にやること（元の画像ファイルか紙を当たる）が変わる。
    """
    blind = drawing.pictures - drawing.picture_alts
    missing = drawing.pictures - len(drawing.media)
    note = (f"貼り付け画像 {drawing.pictures} 枚は絵のままです"
            "（表のスクリーンショットなら、セルの値としては 1 つも取れていません）。"
            f"{holder.reshoot}"
            "図形と違い、画像は絵にする以前から絵です。")
    if drawing.media:
        note += (f"実体は `{holder.prefix}{index}-i1` に出してあります（`images/` の中）。"
                 "開いて読むのは整理層の仕事です。読み取った内容を整理結果へ"
                 f"書くときは、出典に `{holder.prefix}{index}-i1` を指してください。")
        note += _ocr_said(index, drawing, readings, holder)
    if drawing.picture_alts:
        note += (f"うち {drawing.picture_alts} 枚には代替テキストがあるので、"
                 f"何の画像かは `{holder.prefix}{index}-a1` でも分かります。")
    if blind and not drawing.media:
        note += (f"うち {blind} 枚は代替テキストも無く、何の画像かも資料からは"
                 "分かりません。元の画像ファイル（または紙）を当たるか、"
                 "out_of_scope に kind: 未読取 で宣言してください。")
    if missing:
        # **実体の無い画像がある。** リンク画像（``r:link``）は指している先が
        # 作った人の手元のパスなので、こちらからは辿れない ―― 黙ると、出した
        # 枚数と貼ってある枚数の差が誰にも説明されないまま残る。
        note += (f"うち {missing} 枚は実体を取り出せませんでした"
                 "（ブックの外を指すリンク画像です）。元の画像ファイルを"
                 "当たるか、out_of_scope に kind: 未読取 で宣言してください。")
    return note


def _ocr_said(index: int, drawing: Drawing,
              readings: dict[str, ocr.Reading] | None,
              holder: Holder = _SPREADSHEET) -> str:
    """OCR が何枚から字を読めたかの申告。**枚数で言う**（中身は ``o1`` にある）。

    **「読めた」も申告である。** 読めなかったものを数えるのと同じ理由で、
    読めたものも数える ―― 3 枚のうち 1 枚からしか字が出ていないことは、
    「残り 2 枚は開いて見るしかない」という次の一手そのものである。

    **理由は 1 つにまとめる。** 環境が理由（言語パックが無い）なら画像 20 枚
    ぶん同じ文が並んでも分かることは増えないので、代表を 1 つだけ出す。
    """
    if readings is None:
        return ("画像の中の文字は読みにいっていません（`--no-ocr`）。"
                f"`{holder.prefix}{index}-o1` にもそう書いてあります。")
    parts = {part for part, _ in drawing.media}
    got = [readings[part] for part in sorted(parts) if part in readings]
    if not got:
        return ""                                  # 読む相手が 1 枚も無かった
    read = [one for one in got if one.lines]
    troubled = [one for one in got if one.trouble]
    if read:
        note = (f"うち {len(read)} 枚からは Windows OCR が文字を読み出しました"
                f"（`{holder.prefix}{index}-o1`）。読み違えが混ざります（`ORDER-001` が "
                "`ORDER-OOI` になるなど）ので、値として使う前に画像そのものを"
                "確かめてください。")
    else:
        note = ("Windows OCR では文字を 1 つも読めませんでした"
                f"（`{holder.prefix}{index}-o1`）。")
    if troubled:
        note += (f"うち {len(troubled)} 枚は読みにいって失敗しました"
                 f"（{troubled[0].trouble}）。")
    return note


#: 申告に番地を並べる上限。**全部並べると申告が本文より長くなる**ので、
#: 超えたぶんは数だけ言う（何件あるかは必ず言う）。
_MAX_LIST = 20


def _listed(pairs: list[tuple[str, str]]) -> list[tuple[str, str]]:
    """番地の一覧を上限で切る。**切ったことを必ず言う。**

    番地が 20 個で切れているのに黙っていると、「読めなかったのはこの 20 個だけ」
    と読める ―― 申告が実際より小さく見えるのは、申告しないのと同じくらい悪い。

    切るのは**番地だけを並べる申告**（数式・エラー値）である。コメントとリンクは
    切らない ―― あれは「読めなかったものの目録」ではなく**中身そのもの**で、
    セルの値を 20 個で打ち切らないのと同じ理屈である。
    """
    if len(pairs) <= _MAX_LIST:
        return list(pairs)
    return [*pairs[:_MAX_LIST],
            ("…", f"ほか {len(pairs) - _MAX_LIST} 個（番地はここに並べきれません）")]


def _formula_note(index: int, blanks: list[str]) -> str:
    """**空欄に見えるが、資料に値が無いのではない。**

    ``data_only=True`` が返すのは「Excel が最後に計算して保存した値」である。
    ツールが書き出したブック・LibreOffice で保存したブック・計算を手動にした
    まま保存したブックにはその値が入っておらず、openpyxl は黙って None を返す。
    ここを申告しないと、整理層は**合計欄が空の表**を「資料に記載なし」と読む。
    """
    where = "、".join(blanks[:_MAX_LIST])
    if len(blanks) > _MAX_LIST:
        where += f" ほか {len(blanks) - _MAX_LIST} 個"
    return (f"このシートには計算結果が保存されていない数式セルが "
            f"{len(blanks)} 個あります（{where}／アンカー `s{index}-f1`）。"
            "空欄に見えますが、資料に値が無いのではなく機械が読めていません。"
            "Excel で開いて上書き保存すると値が入ります。それができない資料なら、"
            "out_of_scope に kind: 未読取 で宣言してください。")


def _error_note(index: int, errors: list[tuple[str, str]]) -> str:
    """**``#REF!`` は値ではない。** 表にはそう見えているが、仕様は読み取れていない。

    数式の申告（:func:`_formula_note`）と分けてあるのは、**次にやることが
    正反対**だからである ―― あちらは「Excel で開いて保存し直せば値が入る」だが、
    こちらは開き直しても直らない（資料そのものが壊れている）。混ぜて出すと、
    保存し直して直らなかった側が「やっても無駄だった」で放置される。
    """
    where = "、".join(ref for ref, _ in errors[:_MAX_LIST])
    if len(errors) > _MAX_LIST:
        where += f" ほか {len(errors) - _MAX_LIST} 個"
    kinds = "・".join(sorted({value for _, value in errors}))
    return (f"このシートにはエラー値が {len(errors)} 個あります"
            f"（{kinds}／{where}／アンカー `s{index}-e1`）。"
            "表にはそのまま出ていますが値ではありません。参照先が消えた・"
            "引き当てが見つからないという意味で、その欄の仕様は資料から"
            "読み取れていません。開き直しても直りません（数式の未計算とは"
            "別物です）。out_of_scope に kind: 未読取 で宣言してください。")


def _comment_note(index: int, comments: list[tuple[str, str]]) -> str:
    """**コメントは表に出てこない。** セルの上に浮いているので、黙ると丸ごと消える。

    実案件の設計書では、決定の理由・積み残し・レビュー指摘がここに溜まる
    ―― 表の本文より新しいことが珍しくない。
    """
    return (f"セルのコメント（メモ）が {len(comments)} 件あります"
            f"（アンカー `s{index}-m1`）。表には出てこない補足で、"
            "決定の理由・積み残し・レビュー指摘が入っていることがあります"
            "（本文より新しいことがあります）。")


def _link_note(index: int, links: list[tuple[str, str]]) -> str:
    """**表に出ているのは表示文字列だけ。** リンク先は別のブックを指しうる。

    ここを黙ると、目次シートが「語の一覧」にしか見えない ―― 実際には
    **どの資料がまだ手元に無いか**の一覧である。
    """
    outside = sorted({target.split("#")[0] for _, target in links
                      if not target.startswith("#") and target.split("#")[0]})
    note = (f"リンクが {len(links)} 件あります（アンカー `s{index}-l1`）。"
            "表に出ているのは表示文字列だけなので、リンク先はここにしか"
            "ありません。")
    if outside:
        note += (f"うち {len(outside)} 件はこのブックの外を指しています"
                 f"（{'、'.join(outside[:_MAX_LIST])}）"
                 "。まだ集めていない資料なら sources/ に足してください。")
    return note


def _print_note(index: int, setup: list[tuple[str, str]]) -> str:
    """**紙にしか出ないものは、表を読むだけの人には存在しない。**"""
    titles = [value for label, value in setup if label == "印刷タイトル行"]
    note = (f"印刷したときだけ見えるものが {len(setup)} 件あります"
            f"（アンカー `s{index}-p1`）。ヘッダ・フッタはどのページにも"
            "出ているのにセルには 1 つも書かれていません。文書番号・版・"
            "機密区分がここにしか無いことがあります（`&P` `&N` は刷るときに"
            "決まるので、書いてあるまま出しています）。")
    if titles:
        note += (f"繰り返し印刷する見出し行が指定されています"
                 f"（{'、'.join(titles)}）。資料の作成者が「ここまでが見出し」"
                 "と決めた範囲です（機械は当てていません）。")
    return note


def _struck_note(index: int, cells: Cells) -> str:
    """**取り消し線は画面に見えているのに、表では見えなくなる。**

    非表示の行・列と同じ扱いにしてある ―― 落とさず、混ざっていると言う。
    「取り消し線＝廃止」と決めるのは意味の判断なので、機械は言わない。
    """
    struck = cells.struck
    where = "、".join(ref for ref, _ in struck[:_MAX_LIST])
    if len(struck) > _MAX_LIST:
        where += f" ほか {len(struck) - _MAX_LIST} 個"
    partly = (f"うち {cells.partly_struck} 個はセルの一部だけに掛かって"
              "います（1 つのセルの中で書式が変わるもの）。残りの文字は"
              "生きていますので、行ごと廃止とは読めません。消してある文字は"
              "値の隣に書いてあります。" if cells.partly_struck else "")
    return (f"取り消し線の掛かったセルが {len(struck)} 個あります"
            f"（{where}／アンカー `s{index}-d1`）。画面では消してありますが、"
            "表の上では生きた行と見分けが付きません。値はそのまま出して"
            "います。"
            + partly
            + "廃止・消し忘れ・未実装のどれなのかは資料からは決まらない"
            "ので、機械は判断していません。廃止済みなら out_of_scope で"
            "宣言してください。")


def _chart_note(index: int, drawing: Drawing) -> str:
    """**グラフの値は別のシートにある。** どのシートかまで言う。"""
    where = sorted({ref.split("!")[0].strip("'")
                    for _, _, cat, val in drawing.series
                    for ref in (cat, val) if "!" in ref})
    note = (f"グラフ {drawing.charts} 個からタイトル・系列名・参照範囲を"
            f"取り出しました（アンカー `s{index}-k1`）。グラフに描かれている"
            "数そのものはここにありません。系列が指しているのは"
            "セルの範囲です。")
    if where:
        note += (f"参照先は{'・'.join(f'「{one}」' for one in where[:_MAX_LIST])}"
                 "シートなので、値が要るならそちらのパース結果を読んでください。")
    return note


def _hidden_note(cells: Cells) -> str:
    """**非表示の行・列も読んでいる**ことの申告。

    非表示シートは読まないのに行・列は読むのは、**壊れ方が違う**からである。
    非表示シートはブック全体の作業用が大半だが、行・列の非表示はアウトラインの
    折りたたみ（＝生きている仕様が畳まれているだけ）が普通にある。
    「非表示だから廃止された仕様だ」と決めるのは**意味の判断**なので機械はやらない
    ―― 代わりに、混ざっていることを言う。

    **潰れた行だけは別に数える**（:func:`_hidden`）。「再表示」で戻らないのは
    行だけなので、そこにだけ別の案内が要る ―― 幅 0 の列は戻るので足さない。
    """
    where = []
    if cells.hidden_rows:
        where.append(f"行 {cells.hidden_rows} 行")
    if cells.hidden_columns:
        where.append(f"列 {cells.hidden_columns} 列")
    note = (f"非表示の{'・'.join(where)}にある {cells.hidden} セルも読み込んでいます。"
            "画面には出ていない値なので、旧版の残骸・作業用メモが混ざっている"
            "ことがあります（折りたたまれているだけの生きた仕様のこともあるので、"
            "機械は落としません）。廃止済みなら out_of_scope で宣言してください。")
    if cells.crushed_rows:
        note += (f"うち {cells.crushed_rows} 行は高さ 0 で潰されています。"
                 "右クリックの「再表示」では戻りません（行の高さを入れ直して"
                 "ください）。ツールが書き出したブックと、境目を誤って詰めた"
                 "ブックで起きます。")
    return note


def _rel_targets(archive: zipfile.ZipFile, name: str,
                 kind: str = "") -> dict[str, str]:
    """``.rels`` の Id → Target。``kind`` を渡すと**関係の種別で絞る**。"""
    root = ET.fromstring(archive.read(name))
    return {r.get("Id") or "": r.get("Target") or ""
            for r in root.findall("pkg:Relationship", _NS)
            if not kind or (r.get("Type") or "").endswith(kind)}


def _rel_types(archive: zipfile.ZipFile, name: str) -> dict[str, str]:
    """``.rels`` の Id → Type。**1 度読んだ ``.rels`` から複数の種別を拾う**とき用。"""
    root = ET.fromstring(archive.read(name))
    return {r.get("Id") or "": r.get("Type") or ""
            for r in root.findall("pkg:Relationship", _NS)}


def _part_dir(part: str) -> str:
    return part.rsplit("/", 1)[0] if "/" in part else ""


def _resolve(base: str, target: str) -> str:
    """``../drawings/drawing1.xml`` のような相対参照をパッケージ内のパスに直す。

    **返すのは zip の中のキーだけである。** ``..`` がパッケージの外へ出ても
    実体のパスにはならない（そこで頭打ちにする）―― arp4 は zip を**展開せず**、
    ここで作った名前は ``ZipFile.read`` に渡すだけなので、細工した ``.rels`` で
    ディスクの外を読ませることはできない。**確かめたうえで、防御は足していない**
    ―― 足すと「zip の外を指す関係は資料の側にもある（ブックを移動した跡）」を
    エラーとして扱うことになり、読める資料が読めなくなる。
    """
    if not target:
        return ""
    if target.startswith("/"):
        return target.lstrip("/")
    parts = [p for p in base.split("/") if p]
    for piece in target.split("/"):
        if piece in ("", "."):
            continue
        if piece == "..":
            if parts:
                parts.pop()
            continue
        parts.append(piece)
    return "/".join(parts)


@dataclass
class Cells:
    """1 シートぶんの非空セルと、**セルの上に載っていて表に出ないもの**。

    コメント・リンク・エラー値をここで一緒に集めるのは、**実在セルを走るのが
    1 回で済む**ためである。別々に回すと、使用範囲が膨らんだブック（実案件で
    ごく普通にある）でその回数だけ遅くなる。
    """

    values: dict[tuple[int, int], str] = field(default_factory=dict)
    hidden: int = 0                                    # 非表示の行・列にあるセル数
    hidden_rows: int = 0
    hidden_columns: int = 0
    #: そのうち**高さ 0 で潰された**行。**「再表示」では戻らない**（実測）。
    crushed_rows: int = 0
    #: ``(番地, 画面に見えているエラー値)``。値ではないが空欄でもない。
    errors: list[tuple[str, str]] = field(default_factory=list)
    #: ``(番地, 本文)``。**表に出てこない**ので、黙ると丸ごと消える。
    comments: list[tuple[str, str]] = field(default_factory=list)
    #: ``(番地, リンク先)``。表にあるのは表示文字列だけである。
    links: list[tuple[str, str]] = field(default_factory=list)
    #: ``(番地, 値)``。**画面では消してあるのに、表では生きた行と見分けが付かない。**
    struck: list[tuple[str, str]] = field(default_factory=list)
    #: そのうち**セルの一部だけ**に掛かっているもの（リッチテキスト）の数。
    #: 分けて数えるのは**残りが生きている**からで、まるごと消した欄とは
    #: 次に読む人の読み方が変わる。
    partly_struck: int = 0


class _NoMerge(NamedTuple):
    """救出モードの worksheet は結合を持たない（:func:`_open`）。"""

    ranges: tuple = ()


_NO_MERGE = _NoMerge()


def _cells(sheet: Any) -> Cells:
    """非空セル。**縦に広がる結合は下へ展開する。**

    走るのは**実在するセルだけ**である。``iter_rows()`` は使用範囲の矩形を
    端から端まで回すので、遠くに書式だけのセルが 1 つ残っているだけで
    400 万回の空回りになる（実測 7.5 秒 → 0.03 秒）。削除済みデータの残骸で
    使用範囲が膨らんだブックは実案件でごく普通に出てくる。

    **値の無いセルも素通りさせない。** コメントだけが付いたセル・リンクだけが
    張られたセルは値が空なので、値で足切りすると存在ごと消える。
    """
    found = Cells()
    invisible = _hidden(sheet)
    rows, columns = invisible.rows, invisible.columns
    extra: list[tuple[int, int, str, str]] = []        # (行, 列, 種別, 本文)
    for (row, column), cell in _live(sheet).items():
        comment = _comment(cell)
        if comment:
            extra.append((row, column, "m", comment))
        link = _link(cell)
        if link:
            extra.append((row, column, "l", link))
        text = _value(cell)
        if not text:
            continue
        if getattr(cell, "data_type", "") == "e":
            extra.append((row, column, "e", text))
        if _struck(cell):
            extra.append((row, column, "s", text))
        if row in rows or column in columns:
            found.hidden += 1
        found.values[(row, column)] = text

    # **並びは番地の順**（差分をノイズにしない／読み手が表と突き合わせられる）。
    for row, column, kind, text in sorted(extra):
        target = {"e": found.errors, "m": found.comments,
                  "l": found.links, "s": found.struck}[kind]
        target.append((f"{column_name(column)}{row}", text))

    if found.hidden:
        found.hidden_rows = len({r for r, _ in found.values if r in rows})
        found.hidden_columns = len({c for _, c in found.values if c in columns})
        found.crushed_rows = len(
            {r for r, _ in found.values if r in invisible.crushed_rows})

    # **展開するのは値のある範囲までである。** 結合は列を丸ごと選んで掛けられる
    # （`A2:A1048576`）ので、素直に最終行まで埋めると **3 行の表から 100 万行の
    # 表**が生える ―― 画面には 3 行しか見えていないのだから、忠実性の回復では
    # なく捏造である（実測でも 1 シート 1.6 秒・10 万行の Markdown になっていた）。
    # 表の外に行は無いので、他のセルが 1 つでもある最終行で止める。
    last = max((row for row, _ in found.values), default=0)
    for merged in getattr(sheet, "merged_cells", _NO_MERGE).ranges:
        if merged.max_row == merged.min_row:        # 横結合（1 行）は広げない
            continue
        # **広げるのは下だけである。** 区分の列を 2 列ぶんまとめて縦に結合するのは
        # 区分と小区分を持つテーブル定義書でごく普通の書き方で、画面ではその区分が
        # 全行に掛かって見えている ―― 幅 1 だけを展開していた頃は 2 行目以降が
        # 空欄になり、整理層には「区分の無い行」に見えた。横へ広げないのは、
        # **横結合は同上ではなく表題**だからである（値は 1 つしか見えていない）。
        value = found.values.get((merged.min_row, merged.min_col))
        if not value:
            continue
        for row in range(merged.min_row + 1, min(merged.max_row, last) + 1):
            found.values.setdefault((row, merged.min_col), value)
    return found


def _live(sheet: Any) -> dict[tuple[int, int], Any]:
    """実在するセルだけ。openpyxl が内部辞書を持たない形なら素直に回す。"""
    cells = getattr(sheet, "_cells", None)
    if isinstance(cells, dict):
        return cells
    # 救出モード（:func:`_open`）は矩形で返してくるので、**穴埋めの
    # ``EmptyCell`` が混ざる** ―― 番地を持たないので、そこで落とす。
    return {(cell.row, cell.column): cell
            for row in sheet.iter_rows() for cell in row
            if getattr(cell, "row", None)}


class Invisible(NamedTuple):
    """画面に出ていない行・列。**行だけは、潰れたものを分けて持つ。**

    分けるのは**戻し方が違うときだけ**である（:func:`_hidden_note`）。幅 0 の
    列は「再表示」で戻るので、隠した列と同じ扱いでよい。
    """

    rows: set[int]
    columns: set[int]
    #: 高さ 0 で潰れた行（``rows`` にも入っている）。**「再表示」では戻らない。**
    crushed_rows: set[int]


def _hidden(sheet: Any) -> Invisible:
    """非表示の行番号・列番号。**落とすためではなく数えるため**に集める。

    **隠すフラグ（``hidden``）だけを見ていたぶん、高さ 0 の行を「見えている行」
    として出していた。** 行の高さ 0・列の幅 0 は Excel が普通に受け取る値で、
    ``hidden`` は立たないまま画面から消える ―― ツールが書き出した設計書
    （計算結果が保存されていないのと同じ側）と、境目を誤って詰めたブックで
    起きる。非表示行の裏返しの壊れ方で、**申告にも上がらない**ぶんこちらのほうが
    静かである（`未読取` を宣言する先が無い）。

    **数え分けるのは行だけである。** 実物の Excel で確かめたところ、右クリックの
    「再表示」（``Hidden = False``）で幅 0 の列は既定幅（8.11）に戻るのに、
    **高さ 0 の行は 0 のまま戻らない** ―― 行にだけ「高さを入れ直してください」が
    要る。列まで同じ文句を付けると、**戻る操作を「戻りません」と言う**ことになる
    （取れない線 2 種類を同じ文句で申告していたのと同じ失敗である）。
    """
    from openpyxl.utils import column_index_from_string

    rows: set[int] = set()
    crushed_rows: set[int] = set()
    if not hasattr(sheet, "row_dimensions"):        # 救出モード（:func:`_open`）
        return Invisible(rows, set(), crushed_rows)
    for index, dimension in sheet.row_dimensions.items():
        if getattr(dimension, "hidden", False):
            rows.add(index)
        elif getattr(dimension, "height", None) == 0:
            rows.add(index)
            crushed_rows.add(index)
    columns: set[int] = set()
    for key, dimension in sheet.column_dimensions.items():
        if not (getattr(dimension, "width", None) == 0
                or getattr(dimension, "hidden", False)):
            continue
        try:
            first = dimension.min or column_index_from_string(key)
            last = dimension.max or first
        except ValueError:                          # 見たことのないキー
            continue
        columns.update(range(first, last + 1))
    return Invisible(rows, columns, crushed_rows)


#: ヘッダ・フッタの持ち主。**既定は「奇数ページ」**で、先頭ページ・偶数ページは
#: 別に設定されているときだけ中身を持つ（持っていれば出す）。
_PAGE_TEXTS = (("ヘッダ", "oddHeader"), ("フッタ", "oddFooter"),
               ("ヘッダ（偶数ページ）", "evenHeader"),
               ("フッタ（偶数ページ）", "evenFooter"),
               ("ヘッダ（先頭ページ）", "firstHeader"),
               ("フッタ（先頭ページ）", "firstFooter"))


def _print_setup(sheet: Any) -> list[tuple[str, str]]:
    """``(場所, 書いてあること)``。**紙にしか出ないのに、そこにしか無い。**

    日本の設計書は綴じて配るものなので、文書番号・版・機密区分・ページ番号は
    フッタに置かれる ―― **どのページにも出ている**のに、セルには 1 つも
    書かれていない。表だけを読むと、その 1 冊がどの文書なのかが落ちる。

    ``&P`` ``&N`` ``&D`` は**書いてあるまま出す**。ページ番号は刷ってみないと
    決まらない値で、機械が埋めれば資料に無い数を書くことになる。

    **印刷タイトル行は「どこまでが見出しか」の申告である。** 機械が当てるのは
    意味の判断なので :mod:`arp4.render` は ``--title-rows`` を人に聞くが、
    ここに書いてあるのは**資料の作成者がそう決めた**という事実なので取る
    （何行目までかを決めるのは、それを読んだ整理層である）。
    """
    found: list[tuple[str, str]] = []
    for label, attribute in _PAGE_TEXTS:
        holder = getattr(sheet, attribute, None)
        for side, part in (("左", "left"), ("中", "center"), ("右", "right")):
            text = str(getattr(getattr(holder, part, None), "text", "")
                       or "").strip()
            if text:
                found.append((f"{label}{side}", text))
    for label, value in (("印刷タイトル行", getattr(sheet, "print_title_rows", None)),
                         ("印刷タイトル列", getattr(sheet, "print_title_cols", None)),
                         ("印刷範囲", getattr(sheet, "print_area", None))):
        if isinstance(value, (list, tuple)):        # 印刷範囲は複数持てる
            value = ",".join(str(one) for one in value)
        text = str(value or "").strip()
        if text:
            found.append((label, text))
    return found


def _struck(cell: Any) -> bool:
    """取り消し線が掛かっているか。**書式のうち、ここだけは値を偽る。**

    ここには長く「罫線・塗り・フォント・取り消し線は取らない ―― 書式から意味を
    決める（取り消し線＝廃止）のは整理層の仕事である」と書いてあった。
    **前半は正しいが、後半が成り立っていなかった。** 整理層に届いているのは
    ``受注区分`` という文字だけで、画面に見えている ``~~受注区分~~`` ではない
    ―― 判断を任せた先が、判断の材料を受け取っていない。

    塗り・太字・色と違うのは、取り消し線が**その文字が消されていることを表す
    唯一の書式**だという点である。強調は値を偽らないが、取り消し線は偽る
    ―― 日付を ``2026-08-02 00:00:00`` と出すのをやめたのと同じ理由で取る。

    **それでも「廃止」とは言わない。** 出すのは「この番地に取り消し線が
    掛かっている」という転記だけで、消し忘れなのか廃止なのか未実装なのかは
    整理層が決める。
    """
    if not getattr(cell, "has_style", False):
        return False                                # 既定書式のセルは見に行かない
    font = getattr(cell, "font", None)
    return bool(getattr(font, "strike", False))


def _comment(cell: Any) -> str:
    """セルのコメント（メモ）。**記入者も出す** ―― 誰の指摘かで重みが変わる。"""
    note = getattr(cell, "comment", None)
    text = (getattr(note, "text", "") or "").strip() if note is not None else ""
    if not text:
        return ""
    author = (getattr(note, "author", "") or "").strip()
    return f"{author}: {text}" if author else text


def _link(cell: Any) -> str:
    """ハイパーリンクの**行き先**。表示文字列（＝セルの値）とは別物である。

    ブックの中を指す ``#シート!A1`` は ``location`` に、外のファイル・URL は
    ``target`` に入る。**両方あることもある**（別ブックの特定シートを指す形）ので、
    片方だけを見ると行き先が半分になる。
    """
    link = getattr(cell, "hyperlink", None)
    if link is None:
        return ""
    target = (getattr(link, "target", "") or "").strip()
    location = (getattr(link, "location", "") or "").strip()
    if target and location:
        return f"{target}#{location}"
    if location:
        return location if location.startswith("#") else f"#{location}"
    return target


def _value(cell: Any) -> str:
    """**画面に見えている表記**へ直す。値の解釈ではない。

    ``str()`` に任せると日付が ``2026-08-02 00:00:00``、真偽が ``True``、
    パーセントが ``0.153`` になる ―― どれも**資料にそう書いてあったことは無い**。
    とくにパーセントは、``0.153`` と ``15.3%`` で整理層が別の値を読む。

    直すのは**表記が値そのものを偽っているものだけ**である。``￥1,200`` を
    ``1200`` と出すのは桁区切りと通貨記号が落ちるだけで数は同じなので触らない
    （書式の再現を始めると、どこまでやるかの線が引けなくなる）。

    **行頭の空白は落とさない。** 項目定義書の「項目名」列は、字下げで親子を
    表すのが日本の設計書の慣習である（``受注ヘッダ`` / ``　　受注番号``）。
    まとめて ``strip()`` していたぶん、**階層がまるごと平らになっていた** ――
    親と子が同じ深さの項目として並ぶ。行末の空白は画面に出ないので落とす。
    """
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):                     # bool は int より先に見る
        return "TRUE" if value else "FALSE"
    if isinstance(value, dt.datetime):
        if value.time() == dt.time(0):              # 日付だけのセル
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S" if value.second else "%H:%M")
    if isinstance(value, dt.timedelta):             # [h]:mm の経過時間
        minutes, seconds = divmod(int(value.total_seconds()), 60)
        hours, minutes = divmod(minutes, 60)
        return f"{hours}:{minutes:02d}" + (f":{seconds:02d}" if seconds else "")
    if isinstance(value, (int, float)):
        percent = _percent(value, getattr(cell, "number_format", "") or "")
        if percent:
            return percent
        return _number(value)
    return str(value).rstrip().lstrip("\r\n")


#: Excel 自身が持てる有効桁。**画面に出るのもここまで**である。
_SHOWN_DIGITS = 15


def _number(value: float | int) -> str:
    """**Excel が保存する数と、Excel が画面に出す数は違う。**

    二進の浮動小数点なので ``=1200*0.08`` の結果は ``96.00000000000001`` として
    保存されるが、Excel は有効数字 15 桁までしか持たない（持てない）ので画面には
    ``96`` と出ている ―― 素の ``str()`` で出すと、**資料に一度も書かれていない
    桁が仕様になる**。日付を ``2026-08-02 00:00:00`` と出すのをやめたのと同じ
    基準（表記が値そのものを偽っている）である。

    15 桁を超える整数（会員番号を数値で持った表）では、Excel は下位の桁を 0 で
    埋めて見せる ―― ``1234567890123456789`` は画面上も ``1234567890123460000``
    である。**指数表記へ落とさない**のはそのためで、小さいほうの数
    （``1e-07``）はもともと画面も指数なので触らない。

    **整数にも同じ規律を掛ける。** 15 桁を超える数が整数として保存されている
    ことはあるが（``9007199254740992``）、Excel はそれも 15 桁までしか見せない。
    型で扱いを変えると、**同じ表の隣り合う 2 行が別の桁数で出る**。文字列として
    持たれた番号（``'1234567890123456789'``）は数ではないので触らない ――
    そちらは 19 桁のまま画面に出ている。
    """
    # **ほとんどのセルはここで返る。** 15 桁までの整数（桁・件数・コード）は
    # Excel も丸めないので、浮動小数点を経由させる意味が無い。
    if isinstance(value, int) and abs(value) < 10 ** _SHOWN_DIGITS:
        return str(value)
    try:
        text = f"{value:.{_SHOWN_DIGITS}g}"
    except (OverflowError, ValueError):
        return str(value)               # 資料が壊れていてもパースを止めない
    _, _, exponent = text.partition("e")
    if exponent and int(exponent) > 0:
        return format(Decimal(text), "f")
    return text


#: パーセント書式（``0.0%`` / ``0%``）。**書式の ``%`` はパーセントとは限らない。**
#: 引用符の中（``#,##0"%表示ではない"``）は塞いであったが、``\`` の**エスケープ**
#: （``0\%`` ―― 次の 1 文字をそのまま出す）は素通りしていた ―― 画面に ``15%`` と
#: 出ている欄が ``1500%`` になり、**資料に無い数**が仕様になる。
_PERCENT = re.compile(r'^(?:\\.|"[^"]*"|[^"%\\])*%')


def _percent(value: float, number_format: str) -> str:
    """``0.153`` ＋ ``0.0%`` → ``15.3%``。**小数点以下の桁数も書式に合わせる。**"""
    if not _PERCENT.search(number_format):
        return ""
    fraction = number_format.split("%")[0]
    digits = len(fraction.split(".")[-1].rstrip()) if "." in fraction else 0
    digits = min(digits, 10)
    return f"{round(value * 100, digits):.{digits}f}%"


def _regions(cells: dict[tuple[int, int], str]) -> list[list[tuple[int, int]]]:
    """連結成分。空白 1 行／列まではスペーサーとみなして繋ぐ。

    **並びは 1 度だけ作る。** 塊の起点を毎回 ``min(remaining)`` で探していたが、
    それは残りセル全部を走るので、**塊の数 × セルの数**になる ―― 方眼紙で描いた
    画面レイアウトや、印が飛び飛びに散る CRUD マトリクスは 1 枚で数千の塊になり、
    5,000 セルのシート 1 枚に 0.30 秒かかっていた（0.02 秒になる）。塊の中身は
    変わらないので、出来上がりは 1 セルも違わない。
    """
    remaining = set(cells)
    regions: list[list[tuple[int, int]]] = []

    for start in sorted(cells):
        if start not in remaining:
            continue
        stack, region = [start], []
        remaining.discard(start)
        while stack:
            row, column = stack.pop()
            region.append((row, column))
            for dr in range(-_GAP, _GAP + 1):
                for dc in range(-_GAP, _GAP + 1):
                    neighbour = (row + dr, column + dc)
                    if neighbour in remaining:
                        remaining.discard(neighbour)
                        stack.append(neighbour)
        regions.append(region)

    regions.sort(key=lambda r: (min(c[0] for c in r), min(c[1] for c in r)))
    return regions


class Frame(NamedTuple):
    """塊の**枠と中身**。格子（``|  |`` の並び）はまだ作っていない。

    枠だけを先に決めるのは、**すかすかな塊で格子を作らないため**である。
    以前は :func:`_grid` がいきなり ``行数 × 列数`` の格子を組み、そのあとで
    :attr:`sparse` を見て捨てていた ―― 判定に届く前に枠のぶんだけメモリと
    時間を使う。工程表を年単位に伸ばした 1 枚（対角線に 5,000 セル）で
    **12.3 秒・211MB** かかり、出来上がりは箇条書き 5,000 行である。
    枠が大きいほど中身が薄いのだから、**いちばん捨てる塊にいちばん払っていた。**
    """

    rows: list[int]                    # 枠の行番号（間の空行も残す）
    columns: list[int]                 # 枠の列番号（間の空列も残す）
    at: str                            # ``B8:J20``
    addresses: list[tuple[str, str]]   # 実セルだけ（番地, 値）

    @property
    def height(self) -> int:
        return len(self.rows)

    @property
    def width(self) -> int:
        return len(self.columns)

    @property
    def sparse(self) -> bool:
        """**表の格子で出す意味が無いほどすかすかか。** 大きさと密度の両方を見る。"""
        area = self.height * self.width
        return area >= _SPARSE_AREA and area > len(self.addresses) * _SPARSE


def _frame(cells: dict[tuple[int, int], str], region: list[tuple[int, int]]
           ) -> Frame:
    """領域を矩形に切り出す。**間の空行・空列は残す。**

    残すのは、**``at=`` の番地から表の中の位置を割り出せるようにする**ためである。
    空の行・列を詰めていたときは ``at=B3:F5`` と書いてある表の列が 4 本しか
    なく、3 本目が D 列なのか E 列なのかを読み手が決められなかった ―― 番地を
    併記していれば読み直せる、という区切りの前提（:mod:`arp4.parse` の 4）が
    そこだけ成り立っていなかった。

    副作用として、**空列 1 本で隔てて横に並べた 2 つの表**（実物の一覧シートに
    ごく普通にある）が地続きの 1 枚に見えなくなる。詰めていたときは
    ``| 項目ID | 項目名 | ボタン | 動作 |`` と**同じ行に無いものが同じ行に**
    並び、整理層はそれを 1 つの表として読んでいた。

    **番地は実セルだけを走って作る。** 枠を端から端まで舐めて
    ``if (row, column) in cells`` で拾っていたので、ここも枠の面積ぶんかかって
    いた ―― 中身は 1 セルも変わらない。
    """
    rows = _span(sorted({r for r, _ in region}))
    columns = _span(sorted({c for _, c in region}))
    if not rows or not columns:
        return Frame([], [], "", [])

    at = f"{column_name(columns[0])}{rows[0]}"
    if len(rows) > 1 or len(columns) > 1:
        at += f":{column_name(columns[-1])}{rows[-1]}"

    addresses = [(f"{column_name(column)}{row}", cells[(row, column)])
                 for row, column in sorted(region)]
    return Frame(rows, columns, at, addresses)


def _grid(cells: dict[tuple[int, int], str], frame: Frame) -> list[list[str]]:
    """枠を格子にする。**表として出すと決めたあとにだけ呼ぶ。**"""
    return [[cells.get((row, column), "") for column in frame.columns]
            for row in frame.rows]


#: 塊 1 つの**大きさを申告する**行数・セル数。ここを超えると、読み手（整理層）は
#: **先頭だけ読んで「読んだ」と思う** ―― 抽出ツールが吐いた一覧は 1 枚で
#: 数千行あり、見出し行と最初の 20 行を見たかぎりでは小さい表と見分けが付かない。
#:
#: 値を落としているわけではないので「読めなかったもの」の申告ではない。
#: :func:`_sparse_note` と同じ**提示上の申告**である（黙ると読み手が
#: 事実と違うものを読む、というところが同じ）。検体でいちばん大きい塊は
#: 362 行なので、**普通の設計書では 1 度も出ない**。
_BIG = 1000


def _big_note(index: int, anchor: str, frame: Frame) -> str:
    """**大きい塊は大きいと言う。** 黙ると先頭だけ読まれる。"""
    return (f"アンカー `s{index}-{anchor}`（{frame.at}）は "
            f"{frame.height} 行 × {frame.width} 列 "
            f"／ {len(frame.addresses)} セルあります。先頭だけ読むと"
            "残りを見落とします（抽出ツールが吐いた一覧はこの形になります）。"
            "落としている値はありません。")


def _sparse_note(index: int, texts: int, frame: Frame) -> str:
    """**表の形で出していないことを言う。** 黙ると「表が無いシート」に見える。

    工程表・マトリクス（機能 × 画面の丸印）はこの形になる。番地は全部付いて
    いるので情報は落ちていないが、それを言わないと**読み手は表を探しに行く**。
    """
    return (f"アンカー `s{index}-x{texts}`（{frame.at}）は "
            f"{frame.height} 行 × {frame.width} 列の枠に "
            f"{len(frame.addresses)} セルしか無いので、表ではなく番地付きの箇条書き"
            "で出しました（工程表・マトリクスはこの形になります）。"
            "空欄まで格子で並べると中身の何十倍にもなり、読む予算をそれだけで"
            "使い切ります。番地は全部付いているので、落ちている値はありません。")


def _span(present: list[int]) -> list[int]:
    """並んだ行番号・列番号の**間を埋める**（両端は広げない）。"""
    if not present:
        return []
    return list(range(present[0], present[-1] + 1))


def column_name(index: int) -> str:
    name = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        name = chr(ord("A") + remainder) + name
    return name


def safe_name(name: str) -> str:
    """ファイル名にできない文字を落とす。**Windows の予約名も避ける。**

    :mod:`arp4.render` も同じ規則で PNG に名前を付ける（パース結果と画像で
    名前がずれると、どのシートの絵かを人が突き合わせる羽目になる）。
    """
    cleaned = re.sub(r'[\\/:*?"<>|]', "_", name).strip().rstrip(".") or "sheet"
    if cleaned.upper().split(".")[0] in {
            "CON", "PRN", "AUX", "NUL", *(f"COM{i}" for i in range(1, 10)),
            *(f"LPT{i}" for i in range(1, 10))}:
        cleaned = f"_{cleaned}"
    return cleaned


# ── コード（AST） ───────────────────────────────────────────────
#: 宣言の無いソースを読み直す文字コード。**既存資産の日本語は cp932 である**
#: ―― 棚卸しの対象になるのは 10 年もののコードで、UTF-8 で書かれていることの
#: ほうが珍しい。ここが無かったぶん、**コメントに日本語が 1 文字あるだけで
#: そのファイルは骨格ごと落ちていた**（`P010`）。クラスもメソッドも取れない。
_FALLBACK = "cp932"


def _source(path: Path) -> str:
    """ソースを文字列に。**宣言があればそれに従う**（PEP 263 / BOM）。

    ``tokenize.detect_encoding`` が見るのは先頭 2 行の ``# -*- coding: … -*-``
    と BOM で、どちらも**ファイルに書いてある**ものである（既定は UTF-8）。
    無いときだけ :data:`_FALLBACK` で読み直す ―― 日本語が cp932 と UTF-8 の
    両方で読める並びはまず無いので、取り違えは起きない。
    """
    import io
    import tokenize

    body = path.read_bytes()
    encoding = "utf-8"
    try:
        encoding, _ = tokenize.detect_encoding(io.BytesIO(body).readline)
    except SyntaxError:                        # 宣言が壊れている
        pass
    try:
        return body.decode(encoding)
    except UnicodeDecodeError:
        return body.decode(_FALLBACK)


def _is_test(relative: Path) -> bool:
    """テストのファイルか。**pytest が集める規則をそのまま写す**（``test_*.py``
    と ``*_test.py``）―― 中身を読んで「テストらしい」と決めるのは意味の判断だが、
    この 2 つは**規約であって解釈が要らない**（``~$`` を Excel の一時ファイルと
    見るのと同じ）。呼ぶ側が名前で決めているものを、こちらだけ疑う理由が無い。
    """
    stem = relative.stem
    return relative.suffix.lower() == ".py" and (
        stem.startswith("test_") or stem.endswith("_test"))


def _imports(tree: ast.Module) -> list[list[str]]:
    """取り込みの一覧。**AST に書いてあるものを転記するだけ。**

    これを出していなかったあいだ、呼出関係（``calls``）は整理層が原本を読んで
    手で起こしていた ―― **``import`` は構文木からゼロ曖昧に取れる**もので、
    「構文木から取れるものに LLM を使っても精度は上がらず、コストと見落としだけが
    増える」と決めた当のものである（決定 2）。114 本・200 本の規模では手作業が
    先に破綻する。

    **関数の中の取り込みも拾う**（重い依存を遅らせる書き方は普通にある）。
    どれが自分たちのモジュールかは決めない ―― 行と文をそのまま並べる。

    **1 行 = 1 名前**にする。``from arp4 import mdio, yamlio`` を 1 行で出して
    いたあいだ、「元」に書けるのは ``arp4`` だけで、**2 本の依存が 1 つの升に
    畳まれていた**。整理層はそこから ``calls`` を 2 本起こすことになり、実測で
    `arp4.parse → arp4.yamlio` が落ちた（92 本のうち 91 本しか出ていなかった）。
    落ちても機械には言えない ―― 升の中身と関係の本数が対応していないので、
    **数えることも突き合わせることもできない**。

    展開は忠実性の側である。文には名前が 2 つ**書いてある**のを 1 升に畳んで
    いただけで、幅 1 の縦結合を全行へ展開するのと同じ ―― 画面に見えているものへ
    寄せるだけで、**どこを指すかは相変わらず解かない**（``from arp4 import mdio``
    の ``mdio`` がモジュールなのかクラスなのかは、置き場を知っている側の仕事）。
    """
    rows: list[tuple[int, str, str, str]] = []
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                # `import a.b` が束ねるのはモジュールそのもの。**取り出した名前は
                # 無い**ので「名前」は空にする（`a.b` を書くと、`from a import b`
                # と同じ形になって区別が付かなくなる）。
                rows.append((node.lineno, ast.unparse(node), alias.name, ""))
        elif isinstance(node, ast.ImportFrom):
            # `from . import x` は module が None。**点の数は書いてあるとおり**に
            # 出す（どこを指すかを解くのは、置き場を知っている側の仕事である）。
            source = "." * node.level + (node.module or "")
            for alias in node.names:
                rows.append((node.lineno, ast.unparse(node), source, alias.name))
    seen: set[tuple[int, str, str, str]] = set()
    ordered = [r for r in sorted(rows) if not (r in seen or seen.add(r))]
    return [["取り込み", "元", "名前", "行"],
            *[[text, source, name, str(line)]
              for line, text, source, name in ordered]]


#: メンバの表の見出し。Java（``_java``）と**同じ並びに揃えてある** ―― 整理層は
#: 出自を気にしないので、同じものを 2 つの形で出す理由が無い。
_MEMBER = ["メンバ", "種類", "注釈", "シグネチャ", "戻り値", "例外", "行"]

#: コマンドの骨格の表の見出し（``p1``）。名前を左端に置くのはメンバの表と同じ
#: 理由 ―― 照合（``G018`` の ``_member``）は左端の欄で行を当てる。
_COMMAND = ["名前", "種類", "help", "受け手", "行"]


def _parser_wrappers(tree: ast.Module) -> dict[str, tuple[int, int | None]]:
    """``add_parser`` を包む補助関数 → ``(名前の引数の位置, help の位置)``。

    実務の CLI はコマンド定義を共通の補助関数で包む（arp4 自身の ``add(name,
    help_text, …)`` がそれ）。包まれると ``add_parser`` の実引数は変数になり、
    コマンド 16 種が 1 行も出ない ―― **宣言された引数の流れを 1 ホップ辿る**
    （:func:`_factories` が戻り値の注釈を辿るのと同じ規律）。仮引数がそのまま
    ``add_parser`` の名前・``help=`` に渡っているときだけ、その補助関数の
    呼び出しをコマンドの宣言として読む。当てにいくところは 1 つも無い。
    """
    wrappers: dict[str, tuple[int, int | None]] = {}
    for node in ast.walk(tree):
        if not isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            continue
        params = [a.arg for a in (*node.args.posonlyargs, *node.args.args)]
        for call in ast.walk(node):
            if not (isinstance(call, ast.Call)
                    and isinstance(call.func, ast.Attribute)
                    and call.func.attr == "add_parser" and call.args
                    and isinstance(call.args[0], ast.Name)
                    and call.args[0].id in params):
                continue
            help_index = None
            for keyword in call.keywords:
                if (keyword.arg == "help" and isinstance(keyword.value, ast.Name)
                        and keyword.value.id in params):
                    help_index = params.index(keyword.value.id)
            wrappers[node.name] = (params.index(call.args[0].id), help_index)
    return wrappers


def _argparse_rows(tree: ast.Module) -> list[list[str]]:
    """argparse の宣言（``add_parser`` / ``add_argument``）。**転記だけである。**

    CLI のコマンドは利用者向けの仕様だが、実装は先頭 ``_`` の内部関数なので
    「公開名だけ起こす」の規約では 1 件も出ない ―― 利用者向け仕様が正本から
    消える（r001 実測: コマンド 16 種）。宣言そのものは AST に書いてあるので、
    取り出すのは忠実性の回復である（Phase 1-5）。

    **名前が定数でないものも落とさない。** 補助関数で包んである CLI では
    ``add_parser(name, …)`` の実引数が変数になる ―― まず :func:`_parser_wrappers`
    が引数の流れを 1 ホップ辿り、それでも解けないものは書いてあるとおりの式
    （``name``）を出す。何を指すかを解くのは整理層の仕事で、機械が黙って落とすと
    「argparse を使っていない」ように見える。
    """
    wrappers = _parser_wrappers(tree)
    rows: list[tuple[int, str, str, str, str]] = []
    for node in ast.walk(tree):
        if not isinstance(node, ast.Call):
            continue
        if (isinstance(node.func, ast.Name) and node.func.id in wrappers):
            name_at, help_at = wrappers[node.func.id]
            if name_at >= len(node.args):
                continue
            help_text = ""
            if help_at is not None and help_at < len(node.args):
                value = node.args[help_at]
                help_text = (str(value.value) if isinstance(value, ast.Constant)
                             else ast.unparse(value))
            rows.append((node.lineno, ast.unparse(node.args[name_at]),
                         "コマンド", help_text, node.func.id))
            continue
        if not isinstance(node.func, ast.Attribute):
            continue
        kind = {"add_parser": "コマンド", "add_argument": "引数"}.get(node.func.attr)
        if kind is None:
            continue
        # 位置引数が名前である（add_argument は "--round" のような別名も並ぶ）。
        # **書いてあるとおりに出す**（文字列は引用符ごと）。引用符を剥ぐと、
        # 定数 `"parse"` と変数 `name` が表の上で区別できなくなる ―― 下流
        # （draft）は引用符の有無で「名前が確定している宣言」だけを拾う。
        names = ", ".join(ast.unparse(arg) for arg in node.args)
        help_text = ""
        for keyword in node.keywords:
            if keyword.arg == "help":
                help_text = (str(keyword.value.value)
                             if isinstance(keyword.value, ast.Constant)
                             else ast.unparse(keyword.value))
        receiver = ast.unparse(node.func.value)
        rows.append((node.lineno, names, kind, help_text, receiver))
    if not any(kind == "コマンド" for _, _, kind, _, _ in rows):
        # add_parser が 1 つも無いファイルはコマンドの骨格を持たない ――
        # 引数だけの表を出すと、単発スクリプトの全部に塊が生え、整理か対象外
        # 宣言かを迫られる（宣言の量産は宣言を読まれなくする）。
        return []
    return [list(_COMMAND),
            *[[names, kind, help_text, receiver, str(line)]
              for line, names, kind, help_text, receiver in sorted(rows)]]


def _decorators(node: ast.AST) -> str:
    """デコレータ。**種別を決める手がかりがここにしか無い。**

    ``@dataclass`` が付いているかどうかは、そのクラスが値の入れ物なのか処理を
    持つ層なのかを分ける ―― 詳細設計書の ``tier``（Common / Repository /
    Service / Validator）を決める最大の手がかりである。出していなかったあいだ、
    57 モジュールの層分けを整理層が**全部推測で**付けていた。

    ``@property`` も同じで、**呼び出しではなく値に見える**メンバかどうかが
    変わる。どちらも AST に書いてあるので、当てにいく作業が 1 つも無い。
    """
    marks = getattr(node, "decorator_list", []) or []
    return " ".join(f"@{ast.unparse(mark)}" for mark in marks)


def _class_head(node: ast.ClassDef) -> str:
    """``class Round(Paths)`` の形。**継承は書いてあるとおりに出す。**"""
    bases = [ast.unparse(base) for base in node.bases]
    bases += [f"{kw.arg}={ast.unparse(kw.value)}" for kw in node.keywords]
    return f"class {node.name}" + (f"({', '.join(bases)})" if bases else "")


def _fields(node: ast.ClassDef) -> list[tuple[ast.AST, list[str]]]:
    """クラス直下の属性。**``@dataclass`` はここが型の定義そのものである。**

    出していなかったあいだ、``@dataclass`` だと分かってもそれが**何を持つ入れ物
    なのか**はどこにも書いていなかった ―― Java の ``private String orderNo`` に
    あたるものが Python 側だけ落ちていた形である。
    """
    out: list[tuple[ast.AST, list[str]]] = []
    for child in node.body:
        if isinstance(child, ast.AnnAssign) and isinstance(child.target, ast.Name):
            out.append((child, [child.target.id, "フィールド", "",
                                ast.unparse(child), "", ""]))
        elif isinstance(child, ast.Assign):
            for target in child.targets:
                if isinstance(target, ast.Name):
                    out.append((child, [target.id, "フィールド", "",
                                        ast.unparse(child), "", ""]))
    return out


def _module_values(tree: ast.Module) -> list[tuple[ast.AST, list[str]]]:
    """モジュール直下の代入。**「決まっていること」がここにしか無い。**

    ``__version__`` ・拡張子（``EXT = ".yml"``）・予約名（``SPECIAL``）・置き場
    （``PACKS_DIR``）は、資料で言えば「区分の一覧」にあたる ―― クラスと関数と
    取り込みしか見ていなかったあいだ、**それが 1 行も出ていなかった。**
    自身のソースでは 23 本中 1 本（``__init__.py``）が丸ごと空になり、
    ``__version__`` が黙って消えた。

    **私物かどうかで選り分けない**（``_MEMBER`` のような先頭 ``_`` も出す）――
    クラス直下（:func:`_fields`）が選り分けていないのと同じで、どれが仕様かを
    決めるのは整理層である。
    """
    out: list[tuple[ast.AST, list[str]]] = []
    for node in tree.body:
        if isinstance(node, ast.AnnAssign) and isinstance(node.target, ast.Name):
            targets = [node.target]
        elif isinstance(node, ast.Assign):
            targets = [t for t in node.targets if isinstance(t, ast.Name)]
        else:
            continue
        for target in targets:
            out.append((node, [target.id, "定数", "", _value_of(node), "", ""]))
    return out


#: 値をそのまま出すのをやめる長さ。**画面で読める量に寄せる。**
_VALUE_MAX = 200


def _value_of(node: ast.AST) -> str:
    """代入 1 行。**長いものは切って、切ったと言う。**

    正規表現の塊や数十行の辞書をそのまま出すと、パース結果が原本より読みにくく
    なる（`parse.py` の `_MAGIC` は 1 つで 20 行ある）。切った印を残すのは、
    **「資料に無い」と「機械が省いた」を混ぜない**ためである ―― 印が無いと、
    続きが無いのか省かれたのかを整理層が判断できない。
    """
    text = " ".join(ast.unparse(node).split())
    if len(text) <= _VALUE_MAX:
        return text
    return f"{text[:_VALUE_MAX]}…（以下略・原本を見ること）"


def _python(path: Path, relative: Path) -> list[tuple[Path, mdio.Doc]]:
    """骨格だけ。**意図の層は出さない**（整理層が原本を直接読む）。

    塊の頭文字は出自を表す ―― ``m`` は骨格、``i`` は取り込み、``t`` はテスト。
    番号を通しにすると、**取り込みを足しただけで既存の出典が全部ずれる**。
    """
    tree = ast.parse(_source(path), filename=str(path))
    posix = relative.as_posix()
    doc = mdio.Doc(title=posix, source=posix)
    testing = _is_test(relative)
    factories = _factories(tree)

    index = 0
    functions: list[tuple[ast.AST, list[str]]] = []
    cases: list[tuple[ast.AST, list[str]]] = []
    for node in tree.body:
        if isinstance(node, (ast.ClassDef,)):
            index += 1
            rows = [list(_MEMBER),
                    [node.name, "クラス", _decorators(node), _class_head(node),
                     "", "", str(node.lineno)]]
            rows += [[*row, str(child.lineno)] for child, row in _fields(node)]
            rows += [[child.name, "メソッド", _decorators(child),
                      *_method_row(child, factories)[1:], str(child.lineno)]
                     for child in node.body
                     if isinstance(child, (ast.FunctionDef, ast.AsyncFunctionDef))]
            doc.chunks.append(mdio.Chunk(
                anchor=f"m{index}",
                at=f"{posix}#L{node.lineno}-L{_end(node)}",
                heading=("テストクラス: " if testing else "クラス: ") + node.name,
                rows=rows))
        elif isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            # pytest が集めるのは `test` で始まる関数だけ。**残り（fixture・
            # 補助）と混ぜない** ―― 混ぜると整理層がテストケースを数えられない。
            bucket = cases if testing and node.name.startswith("test") else functions
            bucket.append((node, [node.name, "関数", _decorators(node),
                                  *_method_row(node, factories)[1:]]))

    for anchor, heading, kind, group in (
            (f"m{index + 1}", "モジュール関数", "関数", functions),
            ("t1", "テスト", "テスト", cases)):
        if not group:
            continue
        # **1 本ずつを塊にしない。** 出典の精度は上がるが、`freeze` は塊 1 個ずつに
        # 整理か対象外宣言かを迫るので、補助関数 30 本のモジュールが未整理 30 件に
        # なる ―― 通すために宣言を量産させると、**宣言そのものが読まれなくなる。**
        # 行は列で持たせるので、原本のどこかは 1 行ずつ読み直せる。
        doc.chunks.append(mdio.Chunk(
            anchor=anchor, heading=heading,
            at=f"{posix}#L{group[0][0].lineno}-L{_end(group[-1][0])}",
            rows=[list(_MEMBER),
                  *[[row[0], kind, *row[2:], str(node.lineno)]
                    for node, row in group]]))

    # **頭文字を分ける**（`v` は定数）。`m` に混ぜて通し番号にすると、定数を
    # 足しただけで既存のクラスの出典が全部ずれる ―― 塊の種類を足すたびに過去の
    # ラウンドが読み直せなくなるので、番号は種類ごとに独立させてある。
    values = _module_values(tree)
    if values:
        doc.chunks.append(mdio.Chunk(
            anchor="v1", heading="定数",
            at=f"{posix}#L{values[0][0].lineno}-L{_end(values[-1][0])}",
            rows=[list(_MEMBER),
                  *[[*row, str(node.lineno)] for node, row in values]]))

    # コマンドの骨格（argparse の宣言）。頭文字は `p`（`m` / `v` に混ぜると、
    # 塊の種類を足しただけで既存の出典がずれる ―― 番号は種類ごとに独立させる）。
    commands = _argparse_rows(tree)
    if commands:
        lines = sorted(int(row[-1]) for row in commands[1:])
        doc.chunks.append(mdio.Chunk(
            anchor="p1", heading="コマンド（argparse）",
            at=f"{posix}#L{lines[0]}-L{lines[-1]}",
            rows=commands))

    # **取り込みが 0 本でも塊を出す**（中身のあるファイルに限る）。``at`` が
    # ファイルそのものなので、これは取り込みの一覧であると同時に**モジュール
    # そのものを指す唯一のアンカー**である ―― 0 本のとき落としていたので、
    # ``import`` を 1 つも持たないモジュールには指す先が無かった。
    #
    # 実測では自身の ``__init__.py`` がそれで、パッケージの冒頭に書いてある
    # 「意味の判断は整理層だけが行う」が**モジュールとして 1 件も起きなかった**。
    # 未読取でも対象外でもなく、アンカーが無いので ``G001`` にすら上がらない
    # ―― 規律の言う「読めていないものほど静かに消える」がそのまま出た形である。
    #
    # 専用の頭文字（``f1``）を足すほうが名前としては正しいが、**既存のラウンドの
    # 全ファイルに未整理のアンカーが 1 個ずつ増える**。``i1`` は元から
    # ``at=<ファイル>`` でファイル全体を指しており、モジュールの出典としては
    # 既にこれが使われているので、無条件化のほうが動くものが少ない。
    #
    # 空のファイルには出さない。書くことが無いものにアンカーを立てると、
    # 対象外宣言だけが量産されて**宣言そのものが読まれなくなる**（上と同じ理由）。
    rows = _imports(tree)
    if len(rows) > 1 or doc.chunks:
        doc.chunks.append(mdio.Chunk(anchor="i1", at=posix, heading="取り込み",
                                     rows=rows))

    if not doc.chunks:
        return []
    return [(Path(*relative.parts[:-1]) / f"{relative.name}{mdio.EXT}", doc)]


def _args(spec: ast.arguments) -> list[str]:
    """引数の並び。**``args`` だけを見ていたぶん、実装と違うものを出していた。**

    ``ast.arguments`` は引数を 4 つの入れ物に分けて持つ ―― 位置専用（``/`` の
    前）・普通・可変長（``*args``）・キーワード専用（``*`` の後ろ）。``args``
    だけを読むと**キーワード専用引数がまるごと消える**。詳細設計書に
    ``plan_worksheet(worksheet, extent)`` と出ていたものの実物は
    ``(worksheet, extent=(0, 0), *, boxes=(), target_px=1600)`` で、**その表を
    見て書いた呼び出しは通らない。** 読み落としではなく、**書いてある数と
    違うものを書いている**ので、申告のしようもない。

    既定値も出す ―― 「省略できるか」は呼ぶ側がまず知りたいことで、AST に
    書いてある（当てにいく必要が 1 つも無い）。
    """
    def one(arg: ast.arg, default: ast.expr | None = None) -> str:
        text = arg.arg + (f": {ast.unparse(arg.annotation)}" if arg.annotation else "")
        if default is None:
            return text
        # 型註釈があるときだけ `=` の周りを空ける（原本の書き方に合わせる ――
        # 詳細設計書のシグネチャは、そのまま呼び出しに写せることに値がある）。
        return text + (" = " if arg.annotation else "=") + ast.unparse(default)

    positional = [*spec.posonlyargs, *spec.args]
    # 既定値は**後ろから**埋まる（`def f(a, b=1)` の defaults は `[1]`）。
    pad: list[ast.expr | None] = [None] * (len(positional) - len(spec.defaults))
    out = [one(a, d) for a, d in zip(positional, [*pad, *spec.defaults])
           if a.arg not in ("self", "cls")]
    # `/` は**残った**位置専用引数の後ろ。`self` を落としたぶんを数え直さないと
    # 区切りが 1 つ右へずれ、普通の引数まで位置専用に見える。
    posonly = [a for a in spec.posonlyargs if a.arg not in ("self", "cls")]
    if posonly:
        out.insert(len(posonly), "/")

    if spec.vararg:
        out.append("*" + one(spec.vararg))
    elif spec.kwonlyargs:
        out.append("*")                     # キーワード専用の始まりは `*` で示す
    out += [one(a, d) for a, d in zip(spec.kwonlyargs, spec.kw_defaults)]
    if spec.kwarg:
        out.append("**" + one(spec.kwarg))
    return out


def _method_row(node: ast.FunctionDef | ast.AsyncFunctionDef,
                factories: dict[str, str] | None = None) -> list[str]:
    args = _args(node.args)
    returns = ast.unparse(node.returns) if node.returns else ""
    raises = sorted({_raised(child, factories) for child in ast.walk(node)
                     if isinstance(child, ast.Raise)} - {""})
    return [node.name, f"{node.name}({', '.join(args)})", returns, ", ".join(raises)]


def _factories(tree: ast.Module) -> dict[str, str]:
    """``raise` に渡される補助関数 → その**戻り値の型注釈**。

    ``raise _broken(path, exc)`` と書いてあるとき、``例外`` 欄に出るのは
    ``_broken`` である。これは private なヘルパの名前であって例外の型ではないので、
    整理層は書き写せず（``raises: _broken`` は嘘になる）、かといって落とすと
    ``G018`` が「欄が埋まっているのに落ちている」と鳴る ―― **どちらにも
    正解が無い欄**を出していた。実測で ``yamlio.marked`` がこれに当たり、
    詳細設計書の例外欄は空のまま、``load`` / ``load_marked`` の唯一の送出点が
    「投げない」と読める状態になっていた。

    やっているのは**宣言された戻り値を 1 ホップ辿るだけ**である。原本には
    ``def _broken(...) -> YamlError:`` と書いてあるので、当てにいっている
    ところが 1 つも無い（規律の表でいう忠実性の回復）。注釈が無ければ何もせず
    名前のまま出す ―― **読めなかったものは読めなかったと出す。**

    ``-> None`` は入れない。``raise`` の位置に出てくるはずがないので、
    拾うと注釈の書き間違いを例外の型として広めることになる。
    """
    out: dict[str, str] = {}
    for node in tree.body:
        if not isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            continue
        if node.returns is None:
            continue
        text = ast.unparse(node.returns)
        if text and text != "None":
            out[node.name] = text
    return out


def _raised(node: ast.Raise, factories: dict[str, str] | None = None) -> str:
    target = node.exc
    called = isinstance(target, ast.Call)
    if isinstance(target, ast.Call):
        target = target.func
    if isinstance(target, ast.Name):
        # **呼び出しの形のときだけ**辿る。``raise saved`` のように変数を投げる
        # 書き方では、名前がたまたま関数と同じでも別物である。
        if called and factories and target.id in factories:
            return factories[target.id]
        return target.id
    if isinstance(target, ast.Attribute):
        return target.attr
    return ""


def _end(node: ast.AST) -> int:
    return int(getattr(node, "end_lineno", 0) or getattr(node, "lineno", 0))


# ── テキスト ────────────────────────────────────────────────────
#: 見出し行。``#`` のあとに空白が要る（``#!/usr/bin/env`` を見出しにしない）。
_HEADING = re.compile(r"^(#{1,6})\s+(?P<title>.*?)\s*#*\s*$")

#: コードブロックの囲い。**中の ``# コメント`` を見出しに取らない** ―― 手順書は
#: シェルの例を必ず載せるので、取ると塊が本文の途中で切れ、``at`` の行範囲も
#: 本文とずれる。「読めた」と言いながら中身が入れ替わるのがいちばん悪い。
_FENCE = ("```", "~~~")


def _text_of(path: Path) -> str:
    """UTF-8 で読み、駄目なら cp932。**BOM は落とす**（``\\ufeff`` が見出しの
    先頭に残ると ``# `` にマッチせず、1 冊まるごと塊 1 個になる）。"""
    body = path.read_bytes()
    try:
        return body.decode("utf-8-sig")
    except UnicodeDecodeError:
        return body.decode(_FALLBACK, errors="replace")


def _headings(lines: list[str]) -> list[tuple[int, str]]:
    """見出しの ``(行番号, 表題)``。コードブロックの中は見ない。"""
    found: list[tuple[int, str]] = []
    fence = ""
    for index, line in enumerate(lines):
        stripped = line.lstrip()
        if fence:
            if stripped.startswith(fence):
                fence = ""
            continue
        if stripped.startswith(_FENCE):
            fence = stripped[:3]
            continue
        marker = _HEADING.match(line)
        if marker:
            found.append((index, marker.group("title")))
    return found


def _markdown(path: Path, relative: Path) -> list[tuple[Path, mdio.Doc]]:
    """見出しごとに 1 塊。**本文はそのまま**。

    Excel は「セルの面」を紙の見た目へ寄せ直す必要があるが、**テキストの資料は
    原本が既に読める形**なので、寄せ直すところが 1 つも無い。表に組み直したり
    箇条書きに畳んだりすると、入れ子・コードブロック・引用が平らになって
    **資料に書いてあったことが減る**だけである。

    見出しの深さ（``##`` か ``###`` か）は塊の見出しからは落ちる ―― 塊は並列に
    並ぶものなので階層を持てない。**代わりに ``at`` が行範囲を持つ**ので、原本を
    開けばどの深さの見出しだったかは 1 手で読み直せる。
    """
    text = _text_of(path)
    lines = text.splitlines()
    posix = relative.as_posix()
    doc = mdio.Doc(title=posix, source=posix)

    found = _headings(lines)
    # ``(行番号, 表題, 見出し行か)``。見出しの前にある本文（前書き）も塊にする
    # ―― **見出しが 1 つも無い資料（.txt の覚書）でここが効く**。落とすと 1 冊
    # まるごとアンカーが 0 個になり、`freeze` の未整理一覧に上がらないまま消える。
    starts: list[tuple[int, str, bool]] = [(no, title, True) for no, title in found]
    if not found or found[0][0] > 0:
        starts.insert(0, (0, "（前書き）" if found else "（本文）", False))

    index = 0
    for order, (line_no, title, is_heading) in enumerate(starts):
        stop = starts[order + 1][0] if order + 1 < len(starts) else len(lines)
        body = "\n".join(lines[line_no + 1 if is_heading else line_no:stop]).strip()
        if not is_heading and not body:
            continue                                   # 空の前書きは出さない
        index += 1
        doc.chunks.append(mdio.Chunk(
            anchor=f"h{index}",
            at=f"{posix}#L{line_no + 1}-L{stop}",
            heading=title, text=body))

    if not doc.chunks:
        return []
    return [(Path(*relative.parts[:-1]) / f"{relative.name}{mdio.EXT}", doc)]


# ── CSV ─────────────────────────────────────────────────────────
#: 区切りの候補。**この 4 つに絞る**のは、当てにいく幅がそのまま「値が変わる
#: 危険」だからである ―― 空白区切りまで候補に入れると、住所やコメントの中の
#: 空白で列が割れた表が「読めた」顔で出てくる。
_DELIMITERS = (",", "\t", ";", "|")

#: 先頭に BOM があるとき、**文字コードは資料自身が名乗っている**（当てにいく
#: 必要が無い）。Excel の「CSV UTF-8」は BOM 付きで書き出し、「Unicode テキスト」は
#: UTF-16LE ＋ タブ区切りで書き出す ―― どちらも実物にごく普通にある。
_CSV_BOMS = (
    (b"\xef\xbb\xbf", "utf-8-sig", "UTF-8（BOM 付き）"),
    (b"\xff\xfe", "utf-16", "UTF-16 LE（BOM 付き）"),
    (b"\xfe\xff", "utf-16", "UTF-16 BE（BOM 付き）"),
)

#: BOM が無いときに試す順。**UTF-8 を厳密に試してから cp932** へ落ちる
#: （:func:`_source` と同じ順）―― 日本語は cp932 と UTF-8 のどちらでも「読めて
#: しまう」ことがあるが、UTF-8 として妥当なバイト列が cp932 でもあることは
#: まず無いので、この順なら取り違えない。
_CSV_ENCODINGS = (("utf-8", "UTF-8"), ("cp932", "cp932（Shift_JIS）"))

#: 区切りが決まったと言ってよい**揃い方**。全行一致を求めないのは、実物の CSV に
#: 注記の行（``※ 桁あふれは切り捨て``）や小計の行が 1 本混ざるからである ――
#: そこで表にしないと、**揃っている 500 行まで道連れになる。**
#:
#: 5 行に 1 行が違う（8 割）ところで切るのは、そこから先は「区切りがこの資料の
#: 構造を言い当てていない」ほうが疑わしいからである。**通した後も割合はそのまま
#: 申告する**（``80% が 4 列``）ので、最後に決めるのは読む側である。
_CSV_AGREE = 0.8

#: 「大きい」と申告する行数。Excel の塊と同じ理由で**切り詰めはしない**が、
#: 先頭だけ読んで終わりにされると残りが `未読取` にも上がらないまま消える。
_CSV_BIG_ROWS = 1000


#: 半角カタカナ（``｡`` 〜 ``ﾟ``）。**cp932 では 1 バイト**（0xA1〜0xDF）である。
_HALFWIDTH_KANA = re.compile(r"[｡-ﾟ]")

#: 半角カタカナがこれ以上を占めたら「化けている」と言う。実物の資料に半角
#: カタカナが混ざることはあるが（旧システムの項目名）、**本文の 3 割を超える
#: ことは無い** ―― 超えるのは EUC-JP を cp932 として読んだときだけである。
_KANA_MOJIBAKE = 0.3


def _mojibake(text: str) -> str:
    """cp932 として**読めてしまった**化けの申告。**空なら化けていない。**

    ここは CSV でいちばん静かに壊れるところである ―― EUC-JP の日本語は
    0xA1〜0xFE を使い、cp932 ではその範囲が**半角カタカナ 1 文字ずつ**に当たる。
    つまり ``受注番号`` は例外を出さずに ``ｼｳﾃﾞﾁﾔﾝｺﾞｳ`` のような字に化けて、
    **「読めました」という顔で表に入る**。:func:`_csv_text` は例外でしか異常を
    知れないので、ここだけは読めた中身を見る必要がある。

    **判定しているのは字種の割合だけ**である（意味は見ていない）ので、
    「これは EUC-JP だ」とは言わない ―― 言うのは「半角カタカナばかりです」と
    いう数えた事実と、それが起きる典型的な理由である。決めるのは人である。
    """
    body = "".join(text.split())
    if not body:
        return ""
    if len(_HALFWIDTH_KANA.findall(body)) / len(body) < _KANA_MOJIBAKE:
        return ""
    return ("cp932 として読めましたが、**中身が半角カタカナばかりです**"
            "（本文の 3 割以上）。EUC-JP で保存された資料を cp932 として読むと"
            "この形になります ―― 例外が出ないので「読めた」ように見えますが、"
            "**ここに並んでいる字は資料の字ではありません**。"
            "UTF-8 で保存し直してから取り込み直してください。")


def _csv_text(path: Path) -> tuple[str, str, str]:
    """``(本文, 文字コードの名前, 申告)``。**当てた結果を必ず言う。**

    ここが CSV でいちばん危ないところである ―― `,` 区切りだと思って読んだ住所は
    列がずれ、cp932 を UTF-8 として読んだ品名は化けるが、**どちらも「読めた」
    顔で出てくる**。だから決めた結果を捨てずに持ち回り、パース結果の頭に書く。

    最後の砦（``errors="replace"``）まで落ちたときだけ申告を返す ―― 化けた字は
    資料の字ではないので、黙って表に入れてはいけない。
    """
    body = path.read_bytes()
    for mark, encoding, name in _CSV_BOMS:
        # **BOM があるなら当てにいかない。** 資料が自分で名乗っているものを
        # 機械が疑う理由は 1 つも無い。
        if body.startswith(mark):
            try:
                return body.decode(encoding), name, ""
            except UnicodeDecodeError:
                break                      # 名乗りと中身が違う。当てにいく
    for encoding, name in _CSV_ENCODINGS:
        try:
            text = body.decode(encoding)
        except UnicodeDecodeError:
            continue
        # **例外の出ない失敗がある。** cp932 は EUC-JP のバイト列を半角カタカナ
        # として読み切ってしまうので、ここだけは読めた中身を見る。
        return text, name, (_mojibake(text) if encoding == "cp932" else "")
    return (body.decode(_FALLBACK, errors="replace"), f"{_FALLBACK}（読めない字あり）",
            "UTF-8 でも cp932 でも読めないバイトがあります"
            "（EUC-JP・UTF-16 で保存された資料がこの形になります）。"
            "読めなかった字は `�` に置き換わっているので、"
            "**その欄の値は資料の値ではありません**。")


def _csv_rows(text: str, delimiter: str) -> list[list[str]]:
    """``csv`` 標準ライブラリで割る。**引用符の中の改行と区切りを守る。**"""
    import csv as csv_module

    return [row for row in csv_module.reader(text.splitlines(True),
                                             delimiter=delimiter)]


def _csv_fit(rows: list[list[str]]) -> tuple[int, float]:
    """``(いちばん多い列数, その割合)``。**空行は数えない。**"""
    counts = [len(row) for row in rows if any(cell.strip() for cell in row)]
    if not counts:
        return 0, 0.0
    modal = max(set(counts), key=counts.count)
    return modal, counts.count(modal) / len(counts)


def _csv_delimiter(text: str, path: Path) -> tuple[str, str, str]:
    """``(区切り, 呼び名, 決められなかった理由)``。**当たらなければ当てない。**

    `.tsv` は**拡張子が区切りを名乗っている**ので当てにいかない（資料が言って
    いることを機械が疑う理由が無い）。`.csv` は候補を 1 つずつ実際に割ってみて、
    **行ごとの列数が揃うか**で決める ―― 区切り文字の出現数だけで決めると、
    コメント欄にセミコロンの多い 1 冊で `;` を選ぶ。

    2 つ以上の候補が同じだけ揃ったら**決めない** ―― 1 列しか無い CSV（コードの
    一覧）はどの区切りでも「揃って」見えるので、そこで先頭の候補を選ぶのは
    偶然に賭けているだけである。
    """
    if path.suffix.lower() == ".tsv":
        return "\t", "タブ", ""

    scored: list[tuple[float, int, str]] = []
    for candidate in _DELIMITERS:
        columns, agreement = _csv_fit(_csv_rows(text, candidate))
        if columns >= _MIN_TABLE and agreement >= _CSV_AGREE:
            scored.append((agreement, columns, candidate))
    if not scored:
        return "", "", ("どの区切り（" + "・".join(f"`{d}`" for d in _DELIMITERS)
                        + "）でも、行ごとの列数が揃いませんでした")
    best = max(scored)
    tied = [one for one in scored if one[0] == best[0] and one[1] == best[1]]
    if len(tied) > 1:
        return "", "", ("区切りの候補が絞れませんでした（"
                        + "・".join(f"`{one[2]}`" for one in tied)
                        + " のどれでも同じだけ揃います）")
    return best[2], _DELIMITER_NAMES.get(best[2], f"`{best[2]}`"), ""


#: 申告に出す区切りの呼び名。**画面に見えない文字は名前で言う**（`\t` と書いても
#: 読み手には伝わらない）。
_DELIMITER_NAMES = {",": "`,`（カンマ）", "\t": "タブ", ";": "`;`（セミコロン）",
                    "|": "`|`（縦棒）"}


def _csv(path: Path, relative: Path) -> tuple[list[tuple[Path, mdio.Doc]],
                                              list[Finding]]:
    """CSV / TSV を表 1 枚にする。**当てた区切りと文字コードを必ず申告する。**

    長らくここは「Excel で開いて .xlsx として保存し直してください」だけを返して
    いた。理由は正しい（区切りと文字コードは資料ごとに違い、当てにいけば値が
    変わる）が、**その助言に従える人がいる現場ばかりではない** ―― 移行データの
    一覧・コード値の一覧は数百本の CSV で配られ、1 本ずつ Excel で開き直す作業を
    誰も引き受けない。実際には、その数百本が 1 件も仕様にならないまま終わる。

    **当てるのをやめたのではなく、当てた結果を黙らないことにした。** 決めた区切り
    と文字コードと、行ごとの列数がどれだけ揃ったかをパース結果の頭に書く ――
    整理層はそれを見て、値を信じるか原本を開くかを決められる。**決められなかった
    ときは表にしない**（`P018`）―― 揃わない表を幅だけ揃えて出すと、**足りない列が
    「資料が空欄」に見える**。
    """
    posix = relative.as_posix()
    doc = mdio.Doc(title=posix, source=posix)
    text, encoding, trouble = _csv_text(path)
    if not text.strip():
        return [], []                                  # 空なら `_empty_note` が言う

    lines = len(text.splitlines())
    delimiter, called, why = _csv_delimiter(text, path)
    findings: list[Finding] = []
    if trouble:
        findings.append(Finding("warn", "P018", path.name, trouble))
        doc.notes.append(trouble)

    if not delimiter:
        # **表にしない。** 割れていない事実のほうが、割れた顔の表より役に立つ。
        doc.notes.append(
            f"文字コードは {encoding} として読みました。**区切りは決められません**"
            f"（{why}）。表にせず原文のまま出しています ―― 機械が当てて割ると、"
            "列がずれた表が「読めた」顔で出ます。Excel で開いて確かめ、"
            "`.xlsx` として保存し直すか、この写しを直接編集してください。")
        doc.chunks.append(mdio.Chunk(anchor="x1", at=f"{posix}#L1-L{lines}",
                                     heading="原文（区切りが決まっていません）",
                                     text=text.strip()))
        findings.append(Finding(
            "warn", "P018", path.name,
            f"区切りが決められないので表にしていません（{why}）。"
            "原文のまま出してあるので、値は 1 つも落ちていません"))
        return [(Path(*relative.parts[:-1]) / f"{relative.name}{mdio.EXT}", doc)], \
            findings

    rows = _csv_rows(text, delimiter)
    columns, agreement = _csv_fit(rows)
    rows = [row for row in rows if any(cell.strip() for cell in row)]
    doc.notes.append(
        f"文字コードは {encoding}、区切りは {called} として読みました"
        f"（{len(rows)} 行のうち {round(agreement * 100)}% が {columns} 列）。"
        "**どちらも機械が決めたものです** ―― 値そのもの（桁・コード値）を"
        "使う前に、1 行だけでも原本と読み比べてください。")
    if agreement < 1.0:
        # **幅を揃えて出す以上、揃っていなかったことは言わなければならない。**
        # `mdio._table` は短い行を空欄で埋めるので、黙ると資料が空欄に見える。
        doc.notes.append(
            f"{columns} 列でない行が {len(rows) - round(agreement * len(rows))} "
            "行あります。表は幅を揃えて出しているので、**足りない列は空欄に"
            "見えます** ―― そこは「資料が空欄」ではありません。")
    if len(rows) >= _CSV_BIG_ROWS:
        doc.notes.append(
            f"{len(rows)} 行あります（大きい塊です）。**値は 1 つも切り詰めて"
            "いません** ―― 先頭だけ読んで終わりにすると、残りが `未読取` にも"
            "上がらないまま消えます。")
    doc.chunks.append(mdio.Chunk(anchor="t1", at=f"{posix}#L1-L{lines}",
                                 heading=f"表 {len(rows)} 行 × {columns} 列",
                                 rows=rows))
    return [(Path(*relative.parts[:-1]) / f"{relative.name}{mdio.EXT}", doc)], findings


# ── DDL ─────────────────────────────────────────────────────────
#: 引用の始まりと終わり。方言ごとに囲い方が違う（Oracle/PG は ``"``、
#: SQL Server は ``[]``、MySQL は backquote）。**どれも識別子の囲いであって
#: 中身ではない**ので、外して名前だけを出す。
_QUOTES = {"'": "'", '"': '"', "`": "`", "[": "]"}

#: 列の定義ではなく**表そのものに掛かる制約**。行の頭がこれなら列ではない。
_TABLE_CONSTRAINT = ("CONSTRAINT", "PRIMARY", "UNIQUE", "FOREIGN", "CHECK",
                     "KEY", "INDEX", "EXCLUDE", "PERIOD")

#: 列の定義で、型の終わりを告げる語。ここから後ろは既定値と制約である。
_AFTER_TYPE = ("NOT", "NULL", "DEFAULT", "PRIMARY", "UNIQUE", "REFERENCES",
               "CHECK", "GENERATED", "IDENTITY", "AUTO_INCREMENT", "COMMENT",
               "COLLATE", "CONSTRAINT", "VISIBLE", "INVISIBLE")


def _sql_scan(text: str) -> list[tuple[int, str]]:
    """``;`` で文に切る。**引用符とコメントの中では切らない。**

    ``INSERT`` の値に ``;`` が入っている・コメントに ``--`` で仕様が書いてある
    ―― どちらも実物の DDL では普通で、素朴に切ると**表 1 つが 2 つに割れる**。
    戻すのは ``(始まりの行, 文)``。行を持たないと ``at`` が書けない。
    """
    statements: list[tuple[int, str]] = []
    buffer: list[str] = []
    line = start = 1
    index = 0
    while index < len(text):
        char = text[index]
        if not "".join(buffer).strip():
            start = line                              # 空白を跨いだら数え直す
        pair = text[index:index + 2]
        if pair == "--":
            while index < len(text) and text[index] != "\n":
                index += 1
            continue
        if pair == "/*":
            index += 2
            while index < len(text) and text[index:index + 2] != "*/":
                line += text[index] == "\n"
                index += 1
            index += 2
            continue
        if char in _QUOTES:
            close = _QUOTES[char]
            buffer.append(char)
            index += 1
            while index < len(text):
                if text[index] == close:
                    # `''` は引用符そのもの（閉じてすぐ開き直す書き方）
                    if text[index:index + 2] == close * 2:
                        buffer.append(close * 2)
                        index += 2
                        continue
                    break
                line += text[index] == "\n"
                buffer.append(text[index])
                index += 1
            buffer.append(close)
            index += 1
            continue
        if char == ";":
            if "".join(buffer).strip():
                statements.append((start, "".join(buffer).strip()))
            buffer = []
            index += 1
            continue
        line += char == "\n"
        buffer.append(char)
        index += 1
    if "".join(buffer).strip():
        statements.append((start, "".join(buffer).strip()))
    return statements


def _sql_split(text: str) -> list[str]:
    """深さ 0 の ``,`` で切る。``NUMERIC(11, 2)`` の中では切らない。"""
    parts: list[str] = []
    buffer: list[str] = []
    depth = 0
    index = 0
    while index < len(text):
        char = text[index]
        if char in _QUOTES:
            close = _QUOTES[char]
            buffer.append(char)
            index += 1
            while index < len(text) and text[index] != close:
                buffer.append(text[index])
                index += 1
            buffer.append(close)
            index += 1
            continue
        if char == "(":
            depth += 1
        elif char == ")":
            depth -= 1
        elif char == "," and depth == 0:
            parts.append("".join(buffer).strip())
            buffer = []
            index += 1
            continue
        buffer.append(char)
        index += 1
    if "".join(buffer).strip():
        parts.append("".join(buffer).strip())
    return parts


def _sql_body(statement: str) -> tuple[str, str]:
    """``CREATE TABLE 名前 ( 中身 ) 後ろ`` を ``(前, 中身)`` に割る。

    後ろに付く ``ENGINE=InnoDB`` や ``TABLESPACE`` は方言ごとに違うので**読まない**
    ―― 表の中身は括弧の対応だけで取れる（数えるのに方言の知識が要らない）。
    """
    start = statement.find("(")
    if start < 0:
        return statement, ""
    depth = 0
    for index in range(start, len(statement)):
        if statement[index] == "(":
            depth += 1
        elif statement[index] == ")":
            depth -= 1
            if depth == 0:
                return statement[:start], statement[start + 1:index]
    return statement[:start], statement[start + 1:]


def _sql_name(text: str) -> str:
    """識別子から囲いを外す。**囲いは区切りであって名前ではない。**"""
    name = text.strip().rstrip(",").strip()
    for open_, close in _QUOTES.items():
        if len(name) > 1 and name.startswith(open_) and name.endswith(close):
            return name[1:-1]
    return name


def _sql_column(text: str) -> list[str] | None:
    """列の定義 1 本を ``[列, 型, 既定値, 制約]`` に割る。制約行なら ``None``。

    **型は書いてあるものをそのまま出す。** ``NUMERIC(11, 2)`` を「数値」へ
    寄せるのはメタモデルの enum に当てはめる作業で、それは意味の判断である
    ―― 方言ごとの型（``NUMBER`` / ``NUMERIC`` / ``DECIMAL``）が同じものかは、
    資料と実装を見た人にしか決められない。
    """
    words = text.split()
    if not words or words[0].upper() in _TABLE_CONSTRAINT:
        return None
    name = _sql_name(words[0])
    rest = text[len(words[0]):].strip()

    type_words: list[str] = []
    for word in rest.split():
        if word.upper().rstrip("(,") in _AFTER_TYPE:
            break
        type_words.append(word)
    data_type = " ".join(type_words)
    tail = rest[len(data_type):].strip() if data_type else rest

    default = ""
    found = re.search(r"\bDEFAULT\b\s+(.+?)(?=\s+(?:NOT\s+NULL|NULL|PRIMARY|"
                      r"UNIQUE|REFERENCES|CHECK|COMMENT|COLLATE|CONSTRAINT)\b|$)",
                      tail, re.I | re.S)
    if found:
        default = found.group(1).strip()
        tail = (tail[:found.start()] + tail[found.end():]).strip()
    return [name, data_type, default, " ".join(tail.split())]


def _ddl(path: Path, relative: Path) -> list[tuple[Path, mdio.Doc]]:
    """DDL の骨格。**読まなかった文は数えて申告する。**

    方言は無数にあるので、読むのは ``CREATE TABLE`` と ``CREATE INDEX`` と
    ``COMMENT ON`` だけにする ―― **全部を読もうとすると、当てにいく作業が
    永久に続く**（決定 1 が捨てた側である）。読まなかった文を黙って落とすと
    「資料に無い」と「機械が読めていない」が混ざるので、必ず数えて名指しする。
    """
    text = _text_of(path)
    posix = relative.as_posix()
    doc = mdio.Doc(title=posix, source=posix)

    tables = 0
    others: list[str] = []
    index_rows: list[list[str]] = []
    comment_rows: list[list[str]] = []

    for line, statement in _sql_scan(text):
        flat = " ".join(statement.split())
        head = flat.upper()
        at = f"{posix}#L{line}"
        if re.match(r"CREATE\s+(?:\w+\s+)*TABLE\b", head):
            before, body = _sql_body(statement)
            name = _sql_name(re.sub(r"(?i)^CREATE\s+(?:\w+\s+)*TABLE\s+"
                                    r"(?:IF\s+NOT\s+EXISTS\s+)?", "",
                                    " ".join(before.split())))
            if not name:
                others.append(flat[:60])
                continue
            tables += 1
            columns: list[list[str]] = [["列", "型", "既定値", "制約"]]
            constraints: list[list[str]] = [["制約"]]
            for item in _sql_split(body):
                row = _sql_column(item)
                if row is None:
                    constraints.append([" ".join(item.split())])
                else:
                    columns.append(row)
            doc.chunks.append(mdio.Chunk(anchor=f"t{tables}", at=at,
                                         heading=f"テーブル: {name}",
                                         rows=columns))
            if len(constraints) > 1:
                # **表に掛かる制約は列とは別のアンカー**にする。主キーと外部キーは
                # テーブル間の関係の根拠になるので、列 1 本の話と混ぜると
                # 「どの事実を出典にしたいか」が選べない。
                doc.chunks.append(mdio.Chunk(anchor=f"k{tables}", at=at,
                                             heading=f"テーブルの制約: {name}",
                                             rows=constraints))
        elif re.match(r"CREATE\s+(?:UNIQUE\s+)?(?:\w+\s+)*INDEX\b", head):
            before, body = _sql_body(statement)
            found = re.match(r"(?i)CREATE\s+(UNIQUE\s+)?(?:\w+\s+)*INDEX\s+"
                             r"(?:IF\s+NOT\s+EXISTS\s+)?(?P<name>\S+)\s+"
                             r"ON\s+(?P<table>\S+)", " ".join(before.split()))
            if not found:
                others.append(flat[:60])
                continue
            index_rows.append([_sql_name(found.group("name")),
                               _sql_name(found.group("table")),
                               ", ".join(_sql_split(body)),
                               "UNIQUE" if found.group(1) else "", str(line)])
        elif head.startswith("COMMENT ON"):
            found = re.match(r"(?i)COMMENT\s+ON\s+(?P<kind>\w+)\s+(?P<target>\S+)"
                             r"\s+IS\s+(?P<body>.+)$", flat)
            if not found:
                others.append(flat[:60])
                continue
            comment_rows.append([found.group("kind").upper(),
                                 _sql_name(found.group("target")),
                                 _sql_name(found.group("body")),
                                 str(line)])
        else:
            others.append(flat[:60])

    if index_rows:
        doc.chunks.append(mdio.Chunk(
            anchor="x1", at=posix, heading="索引",
            rows=[["索引", "テーブル", "構成列", "一意", "行"], *index_rows]))
    if comment_rows:
        # **列の論理名はここにしか無いことがある。** Oracle/PG の現場では、
        # 日本語の項目名を `COMMENT ON` だけで持つ運用が普通にある。
        doc.chunks.append(mdio.Chunk(
            anchor="c1", at=posix, heading="コメント",
            rows=[["対象", "名前", "コメント", "行"], *comment_rows]))
    if others:
        doc.notes.append(
            f"読まなかった文が {len(others)} 本あります（読むのは "
            "`CREATE TABLE`・`CREATE INDEX`・`COMMENT ON` だけです）。"
            + "／".join(f"`{text}`" for text in others[:5])
            + ("ほか" if len(others) > 5 else "")
            + "。ビューとストアドの中に仕様が入っていることがあるので、"
              "資料として要るかどうかは人が決めてください。")

    if not doc.chunks and not doc.notes:
        return []
    return [(Path(*relative.parts[:-1]) / f"{relative.name}{mdio.EXT}", doc)]


# ── Java ────────────────────────────────────────────────────────
#: 型の宣言を始める語。**``@interface`` を先に見る** ―― 後ろに回すと
#: ``public @interface Service`` が ``interface`` に当たって注釈型が消える。
_JAVA_TYPE = {"@interface": "注釈型", "class": "クラス",
              "interface": "インタフェース", "enum": "列挙", "record": "レコード"}

#: 識別子。Java は全角も識別子に使えるので ``\w`` で取る（日本語のクラス名は
#: 実案件にある）。
_JAVA_IDENT = re.compile(r"[\w$]+")

#: 注釈 1 つ。引数を持つ（``@Column(name = "ORDER_NO")``）ものも 1 つに数える。
#: ``@interface`` は**注釈ではなく注釈型の宣言**なので外す ―― 一緒に落とすと、
#: 自前で宣言した注釈（Spring を持ち込めない現場では普通にある）が
#: 型として 1 つも出なくなる。
_JAVA_ANNOTATION = re.compile(
    r"@(?!interface\b)[\w.$]+(?:\s*\([^()]*(?:\([^()]*\)[^()]*)*\))?")


def _java_clean(text: str) -> tuple[str, str, list[int]]:
    """``(数える用, 出す用, 1 文字ごとの行番号)``。**どれも長さは元のまま。**

    ``{`` の対応だけで宣言を切り出すので、``"}"`` という文字列や ``// }`` という
    コメントが 1 つ混ざるだけで**そこから先の型が全部ずれる**。数える用は文字列も
    コメントも潰す。

    出す用は**コメントだけ**を潰す。元のまま出すと javadoc が宣言の欄へ丸ごと
    入り（意図の層は出さないと決めてある）、数える用をそのまま出すと
    ``@Column(name = "ORDER_NO")`` の中身が消える ―― 潰す理由が 2 つあって
    範囲が違うので、写しも 2 つ要る。
    """
    out: list[str] = []
    shown: list[str] = []
    lines: list[int] = []
    line = 1
    index = 0
    size = len(text)
    while index < size:
        char = text[index]
        pair = text[index:index + 2]
        if pair == "//":
            while index < size and text[index] != "\n":
                out.append(" ")
                shown.append(" ")
                lines.append(line)
                index += 1
            continue
        if pair == "/*":
            while index < size and text[index:index + 2] != "*/":
                out.append(" ")
                shown.append(" ")
                lines.append(line)
                line += text[index] == "\n"
                index += 1
            for _ in range(min(2, size - index)):
                out.append(" ")
                shown.append(" ")
                lines.append(line)
                index += 1
            continue
        if char in ('"', "'"):
            close = char
            out.append(" ")
            shown.append(char)
            lines.append(line)
            index += 1
            while index < size and text[index] != close:
                if text[index] == "\\":
                    out.append(" ")
                    shown.append(text[index])
                    lines.append(line)
                    index += 1
                    if index >= size:
                        break
                out.append(" ")
                shown.append(text[index])
                lines.append(line)
                line += text[index] == "\n"
                index += 1
            out.append(" ")
            shown.append(close)
            lines.append(line)
            index += 1
            continue
        out.append(char)
        shown.append(char)
        lines.append(line)
        line += char == "\n"
        index += 1
    return "".join(out), "".join(shown), lines


def _java_split(head: str) -> tuple[str, str]:
    """宣言から注釈を切り離す。**意味はここにしか無い。**

    ``@Entity`` はエンティティ、``@RestController`` は外部インターフェース、
    ``@Column`` はデータ項目 ―― Java の現場では、種別を決める手がかりが本体では
    なく注釈に載っている。本体と混ぜると整理層が正規表現を書くことになる。
    """
    marks = _JAVA_ANNOTATION.findall(head)
    body = _JAVA_ANNOTATION.sub(" ", head)
    return " ".join(" ".join(marks).split()), " ".join(body.split())


def _java_name(body: str, kind: str) -> str:
    """宣言から名前を取る。**取れなければ空**（当てにいかない）。"""
    if kind in _JAVA_TYPE.values():
        for keyword, label in _JAVA_TYPE.items():
            found = re.search(rf"(?:^|\s){re.escape(keyword)}\s+([\w$]+)", body)
            if label == kind and found:
                return found.group(1)
        return ""
    if kind == "メソッド":
        head = body.split("(")[0]
        names = _JAVA_IDENT.findall(head)
        return names[-1] if names else ""
    head = body.split("=")[0]
    names = _JAVA_IDENT.findall(head)
    return names[-1] if names else ""


def _java_kind(body: str) -> str:
    """宣言の種類。``{`` の前に何が書いてあるかだけで決まる。"""
    for keyword, label in _JAVA_TYPE.items():
        if re.search(rf"(?:^|\s){re.escape(keyword)}\s+[\w$]+", body):
            return label
    return "メソッド" if "(" in body else "フィールド"


#: 宣言の頭に見えるが宣言ではないもの。``static { … }`` の初期化子・制御構文の
#: 頭がここに来る（メソッドの中は見ないので残るのは型の直下だけである）。
_JAVA_SKIP = {"static", "synchronized", "if", "for", "while", "switch", "try",
              "catch", "finally", "do", "else", "return", "new", "package",
              "import", "this", "super"}


def _java_members(text: str) -> list[tuple[int, int, str, str, str, str]]:
    """``(深さ, 行, 種類, 名前, 注釈, 宣言)`` を上から順に。

    **中身は読まない。** 業務ルールや条件分岐の意味は javadoc と本体にあり、
    そこは整理層が原本を読む（決定 2）―― 機械が出すのは宣言の頭だけである。

    どこまでが「宣言の頭」かは深さでは決まらない。``{`` を数えるだけだと
    **メソッドの中の文が深さ 2 のメンバに見える**（`int matched = 0;` が
    フィールドとして出ていた）。波括弧 1 つずつに**それが型の本体かどうか**を
    覚えさせ、**型の本体だけを通ってきた宣言**を members とする。
    """
    clean, visible, lines = _java_clean(text)
    found: list[tuple[int, int, str, str, str, str]] = []
    inside: list[bool] = []                     # その波括弧は型の本体か
    start = 0
    for index, char in enumerate(clean):
        if char not in "{};":
            continue
        head, raw = clean[start:index], visible[start:index]
        start = index + 1
        if char == "}":
            if inside:
                inside.pop()
            continue

        body = _java_split(head)[1]
        # 注釈と宣言は**文字列の残っているほう**から取る。`@Column(name = "ORDER_NO")`
        # の中身が消えると、物理名の根拠がどこにも無くなる（列の論理名を
        # `COMMENT ON` からしか取れない DDL と同じ形の穴である）。
        marks, shown = _java_split(raw) if head.strip() else ("", "")
        kind = _java_kind(body) if body else ""
        # `int[] a = {1, 2}` の初期化子は宣言ではない（`=` が `(` より先に来る）
        initializer = char == "{" and "=" in body.split("(")[0]
        is_type = (kind in _JAVA_TYPE.values() and char == "{"
                   and not initializer)

        member = all(inside) and body and not initializer
        if char == "{":
            inside.append(bool(is_type))
        if not member:
            continue
        name = _java_name(body, kind)
        if not name or name in _JAVA_SKIP or body.split()[0] in _JAVA_SKIP:
            continue
        if char == ";" and kind == "メソッド" and "(" not in body:
            continue
        found.append((len(inside) - (1 if char == "{" else 0),
                      lines[min(index, len(lines) - 1)], kind, name,
                      marks, shown or body))
    return found


def _java(path: Path, relative: Path) -> list[tuple[Path, mdio.Doc]]:
    """Java の宣言と注釈。**構文解析まで踏み込まない。**

    全文を読もうとすると方言（世代ごとの文法・生成コード・ロンボク）に付き合う
    作業が永久に続く ―― 決定 1 が捨てた側である。読むのは ``{`` の対応だけで
    取れるところ、つまり**宣言の頭と注釈**に限る。

    注釈に寄せるのは、**Java の現場では種別を決める手がかりが本体ではなく注釈に
    載っている**からである（``@Entity`` / ``@RestController`` / ``@Column``）。
    """
    text = _text_of(path)
    posix = relative.as_posix()
    doc = mdio.Doc(title=posix, source=posix)
    clean, _, lines = _java_clean(text)

    rows: list[list[str]] = [["取り込み", "元", "行"]]
    for found in re.finditer(r"^\s*(package|import)\s+([^;]+);", clean, re.M):
        # 行は**語のところ**で数える。`^\s*` は前の空行まで飲むので、`found.start()`
        # だと `package` が 1 行前に、コメント直後の import が 4 行前に出ていた。
        line = lines[min(found.start(1), len(lines) - 1)]
        source = " ".join(found.group(2).split())
        rows.append([f"{found.group(1)} {source}", source, str(line)])

    members = _java_members(text)
    index = 0
    current: mdio.Chunk | None = None
    for depth, line, kind, name, marks, body in members:
        if kind in _JAVA_TYPE.values() and depth == 0:
            index += 1
            current = mdio.Chunk(
                anchor=f"j{index}", at=f"{posix}#L{line}",
                heading=f"{kind}: {name}",
                rows=[["メンバ", "種類", "注釈", "宣言", "行"],
                      [name, kind, marks, body, str(line)]])
            doc.chunks.append(current)
            continue
        if current is None:
            continue                            # 型の外にある宣言は持ち主が無い
        current.rows.append([name, kind, marks, body, str(line)])

    if len(rows) > 1:
        doc.chunks.append(mdio.Chunk(anchor="i1", at=posix, heading="取り込み",
                                     rows=rows))
    if not doc.chunks:
        return []
    return [(Path(*relative.parts[:-1]) / f"{relative.name}{mdio.EXT}", doc)]
