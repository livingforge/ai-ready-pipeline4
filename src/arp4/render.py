"""シートを画像にする ―― **テキストに落ちなかったものを、見て確かめるため。**

パース結果は「機械が読めた分」しか持たない。図形は中の文字までで**線の接続と配置は
取れず**（→ :mod:`arp4.parse`）、画像として貼られた表に至っては 1 文字も出ない。
読めていないと分かっているのに手が無いのが困るので、その範囲を**見た目そのまま**
画像にして、目のある相手（LLM）へ渡せるようにする。

**なぜ分割するか。** 1 シートを 1 枚に収めると、長い一覧表は縦に数千ピクセルになる。
受け取り側は長辺を 1500 px 前後へ縮めてから見るので、縮小率が上がるほど字が潰れる
―― **読めない 1 枚より、読める 3 枚のほうがよい。** 分割は「入るように切る」のでは
なく「**縮小されても字が残るように切る**」ためのものなので、しきい値は行数ではなく
ピクセルで持つ。

**横は簡単には切らない。** 縦に切っても表は続きとして読めるが、横に切ると 1 行が
2 枚に分かれて対応が取れなくなる。だから横は「縮小しても読めない」ほど広いときだけ
切る（:data:`WIDE_PX`）。

**切り口は結合セルを避ける。** 日本の設計書は結合セルで見出しを作るので、そこを
跨いで切ると見出しだけが前のタイルに残る。境界が結合を割るときは結合の手前まで
戻す。加えて数行ぶん重ねて撮る ―― **重複はレビューで捨てられるが、切れた行は
そこに何かあったことすら分からない。**

**図形と貼り付け画像も同じく避ける ―― こちらのほうが損が大きい。撮る理由その
ものが図形だからである。** 実測（`受注管理システム基本設計書` の `画面レイアウト`）：
セルは AB 列までしか無いのに図形が AO 列まで届くので横に割れ、切れ目（AH ｜ AI）が
注記の箱と貼り付け画像 2 枚を縦に貫いた。**左側が 1 枚目・右側が 2 枚目に写り、
どちらの絵にも画像が丸ごとは出ない。** 重ね幅（:data:`OVERLAP` の 2 列）はここでは
128 px しかなく、重ねても繋がらない ―― しかも**絵は 2 枚とも「写ってはいる」ので、
撮れなかったことにすらならない**。だから境界は図形のアンカー（:func:`drawing_boxes`）
も避ける。避けきれない（図形 1 個が予算より大きい）ときだけ切るのは結合と同じ。

**広がりは図形も見て決める。** セルが空でも図形だけが置かれたシートがある
（業務フローを図形で描いたシートがまさにそれで、**画像化したいのは大抵これ**）。
openpyxl の使用範囲は図形を知らないので、描画のアンカーから届く先を足す。

**Excel COM を使う。** 図形・罫線・塗りを見た目どおり撮れるのは Excel 自身だけで、
openpyxl では描画を復元できない。よって Windows + Excel + pywin32 が要る
（``pip install "ai-ready-pipeline4[render]"``）。無い環境では**黙って劣化させず**、
「Excel が要る」ことを含むエラーで止める ―― 中身の無い PNG が出るほうが後で高くつく。

**撮り方は「印刷」である。** 2 側（docextract）は範囲を ``CopyPicture`` してチャートへ
貼る手を使っており、こちらも最初はそれを写した。**画面の外にある範囲が撮れない。**
``CopyPicture(xlScreen)`` は「画面に見えているもの」を撮るので、A1 から始まる範囲は
何行でも撮れるのに、``A60:BF77`` のように途中から始まる範囲は貼り付けが必ず失敗する
（``Visible`` を立てても ``ScrollRow`` で送っても直らなかった）。**分割はまさに
「途中から始まる範囲」を作る操作**なので、この経路では分割そのものが成り立たない。

そこで印刷範囲を 1 ページに収めて PDF へ書き出し、それをラスタ化する。画面に依存
しないので何行目からでも撮れ、**クリップボードを使わないので並列に撮れる**（``CopyPicture``
は共有資源を掴むため直列にしか撮れなかった）。文字はベクタのまま焼くので、同じ
ピクセル数なら画面のコピーより読みやすい。紙は内容より大きいので、**焼いたあとに
余白を切り落とす** ―― 余白は受け取り側でトークンを食うだけで何も伝えない。
"""

from __future__ import annotations

import math
import os
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence

from arp4 import mdio
from arp4.parse import column_name, drawing_parts, safe_name

#: ポイント → ピクセル（96 dpi の画面解像度で撮るため）。
PX_PER_POINT = 4 / 3

#: 1 枚の縦の上限（px）。受け取り側が長辺 1500 px 前後へ縮めるので、そこを超えると
#: 縮小率が上がって字が潰れる。**読めるかどうかの基準**であって器の都合ではない。
MAX_PX = 1400

#: 横をやむなく切る幅（px）。ここまでは縮小に耐える（1 行の対応が取れるほうが大事）。
WIDE_PX = 2200

#: 重ねて撮る行数・列数。境界で切れた行を拾い直すための保険。
OVERLAP = 2

#: 書き出す PNG の長辺（px）。受け取り側がここまで縮めるので、それ以上は**送る量が
#: 増えるだけで情報は増えない**。
TARGET_PX = 1400

#: 右端の逃げ（px 相当）。セルの文字は右へはみ出して表示されるが、印刷範囲は列で
#: 切れるので、**画面では読めていた凡例が絵から消える**。数列ぶん余分に撮る。
MARGIN_PX = 240

#: 高さ・幅が宣言されていないときの既定（Excel の既定値）。
DEFAULT_ROW_POINTS = 15.0
DEFAULT_COL_CHARS = 8.43

#: 列幅（文字数）→ ピクセル。Calibri 11pt での Excel の換算式。
_COL_PX_PER_CHAR = 7.0
_COL_PX_PADDING = 5.0

#: EMU（Office の内部単位）→ ポイント。
_EMU_PER_POINT = 12700

#: 図形 1 個が占める区間 ``(行, 列, 行, 列)``。**1 始まり・両端含む。**
Box = tuple[int, int, int, int]

#: 描画のアンカー（parse と同じ 3 種）。
_ANCHORS = ("twoCellAnchor", "oneCellAnchor", "absoluteAnchor")
_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_NS = {"xdr": _XDR}


class RenderUnavailableError(RuntimeError):
    """Excel COM が使えないとき。**次の一手まで書く。**"""


@dataclass(frozen=True)
class Tile:
    """撮る 1 枚。**まだ撮っていない**（計画と実行を分ける）。"""

    rows: tuple[int, int]                  # 1 始まり・両端含む
    cols: tuple[int, int]
    width_px: int
    height_px: int

    @property
    def range(self) -> str:
        """A1 レンジ（``B2:L40``）。"""
        return (f"{column_name(self.cols[0])}{self.rows[0]}"
                f":{column_name(self.cols[1])}{self.rows[1]}")


@dataclass(frozen=True)
class Shot:
    """撮った 1 枚。"""

    tile: Tile
    path: Path
    size: tuple[int, int]                  # 実際の PNG の (幅, 高さ) px


@dataclass(frozen=True)
class Job:
    """撮る依頼 1 件（シート 1 枚ぶん）。"""

    xlsx: Path
    sheet: str
    out_dir: Path
    tiles: tuple[Tile, ...]
    stem: str = ""                         # PNG の名前（既定はシート名）
    #: 撮った絵を貼り戻すパース結果と、その中のアンカー。
    #: **絵は撮っただけでは使われない** ―― 整理層が読むのは ``parsed/**.md`` なので、
    #: そこから辿れないと「撮ったのに誰も見ない」が起きる。
    parsed: Path | None = None
    anchor: str = ""
    #: 全部の絵に載せる見出し行（``3`` なら 1〜3 行目）。**既定は 0＝載せない。**
    #: 長い一覧表を割ると 2 枚目以降に見出しが無く、どの列が何かが読めなくなる。
    #: ただし**どこまでが見出しかは意味の判断**なので、機械は当てない ―― 整理層が
    #: パース結果を見て指定する（`arp4 render --title-rows 3`）。
    title_rows: int = 0


@dataclass
class Result:
    """依頼 1 件の結果。**撮れなかったことも結果である。**"""

    job: Job
    shots: list[Shot] = field(default_factory=list)
    blank: list[Tile] = field(default_factory=list)
    error: str = ""                        # 空でなければ撮れていない
    attached: Path | None = None           # 絵を貼り戻したパース結果


# ── 計画（Excel を起動しない） ──────────────────────────────────
def plan_sheet(xlsx_path: str | Path, sheet: str, *, max_px: int = MAX_PX,
               wide_px: int = WIDE_PX, overlap: int = OVERLAP) -> list[Tile]:
    """1 シートの撮り方を決める。**Excel を起動しない**ので一括で見積もれる。"""
    import warnings

    from openpyxl import load_workbook          # 依存は使うときだけ読む

    with warnings.catch_warnings():
        # **図形は失われない。** ここは図形を撮るためのコマンドで、広がりは
        # 描画パートから自前で読んでいる（:func:`drawing_extent`）。openpyxl の
        # 「Shapes and drawings will be lost」を出したままにすると、いちばん
        # 撮ってほしいシートで「撮っても無駄だ」と読める。
        warnings.filterwarnings("ignore", message=".*DrawingML.*")
        book = load_workbook(xlsx_path, data_only=True)
    try:
        if sheet not in book.sheetnames:
            raise KeyError(f"シートがありません: {sheet}")
        worksheet = book[sheet]
        boxes = drawing_boxes(Path(xlsx_path)).get(sheet, [])
        return plan_worksheet(worksheet, extent_of(boxes), boxes=boxes,
                              max_px=max_px, wide_px=wide_px, overlap=overlap)
    finally:
        book.close()


def plan_worksheet(worksheet: Any, extent: tuple[int, int] = (0, 0), *,
                   boxes: Sequence[Box] = (), max_px: int = MAX_PX,
                   wide_px: int = WIDE_PX,
                   overlap: int = OVERLAP) -> list[Tile]:
    """ワークシート（openpyxl）と図形の届く先から、撮る枚数と範囲を決める。

    ``boxes`` は図形 1 個ずつの占める区間（:func:`drawing_boxes`）で、**切り口を
    決めるために要る** ―― ``extent`` は右下しか持たないので「どこからどこまでが
    1 個の図形か」が分からず、境界がその真ん中へ落ちても気付けない。
    """
    last_row = max(int(worksheet.max_row or 0), extent[0])
    last_col = max(int(worksheet.max_column or 0), extent[1])
    if last_row < 1 or last_col < 1:
        return []
    # openpyxl は**空のシートでも 1 行 1 列**と答える。A1 まで見ないと「何も無い
    # シート」と「A1 だけのシート」が区別できず、白紙を 1 枚撮ることになる。
    if last_row == 1 and last_col == 1 and not str(worksheet["A1"].value or ""):
        return []

    row_px = [_row_px(worksheet, r) for r in range(1, last_row + 1)]
    col_px = _col_px(worksheet, last_col + _margin_cols(worksheet, last_col))
    keep = _keep(worksheet, boxes)

    row_spans = _split(row_px, max_px, overlap, keep[0])
    # **横は最後の手段。** 1 行が 2 枚に分かれると対応が取れなくなる。
    col_spans = _split(col_px, wide_px, overlap, keep[1])

    return [Tile(rows=rows, cols=cols,
                 width_px=round(sum(col_px[cols[0] - 1:cols[1]])),
                 height_px=round(sum(row_px[rows[0] - 1:rows[1]])))
            for rows in row_spans for cols in col_spans]


def plan_range(worksheet: Any, cell_range: str, *, boxes: Sequence[Box] = (),
               max_px: int = MAX_PX, wide_px: int = WIDE_PX,
               overlap: int = OVERLAP) -> list[Tile]:
    """**名指しされた範囲だけ**を撮る計画。

    自動の :func:`plan_worksheet` と違うのは 3 つで、どれも「人が範囲を決めた」
    ことの尊重である。

    - 図形の広がりを足さない（**広げてほしいなら人がそう書く**）
    - 右のはみ出し逃げ（:data:`MARGIN_PX`）を足さない ―― 指定は指定である
    - 空でも撮る（何も無いことを目で確かめたいことがある）

    **切り方は同じ**。狭い範囲でも縦に長ければ割る（拡大したいのに 1 枚へ詰めたら
    元の木阿弥）し、割るなら結合セルも図形も避ける ―― ここは寄って**図形を読む**
    ための入口なので、割った拍子に図形が切れたら来た意味が無い。
    """
    from openpyxl.utils import range_boundaries

    try:
        first_col, first_row, last_col, last_row = range_boundaries(cell_range)
    except Exception as exc:                        # noqa: BLE001 書式違い
        raise ValueError(f"範囲として読めません: {cell_range}（例: B2:L20）") from exc
    if not all((first_col, first_row, last_col, last_row)):
        raise ValueError(f"行と列の両方を書いてください: {cell_range}（例: B2:L20）")

    row_px = [_row_px(worksheet, r) for r in range(first_row, last_row + 1)]
    col_px = _col_px(worksheet, last_col)[first_col - 1:]
    keep = _keep(worksheet, boxes)

    row_spans = _split(row_px, max_px, overlap, keep[0], first_row)
    col_spans = _split(col_px, wide_px, overlap, keep[1], first_col)
    return [Tile(rows=rows, cols=cols,
                 width_px=round(sum(col_px[cols[0] - first_col:
                                           cols[1] - first_col + 1])),
                 height_px=round(sum(row_px[rows[0] - first_row:
                                            rows[1] - first_row + 1])))
            for rows in row_spans for cols in col_spans]


def _split(sizes: Sequence[float], budget: float, overlap: int,
           keep: Iterable[tuple[int, int]], offset: int = 1
           ) -> list[tuple[int, int]]:
    """1 軸を ``budget`` px 以下の区間へ切る。返り値は **1 始まり・両端含む**。

    ``sizes[i]`` は行（列）番号 ``offset + i`` の大きさである。``offset`` があるのは、
    **シートの途中だけを撮る**（``--range``）ときに 1 行目から数え直さないため。

    ``keep`` は割ってはいけない区間（結合セル）。境界がそこへ落ちたら**手前へ戻す**
    ―― 先へ送ると 1 枚が予算を超え、超えたぶんだけ字が潰れる。戻せない
    （その結合が予算より長い）ときは諦めて切る。**進まないより切れたほうがよい。**
    """
    total = len(sizes)
    if total == 0:
        return []
    blocks = sorted(keep)
    spans: list[tuple[int, int]] = []
    start = 0                                       # 0 始まりの索引
    while start < total:
        used = 0.0
        end = start                                 # end は「次の開始」＝終端 +1
        while end < total and (end == start or used + sizes[end] <= budget):
            used += sizes[end]
            end += 1
        if end < total:
            end = _snap(end, start, blocks, offset)
        spans.append((start + offset, end + offset - 1))
        if end >= total:
            break
        start = max(end - overlap, start + 1)       # **必ず前へ進む**
    return spans


def _snap(end: int, start: int, blocks: Sequence[tuple[int, int]],
          offset: int) -> int:
    """境界（0 始まり索引 ``end`` の直前で切る）が結合を割るなら手前へ戻す。"""
    boundary = end + offset                         # 次のタイルの先頭（1 始まり）
    candidates = [first for first, last in blocks
                  if first < boundary <= last and first - offset > start]
    return min(candidates) - offset if candidates else end


# ── 寸法 ────────────────────────────────────────────────────────
def _row_px(worksheet: Any, row: int) -> float:
    dimension = worksheet.row_dimensions.get(row)
    if dimension is not None and getattr(dimension, "hidden", False):
        return 0.0
    height = getattr(dimension, "height", None) if dimension is not None else None
    if height is None:
        height = getattr(worksheet.sheet_format, "defaultRowHeight", None)
    return float(height or DEFAULT_ROW_POINTS) * PX_PER_POINT


def _col_px(worksheet: Any, last_col: int) -> list[float]:
    """1 列目から ``last_col`` までの幅（px）。

    **1 列ずつ引いてはいけない。** openpyxl は ``<col min="1" max="12" width="3"/>``
    を**先頭の列の名前にしか登録しない**ので、``column_dimensions["D"]`` は None を
    返す。それを既定幅で埋めると、狭い列が 12 本並ぶ方眼シートで幅を 3 倍に
    見積もり、要らない横分割が出る。宣言を ``min``〜``max`` へ広げてから引く。
    """
    default = getattr(worksheet.sheet_format, "defaultColWidth", None)
    widths = [float(default or DEFAULT_COL_CHARS)] * (last_col + 1)
    hidden = [False] * (last_col + 1)
    for dimension in worksheet.column_dimensions.values():
        first = int(getattr(dimension, "min", 0) or 0)
        last = int(getattr(dimension, "max", 0) or 0) or first
        if not first:
            continue
        for column in range(first, min(last, last_col) + 1):
            if getattr(dimension, "width", None) is not None:
                widths[column] = float(dimension.width)
            hidden[column] = bool(getattr(dimension, "hidden", False))
    return [0.0 if hidden[c] else widths[c] * _COL_PX_PER_CHAR + _COL_PX_PADDING
            for c in range(1, last_col + 1)]


def _margin_cols(worksheet: Any, last_col: int) -> int:
    """右端に足す列数。**はみ出した文字を撮り逃さないための逃げ**である。

    幅は列によって違うので「何列」ではなく :data:`MARGIN_PX` ぶんで数える
    （方眼シートの 1 列は 17 px 前後、素のシートなら 64 px 前後）。
    """
    widths = _col_px(worksheet, last_col + 32)[last_col:]
    used = 0.0
    for count, width in enumerate(widths, start=1):
        used += width
        if used >= MARGIN_PX:
            return count
    return len(widths)


def _keep(worksheet: Any, boxes: Iterable[Box] = ()
          ) -> tuple[list[tuple[int, int]], list[tuple[int, int]]]:
    """割ってはいけない区間 ``(行, 列)``。**結合セルと図形を同じ扱いにする。**

    1 行（1 列）に収まるものは境界にならない ―― どこで切っても割れないので、
    候補に混ぜると :func:`_snap` が要らない戻しをする。
    """
    rows: list[tuple[int, int]] = []
    columns: list[tuple[int, int]] = []
    merged_cells = getattr(worksheet, "merged_cells", None)
    for merged in getattr(merged_cells, "ranges", []):
        if merged.max_row > merged.min_row:
            rows.append((merged.min_row, merged.max_row))
        if merged.max_col > merged.min_col:
            columns.append((merged.min_col, merged.max_col))
    for first_row, first_col, last_row, last_col in boxes:
        if last_row > first_row:
            rows.append((first_row, last_row))
        if last_col > first_col:
            columns.append((first_col, last_col))
    return rows, columns


def drawing_boxes(path: Path) -> dict[str, list[Box]]:
    """シート名 → 図形 1 個ずつが占める区間。**セルが空でも図形はそこにある。**

    :func:`drawing_extent` は右下しか返さないので、**切り口を決める側には使え
    ない** ―― 「どこからどこまでが 1 個か」が分からないと、境界がその真ん中へ
    落ちても気付けない。読むのは同じ描画パートなので、こちらを正にして広がりの
    ほうを導く。

    **群（group）は 1 個として数える。** アンカーが 1 つなので自然にそうなるが、
    意味の上でもそれでよい ―― ゾーンごとに括られた業務フローは、群の途中で
    切ったら図として読めない。
    """
    found: dict[str, list[Box]] = {}
    for title, bodies in drawing_parts(path).items():
        boxes: list[Box] = []
        for body in bodies:
            try:
                root = ET.fromstring(body)
            except ET.ParseError:
                continue
            for tag in _ANCHORS:
                boxes += [_anchor_box(anchor)
                          for anchor in root.findall(f"xdr:{tag}", _NS)]
        if boxes:
            found[title] = boxes
    return found


def extent_of(boxes: Iterable[Box]) -> tuple[int, int]:
    """図形が届く ``(行, 列)``。図形が無ければ ``(0, 0)``。"""
    ends = list(boxes)
    return (max((box[2] for box in ends), default=0),
            max((box[3] for box in ends), default=0))


def drawing_extent(path: Path) -> dict[str, tuple[int, int]]:
    """シート名 → 図形が届く ``(行, 列)``。**セルが空でも図形はそこにある。**"""
    found: dict[str, tuple[int, int]] = {}
    for title, boxes in drawing_boxes(path).items():
        extent = extent_of(boxes)
        if any(extent):
            found[title] = extent
    return found


def _anchor_box(anchor: ET.Element) -> Box:
    """アンカー 1 個の占める区間（1 始まり・両端含む）。

    ``to`` が無い形（``oneCellAnchor`` / ``absoluteAnchor``）は大きさから
    見積もる。
    """
    start = anchor.find("xdr:from", _NS)
    row = _cell_index(start, "row") if start is not None else 1
    col = _cell_index(start, "col") if start is not None else 1

    to = anchor.find("xdr:to", _NS)
    if to is not None:
        return (row, col,
                max(row, _cell_index(to, "row")),
                max(col, _cell_index(to, "col")))

    last_row, last_col = row, col
    extent = anchor.find("xdr:ext", _NS)
    if extent is not None:
        # 既定の行高・列幅で割るのは**当てずっぽうではなく安全側**の見積もりである
        # （細かい行なら実際はもっと下まで届くが、撮り逃すより広く撮るほうが安い）。
        last_row += math.ceil(_emu_points(extent.get("cy")) / DEFAULT_ROW_POINTS)
        last_col += math.ceil(_emu_points(extent.get("cx"))
                              / (DEFAULT_COL_CHARS * _COL_PX_PER_CHAR
                                 / PX_PER_POINT))
    return (row, col, last_row, last_col)


def _cell_index(element: ET.Element, tag: str) -> int:
    """``xdr:row`` / ``xdr:col`` は 0 始まりなので 1 始まりへ直す。"""
    try:
        return int((element.findtext(f"xdr:{tag}", "0", _NS) or "0").strip()) + 1
    except ValueError:
        return 1


def _emu_points(value: str | None) -> float:
    try:
        return int(value or "0") / _EMU_PER_POINT
    except ValueError:
        return 0.0


# ── 撮影（Excel COM → PDF → ラスタ） ──────────────────────────
_XL_PDF = 0
_XL_PORTRAIT, _XL_LANDSCAPE = 1, 2

#: PDF を焼く解像度。**切り落とす前**の値なので、仕上がり（:data:`TARGET_PX`）より
#: 高くしておく ―― 内容が紙の一部しか占めないとき、低い解像度で焼くと切り出した
#: あとに拡大する羽目になる。
_RASTER_DPI = 300

#: 白とみなす差。JPEG 由来の紙のノイズで外接矩形が紙いっぱいに広がるのを防ぐ。
_WHITE_TOLERANCE = 8


def acquire_excel() -> tuple[Any, Any]:
    """COM を初期化して ``(pythoncom, Excel.Application)`` を返す。

    実際の COM 取得をこの 1 関数へ閉じ込めてあるのは、**Excel を起動しない
    テスト**が差し替えられるようにするためである。
    """
    try:
        import pythoncom
        import win32com.client as com
    except ImportError as exc:                      # pywin32 が無い
        raise RenderUnavailableError(
            "シートの画像化には Excel と pywin32 が要ります"
            '（pip install "ai-ready-pipeline4[render]"）') from exc

    pythoncom.CoInitialize()
    try:
        app = com.DispatchEx("Excel.Application")
    except Exception as exc:                        # noqa: BLE001 Excel 未導入等
        pythoncom.CoUninitialize()
        raise RenderUnavailableError(
            f"Excel を起動できませんでした: {exc}"
            "（Windows と Microsoft Excel が要ります）") from exc
    app.Visible = False
    app.DisplayAlerts = False
    return pythoncom, app


def render_all(jobs: Sequence[Job], *, target_px: int = TARGET_PX,
               acquire: Callable[[], tuple[Any, Any]] | None = None
               ) -> list[Result]:
    """まとめて撮る。**Excel の起動は全体で 1 回、ブックを開くのも 1 冊 1 回。**

    起動と読み込みが所要のほとんどを占めるので、シートごとに開き直すと 15 シートで
    分単位になる。同じブックのシートは隣り合わせて処理する。

    **1 冊の失敗で全部を落とさない。** Excel が開けないブックは実際にあり
    （openpyxl は読めるのに Excel が拒む壊れ方をした資料に当たった）、そこで
    止まると**それまでに撮れた絵まで無かったことになる**。撮れなかった理由は
    :attr:`Result.error` に残して先へ進む ―― 黙って飛ばすのとは違う。
    """
    if not jobs:
        return []

    pythoncom, app = (acquire or acquire_excel)()
    results: list[Result] = []
    try:
        for source, group in _by_book(jobs):
            try:
                book = app.Workbooks.Open(str(source), ReadOnly=True)
            except Exception as exc:                # noqa: BLE001 壊れた資料で止めない
                reason = f"Excel がこのブックを開けませんでした: {_com_message(exc)}"
                results += [Result(job=job, error=reason) for job in group]
                continue
            try:
                for job in group:
                    results.append(_render_one(app, book, job, target_px))
            finally:
                try:
                    book.Close(SaveChanges=False)
                except Exception:                   # noqa: BLE001
                    pass
    finally:
        try:
            app.Quit()
        finally:
            pythoncom.CoUninitialize()
    return results


def _by_book(jobs: Sequence[Job]) -> list[tuple[Path, list[Job]]]:
    """ブックごとにまとめる（**順番は与えられたまま**）。"""
    grouped: dict[Path, list[Job]] = {}
    for job in jobs:
        grouped.setdefault(Path(job.xlsx).resolve(), []).append(job)
    return list(grouped.items())


def _render_one(app: Any, book: Any, job: Job, target_px: int) -> Result:
    """1 シートぶん。**シート 1 枚の失敗も、そのシートだけの失敗に留める。**"""
    result = Result(job=job)
    try:
        worksheet = book.Worksheets(job.sheet)
    except Exception as exc:                        # noqa: BLE001
        result.error = f"シートを開けませんでした: {_com_message(exc)}"
        return result

    directory = Path(job.out_dir)
    directory.mkdir(parents=True, exist_ok=True)
    base = job.stem or safe_name(job.sheet)
    plan = list(job.tiles)
    for index, tile in enumerate(plan, start=1):
        name = f"{base}.png" if len(plan) == 1 else f"{base}-{index}.png"
        out = directory / name
        # 見出しは**元から入っている枚には重ねない**（同じ行が 2 度出る）。
        titles = job.title_rows if tile.rows[0] > job.title_rows else 0
        try:
            drawn = _shoot(app, worksheet, tile, out, target_px, titles)
        except RenderUnavailableError:
            raise
        except Exception as exc:                    # noqa: BLE001
            result.error = f"{tile.range} を撮れませんでした: {_com_message(exc)}"
            return result
        if drawn:
            result.shots.append(Shot(tile=tile, path=out, size=png_size(out)))
        else:
            result.blank.append(tile)

    if result.shots:
        result.attached = attach(job, result.shots)
    return result


def attach(job: Job, shots: Sequence[Shot]) -> Path | None:
    """撮った絵をパース結果のアンカーへ貼る。貼れなければ None。

    リンクは**パース結果からの相対パス**にする。``rounds/r001/`` の中で
    ``parsed/`` と ``images/`` が並んでいるので、エディタでそのまま開ける。
    """
    if job.parsed is None or not job.anchor:
        return None
    links = [(f"{job.sheet}（{shot.tile.range}）", _link(job.parsed, shot.path))
             for shot in shots]
    return job.parsed if mdio.attach(job.parsed, job.anchor, links) else None


def _link(parsed: Path, image: Path) -> str:
    """``../../images/…`` の形へ。**辿れない相対は作らない**（絶対で逃がす）。"""
    try:
        relative = os.path.relpath(image.resolve(), parsed.resolve().parent)
    except ValueError:                              # 別ドライブ
        return image.resolve().as_posix()
    return Path(relative).as_posix()


def _com_message(exc: Exception) -> str:
    """COM の例外は「例外が発生しました」だけになりがちなので、中身まで拾う。"""
    details = getattr(exc, "excepinfo", None)
    if isinstance(details, tuple):
        text = " ".join(str(part) for part in details[1:3] if part)
        if text.strip():
            return text.strip()
    return str(exc)


def render_sheet(xlsx_path: str | Path, sheet: str, out_dir: str | Path, *,
                 stem: str = "", tiles: Sequence[Tile] | None = None,
                 target_px: int = TARGET_PX,
                 acquire: Callable[[], tuple[Any, Any]] | None = None
                 ) -> tuple[list[Shot], list[Tile]]:
    """1 シートだけ撮り、``(撮れた枚, 白紙だった区画)`` を返す。

    まとめて撮るなら :func:`render_all`（起動を分け合える）。撮れなかったときは
    ここでは例外にする ―― 1 枚しか頼んでいない呼び出しに、握り潰した結果を
    返しても仕方がない。
    """
    plan = list(tiles) if tiles is not None else plan_sheet(xlsx_path, sheet)
    if not plan:
        return [], []
    job = Job(xlsx=Path(xlsx_path), sheet=sheet, out_dir=Path(out_dir),
              tiles=tuple(plan), stem=stem)
    result = render_all([job], target_px=target_px, acquire=acquire)[0]
    if result.error:
        raise RenderUnavailableError(result.error)
    return result.shots, result.blank


def _shoot(app: Any, worksheet: Any, tile: Tile, out: Path,
           target_px: int, title_rows: int = 0) -> bool:
    """1 区画を撮って PNG に落とす。白紙なら**何も書かず** False を返す。"""
    import tempfile

    with tempfile.TemporaryDirectory() as work:
        pdf = Path(work) / "tile.pdf"
        _print_area(app, worksheet, tile, title_rows)
        try:
            worksheet.ExportAsFixedFormat(_XL_PDF, str(pdf))
        except Exception as exc:                    # noqa: BLE001
            raise RenderUnavailableError(
                f"{tile.range} を PDF に書き出せませんでした: {exc}") from exc
        image = _rasterize(pdf, target_px)

    if image is None:
        return False
    image.save(out)
    return True


def _print_area(app: Any, worksheet: Any, tile: Tile,
                title_rows: int = 0) -> None:
    """印刷設定を**1 区画 = 1 ページ**に固める。

    ``PrintCommunication`` を落としてからまとめて設定するのは、1 項目ごとに
    プリンタドライバへ往復して**設定だけで数秒かかる**ため。

    **ヘッダ・フッタの消去だけはそのまとめの外でやる。** ``PrintCommunication``
    を落としている間、Excel は**ヘッダ・フッタへの代入だけを黙って捨てる**
    ―― 余白もズームも印刷範囲も見出し行も効くので、まとめて書いているうちの
    1 種類だけが通っていないことに気付けない（実測 ―― バッチの中では
    ``LeftHeader`` が元の文字列のまま残り、外に出すと空になる）。

    落ちていると**紙の装飾が絵に焼き込まれる**。しかも余白を 0 にしてあるので
    ヘッダは 1 行目と同じ高さに来て、**表の見出しへ重ね刷りになる**
    ―― 撮る対象そのものを潰すので、雑音より悪い。文書番号・版・機密区分は
    パースが `s?-p1` で別に申告しているから、絵に要るものは 1 つも無い。
    """
    setup = worksheet.PageSetup
    # **まとめの前に消す**（中で消しても通らない ―― docstring の理由）。
    for attribute in ("LeftHeader", "CenterHeader", "RightHeader",
                      "LeftFooter", "CenterFooter", "RightFooter"):
        try:
            setattr(setup, attribute, "")
        except Exception:                           # noqa: BLE001 版差なら素で続行
            pass
    try:
        app.PrintCommunication = False
    except Exception:                               # noqa: BLE001 版差なら素で続行
        pass
    try:
        setup.PrintArea = tile.range
        setup.PrintTitleRows = f"$1:${title_rows}" if title_rows else ""
        setup.Zoom = False
        setup.FitToPagesWide = 1
        setup.FitToPagesTall = 1                    # **1 区画は必ず 1 ページ**
        setup.LeftMargin = setup.RightMargin = 0
        setup.TopMargin = setup.BottomMargin = 0
        setup.HeaderMargin = setup.FooterMargin = 0
        setup.PrintGridlines = False                # 罫線は資料の情報、目盛は違う
        setup.PrintHeadings = False
        setup.CenterHorizontally = False
        setup.CenterVertically = False
        setup.Orientation = (_XL_LANDSCAPE if tile.width_px > tile.height_px
                             else _XL_PORTRAIT)
    finally:
        try:
            app.PrintCommunication = True
        except Exception:                           # noqa: BLE001
            pass


def _rasterize(pdf: Path, target_px: int) -> Any:
    """PDF を焼いて**余白を切り落とし**、長辺を ``target_px`` に揃える。

    複数ページになったら**縦に繋ぐ**。1 区画 1 ページに固めてあるので普通は
    起きないが、手動の改ページが残っていると割れることがあり、そのとき捨てると
    区画の一部が黙って消える。
    """
    import pypdfium2 as pdfium                      # 依存は使うときだけ読む
    from PIL import Image

    document = pdfium.PdfDocument(str(pdf))
    try:
        pages = [_trim(document[i].render(scale=_RASTER_DPI / 72).to_pil()
                       .convert("RGB")) for i in range(len(document))]
    finally:
        document.close()

    parts = [page for page in pages if page is not None]
    if not parts:
        return None
    image = parts[0] if len(parts) == 1 else _stack(parts)

    scale = target_px / max(image.size)
    if scale < 1:
        image = image.resize((max(1, round(image.width * scale)),
                              max(1, round(image.height * scale))),
                             Image.LANCZOS)
    return image


def _trim(image: Any) -> Any:
    """紙の余白を落とす。**真っ白なら None**（＝そこには何も無かった）。"""
    from PIL import Image, ImageChops

    white = Image.new("RGB", image.size, (255, 255, 255))
    mask = ImageChops.difference(image, white).convert("L")
    box = mask.point(lambda v: 255 if v > _WHITE_TOLERANCE else 0).getbbox()
    return image.crop(box) if box else None


def _stack(parts: list[Any]) -> Any:
    """ページを縦に繋ぐ（幅は最大に合わせ、左揃え）。"""
    from PIL import Image

    width = max(part.width for part in parts)
    canvas = Image.new("RGB", (width, sum(p.height for p in parts)),
                       (255, 255, 255))
    top = 0
    for part in parts:
        canvas.paste(part, (0, top))
        top += part.height
    return canvas


# ── 出来た画像を確かめる ────────────────────────────────────────
_PNG_MAGIC = b"\x89PNG\r\n\x1a\n"


def png_size(path: Path) -> tuple[int, int]:
    """PNG の ``(幅, 高さ)``。**読めなければ (0, 0)**（ヘッダの 24 バイトだけ読む）。

    実測を出すためだけでなく、**空撮の検知**にも使う ―― 貼り付けが空振りしたときは
    ここが極端に小さくなる。
    """
    try:
        head = path.read_bytes()[:24]
    except OSError:
        return (0, 0)
    if len(head) < 24 or not head.startswith(_PNG_MAGIC) or head[12:16] != b"IHDR":
        return (0, 0)
    return (int.from_bytes(head[16:20], "big"), int.from_bytes(head[20:24], "big"))
