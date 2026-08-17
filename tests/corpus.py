"""検証用の資料コーパス ―― **実物の Excel でよくある書かれ方**を 1 か所に集める。

テスト用の資料は、これまで各テストが必要なぶんだけその場で組み立てていた。
それだと「実物にはあるがテストには無い書かれ方」が誰の目にも触れない ――
数式の合計欄・非表示の行・貼り付け画像・マクロ付きのブックは、どれも実案件の
設計書に普通にあるのに、1 度もパースに通されていなかった。

ここに置いたものは **2 通りに使う**。

* テスト（:mod:`test_parse_corpus`）が丸ごとパースして、落ちるものが無いか見る
* ``python tests/corpus.py <ディレクトリ>`` で**実ファイルとして書き出す**
  ―― 実物のコーパス（別リポ）に混ぜて ``arp4 parse`` を手で回すため

**検体のバイナリはリポジトリに置かない。** 中身がコードで読める形で残っていないと、
「この資料は何を写したものか」が半年後に誰にも分からなくなる。

**掛かるのは検体（`tests/`）にだけである。** `examples/*/資料/` の見本は
**実ファイルを git に入れてある** ―― `arp4` が何を読めるのかは、生成器の
コードより開いた 1 冊のほうが早く分かるからで、Python を動かさずに確かめ
られることに価値がある。**扱いが違うのは、数と変更頻度が違うから**である ――
検体は 61 本・11MB あって変更のたびに全部差し替わるが、見本は 6 冊・21KB で
めったに動かない。
"""

from __future__ import annotations

import datetime as dt
import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

#: 関係の種別（``.rels`` の Type）の頭。ISO のものと Microsoft の独自拡張。
_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_MS = "http://schemas.microsoft.com/office/2017/10/relationships"

# ── 図形（openpyxl は書けないので、保存した xlsx へ差し込む） ────
_RELS = """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Target="../drawings/drawing1.xml"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"/>
</Relationships>
"""

#: 業務フロー ―― **グループ化された箱 3 個**（アンカーは 1 個しかない）と、
#: 矢羽根が終点側の線・始点側の線・**どこにも繋がっていない線**。実物の
#: 業務フローはこの 3 種類の線が混ざる。
_FLOW = """<?xml version="1.0" encoding="UTF-8"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:twoCellAnchor>
    <xdr:grpSp>
      <xdr:sp><xdr:nvSpPr><xdr:cNvPr id="2" name="box1"/></xdr:nvSpPr>
        <xdr:txBody><a:p><a:r><a:t>受注登録</a:t></a:r></a:p></xdr:txBody></xdr:sp>
      <xdr:sp><xdr:nvSpPr><xdr:cNvPr id="3" name="box2"/></xdr:nvSpPr>
        <xdr:txBody><a:p><a:r><a:t>与信</a:t></a:r><a:r><a:t>判定</a:t></a:r></a:p>
          <a:p><a:r><a:t>NG は差戻し</a:t></a:r></a:p></xdr:txBody></xdr:sp>
      <xdr:sp><xdr:nvSpPr><xdr:cNvPr id="4" name="box3"/></xdr:nvSpPr>
        <xdr:txBody><a:p><a:r><a:t>出荷指示</a:t></a:r></a:p></xdr:txBody></xdr:sp>
    </xdr:grpSp>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
  <xdr:twoCellAnchor>
    <xdr:cxnSp><xdr:nvCxnSpPr><xdr:cNvCxnSpPr>
      <a:stCxn id="2" idx="3"/><a:endCxn id="3" idx="1"/>
      </xdr:cNvCxnSpPr></xdr:nvCxnSpPr>
      <xdr:spPr><a:ln><a:tailEnd type="triangle"/></a:ln></xdr:spPr></xdr:cxnSp>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
  <xdr:twoCellAnchor>
    <xdr:cxnSp><xdr:nvCxnSpPr><xdr:cNvCxnSpPr>
      <a:stCxn id="4" idx="1"/><a:endCxn id="3" idx="3"/>
      </xdr:cNvCxnSpPr></xdr:nvCxnSpPr>
      <xdr:spPr><a:ln><a:headEnd type="triangle"/></a:ln></xdr:spPr></xdr:cxnSp>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
  <xdr:twoCellAnchor>
    <xdr:cxnSp><xdr:nvCxnSpPr><xdr:cNvCxnSpPr/></xdr:nvCxnSpPr>
      <xdr:spPr><a:ln><a:tailEnd type="triangle"/></a:ln></xdr:spPr></xdr:cxnSp>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
</xdr:wsDr>
"""

#: 画面レイアウト ―― **表のスクリーンショットを貼っただけ**のシート。
#: 実案件でいちばん多い「読めないシート」で、図形と違って撮り直しても読めない。
#:
#: **1 枚目には代替テキストが入っている。** Excel が自動で振るのは名前
#: （``Picture 1``）だけで、``descr`` は人が書いたときにしか入らない ――
#: 中身は取れないままだが、**何の画像かは分かる**。2 枚目は入っていない側の
#: 検体である（実物では入っているほうが少ない）。
#:
#: 埋め込みオブジェクト（Word の貼り込み）も 1 つ載せてある。中身は別形式の
#: ファイルが丸ごと入っているので開けないが、代替テキストがあれば
#: **`sources/` に足すべき資料の名前**が分かる。
_PICTURES = """<?xml version="1.0" encoding="UTF-8"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:twoCellAnchor>
    <xdr:pic><xdr:nvPicPr><xdr:cNvPr id="2" name="Picture 1"
      descr="受注入力画面のイメージ（明細は 20 行まで）"/></xdr:nvPicPr></xdr:pic>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
  <xdr:twoCellAnchor>
    <xdr:pic><xdr:nvPicPr><xdr:cNvPr id="3" name="Picture 2"/></xdr:nvPicPr></xdr:pic>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
  <xdr:twoCellAnchor>
    <xdr:graphicFrame>
      <xdr:nvGraphicFramePr><xdr:cNvPr id="4" name="Object 1"
        descr="項目定義（受注入力）.docx"/></xdr:nvGraphicFramePr>
      <a:graphic><a:graphicData
        uri="http://schemas.openxmlformats.org/presentationml/2006/ole"/></a:graphic>
    </xdr:graphicFrame>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
</xdr:wsDr>
"""


#: SmartArt（組織図・手順・循環）―― **``xdr:sp`` を 1 つも持たない図**である。
#: 描画パートにあるのは「ここに SmartArt が載っている」という枠だけで、箱の文字は
#: ``xl/diagrams/data1.xml`` に別建てで入っている。近年の設計書の体制図・業務
#: フローはこれで描かれていることが多く、図形として数えると 0 個になる。
_SMARTART = """<?xml version="1.0" encoding="UTF-8"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <xdr:twoCellAnchor>
    <xdr:graphicFrame>
      <xdr:nvGraphicFramePr><xdr:cNvPr id="2" name="Diagram 1"/></xdr:nvGraphicFramePr>
      <a:graphic><a:graphicData
        uri="http://schemas.openxmlformats.org/drawingml/2006/diagram"/></a:graphic>
    </xdr:graphicFrame>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
</xdr:wsDr>
"""

#: SmartArt の中身。**箱・繋ぎの点・見た目用の複製が同じ一覧に並ぶ** ――
#: ``parTrans``（親子の線）と ``pres``（見た目の複製）まで拾うと、同じ語が
#: 2 回出て箱の数が合わなくなる。
_DIAGRAM = """<?xml version="1.0" encoding="UTF-8"?>
<dgm:dataModel xmlns:dgm="http://schemas.openxmlformats.org/drawingml/2006/diagram"
               xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <dgm:ptLst>
    <dgm:pt modelId="0" type="doc"/>
    <dgm:pt modelId="1"><dgm:t><a:p><a:r><a:t>PM</a:t></a:r></a:p></dgm:t></dgm:pt>
    <dgm:pt modelId="2"><dgm:t>
      <a:p><a:r><a:t>業務</a:t></a:r><a:r><a:t>チーム</a:t></a:r></a:p>
      <a:p><a:r><a:t>受注・請求</a:t></a:r></a:p></dgm:t></dgm:pt>
    <dgm:pt modelId="3"><dgm:t><a:p><a:r><a:t>基盤チーム</a:t></a:r></a:p></dgm:t></dgm:pt>
    <dgm:pt modelId="4" type="parTrans"><dgm:t><a:p/></dgm:t></dgm:pt>
    <dgm:pt modelId="5" type="sibTrans"><dgm:t><a:p/></dgm:t></dgm:pt>
    <dgm:pt modelId="6" type="pres"><dgm:t><a:p><a:r><a:t>PM</a:t></a:r></a:p></dgm:t></dgm:pt>
  </dgm:ptLst>
</dgm:dataModel>
"""

#: 描画パートから SmartArt のデータパートへの関係。
_DRAWING_RELS = """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Target="../diagrams/data1.xml"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/diagramData"/>
</Relationships>
"""

#: スレッドコメント ―― **2018 年以降の Excel が書くコメント**である。
#: レビュー指摘・決定の経緯はここに溜まるようになった（旧形式は「メモ」に
#: 名前が変わり、既定では作られない）。持っているものが 3 つ増えている ――
#: **返信**（指摘への回答）、**記入者の表示名**、**解決済みかどうか**。
_THREADED = """<?xml version="1.0" encoding="UTF-8"?>
<ThreadedComments
    xmlns="http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments">
  <threadedComment ref="C3" dT="2026-07-01T10:12:00.00" personId="{P1}" id="{T1}">
    <text>採番は 2026 年度から 10 桁へ。旧 8 桁は移行対象</text>
  </threadedComment>
  <threadedComment ref="C3" dT="2026-07-02T09:03:00.00" personId="{P2}" id="{T2}"
                   parentId="{T1}">
    <text>移行対象の洗い出しは 8/20 まで</text>
  </threadedComment>
  <threadedComment ref="C4" dT="2026-06-20T14:40:00.00" personId="{P2}" id="{T3}"
                   done="1">
    <text>桁あふれの検討結果は別紙のとおり</text>
  </threadedComment>
  <threadedComment ref="D5" dT="2026-07-05T11:00:00.00" personId="{P3}" id="{T4}">
    <text>ここは仕様が固まっていない</text>
  </threadedComment>
</ThreadedComments>
"""

#: 記入者の名簿。**コメント側は id しか持たない**ので、名前はここにしか無い。
#: ``{P3}`` をわざと載せていないのは、実物で**退職者・外部の記入者が名簿から
#: 消えている**ことがあるためである（名前が引けないコメントを落とさない）。
_PERSONS = """<?xml version="1.0" encoding="UTF-8"?>
<personList
    xmlns="http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments">
  <person displayName="設計者" id="{P1}" userId="sekkei" providerId="AD"/>
  <person displayName="レビュア" id="{P2}" userId="review" providerId="AD"/>
</personList>
"""

#: 旧形式へのなりすまし ―― スレッドコメントを付けた Excel は、**古い Excel でも
#: 読めるように**同じ内容を旧形式でも書く。openpyxl が返すのはこちらだけなので、
#: 素直に出すと 1 件ごとに 200 字の但し書きが載り、本文はその末尾に埋もれる。
_LEGACY = (
    "[Threaded comment]\n\nYour version of Excel allows you to read this threaded "
    "comment; however, any edits to it will get removed if the file is opened in "
    "a newer version of Excel. Learn more about threaded comments: "
    "https://go.microsoft.com/fwlink/?linkid=870924\n\nComment:\n    {body}")

#: 外部ブック参照 ―― ``='[単価表.xlsx]単価'!B2`` の参照先。**値はキャッシュ
#: される**ので、画面にはただの数が出ている（別ブック由来だと誰も気付かない）。
_EXTERNAL = """<?xml version="1.0" encoding="UTF-8"?>
<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <externalBook
      xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
      r:id="rId1">
    <sheetNames><sheetName val="単価"/><sheetName val="改定履歴"/></sheetNames>
    <sheetDataSet>
      <sheetData sheetId="0"><row r="2"><cell r="B2"><v>1200</v></cell></row></sheetData>
    </sheetDataSet>
  </externalBook>
</externalLink>
"""

#: 参照先ブックの置き場。**本文ではなく ``.rels`` の Target に書いてある。**
_EXTERNAL_RELS = f"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" TargetMode="External"
    Target="../../外部/単価表.xlsx" Type="{_TYPE}/externalLinkPath"/>
</Relationships>
"""


def inject_drawing(path: Path, xml: str, sheet: int = 1,
                   extra: dict[str, str] | None = None) -> Path:
    """保存済みの xlsx へ描画パートを差し込む。**openpyxl は図形を書けない。**

    ``extra`` は追加のパート（SmartArt のデータと、そこへの関係）。
    """
    original = path.with_suffix(".orig")
    shutil.move(path, original)
    rels = f"xl/worksheets/_rels/sheet{sheet}.xml.rels"
    with zipfile.ZipFile(original) as source, zipfile.ZipFile(path, "w") as target:
        for entry in source.infolist():
            if entry.filename != rels:
                target.writestr(entry, source.read(entry.filename))
        target.writestr(rels, _RELS)
        target.writestr("xl/drawings/drawing1.xml", xml)
        for name, body in (extra or {}).items():
            target.writestr(name, body)
    original.unlink()
    return path


def inject(path: Path, parts: dict[str, str],
           links: dict[str, list[tuple[str, str, str]]]) -> Path:
    """パートと関係を差し込む。**既にある関係は残す。**

    :func:`inject_drawing` は ``.rels`` を丸ごと書き換えてよかった（対象の
    シートは他に関係を持っていなかった）。スレッドコメントと外部参照はそうは
    いかない ―― 足す先の ``xl/_rels/workbook.xml.rels`` には**シートへの関係が
    既に並んでいる**ので、書き換えるとブックが 1 冊まるごと開けなくなる。

    ``links`` は ``.rels`` のパス → ``[(Id, 種別, Target)]``。
    """
    original = path.with_suffix(".orig")
    shutil.move(path, original)
    with zipfile.ZipFile(original) as source, zipfile.ZipFile(path, "w") as target:
        existing = {entry.filename: source.read(entry.filename)
                    for entry in source.infolist()}
        for where, added in links.items():
            existing[where] = _with_links(existing.get(where), added)
        for where, body in parts.items():
            existing[where] = body.encode("utf-8")
        for where, body in existing.items():
            target.writestr(where, body)
    original.unlink()
    return path


def _with_links(body: bytes | None, added: list[tuple[str, str, str]]) -> bytes:
    """``.rels`` へ関係を足す（無ければ作る）。"""
    rows = "".join(f'<Relationship Id="{i}" Type="{t}" Target="{g}"/>'
                   for i, t, g in added)
    if body is None:
        return ('<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns='
                '"http://schemas.openxmlformats.org/package/2006/relationships">'
                f"{rows}</Relationships>").encode("utf-8")
    return body.decode("utf-8").replace("</Relationships>",
                                        f"{rows}</Relationships>").encode("utf-8")


# ── 資料 ────────────────────────────────────────────────────────
def _put(sheet, origin: str, rows: list[list]) -> None:
    """``origin`` を左上として表を置く（``None`` は空欄のまま）。"""
    from openpyxl.utils.cell import coordinate_to_tuple

    top, left = coordinate_to_tuple(origin)
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            if value is not None:
                sheet.cell(row=top + r, column=left + c, value=value)


def 基本設計書(path: Path) -> Path:
    """**いちばん普通の設計書。** 表紙・定義表・一覧・図形の業務フローが 1 冊に入る。"""
    book = Workbook()

    cover = book.active
    cover.title = "表紙"
    _put(cover, "B2", [["基本設計書"], ["新販売管理システム"], ["2026-08-01 版"]])

    table = book.create_sheet("受注テーブル")
    table["B2"] = "受注テーブル（T_ORDER）定義書  改訂 2.1"
    _put(table, "B5", [
        ["区分", "論理名", "物理名", "型", "桁", "PK", "必須", "備考"],
        ["ヘッダ", "受注番号", "ORDER_NO", "文字列", 10, "○", "○", "採番ルールは別紙"],
        [None, "受注日", "ORDER_DATE", "日付", None, None, "○", None],
        [None, "顧客コード", "CUSTOMER_CD", "文字列", 8, None, "○", None],
        ["明細", "行番号", "LINE_NO", "数値", 3, "○", "○", None],
        [None, "数量", "QTY", "数値", 5, None, "○", "0 は不可"]])
    # **幅 1 の縦結合＝「同上」**（日本の表の慣習）。画面上は全行に値が見えている。
    table.merge_cells("B6:B8")
    table.merge_cells("B9:B10")

    screen = book.create_sheet("画面項目")
    # **二段見出し**（横結合）。上段は 1 セルにしか値が無く、画面上も 1 つしか見えない。
    screen["B2"] = "受注入力画面"
    screen.merge_cells("B2:E2")
    _put(screen, "B3", [
        ["項目ID", "項目名", "必須", "備考"],
        ["SCR001-01", "受注番号", "×", "採番後に表示"],
        ["SCR001-02", "得意先", "○", None]])
    # **横に並んだ 2 つ目の表**（間は空列 1 本だけ）。実物の一覧シートによくある。
    _put(screen, "G3", [["ボタン", "動作"], ["登録", "受注を確定する"]])

    flow = book.create_sheet("業務フロー")
    flow["A1"] = "受注業務フロー"
    flow["A2"] = "※ 詳細は処理仕様書を参照"
    book.save(path)
    return inject_drawing(path, _FLOW, sheet=4)


def 別フォルダの基本設計書(path: Path) -> Path:
    """**同名・別フォルダ。** 表示名のハッシュで ID を作ると衝突する形。"""
    book = Workbook()
    book.active.title = "受注テーブル"
    _put(book.active, "A1", [["論理名", "物理名"], ["受注番号", "ORDER_NO"]])
    book.save(path)
    return path


def 処理仕様書(path: Path) -> Path:
    """**マクロ付き（.xlsm）。** 中身は .xlsx と同じで、読めるのに弾いていた形。"""
    book = Workbook()
    book.active.title = "受注登録"
    _put(book.active, "A1", [
        ["No", "処理", "内容"],
        [1, "入力チェック", "必須項目が空なら E001 を表示する"],
        [2, "在庫引当", "引当可能数が不足するときは受注を保留にする"]])
    book.save(path)
    return path


def 集計表(path: Path) -> Path:
    """**数式の合計欄。** 計算結果が保存されていないと、表が空欄に見える。

    ツールが書き出したブック・LibreOffice 保存・手動計算のまま保存したブックは
    どれもこうなる。実物では「合計」「率」「件数」の列がまるごと空になる。
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "件数集計"
    _put(sheet, "A1", [
        ["区分", "件数"],
        ["受注", 100],
        ["出荷", 30],
        ["合計", "=SUM(B2:B3)"],
        ["出荷率", "=B3/B2"]])
    sheet["B5"].number_format = "0.0%"

    # **数式しか無いシート。** 落とすとシートが存在したことすら伝わらない。
    derived = book.create_sheet("派生値")
    _put(derived, "A1", [["前月比", "=件数集計!B4/100"]])
    book.save(path)
    return path


def 型の見本(path: Path) -> Path:
    """**画面の見え方と ``str()`` がずれる型**を並べたもの。"""
    book = Workbook()
    sheet = book.active
    sheet.title = "型"
    _put(sheet, "A1", [
        ["項目", "値"],
        ["納期", dt.datetime(2026, 8, 2)],
        ["締切", dt.datetime(2026, 8, 2, 17, 30)],
        ["受付開始", dt.time(9, 30)],
        ["対象外", False],
        ["達成率", 0.153],
        ["消費税率", 0.1],
        ["単価", 1200],
        ["想定件数", 1200000],
        ["処理時間", dt.timedelta(hours=27, minutes=30)],
        ["差額", -1234],
        ["支店コード", "007"],
        ["件数", 15]])
    sheet["B6"].number_format = "0.0%"           # 達成率
    sheet["B7"].number_format = "0%"             # 消費税率
    sheet["B8"].number_format = '"￥"#,##0'       # 単価
    sheet["B10"].number_format = "[h]:mm"        # 24 時間を超える経過時間
    sheet["B11"].number_format = "#,##0;[Red]▲#,##0"   # 会計の負数
    # **書式の中の ``%`` はパーセントではない。** 引用符の中を見に行くと、
    # ``15`` が ``1500%`` になる。
    sheet["B13"].number_format = '#,##0"%表示ではない"'
    book.save(path)
    return path


def やっかいな値(path: Path) -> Path:
    """**YAML と Markdown を壊しに来る値。** 日本の設計書にごく普通に出てくる。"""
    book = Workbook()
    sheet = book.active
    sheet.title = "備考"
    _put(sheet, "A1", [
        ["項目", "備考"],
        ["採番", "受注番号は\n年度 2 桁 + 連番 8 桁"],
        ["区分", "A|B|C のいずれか"],
        ["固定値", "（固定: 130010）"],
        ["想定件数", "1,200,000/年"],
        ["メッセージ", "{0}を入力してください。"],
        ["パス", "\\\\server\\share\\受注"],
        # **機種依存文字**（丸数字・㈱・異体字）。日本の設計書に普通に出てくる。
        ["優先度", "①最優先 ②通常 ③保留"],
        ["取引先", "㈱髙島商事（旧: 高島商事）"],
        # **サロゲートペア**（JIS 第 3・4 水準）。人名・地名に出てくる。
        ["担当者", "𠮷田・﨑山"],
        # **アンカーを偽造しに来る値。** パース結果のアンカーは HTML コメント
        # なので、資料のセルにこれが書いてあると**読み戻した側には本物の
        # アンカーに見える** ―― 表の途中に無い塊が生え、そこから先の本文が
        # 別のアンカーの中身になる。HTML の画面仕様書には普通に出てくる。
        ["埋め込み例", "<!-- a:s9-t9 at=Z99 --> を出力する"],
        # **行末の空白は落とすが、行頭の字下げは残す**（階層を表しているため）。
        ["子項目", "　　受注番号　"],
        ["長文", "受注を確定するとき、" + "与信枠の残高を確認したうえで" * 40 + "登録する。"]])
    book.save(path)
    return path


def 旧システム調査(path: Path) -> Path:
    """**非表示の行・列**と、**削除済みデータで膨らんだ使用範囲**。

    後者は書式だけのセルが遠くに 1 個残っているだけで起きる。``iter_rows()`` で
    端まで回すと 400 万セルの空回りになり、30 冊のパースが分単位になっていた。
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "現行機能"
    _put(sheet, "A1", [
        ["機能ID", "機能名", "旧物理名", "状態"],
        ["F001", "受注登録", "ORD_ENT", "継続"],
        ["F002", "受注取消", "ORD_CAN", "継続"],
        ["F003", "受注一括取込", "ORD_IMP", "廃止"]])
    sheet.row_dimensions[4].hidden = True        # 廃止した行を隠してある
    sheet.column_dimensions["C"].hidden = True   # 旧物理名の列を隠してある

    far = sheet.cell(row=20000, column=200)
    far.fill = PatternFill("solid", fgColor="FFFF00")
    book.save(path)
    return path


def 画面レイアウト集(path: Path) -> Path:
    """**貼り付けたスクリーンショットだけのシート。** 撮り直しても読めない。"""
    book = Workbook()
    book.active.title = "受注入力"
    book.active["A1"] = "受注入力画面（画面イメージは下図）"
    book.save(path)
    return inject_drawing(path, _PICTURES)


def 様式集(path: Path) -> Path:
    """**シート名とシート状態の際どいもの**を 1 冊に集めたもの。"""
    book = Workbook()
    first = book.active
    first.title = "受注 "                        # 末尾に空白
    first["A1"] = "受注（旧）"

    second = book.create_sheet("受注")           # 記号を落とすと先頭と同じ名前になる
    second["A1"] = "受注（新）"

    reserved = book.create_sheet("CON")          # Windows の予約名
    reserved["A1"] = "予約名のシート"

    book.create_sheet("白紙")                    # 空のシートは出さない

    hidden = book.create_sheet("作業用")
    hidden["A1"] = "集計メモ"
    hidden.sheet_state = "hidden"

    single = book.create_sheet("注記")
    single["C5"] = "本書は 2026 年度版である"    # 1x1 は表ではなくテキスト
    book.save(path)
    return path


def 目次(path: Path) -> Path:
    """**ハイパーリンクだけが情報のシート。** 表には表示文字列しか出ていない。

    実物の設計書は 1 冊目が目次で、そこから他のブックへ飛ぶ。リンク先を落とすと
    **まだ集めていない資料があること自体が分からない** ―― 目次が「語の一覧」に
    見えるだけになる。
    """
    from openpyxl.worksheet.hyperlink import Hyperlink

    book = Workbook()
    sheet = book.active
    sheet.title = "目次"
    _put(sheet, "A1", [
        ["No", "資料", "備考"],
        [1, "受注テーブル定義", "本書内"],
        [2, "外部インタフェース仕様", "別ブック"],
        [3, "社内標準", "社内ポータル"],
        [4, "問い合わせ先", "設計チーム"]])
    # **ブックの中を指すリンクは ``location`` にしか入らない**（``target`` は空）。
    sheet["B2"].hyperlink = Hyperlink(ref="B2", location="受注テーブル!A1")
    sheet["B3"].hyperlink = "../外部/外部インタフェース仕様書.xlsx"
    sheet["B4"].hyperlink = "https://example.invalid/standards"
    sheet["B5"].hyperlink = "mailto:sekkei@example.invalid"

    inner = book.create_sheet("受注テーブル")
    _put(inner, "A1", [["論理名", "物理名"], ["受注番号", "ORDER_NO"]])
    book.save(path)
    return path


def 項目定義書(path: Path) -> Path:
    """**字下げで親子を表す表**と、**セルのコメント**と、**入力規則**。

    どれも「セルの値」ではないところに仕様が乗っている書かれ方である。
    """
    from openpyxl.comments import Comment
    from openpyxl.worksheet.datavalidation import DataValidation

    book = Workbook()
    sheet = book.active
    sheet.title = "項目"
    # **全角空白 2 つの字下げが階層**（日本の項目定義書の慣習）。
    _put(sheet, "A2", [
        ["項目名", "型", "桁", "必須", "区分"],
        ["受注ヘッダ", None, None, None, None],
        ["　　受注番号", "文字列", 10, "○", "通常"],
        ["　　受注日", "日付", None, "○", None],
        ["明細", None, None, None, None],
        ["　　行番号", "数値", 3, "○", None]])
    sheet["A5"].comment = Comment(
        "採番は 2026 年度から 10 桁へ。旧 8 桁は移行対象", "設計者")
    # **値の無いセルに付いたコメント。** 値で足切りすると存在ごと消える。
    sheet["E6"].comment = Comment("区分は入力規則から選ぶ", "レビュア")
    # **入力規則の選択肢は取らない**（クリックするまで画面に出ない）。
    rule = DataValidation(type="list", formula1='"通常,返品,見積"', allow_blank=True)
    sheet.add_data_validation(rule)
    rule.add("E4:E20")
    book.save(path)
    return path


def 実績報告(path: Path) -> Path:
    """**グラフだけのシート**と**グラフシート**。どちらもセルが 0 個になる。

    図形（``xdr:sp``）を数えていただけの頃は、この 2 枚はファイルが 1 本も
    出ずに消えていた ―― シートが存在したことすら整理層に伝わらない。

    数えるだけにしていた頃も、出るのは「グラフ 1 個」の 1 行だけだった。
    **タイトルは人が書いた文字**であり、系列が**どのシートのどの範囲**を
    指しているかは ``c:f`` にそのまま書いてある ―― どちらも転記で取れる。
    グラフシートに至っては、それ以外に書くことが 1 つも無い。
    """
    from openpyxl.chart import BarChart, Reference

    book = Workbook()
    data = book.active
    data.title = "元データ"
    _put(data, "A1", [["月", "受注件数", "出荷件数"],
                      ["4月", 120, 110], ["5月", 145, 138]])

    graph = book.create_sheet("推移グラフ")
    chart = BarChart()
    chart.title = "月別の受注・出荷件数"
    chart.add_data(Reference(data, min_col=2, max_col=3, min_row=1, max_row=3),
                   titles_from_data=True)
    chart.set_categories(Reference(data, min_col=1, min_row=2, max_row=3))
    graph.add_chart(chart, "B2")

    only = book.create_chartsheet("売上グラフ")
    second = BarChart()
    second.title = "受注件数の推移"
    second.add_data(Reference(data, min_col=2, min_row=2, max_row=3))
    only.add_chart(second)
    book.save(path)
    return path


def 体制図(path: Path) -> Path:
    """**SmartArt で描いた図。** 図形は 0 個で、文字は別のパートにある。"""
    book = Workbook()
    book.active.title = "体制"
    book.active["A1"] = "プロジェクト体制（詳細は下図）"
    book.save(path)
    return inject_drawing(path, _SMARTART, extra={
        "xl/drawings/_rels/drawing1.xml.rels": _DRAWING_RELS,
        "xl/diagrams/data1.xml": _DIAGRAM})


def 移行計画(path: Path) -> Path:
    """**エラー値の残った表**と**非表示シート**。どちらも黙ると誤読される。

    ``#REF!`` は表の上では値のように見えるが、その欄の仕様は資料から読み取れて
    いない。非表示シートは「旧版」を隠しただけのことがある。
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "課題一覧"
    _put(sheet, "A1", [
        ["No", "課題", "移行件数", "担当"],
        [1, "顧客マスタの名寄せ", 12000, "業務"],
        [2, "受注履歴の移送", "#REF!", "基盤"],
        [3, "単価表の突合", "#N/A", "業務"]])

    work = book.create_sheet("作業用")
    work["A1"] = "件数の下書き"
    work.sheet_state = "hidden"

    old = book.create_sheet("旧版_v1.0")
    _put(old, "A1", [["No", "課題"], [1, "旧レイアウトの調査"]])
    old.sheet_state = "hidden"
    book.save(path)
    return path


def 一覧表(path: Path) -> Path:
    """**1 枚に表が 4 つ。** 横に 2 つ・縦に 2 つ並び、間は空列／空行 1 本だけ。

    実物の一覧シートでいちばん多い形である。空を詰めて出すと、横に並んだ別々の
    表が**同じ行に無いものを同じ行に**並べた 1 枚の表になり、番地から位置を
    割り出せなくなる（``at=B3:F6`` と書いてあるのに列が 4 本しかない）。
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "画面一覧"
    _put(sheet, "B3", [
        ["画面ID", "画面名"],
        ["SCR001", "受注入力"],
        ["SCR002", "受注一覧"]])
    _put(sheet, "E3", [                          # 間は D 列 1 本だけ
        ["ボタン", "動作"],
        ["登録", "受注を確定する"]])
    _put(sheet, "B7", [                          # 間は 6 行目 1 本だけ
        ["帳票ID", "帳票名"],
        ["RPT001", "受注一覧表"]])
    # **2 行空ければ別の塊。** ここは繋がらないことを確かめる側の検体である。
    _put(sheet, "B11", [["※ 画面遷移は業務フローを参照"]])
    book.save(path)
    return path


def レビュー記録(path: Path) -> Path:
    """**スレッドコメント**（返信・記入者・解決済み）と、旧形式のメモの混在。

    2018 年以降の Excel が書くコメントである。openpyxl が返すのは古い Excel
    向けの**なりすまし**だけで、本文は 200 字の但し書きの末尾に埋もれ、記入者は
    ``tc={GUID}`` になる ―― レビュー指摘が資料のいちばん新しい情報であることが
    多いのに、そこがいちばん読めない形で出ていた。

    3 枚目は**スレッドのパートを落として配られた資料**の検体である（メールに
    貼られた抜粋・別ツールで再保存されたブックはこうなる）。なりすましだけが
    残るので、但し書きを剥がして本文を出すしかない。
    """
    from openpyxl.comments import Comment

    book = Workbook()
    sheet = book.active
    sheet.title = "指摘"
    _put(sheet, "A2", [
        ["No", "対象", "項目", "状態"],
        [1, "受注テーブル", "受注番号", "確認中"],
        [2, "受注テーブル", "受注金額", "確定"],
        [3, "受注明細", "数量", None]])
    # **なりすましも一緒に置く。** 実物では必ず両方入っているので、混ぜて
    # 2 件に見せないところまでが検体である。
    sheet["C3"].comment = Comment(
        _LEGACY.format(body="採番は 2026 年度から 10 桁へ。旧 8 桁は移行対象"),
        "tc={7A1B2C3D-0000-0000-0000-000000000001}")
    sheet["C4"].comment = Comment(
        _LEGACY.format(body="桁あふれの検討結果は別紙のとおり"),
        "tc={7A1B2C3D-0000-0000-0000-000000000002}")
    # **旧形式のメモも生き残っている**（同じブックに両方あるのが普通である）。
    sheet["A2"].comment = Comment("この表は 2026-07 版", "とりまとめ")

    old = book.create_sheet("旧メモ")
    _put(old, "A1", [["項目", "内容"], ["採番", "8 桁"]])
    old["B2"].comment = Comment(
        _LEGACY.format(body="10 桁化は次期対応"), "tc={7A1B2C3D-0000-0000-0000-000000000003}")

    book.save(path)
    return inject(path, {
        "xl/threadedComments/threadedComment1.xml": _THREADED,
        "xl/persons/person1.xml": _PERSONS,
    }, {
        "xl/worksheets/_rels/sheet1.xml.rels": [
            ("rIdTc1", f"{_MS}/threadedComment",
             "../threadedComments/threadedComment1.xml")],
        "xl/_rels/workbook.xml.rels": [
            ("rIdPerson1", f"{_MS}/person", "persons/person1.xml")],
    })


def 外部参照(path: Path) -> Path:
    """**別のブックを参照している資料。** 参照先の値はキャッシュされている。

    ハイパーリンク（:func:`目次`）と同じく「まだ手元に無い資料の一覧」だが、
    こちらは**表からいっそう見えない** ―― 画面に出ているのはただの数で、
    それが別ブック由来だとは誰も気付かない。参照先が集まっていなければ、
    その数の根拠は次のラウンドでも確かめられない。
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "受注実績"
    _put(sheet, "A1", [
        ["商品コード", "単価", "数量", "金額"],
        ["P001", 1200, 10, 12000],
        ["P002", 800, 5, 4000]])
    book.save(path)
    return inject(path, {
        "xl/externalLinks/externalLink1.xml": _EXTERNAL,
        "xl/externalLinks/_rels/externalLink1.xml.rels": _EXTERNAL_RELS,
    }, {
        "xl/_rels/workbook.xml.rels": [
            ("rIdExt1", f"{_TYPE}/externalLink",
             "externalLinks/externalLink1.xml")],
    })


def 廃止一覧(path: Path) -> Path:
    """**取り消し線で消してある行。** 表の上では生きた行と見分けが付かない。

    既存資産の棚卸しでいちばん多い「消し方」である。行を消さずに線を引くのは、
    **いつ何をやめたかを残すため**なので、資料としてはむしろ丁寧な書き方である
    ―― それが値だけ抜き出すと消える。
    """
    from openpyxl.styles import Font

    book = Workbook()
    sheet = book.active
    sheet.title = "機能一覧"
    _put(sheet, "A1", [
        ["機能ID", "機能名", "備考"],
        ["F001", "受注登録", None],
        ["F002", "受注一括取込", "2026-04 廃止"],
        ["F003", "受注取消", None],
        ["F004", "受注照会", "画面統合"]])
    # **行まるごと消してある**（F002）のと、**1 セルだけ消してある**（F004 の
    # 機能名だけ差し替え途中）のと。実物では後者もごく普通にある。
    for ref in ("A3", "B3", "C3"):
        sheet[ref].font = Font(strike=True)
    sheet["B5"].font = Font(strike=True)
    # **太字・色は取らない側の検体**である（強調は値を偽らない）。
    sheet["B2"].font = Font(bold=True, color="FF0000")
    book.save(path)
    return path


def 工程表(path: Path) -> Path:
    """**表の格子で出すと中身の何十倍にもなるシート**を 2 通り。

    ガントチャートは斜めに ``■`` が並ぶので、空白 1 マスずつで繋がって
    1 つの塊になる ―― 塊の中の空きを詰めない約束（番地から位置を割り出せる
    ようにするため）は、たかだか空行 1 本ぶんの膨らみを想定していた。

    2 枚目は**列を丸ごと選んで掛けた縦結合**である。画面に見えている表は
    3 行しかないのに、素直に最終行まで展開すると 2000 行の表が生える。
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "工程"
    sheet["A1"] = "開発工程表（2026 年度）"
    _put(sheet, "A2", [["工程", *[f"{day}日" for day in range(1, 121)]]])
    for step in range(1, 61):
        sheet.cell(row=step + 2, column=1, value=f"工程{step:02d}")
        for day in range(step, step + 3):             # 斜めに 3 マスずつ
            sheet.cell(row=step + 2, column=day + 1, value="■")

    merged = book.create_sheet("体制")
    _put(merged, "A1", [
        ["区分", "担当"],
        ["業務", "受注チーム"],
        [None, "請求チーム"]])
    merged.merge_cells("A2:A2000")                    # 列を選んで掛けた結合
    book.save(path)
    return path


def 壊れたブック(path: Path) -> Path:
    """**開けない資料でパースを止めない。** 30 冊のうち 1 冊は必ずこうなる。"""
    path.write_bytes(b"PK\x03\x04 not really a workbook")
    return path


#: 資料の置き場 → 作る関数。**フォルダも実物に寄せる**（同名別フォルダを含む）。
CASES: dict[str, object] = {
    "資料/A/基本設計書.xlsx": 基本設計書,
    "資料/B/基本設計書.xlsx": 別フォルダの基本設計書,
    "資料/A/処理仕様書.xlsm": 処理仕様書,
    "資料/A/集計表.xlsx": 集計表,
    "資料/A/目次.xlsx": 目次,
    "資料/A/項目定義書.xlsx": 項目定義書,
    "資料/B/実績報告.xlsx": 実績報告,
    "資料/B/体制図.xlsx": 体制図,
    "資料/C/旧システム調査.xlsx": 旧システム調査,
    "資料/C/画面レイアウト集.xlsx": 画面レイアウト集,
    "資料/C/移行計画.xlsx": 移行計画,
    "資料/D/型の見本.xlsx": 型の見本,
    "資料/D/やっかいな値.xlsx": やっかいな値,
    "資料/D/様式集.xlsx": 様式集,
    "資料/D/一覧表.xlsx": 一覧表,
    "資料/A/レビュー記録.xlsx": レビュー記録,
    "資料/C/外部参照.xlsx": 外部参照,
    "資料/C/廃止一覧.xlsx": 廃止一覧,
    "資料/E/工程表.xlsx": 工程表,
    "資料/E/壊れたブック.xlsx": 壊れたブック,
}


def build(directory: Path) -> list[Path]:
    """コーパスを書き出す。**並びは決定的**（差分をノイズにしない）。"""
    made: list[Path] = []
    for where, make in CASES.items():
        path = directory / where
        path.parent.mkdir(parents=True, exist_ok=True)
        made.append(make(path))                        # type: ignore[operator]
    return made


if __name__ == "__main__":
    import sys

    target = Path(sys.argv[1] if len(sys.argv) > 1 else "corpus").resolve()
    for made in build(target):
        print(made)
