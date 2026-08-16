"""検体を **データとして** 書くための読み込み器 ―― ここに検体は 1 つも無い。

検体（``tests/dataset/*.yml``）と、それを xlsx へ組み立てる手続き（この
ファイル）を分けてある。:mod:`corpus` は 1 冊ごとに関数を書く形で、20 冊まで
はそれで足りていたが、**検体が増えるほど「資料の写し」と「Excel の組み立て方」
が同じ関数の中で混ざる**。混ざると 2 つのことが起きる ――

* **検体を足すのに Python が要る。** 資料の写しは「B5 にこう書いてあった」と
  いう事実の列であって、手続きではない。手続きの形で書くと、書いた本人以外は
  「どのセルに何が入っているか」を関数を実行しないと言えなくなる
* **組み立ての癖が検体に混ざる。** ``inject_drawing`` が ``.rels`` を丸ごと
  書き換えていた（＝ 1 冊に描画が 1 つしか置けない）ような制約が、資料の側の
  都合に見えてしまう

だからここには**資料の知識を 1 つも置かない**。YAML に書いてあることを
そのまま Excel にするだけで、「この検体は何を突くか」は YAML の ``なぜ`` に
書く（:mod:`corpus` の docstring と同じ役目である）。

**バイナリは置かない**という約束は :mod:`corpus` と同じである ―― 中身が
読める形で残っていないと、その資料が何を写したものか半年後に誰にも分からない。

```bash
python tests/dataset.py <ディレクトリ>   # 実ファイルとして書き出す
```
"""

from __future__ import annotations

import datetime as dt
import re
import shutil
import zipfile
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

#: 検体の置き場。**1 ファイルに複数冊**（YAML の ``---`` 区切り）書ける。
SPECS = Path(__file__).with_name("dataset")


class _Loader(yaml.SafeLoader):
    """``No`` を偽と読まない YAML。**設計書の 1 列目はたいてい ``No`` である。**

    YAML 1.1 は ``No`` ``Yes`` ``On`` ``Off`` ``Y`` ``N`` を真偽値と読む。検体を
    素直に書くと、**課題一覧の見出し ``No`` が ``FALSE`` になって Excel に入る**
    ―― パース結果は正しいのに検体のほうが間違っている、といういちばん質の悪い
    失敗である（``ON`` / ``OFF`` の区分値、``Y`` / ``N`` のフラグでも同じ）。

    引用符で囲めば済むが、**囲み忘れが静かに通る**のでは検体として使えない。
    真偽値として読むのは ``true`` / ``false`` だけにする。
    """


_Loader.yaml_implicit_resolvers = {
    first: [(tag, pattern) for tag, pattern in resolvers
            if tag != "tag:yaml.org,2002:bool"]
    for first, resolvers in yaml.SafeLoader.yaml_implicit_resolvers.items()}
_Loader.add_implicit_resolver(
    "tag:yaml.org,2002:bool",
    re.compile(r"^(?:true|True|TRUE|false|False|FALSE)$"), list("tTfF"))

#: 関係の種別。YAML には ``drawing`` と短く書き、ここで正式な URI に直す
#: ―― 200 字の URL が検体に並ぶと、**資料の写しが URL で埋まって読めなくなる。**
_ISO = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_MS = "http://schemas.microsoft.com/office/2017/10/relationships"
_KINDS = {
    "drawing": f"{_ISO}/drawing",
    "diagramData": f"{_ISO}/diagramData",
    "diagramLayout": f"{_ISO}/diagramLayout",
    "diagramQuickStyle": f"{_ISO}/diagramQuickStyle",
    "diagramColors": f"{_ISO}/diagramColors",
    "image": f"{_ISO}/image",
    "chart": f"{_ISO}/chart",
    "externalLink": f"{_ISO}/externalLink",
    "externalLinkPath": f"{_ISO}/externalLinkPath",
    "threadedComment": f"{_MS}/threadedComment",
    "person": f"{_MS}/person",
}


# ── 入口 ────────────────────────────────────────────────────────
def specs() -> list[dict[str, Any]]:
    """検体の定義を読む。**並びは決定的**（差分をノイズにしない）。"""
    found: list[dict[str, Any]] = []
    for path in sorted(SPECS.glob("*.yml")):
        for spec in yaml.load_all(path.read_text(encoding="utf-8"), _Loader):
            if spec:
                found.append(spec)
    return found


def build(directory: Path) -> list[Path]:
    """検体を書き出す。返すのは書いたファイルの一覧。"""
    return [_book(directory, spec) for spec in specs()]


def _book(directory: Path, spec: dict[str, Any]) -> Path:
    path = directory / spec["置き場"]
    path.parent.mkdir(parents=True, exist_ok=True)

    if "生バイト" in spec or "先頭16進" in spec:   # 開けない資料
        # **先頭の数バイトだけ 16 進で書く。** 資料が何であるかは magic bytes に
        # 書いてあり（OLE・PDF・PNG）、パースはそこしか読まない。全体を 16 進で
        # 並べると**検体が読めなくなる**ので、続きは日本語のまま置く。
        head = bytes.fromhex(spec.get("先頭16進", "").replace(" ", ""))
        path.write_bytes(head + spec.get("生バイト", "").encode("utf-8"))
        return path

    if "zip中身" in spec:                        # zip だが Excel ブックではない
        with zipfile.ZipFile(path, "w") as archive:
            for name, body in spec["zip中身"].items():
                archive.writestr(name, body)
        return path

    if "本文" in spec:                           # ソースコード（Excel ではない）
        path.write_bytes(spec["本文"].encode(spec.get("文字コード", "utf-8")))
        return path

    if "そのまま" in spec:
        # **資料ではないものも置き場には入っている。** 客先からもらったフォルダを
        # そのまま置くと `.git` も `node_modules` も付いてくる ―― それを
        # 資料として数えないことは、資料を読むのと同じくらい検体が要る。
        path.mkdir(parents=True, exist_ok=True)
        for name, body in spec["そのまま"].items():
            (path / name).parent.mkdir(parents=True, exist_ok=True)
            (path / name).write_text(str(body), encoding="utf-8")
        return path

    book = Workbook()
    book.remove(book.active)
    for sheet_spec in spec.get("シート", []):
        name = sheet_spec["名前"]
        if sheet_spec.get("種別") == "グラフシート":
            sheet = book.create_chartsheet(name)
        else:
            sheet = book.create_sheet(name)
            _fill(sheet, sheet_spec)
        if sheet_spec.get("状態"):
            sheet.sheet_state = sheet_spec["状態"]
    book.save(path)
    return _inject(path, spec)


# ── シート 1 枚 ─────────────────────────────────────────────────
def _fill(sheet: Any, spec: dict[str, Any]) -> None:
    """値 → 書式 → 結合 → 非表示 の順に置く（後のものが前のものを前提にする）。"""
    for origin, rows in (spec.get("表") or {}).items():
        top, left = coordinate_to_tuple(origin)
        for r, row in enumerate(rows):
            for c, value in enumerate(row):
                if value is not None:
                    sheet.cell(row=top + r, column=left + c, value=_value(value))

    # **規模は書き写せない。** ここまでの ``表`` / ``セル`` は「B5 にこう書いて
    # あった」という事実の列だが、抽出ツールが吐いた 1,200 行の一覧を 1 行ずつ
    # 並べると、**検体が読めなくなるだけで突くものは 1 つも増えない**（1,200 行が
    # 同じ形をしていることこそが、その資料の性質である）。
    #
    # だから**繰り返しであることを宣言する**。ここにも資料の知識は無い ――
    # 何行で、1 行がどういう形かは YAML の側にある。
    # **横の繰り返しも同じ理由で書き写せない。** CRUD 図はテーブル 1 本につき
    # ``C`` ``R`` ``U`` ``D`` の 4 列が並ぶので、20 本で 80 列になる ―― 見出しを
    # 1 セルずつ書くと 3 行で 250 セルになり、**検体が読めなくなる**。
    # ``横回数`` は雛形を右へ敷き詰める回数で、``{m}`` がその番号である
    # （``{n}`` は行の番号）。
    for block in spec.get("生成") or []:
        top, left = coordinate_to_tuple(block["起点"])
        step = int(block.get("列送り", 0))       # 1 行ごとに右へ（対角線を引く）
        wide = int(block.get("横回数", 1))
        span = len(block["雛形"])
        # **2 面付けの右半分は左の続きである**（No. は 51 から始まる）。
        # 番号の始まりを宣言できないと、同じ 1〜50 が 2 度並んだ検体になる。
        first = int(block.get("開始", 1))
        for index, n in enumerate(range(first, first + int(block["行数"]))):
            for m in range(1, wide + 1):
                for offset, template in enumerate(block["雛形"]):
                    if template is None:
                        continue
                    sheet.cell(
                        row=top + index,
                        column=left + (m - 1) * span + offset + step * index,
                        value=str(template).format(n=n, m=m))

    for where, value in (spec.get("セル") or {}).items():
        row, column = coordinate_to_tuple(where)
        sheet.cell(row=row, column=column, value=_value(value))

    for where, number_format in (spec.get("書式") or {}).items():
        for cell in _cells(sheet, where):
            cell.number_format = number_format

    for where in spec.get("取り消し線") or []:
        for cell in _cells(sheet, where):
            cell.font = Font(strike=True)

    # **1 つのセルの中で書式が変わる**（リッチテキスト）。セルの書式は 1 つしか
    # 持てないので、後半だけを消した欄はこの形でしか書けない ―― 宣言するのは
    # 「この文字に線が掛かっていた」という事実だけで、組み立て方は持たせない。
    for where, marked in (spec.get("一部取り消し線") or {}).items():
        row, column = coordinate_to_tuple(where)
        cell = sheet.cell(row=row, column=column)
        cell.value = _rich(str(cell.value or ""), str(marked))

    for where in spec.get("太字") or []:
        for cell in _cells(sheet, where):
            cell.font = Font(bold=True, color="FF0000")

    # **印刷したときだけ見えるもの。** 日本の設計書は綴じて配るので、文書番号・
    # 版・機密区分・ページ番号はフッタに入る ―― どのページにも出ているのに、
    # セルには 1 つも書かれていない。
    for where, text in (spec.get("印刷") or {}).items():
        if where in _PRINT_TEXTS:
            holder, side = _PRINT_TEXTS[where]
            getattr(getattr(sheet, holder), side).text = str(text)
        else:
            setattr(sheet, _PRINT_RANGES[where], str(text))

    # **入力規則。** arp4 は候補を取らないと決めている（クリックするまで画面に
    # 出ない）が、**資料の側には普通にある** ―― 検体に入れておかないと、
    # 「取らないと決めたもの」が実物で効いているかを誰も確かめられない。
    for where, listed in (spec.get("入力規則") or {}).items():
        rule = DataValidation(
            type="list", allow_blank=True,
            formula1=(listed if isinstance(listed, str)
                      else '"' + ",".join(str(one) for one in listed) + '"'))
        rule.add(where)
        sheet.add_data_validation(rule)

    for where, note in (spec.get("コメント") or {}).items():
        body = note if isinstance(note, str) else note["本文"]
        who = "" if isinstance(note, str) else note.get("記入者", "")
        sheet[where].comment = Comment(body, who)

    for where, link in (spec.get("リンク") or {}).items():
        if isinstance(link, str):
            sheet[where].hyperlink = link
        else:
            sheet[where].hyperlink = Hyperlink(
                ref=where, target=link.get("先"), location=link.get("位置"))

    for span in spec.get("結合") or []:
        sheet.merge_cells(span)

    # **列幅は資料の中身ではないが、無いと人が検体を確かめられない。** 既定幅の
    # ままだと日付欄が `########` になり、長い名称が隣の列で切れる ―― パース結果
    # （値そのもの）は変わらないのに、**目で突き合わせる側だけが読めなくなる**。
    for column, width in (spec.get("列幅") or {}).items():
        sheet.column_dimensions[column].width = width
    for row, height in (spec.get("行高") or {}).items():
        sheet.row_dimensions[row].height = height

    # **改行を持つセルは、折り返しを立てないと画面で 1 行に見える。** 値の側は
    # 変わらないので**パース結果は正しいまま**だが、行を高くしてあるのに 1 行
    # しか出ていない検体を開いた人は「この資料は 1 行だったのか」と読む ――
    # 列幅・行高と同じで、**目で突き合わせる側だけが資料と食い違う**。
    # 業務フローの説明欄・処理記述の条件欄は実物では必ず折り返してある。
    for where in spec.get("折り返し") or []:
        for cell in _cells(sheet, where):
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in spec.get("非表示行") or []:
        sheet.row_dimensions[row].hidden = True
    for column in spec.get("非表示列") or []:
        sheet.column_dimensions[column].hidden = True


#: ヘッダ・フッタの置き場。**既定は「奇数ページ」**（全ページに掛かる）。
_PRINT_TEXTS = {
    "ヘッダ左": ("oddHeader", "left"), "ヘッダ中": ("oddHeader", "center"),
    "ヘッダ右": ("oddHeader", "right"), "フッタ左": ("oddFooter", "left"),
    "フッタ中": ("oddFooter", "center"), "フッタ右": ("oddFooter", "right"),
}
#: ページの区切り方。``見出し行`` は「どこまでが見出しか」の作成者の申告である。
_PRINT_RANGES = {"見出し行": "print_title_rows", "見出し列": "print_title_cols",
                 "範囲": "print_area"}


def _cells(sheet: Any, where: str) -> list[Any]:
    """``B3`` でも ``B3:D9`` でも同じように扱う（検体側で書き分けさせない）。"""
    if ":" not in where:
        return [sheet[where]]
    left, top, right, bottom = range_boundaries(where)
    return [sheet.cell(row=r, column=c)
            for r in range(top, bottom + 1) for c in range(left, right + 1)]


def _rich(value: str, marked: str) -> Any:
    """``受注区分（廃止）`` の ``（廃止）`` だけに取り消し線を掛けた値。

    セルの書式（``Font(strike=True)``）は**セル 1 つに 1 つ**なので、後半だけを
    消してある欄はリッチテキストでしか書けない ―― 実物の設計書ではごく普通の
    書き方である（項目名を残して「（廃止）」だけを消す）。
    """
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    before, _, after = value.partition(marked)
    pieces = [TextBlock(InlineFont(), before)] if before else []
    pieces.append(TextBlock(InlineFont(strike=True), marked))
    if after:
        pieces.append(TextBlock(InlineFont(), after))
    return CellRichText(pieces)


def _value(value: Any) -> Any:
    """YAML の値を Excel の値に。**自動解釈に頼らないための明示形**を持つ。

    YAML は ``09:30`` を 570（60 進数）と読み、``2026-08-02`` を日付と読む
    ―― 前者は事故で、後者は都合がいい。**事故のほうだけ**を明示形で塞ぐ。
    """
    if not isinstance(value, dict):
        return value
    # **セル 1 つに 3 万字**（Excel の上限は 32,767 字）。移行の注意事項を
    # 1 つの欄に貼り込んだ資料は実物にあり、**その 1 セルだけで整理層の
    # 読める量を使い切る** ―― 検体に 3 万字を書き写すことはできないので、
    # ここでも**繰り返しであることを宣言する**（``生成`` と同じ理屈）。
    if "繰り返す" in value:
        return str(value["繰り返す"]) * int(value["回数"])
    if "時刻" in value:
        return dt.time.fromisoformat(value["時刻"])
    if "経過" in value:
        hours, _, minutes = value["経過"].partition(":")
        return dt.timedelta(hours=int(hours), minutes=int(minutes))
    if "日時" in value:
        return dt.datetime.fromisoformat(value["日時"])
    return value["値"]


# ── zip への差し込み（openpyxl が書けないもの） ─────────────────
def _inject(path: Path, spec: dict[str, Any]) -> Path:
    """描画・追加パート・関係を差し込む。**既にある関係は残す。**

    ``.rels`` を丸ごと書き換えると、シートへの関係が並んでいる
    ``xl/_rels/workbook.xml.rels`` ではブックが 1 冊まるごと開けなくなる。
    描画も同じで、書き換えていた頃は**1 冊に 1 枚しか置けなかった**。
    """
    parts: dict[str, bytes] = {name: text.encode("utf-8")
                               for name, text in (spec.get("パーツ") or {}).items()}
    links: dict[str, list[tuple[str, str, str, bool]]] = {}

    for where, listed in (spec.get("関係") or {}).items():
        links.setdefault(where, []).extend(
            (one[0], _KINDS.get(one[1], one[1]), one[2], len(one) > 3)
            for one in listed)

    original = path.with_suffix(".orig")
    shutil.move(path, original)
    with zipfile.ZipFile(original) as source:
        body = {entry.filename: source.read(entry.filename)
                for entry in source.infolist()}

    sheets = _sheet_parts(body)
    for sheet_spec in spec.get("シート", []):
        # **1 枚だけ壊す。** 開けない 1 冊（`生バイト`）はブックごと壊すが、
        # こちらは**残りの枚数が読めるか**を見るための検体なので、シート 1 枚の
        # パートだけを途中で切る（ツールが書き出したブック・部分復旧したブックが
        # この形になる）。openpyxl では書けないので zip に直に置く。
        if sheet_spec.get("壊す") and sheet_spec.get("名前") in sheets:
            part = sheets[sheet_spec["名前"]]
            body[part] = body[part][: len(body[part]) // 2]

        rows = [row for row, height in (sheet_spec.get("行高") or {}).items()
                if height == 0]
        columns = [column for column, width in (sheet_spec.get("列幅") or {}).items()
                   if width == 0]
        if (rows or columns) and sheet_spec.get("名前") in sheets:
            body[sheets[sheet_spec["名前"]]] = _crushed(
                body[sheets[sheet_spec["名前"]]], rows, columns)

    for order, sheet_spec in enumerate(spec.get("シート", []), start=1):
        if not sheet_spec.get("描画"):
            continue
        part = f"xl/drawings/drawing{order}.xml"
        xml, added_parts, added_links = _drawing(sheet_spec["描画"], order)
        parts[part] = xml.encode("utf-8")
        parts.update(added_parts)
        links.setdefault(_rels_of(part), []).extend(added_links)
        for one in sheet_spec.get("描画の関係") or []:
            links.setdefault(_rels_of(part), []).append(
                (one[0], _KINDS.get(one[1], one[1]), one[2], len(one) > 3))
        # **ここを忘れると Excel には何も出ない。** パートを置いてシートから
        # 関係を張っても、シート XML の ``<drawing>`` が無ければ Excel は
        # 描画を探しに行かない ―― arp4 は関係だけを辿るので**パースは通り、
        # 実物の Excel で開いたときだけ図が消える**（いちばん気付きにくい）。
        sheet_part = sheets[sheet_spec["名前"]]
        links.setdefault(_rels_of(sheet_part), []).append(
            (f"rIdDraw{order}", _KINDS["drawing"],
             _relative(_dir(sheet_part), part), False))
        body[sheet_part] = _with_drawing(body[sheet_part], order)

    for where, added in links.items():
        body[where] = _with_links(body.get(where), added)
    body["xl/workbook.xml"] = _with_externals(
        body["xl/workbook.xml"], links.get("xl/_rels/workbook.xml.rels", []))
    body.update(parts)
    body["[Content_Types].xml"] = _content_types(
        body["[Content_Types].xml"], parts)

    with zipfile.ZipFile(path, "w") as target:
        for where, blob in body.items():
            target.writestr(where, blob)
    original.unlink()
    return path


def _crushed(body: bytes, rows: list[int], columns: list[str]) -> bytes:
    """高さ 0 / 幅 0 を書き込む。**openpyxl では書けない**ので手で入れる。

    ``ht`` も ``width`` も 0 を偽と見なして落とされる（``if value:`` で絞って
    いる）ので、``行高: {14: 0}`` と書いても**普通の行のまま出てくる**。検体が
    突きたいのはまさにその 0 なので、ここだけは XML に直に書く。

    **``hidden`` は立てない。** 立ててしまうと「隠した行」になり、検体が
    別のものになる ―― 潰れた行は右クリックの「再表示」に出てこないところが
    非表示行との違いで、そこを見るための検体である。
    """
    text = body.decode("utf-8")
    for row in rows:
        # 既にある高さの申告は**捨ててから**書く（openpyxl は ``customHeight``
        # だけ書いていることがあり、二重に付けるとブックごと開けなくなる）。
        def crush(found: re.Match[str]) -> str:
            kept = re.sub(r'\s(?:ht|customHeight)="[^"]*"', "", found.group(1))
            return f'<row r="{row}"{kept} ht="0" customHeight="1">'

        text = re.sub(rf'<row r="{row}"([^>]*)>', crush, text, count=1)

    if columns:
        added = "".join(
            f'<col min="{column_index_from_string(column)}"'
            f' max="{column_index_from_string(column)}"'
            ' width="0" customWidth="1"/>' for column in columns)
        if "<cols>" in text:
            text = text.replace("</cols>", f"{added}</cols>", 1)
        else:                                   # 置き場はスキーマで決まっている
            text = text.replace("<sheetData", f"<cols>{added}</cols><sheetData", 1)
    return text.encode("utf-8")


def _with_externals(body: bytes,
                    added: list[tuple[str, str, str, bool]]) -> bytes:
    """``xl/workbook.xml`` へ ``<externalReferences>`` を足す。

    **``='[1]単価'!B2`` の ``[1]`` はこの並びの番号である。** 関係
    （``workbook.xml.rels``）を張っただけでは Excel は参照先を数えないので、
    番号の指す先が無い数式になり ―― **ブックが丸ごと開けなくなる**。arp4 は
    関係だけを辿るので**パースは通り、実物で開いたときにだけ壊れる**、いつもの
    形の失敗である（``<drawing>`` を書き忘れたときとまったく同じ）。

    置く場所はスキーマで決まっており、``</sheets>`` の直後である。
    """
    ids = [i for i, kind, _, _ in added if kind == _KINDS["externalLink"]]
    if not ids:
        return body
    rows = "".join(f'<externalReference xmlns:r="{_R}" r:id="{one}"/>'
                   for one in ids)
    return body.decode("utf-8").replace(
        "</sheets>", f"</sheets><externalReferences>{rows}</externalReferences>",
        1).encode("utf-8")


def _with_drawing(body: bytes, order: int) -> bytes:
    """シート XML へ ``<drawing>`` を足す。**``<legacyDrawing>`` より前に置く。**

    要素の順番はスキーマで決まっており、コメントの付いたシート（openpyxl が
    ``<legacyDrawing>`` を書く）で後ろに置くと**ブックごと開けなくなる** ――
    「図が出ない」より悪い壊れ方で、しかも図の無いシートでは起きないので
    検体を 1 枚ずつ見ていると気付けない。
    """
    element = f'<drawing xmlns:r="{_R}" r:id="rIdDraw{order}"/>'
    text = body.decode("utf-8")
    for before in ("<legacyDrawing", "</worksheet>"):
        if before in text:
            return text.replace(before, element + before, 1).encode("utf-8")
    return body


#: 差し込んだパートの**種別の申告**。``[Content_Types].xml`` に書いていないと、
#: 既定（``Default Extension="xml"``）の「ただの XML」として扱われる ―― Excel は
#: 描画ともグラフとも思わないので、**黙って無かったことになる**。
#:
#: **申告するのは、こちらが完全な形で作れるものだけ**である。申告した種別の
#: パートは Excel が中身まで検査し、揃っていなければ**ブックごと開けなくなる**
#: ―― 図が出ないより悪い。SmartArt（並べ方・配色・体裁の 3 パートが要る）と
#: スレッドコメント（旧形式のコメント一式との対応が要る）はそちら側なので、
#: パートは置くが申告はしない（arp4 は関係を辿るので読める）。
_DML = "application/vnd.openxmlformats-officedocument.drawingml"
_SML = "application/vnd.openxmlformats-officedocument.spreadsheetml"
_CONTENT = {
    # **外部ブック参照は 3 つ揃って初めて Excel が開ける。** 関係（``.rels``）と
    # ``<externalReferences>``（:func:`_with_externals`）とこの申告で、``[1]`` が
    # 指す先が決まる ―― 1 つでも欠けると番号の指す先が無い数式になり、
    # **ブックが丸ごと開けなくなる**（arp4 は関係だけを辿るのでパースは通る）。
    "xl/externalLinks/externalLink": f"{_SML}.externalLink+xml",
    "xl/drawings/": "application/vnd.openxmlformats-officedocument.drawing+xml",
    "xl/charts/": f"{_DML}.chart+xml",
    "xl/diagrams/data": f"{_DML}.diagramData+xml",
    "xl/diagrams/layout": f"{_DML}.diagramLayout+xml",
    "xl/diagrams/quickStyle": f"{_DML}.diagramStyle+xml",
    "xl/diagrams/colors": f"{_DML}.diagramColors+xml",
}


def _content_types(body: bytes, parts: dict[str, bytes]) -> bytes:
    rows = ""
    if any(name.endswith(".png") for name in parts):
        rows += '<Default Extension="png" ContentType="image/png"/>'
    for name in parts:
        for prefix, kind in _CONTENT.items():
            if name.startswith(prefix):
                rows += f'<Override PartName="/{name}" ContentType="{kind}"/>'
                break
    return body.decode("utf-8").replace("</Types>", f"{rows}</Types>").encode("utf-8")


# ── 図（**Excel が描けるかたち**で出す） ────────────────────────
#
# 描画パートは生の XML で書いていたが、それでは**Excel で開いても図が出ない**
# 検体になっていた。arp4 は zip の関係だけを辿るのでパースは通ってしまい、
# 実物の Excel で開いたときだけ図が消える ―― 検体は「実物によくある書かれ方」
# を写したものであるはずなので、**Excel が描けないものは検体ではない**
# （人が開いて確かめられず、`arp4 render` も空の絵しか撮れない）。
#
# 足りなかったのは 3 つで、どれも**資料の中身ではなく組み立ての作法**である
# ―― だからここ（読み込み器）が埋め、YAML には箱と線だけを書く。
#
# 1. ``[Content_Types].xml`` の申告（:data:`_CONTENT`）
# 2. シート XML の ``<drawing r:id=…>``（:func:`_inject`）
# 3. 図形そのものの必須要素 ―― 位置（``xdr:from`` / ``xdr:to``）、寸法
#    （``a:xfrm``）、形（``a:prstGeom``）、文字枠（``a:bodyPr``）
_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_DGM = "http://schemas.openxmlformats.org/drawingml/2006/diagram"

#: 1 px ぶんの EMU。**貼り付け画像の画素数はここから決める**（検体に px を
#: 書かせない ―― 箱を 1 つ動かすたびに画素数を計算し直すことになる）。
_PX = 9525

#: 既定のセル 1 個ぶん（EMU）。既定幅 8.43 文字 ＝ 64px、既定高 15pt ＝ 20px。
#: **見た目を合わせるためではなく、図が枠を持つため**の数字である。
_COL, _ROW = 64 * _PX, 20 * _PX

#: 図表枠（``xdr:graphicFrame``）の種別 → ``a:graphicData`` の uri。
_FRAME = {
    "グラフ": "http://schemas.openxmlformats.org/drawingml/2006/chart",
    "SmartArt": "http://schemas.openxmlformats.org/drawingml/2006/diagram",
    "埋め込み": "http://schemas.openxmlformats.org/presentationml/2006/ole",
}


def _drawing(shapes: list[dict[str, Any]], order: int
             ) -> tuple[str, dict[str, bytes], list[tuple[str, str, str, bool]]]:
    """図形の宣言 → 描画パート。返すのは ``(XML, 追加パート, 追加の関係)``。"""
    parts: dict[str, bytes] = {}
    links: list[tuple[str, str, str, bool]] = []
    shapes = [made for shape in shapes
              for made in (_spread(shape["生成"]) if "生成" in shape else [shape])]
    boxes = _boxes(shapes)                     # 線の通り道はここから決まる
    anchors = ""
    for index, shape in enumerate(shapes, start=1):
        anchors += (f"<xdr:twoCellAnchor>{_anchor(_rect(shape, boxes))}"
                    f"{_shape(shape, order, index, parts, links, boxes)}"
                    "<xdr:clientData/></xdr:twoCellAnchor>")
    return (f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<xdr:wsDr xmlns:xdr="{_XDR}" xmlns:a="{_A}" xmlns:r="{_R}">'
            f"{anchors}</xdr:wsDr>", parts, links)


def _spread(block: dict[str, Any]) -> list[dict[str, Any]]:
    """**図形も規模は書き写せない。** 120 個の箱を 1 つずつ書いた検体は読めない。

    実物の業務フロー（全体版）は 1 枚に 100 個以上の箱が載り、紙に収めるために
    **段組み**にしてある ―― 縦に 40 個並べたら右へ折り返す。だから並べ方
    （起点・大きさ・1 段の個数）を宣言し、番地はここが決める（``生成`` と同じ）。

    ``繋ぐ: 縦`` は**同じ段の中だけ**を結ぶ。段をまたぐ繋がりに線は引かれて
    おらず、丸印の中の記号でしか表されていない ―― それが実物の描き方なので、
    ここで勝手に結ぶと**資料に無い線**が生える（機械は繋がりを判断しない）。
    """
    from openpyxl.utils import get_column_letter

    place = block["並び"]
    rows = int(place["行数"])
    width, height = int(place["幅"]), int(place["高さ"])
    # **縦と横の空きは別に決まる。** ラベルと入力枠のように 2 組を噛み合わせて
    # 並べるときは、行の間隔を揃えたまま列の間隔だけを空ける（1 つの ``間`` に
    # まとめると、片方を空けたぶん縦にも散って**組どうしがずれる**）。
    down = int(place.get("行間", 1))
    across = int(place.get("列間", 1))
    top, left = coordinate_to_tuple(place["起点"])
    base = int(block.get("id起点", 1000))
    count = int(block["個数"])

    made: list[dict[str, Any]] = []
    for n in range(1, count + 1):
        column, row = divmod(n - 1, rows)
        first = left + column * (width + across)
        start = top + row * (height + down)
        made.append({
            "箱": [str(line).format(n=n) for line in block.get("箱") or []],
            "id": base + n,
            "名前": str(block.get("名前", "図形{n}")).format(n=n),
            "形": block.get("形", "rect"),
            "寄せ": block.get("寄せ", "中"),
            "位置": (f"{get_column_letter(first)}{start}:"
                     f"{get_column_letter(first + width - 1)}{start + height - 1}"),
        })
    if block.get("繋ぐ") == "縦":
        for n in range(1, count):
            if n % rows:                       # 段の変わり目には線が無い
                made.append({"線": {"元": base + n, "先": base + n + 1},
                             "id": base + count + n})
    return made


def _boxes(shapes: list[dict[str, Any]],
           found: dict[Any, tuple[int, int, int, int]] | None = None
           ) -> dict[Any, tuple[int, int, int, int]]:
    """id → その図形の矩形。**線が端をどこに置くかはこれで決まる**（:func:`_route`）。

    群の中まで拾う ―― 実物の業務フローはゾーンごとに group してあり、線は
    その外から中の箱へ繋がっている。
    """
    found = {} if found is None else found
    for shape in shapes:
        if "群" in shape:
            _boxes(shape["群"], found)
        elif "線" not in shape and "id" in shape:
            found[shape["id"]] = _area(shape["位置"])
    return found


def _area(where: str) -> tuple[int, int, int, int]:
    """番地 → EMU の矩形。**検体に EMU を書かせない**ための入口である。"""
    left, top, right, bottom = range_boundaries(where)
    return ((left - 1) * _COL, (top - 1) * _ROW, right * _COL, bottom * _ROW)


def _rect(shape: dict[str, Any],
          boxes: dict[Any, tuple[int, int, int, int]]) -> tuple[int, int, int, int]:
    """図形の占める矩形（EMU の ``(左, 上, 右, 下)``）。

    **群の位置は子の合併**である ―― 枠を別に書かせると、子を 1 つ足すたびに
    親の番地も直すことになり、検体が「資料の写し」から遠ざかる。
    **線の位置は端の図形が決める**（:func:`_route`）。
    """
    if "群" in shape:
        kids = [_rect(child, boxes) for child in shape["群"]]
        return (min(k[0] for k in kids), min(k[1] for k in kids),
                max(k[2] for k in kids), max(k[3] for k in kids))
    if "線" in shape:
        return _route(shape, boxes)[0]
    return _area(shape["位置"])


def _anchor(rect: tuple[int, int, int, int]) -> str:
    x0, y0, x1, y1 = rect
    return f"<xdr:from>{_at(x0, y0)}</xdr:from><xdr:to>{_at(x1, y1)}</xdr:to>"


def _at(x: int, y: int) -> str:
    """EMU → セル番号とその中の位置。

    **端はセルの境目にあるとは限らない。** 図形の接続点は辺の真ん中にあるので、
    ``colOff`` / ``rowOff`` を 0 に固定していた頃は**線が箱の角から角へ斜めに
    走っていた** ―― 番地の格子でしか端を置けなかったためである。
    """
    return (f"<xdr:col>{x // _COL}</xdr:col><xdr:colOff>{x % _COL}</xdr:colOff>"
            f"<xdr:row>{y // _ROW}</xdr:row><xdr:rowOff>{y % _ROW}</xdr:rowOff>")


def _offext(rect: tuple[int, int, int, int]) -> str:
    x0, y0, x1, y1 = rect
    return f'<a:off x="{x0}" y="{y0}"/><a:ext cx="{x1 - x0}" cy="{y1 - y0}"/>'


def _xfrm(rect: tuple[int, int, int, int], child: bool = False,
          flip: tuple[bool, bool] = (False, False)) -> str:
    """絶対座標。**群の中も同じ座標系にする**（``chOff`` / ``chExt`` を親に合わせる）。

    ``flip`` は線の向きである ―― 矩形は左上から右下へしか書けないので、右から
    左へ・下から上へ引いた線は反転で表す。**矢羽根の付く端が入れ替わる**ので、
    これを落とすと資料と逆を向いた矢印が出る。
    """
    x0, y0, x1, y1 = rect
    kid = "" if not child else (f'<a:chOff x="{x0}" y="{y0}"/>'
                                f'<a:chExt cx="{x1 - x0}" cy="{y1 - y0}"/>')
    flips = "".join(f' flip{axis}="1"' for axis, on in zip("HV", flip) if on)
    return f"<a:xfrm{flips}>{_offext(rect)}{kid}</a:xfrm>"


def _shape(shape: dict[str, Any], order: int, index: int,
           parts: dict[str, bytes], links: list[tuple[str, str, str, bool]],
           boxes: dict[Any, tuple[int, int, int, int]]) -> str:
    rect = _rect(shape, boxes)
    ident = shape.get("id", 1000 + index)
    name = shape.get("名前", f"図形{ident}")
    if "群" in shape:
        inner = "".join(_shape(child, order, index * 100 + n, parts, links, boxes)
                        for n, child in enumerate(shape["群"], start=1))
        return (f"<xdr:grpSp><xdr:nvGrpSpPr>"
                f'<xdr:cNvPr id="{ident}" name="{_esc(name)}"{_descr(shape)}/>'
                "<xdr:cNvGrpSpPr/>"
                f"</xdr:nvGrpSpPr><xdr:grpSpPr>{_xfrm(rect, child=True)}"
                f"</xdr:grpSpPr>{inner}</xdr:grpSp>")
    if "箱" in shape:
        return _sp(shape, rect, ident, name)
    if "線" in shape:
        return _cxn(shape, ident, name, boxes)
    if "画像" in shape:
        return _pic(shape["画像"], rect, ident, name, order, index, parts, links)
    return _frame(shape["図表"], rect, ident, name, order, index, parts, links)


def _sp(shape: dict[str, Any], rect: tuple[int, int, int, int],
        ident: int, name: str) -> str:
    """箱。``箱`` は**段落の配列**で、段落は文字列か「行の配列」である。

    * ``[受注登録]`` ―― 1 段落
    * ``[[承認待ち, 3 営業日以内]]`` ―― 1 段落を**行区切り**（``a:br``）で 2 行
    * ``[受注一覧, SCR002]`` ―― 段落（``a:p``）が 2 つ
    * ``[]`` ―― 文字を持たない箱（囲み枠・装飾）
    """
    empty = '<a:p><a:endParaRPr lang="ja-JP"/></a:p>'
    align = _ALIGN[shape.get("寄せ", "中")]
    body = "".join(_paragraph(para, align) for para in shape["箱"]) or empty
    geometry = shape.get("形", "roundRect")
    return ('<xdr:sp macro="" textlink=""><xdr:nvSpPr>'
            f'<xdr:cNvPr id="{ident}" name="{_esc(name)}"{_descr(shape)}/>'
            "<xdr:cNvSpPr/>"
            f"</xdr:nvSpPr><xdr:spPr>{_xfrm(rect)}"
            f'<a:prstGeom prst="{geometry}"><a:avLst/></a:prstGeom>'
            '<a:solidFill><a:srgbClr val="FFFFFF"/></a:solidFill>'
            '<a:ln w="12700"><a:solidFill><a:srgbClr val="000000"/></a:solidFill></a:ln>'
            "</xdr:spPr><xdr:txBody>"
            '<a:bodyPr vertOverflow="clip" horzOverflow="clip" wrap="square"'
            f' rtlCol="0" anchor="ctr"/><a:lstStyle/>{body}'
            "</xdr:txBody></xdr:sp>")


#: 段落の寄せ。**既定は中央**（フロー図の箱はそう書かれている）。
#:
#: 左寄せを持たせてあるのは、**字下げが画面に出るのは左寄せのときだけ**だから
#: である ―― 機能構成図・注記を 1 つのテキストボックスで書くと、親子は段落頭の
#: 全角空白でしか表されない。中央寄せのまま書くと字下げは画面のどこにも出ず、
#: 「画面に見えているものへ寄せる」という規律の掛かる先が無くなる。
_ALIGN = {"中": "ctr", "左": "l", "右": "r"}


def _descr(shape: dict[str, Any]) -> str:
    """代替テキスト（``descr``）。**書いてあるときだけ**属性を出す。

    Excel は自動で入れないので、入っていれば人が書いた文である ―― 画像だけ
    でなく図形・群にも書ける（構成図に 1 文添えるのはこの形になる）。
    """
    alt = shape.get("代替", "")
    return f' descr="{_esc(alt)}"' if alt else ""


def _paragraph(para: Any, align: str = "ctr") -> str:
    lines = para if isinstance(para, list) else [para]
    runs = '<a:br><a:rPr lang="ja-JP"/></a:br>'.join(
        '<a:r><a:rPr lang="ja-JP" sz="1100"><a:solidFill>'
        f'<a:srgbClr val="000000"/></a:solidFill></a:rPr><a:t>{_esc(line)}</a:t></a:r>'
        for line in lines)
    return f'<a:p><a:pPr algn="{align}"/>{runs}</a:p>'


#: 矢羽根。``headEnd`` は線の始点側、``tailEnd`` は終点側に付く。
_ARROWS = {"終点": ("none", "triangle"), "始点": ("triangle", "none"),
           "両方": ("triangle", "triangle"), "なし": ("none", "none")}

#: 接続点の番号。``prstGeom`` は 4 辺の真ん中に 1 つずつ持っている。
_SITES = {"上": 0, "左": 1, "下": 2, "右": 3}

#: 自己ループの通り道 ―― 箱の右から出て、天の中央へ返す。曲がり位置
#: （``adj1``）を**負にする**と縦の一節が枠の外を通り、箱を跨がずに回る。
_LOOP = ("右", "上")
_LOOP_ADJ = -100000


def _route(shape: dict[str, Any], boxes: dict[Any, tuple[int, int, int, int]]
           ) -> tuple[tuple[int, int, int, int],
                      tuple[tuple[int, int], tuple[int, int], tuple[str, str]] | None]:
    """線の通り道。返すのは ``(矩形, (始点, 終点, 接続点))``。

    **繋がっている線の位置は資料の情報ではない。** 端は図形の接続点にあり、
    箱を動かせば Excel が引き直す ―― だから検体は「どの図形とどの図形か」だけを
    書き、通り道はここが決める。番地を持つのは**目分量で置いた線**（``元`` /
    ``先`` を書かない線）だけで、それはそもそも**位置にしか情報が無い**線である。

    どの辺から出るかは位置関係で決まる（右へ流れるなら右 → 左、下へ流れるなら
    下 → 上）。これは意味の判断ではなく、Excel が同じ規則で引き直す。

    **見るのは中心の差ではなく辺と辺の隙間である。** 中心で決めていたぶん、
    真下にある**横長のゾーン枠**（中心が左寄りにある）へ引いた線が「左へ流れる」
    と判定され、箱を突っ切って斜めに走っていた ―― 2 つの箱がどちらの向きに
    離れて置かれているかは、隙間のいちばん広い向きが言っている。
    """
    line = shape["線"]
    src, dst = boxes.get(line.get("元")), boxes.get(line.get("先"))
    if src is None or dst is None:
        return _area(shape["位置"]), None
    if line["元"] == line["先"]:
        sites = _LOOP
    else:
        gaps = {("下", "上"): dst[1] - src[3], ("上", "下"): src[1] - dst[3],
                ("右", "左"): dst[0] - src[2], ("左", "右"): src[0] - dst[2]}
        sites = max(gaps, key=lambda site: gaps[site])
    start, end = _point(src, sites[0]), _point(dst, sites[1])
    return ((min(start[0], end[0]), min(start[1], end[1]),
             max(start[0], end[0]), max(start[1], end[1])), (start, end, sites))


def _point(rect: tuple[int, int, int, int], site: str) -> tuple[int, int]:
    """接続点の座標。**辺の真ん中**である（角ではない）。"""
    x0, y0, x1, y1 = rect
    return {"上": ((x0 + x1) // 2, y0), "下": ((x0 + x1) // 2, y1),
            "左": (x0, (y0 + y1) // 2), "右": (x1, (y0 + y1) // 2)}[site]


def _cxn(shape: dict[str, Any], ident: int, name: str,
         boxes: dict[Any, tuple[int, int, int, int]]) -> str:
    """線。``元`` / ``先`` を書かなければ**目分量で置いた線**（両端が繋がらない）。

    端を番地の格子に置いていた頃は、**Excel で開いても図がフローに見えなかった**
    ―― 線は箱の角から角へ斜めに走り、自己ループは箱と関係の無いところに浮いて
    いた。接続は id で書いてあるのでパースは正しく通り、テストも全部通る。
    つまり**実物で開いた人にだけ壊れて見える**検体で、それでは「この資料は何を
    写したものか」を目で確かめられない（`arp4 render` の絵も同じである）。
    """
    line = shape["線"]
    rect, joined = _route(shape, boxes)
    ends, adjust, prst, flip = "", "", "straightConnector1", (False, False)
    if joined:
        start, end, sites = joined
        flip = (start[0] > end[0], start[1] > end[1])
        ends = "".join(f'<a:{tag} id="{line[key]}" idx="{_SITES[site]}"/>'
                       for key, tag, site in (("元", "stCxn", sites[0]),
                                              ("先", "endCxn", sites[1])))
        if line["元"] == line["先"]:               # 自己ループは枠の外を回す
            prst = "bentConnector3"
            adjust = f'<a:gd name="adj1" fmla="val {_LOOP_ADJ}"/>'
    head, tail = _ARROWS[line.get("矢", "終点")]
    return (f'<xdr:cxnSp macro=""><xdr:nvCxnSpPr>'
            f'<xdr:cNvPr id="{ident}" name="{_esc(name)}"/>'
            f"<xdr:cNvCxnSpPr>{ends}</xdr:cNvCxnSpPr></xdr:nvCxnSpPr>"
            f"<xdr:spPr>{_xfrm(rect, flip=flip)}"
            f'<a:prstGeom prst="{prst}"><a:avLst>{adjust}</a:avLst></a:prstGeom>'
            '<a:ln w="19050"><a:solidFill><a:srgbClr val="000000"/></a:solidFill>'
            f'<a:headEnd type="{head}"/><a:tailEnd type="{tail}"/></a:ln>'
            "</xdr:spPr></xdr:cxnSp>")


def _pic(spec: dict[str, Any], rect: tuple[int, int, int, int], ident: int,
         name: str, order: int, index: int, parts: dict[str, bytes],
         links: list[tuple[str, str, str, bool]]) -> str:
    """貼り付け画像。**実体が要る**うえに、検体によっては**絵柄も要る。**

    実体を置くのは、``r:embed`` の先が無い画像を Excel が**壊れたファイルとして
    修復**してしまい、開いた人が図の有無を確かめられなくなるためである。

    絵柄（``絵柄: 手書き`` など）を書いた検体では、そこがもう一段効く ――
    パースは「撮り直しても読めるようにはなりません」と申告するが、**灰色の
    矩形ではその申告が正しいかを目で判定できない**（「読めない」のではなく
    「何も描いていない」）。何を描くかは :mod:`picture` にある。

    **大きさは図形の枠から決める。** 検体に px を書かせると、箱を 1 つ動かす
    たびに画素数を計算し直すことになる（位置を EMU で書かせないのと同じ）。
    """
    import picture

    x0, y0, x1, y1 = rect
    media = f"xl/media/image{order}_{index}.png"
    parts[media] = picture.draw(spec.get("絵柄", ""),
                                (x1 - x0) // _PX, (y1 - y0) // _PX)
    rid = f"rIdImg{order}_{index}"
    links.append((rid, _KINDS["image"], _relative("xl/drawings", media), False))
    alt = spec.get("代替", "")
    return (f'<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="{ident}" name="{_esc(name)}"'
            + (f' descr="{_esc(alt)}"' if alt else "")
            + '/><xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr>'
            f'</xdr:nvPicPr><xdr:blipFill><a:blip r:embed="{rid}"/>'
            "<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>"
            f"<xdr:spPr>{_xfrm(rect)}"
            '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr></xdr:pic>')


def _frame(frame: dict[str, Any], rect: tuple[int, int, int, int], ident: int,
           name: str, order: int, index: int, parts: dict[str, bytes],
           links: list[tuple[str, str, str, bool]]) -> str:
    """図表枠（グラフ・SmartArt・埋め込みオブジェクト）。"""
    kind = _FRAME[frame["種類"]]
    inner = ""
    if frame["種類"] == "グラフ" and frame.get("関係"):
        inner = ('<c:chart xmlns:c="http://schemas.openxmlformats.org/'
                 f'drawingml/2006/chart" xmlns:r="{_R}" r:id="{frame["関係"]}"/>')
    elif frame["種類"] == "SmartArt":
        inner = _smartart(frame, order, index, parts, links)
    alt = frame.get("代替", "")
    return (f'<xdr:graphicFrame macro=""><xdr:nvGraphicFramePr>'
            f'<xdr:cNvPr id="{ident}" name="{_esc(name)}"'
            + (f' descr="{_esc(alt)}"' if alt else "")
            + "/><xdr:cNvGraphicFramePr/></xdr:nvGraphicFramePr>"
            f"<xdr:xfrm>{_offext(rect)}</xdr:xfrm>"
            f'<a:graphic><a:graphicData uri="{kind}">{inner}</a:graphicData>'
            "</a:graphic></xdr:graphicFrame>")


# ── SmartArt（4 パートが揃って初めて図になる） ──────────────────
#
# SmartArt は箱の文字（``data``）だけでは描かれない ―― 並べ方（``layout``）・
# 体裁（``quickStyle``）・配色（``colors``）の 3 つが要り、``dgm:relIds`` で
# 4 つとも指していなければ Excel は**ブックごと開けなくなる**。
#
# 「並べ方は数百行の図表レイアウト定義なので手で書けない」と一度は諦めたが、
# **それは Microsoft が配っている 130 種類のレイアウトの話**だった。縦に箱を
# 並べるだけなら 20 行で足りる（実測 ―― Excel で開いて図が出ることを確かめた）。
# 検体に要るのは「箱の文字が読めること」であって、体裁の再現ではない。
_LAYOUT_ID = "urn:arp4/officeart/layout/list"
_STYLE_ID = "urn:arp4/officeart/quickstyle/simple"
_COLORS_ID = "urn:arp4/officeart/colors/accent1"

#: 縦に箱を並べるだけのレイアウト定義。``forEach`` が子の点を 1 つずつ箱にする。
_LAYOUT = f"""{{head}}<dgm:layoutDef xmlns:dgm="{{dgm}}" xmlns:a="{{a}}"
 xmlns:r="{_R}" uniqueId="{_LAYOUT_ID}">
<dgm:title val=""/><dgm:desc val=""/>
<dgm:catLst><dgm:cat type="list" pri="400"/></dgm:catLst>
<dgm:sampData><dgm:dataModel><dgm:ptLst/></dgm:dataModel></dgm:sampData>
<dgm:styleData><dgm:dataModel><dgm:ptLst/></dgm:dataModel></dgm:styleData>
<dgm:clrData><dgm:dataModel><dgm:ptLst/></dgm:dataModel></dgm:clrData>
<dgm:layoutNode name="diagram">
  <dgm:varLst><dgm:chMax val="-1"/><dgm:dir val="norm"/><dgm:animLvl val="lvl"/></dgm:varLst>
  <dgm:alg type="lin"><dgm:param type="linDir" val="fromT"/></dgm:alg>
  <dgm:shape type="none" blipPhldr="0"/><dgm:presOf/>
  <dgm:constrLst>
    <dgm:constr type="w" for="ch" ptType="node" refType="w"/>
    <dgm:constr type="h" for="ch" ptType="node" refType="h" fact="0.2"/>
    <dgm:constr type="sibSp" refType="h" refFor="ch" refPtType="node" fact="0.2"/>
  </dgm:constrLst><dgm:ruleLst/>
  <dgm:forEach name="nodes" axis="ch" ptType="node">
    <dgm:layoutNode name="node" styleLbl="node1">
      <dgm:alg type="tx"/><dgm:shape type="roundRect" blipPhldr="0"/>
      <dgm:presOf axis="desOrSelf" ptType="node"/>
      <dgm:constrLst>
        <dgm:constr type="primFontSz" val="65"/>
        <dgm:constr type="lMarg" refType="primFontSz" fact="0.3"/>
        <dgm:constr type="rMarg" refType="primFontSz" fact="0.3"/>
      </dgm:constrLst>
      <dgm:ruleLst><dgm:rule type="primFontSz" val="5" fact="NaN" max="NaN"/></dgm:ruleLst>
    </dgm:layoutNode>
  </dgm:forEach>
</dgm:layoutNode></dgm:layoutDef>"""

_SCENE = ('<dgm:scene3d><a:camera prst="orthographicFront"/>'
          '<a:lightRig rig="threePt" dir="t"/></dgm:scene3d>')
_STYLE_LABEL = (
    '<dgm:styleLbl name="{name}">' + _SCENE + "<dgm:sp3d/><dgm:txPr/><dgm:style>"
    '<a:lnRef idx="2"><a:scrgbClr r="0" g="0" b="0"/></a:lnRef>'
    '<a:fillRef idx="1"><a:scrgbClr r="0" g="0" b="0"/></a:fillRef>'
    '<a:effectRef idx="0"><a:scrgbClr r="0" g="0" b="0"/></a:effectRef>'
    '<a:fontRef idx="minor"><a:schemeClr val="lt1"/></a:fontRef>'
    "</dgm:style></dgm:styleLbl>")
_COLOR_LABEL = (
    '<dgm:styleLbl name="{name}">'
    '<dgm:fillClrLst meth="repeat"><a:schemeClr val="accent1"/></dgm:fillClrLst>'
    '<dgm:linClrLst meth="repeat"><a:schemeClr val="lt1"/></dgm:linClrLst>'
    "<dgm:effectClrLst/><dgm:txLinClrLst/>"
    '<dgm:txFillClrLst meth="repeat"><a:schemeClr val="lt1"/></dgm:txFillClrLst>'
    "<dgm:txEffectClrLst/></dgm:styleLbl>")
#: 体裁と配色は箱に掛かる 3 つのラベルだけ定義する（Excel は残りを既定で埋める）。
_LABELS = ("node0", "node1", "alignNode1")


def _smartart(frame: dict[str, Any], order: int, index: int,
              parts: dict[str, bytes], links: list[tuple[str, str, str, bool]]) -> str:
    """箱の一覧 → SmartArt 一式。返すのは ``graphicData`` の中身（``dgm:relIds``）。"""
    head = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    fixed = {"head": head, "dgm": _NS_DGM, "a": _A}
    parts["xl/diagrams/layout1.xml"] = _LAYOUT.format(**fixed).encode("utf-8")
    parts["xl/diagrams/quickStyle1.xml"] = (
        f'{head}<dgm:styleDef xmlns:dgm="{_NS_DGM}" xmlns:a="{_A}"'
        f' uniqueId="{_STYLE_ID}"><dgm:title val=""/><dgm:desc val=""/>'
        '<dgm:catLst><dgm:cat type="simple" pri="10100"/></dgm:catLst>' + _SCENE
        + "".join(_STYLE_LABEL.format(name=n) for n in _LABELS)
        + "</dgm:styleDef>").encode("utf-8")
    parts["xl/diagrams/colors1.xml"] = (
        f'{head}<dgm:colorsDef xmlns:dgm="{_NS_DGM}" xmlns:a="{_A}"'
        f' uniqueId="{_COLORS_ID}"><dgm:title val=""/><dgm:desc val=""/>'
        '<dgm:catLst><dgm:cat type="accent1" pri="11100"/></dgm:catLst>'
        + "".join(_COLOR_LABEL.format(name=n) for n in _LABELS)
        + "</dgm:colorsDef>").encode("utf-8")

    data = f"xl/diagrams/data{order}{index}.xml"
    parts[data] = _diagram_data(frame["箱"], head).encode("utf-8")
    ids = {"dm": f"rIdDm{order}{index}", "lo": f"rIdLo{order}{index}",
           "qs": f"rIdQs{order}{index}", "cs": f"rIdCs{order}{index}"}
    for key, kind, target in (
            ("dm", "diagramData", _relative("xl/drawings", data)),
            ("lo", "diagramLayout", "../diagrams/layout1.xml"),
            ("qs", "diagramQuickStyle", "../diagrams/quickStyle1.xml"),
            ("cs", "diagramColors", "../diagrams/colors1.xml")):
        links.append((ids[key], _KINDS[kind], target, False))
    return (f'<dgm:relIds xmlns:dgm="{_NS_DGM}" xmlns:r="{_R}"'
            + "".join(f' r:{key}="{value}"' for key, value in ids.items()) + "/>")


def _diagram_data(boxes: list[Any], head: str) -> str:
    """箱の一覧 → データモデル。

    **箱だけでは成り立たない。** どの箱がどの親の何番目かは ``dgm:cxn`` が持ち、
    箱と箱の間には繋ぎの点（``parTrans`` / ``sibTrans``）が要る ―― 見た目用の
    複製（``pres``）も Excel は必ず書く。arp4 が数えるのは**箱だけ**で、
    ここで一緒に並ぶ 3 種類は数えてはいけない（同じ語が 2 回出る）。
    """
    points = ['<dgm:pt modelId="1" type="doc">'
              f'<dgm:prSet loTypeId="{_LAYOUT_ID}" loCatId="list"'
              f' qsTypeId="{_STYLE_ID}" qsCatId="simple"'
              f' csTypeId="{_COLORS_ID}" csCatId="accent1"/>'
              "<dgm:spPr/><dgm:t><a:bodyPr/><a:lstStyle/>"
              '<a:p><a:endParaRPr lang="ja-JP"/></a:p></dgm:t></dgm:pt>']
    connections = []
    for order, box in enumerate(boxes):
        text = box["文字"] if isinstance(box, dict) else box
        kind = f' type="{box["種別"]}"' if isinstance(box, dict) and "種別" in box else ""
        node, par, sib, cxn = 10 + order, 100 + order, 200 + order, 300 + order
        points.append(f'<dgm:pt modelId="{node}"{kind}><dgm:prSet/><dgm:spPr/>'
                      f"<dgm:t><a:bodyPr/><a:lstStyle/>{_diagram_text(text)}"
                      "</dgm:t></dgm:pt>")
        for identity, role in ((par, "parTrans"), (sib, "sibTrans")):
            points.append(
                f'<dgm:pt modelId="{identity}" type="{role}" cxnId="{cxn}">'
                "<dgm:prSet/><dgm:spPr/><dgm:t><a:bodyPr/><a:lstStyle/>"
                '<a:p><a:endParaRPr lang="ja-JP"/></a:p></dgm:t></dgm:pt>')
        # **見た目用の複製。** Excel は箱と同じ文字をもう 1 つ書く
        points.append(f'<dgm:pt modelId="{400 + order}" type="pres">'
                      "<dgm:prSet/><dgm:spPr/><dgm:t><a:bodyPr/><a:lstStyle/>"
                      f"{_diagram_text(text)}</dgm:t></dgm:pt>")
        connections.append(
            f'<dgm:cxn modelId="{cxn}" srcId="1" destId="{node}" srcOrd="{order}"'
            f' destOrd="0" parTransId="{par}" sibTransId="{sib}"/>')
    return (f'{head}<dgm:dataModel xmlns:dgm="{_NS_DGM}" xmlns:a="{_A}">'
            f'<dgm:ptLst>{"".join(points)}</dgm:ptLst>'
            f'<dgm:cxnLst>{"".join(connections)}</dgm:cxnLst>'
            "<dgm:bg/><dgm:whole/></dgm:dataModel>")


def _diagram_text(text: Any) -> str:
    """箱の中の文字。**書き方は :func:`_sp` と同じ**（段落の配列）。"""
    return "".join(_paragraph(para) for para in text) or (
        '<a:p><a:endParaRPr lang="ja-JP"/></a:p>')


def _esc(text: str) -> str:
    return (str(text).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def _sheet_parts(body: dict[str, bytes]) -> dict[str, str]:
    """シート名 → パッケージ内のシート XML。**並び順に頼らない。**

    ``xl/worksheets/sheet1.xml`` が 1 枚目とは限らない（グラフシートは
    ``xl/chartsheets/`` に行く）ので、``workbook.xml`` の関係から引く
    ―― :func:`arp4.parse._sheet_parts` と同じ道である。
    """
    import xml.etree.ElementTree as ET

    main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    pkg = "http://schemas.openxmlformats.org/package/2006/relationships"

    rels = ET.fromstring(body["xl/_rels/workbook.xml.rels"])
    targets = {r.get("Id"): r.get("Target") for r in
               rels.findall(f"{{{pkg}}}Relationship")}
    book = ET.fromstring(body["xl/workbook.xml"])
    found: dict[str, str] = {}
    for sheet in book.findall(f"{{{main}}}sheets/{{{main}}}sheet"):
        target = targets.get(sheet.get(f"{{{rel}}}id")) or ""
        # **``Target`` はパッケージ絶対のことがある**（openpyxl は
        # ``/xl/worksheets/sheet1.xml`` と書く）。``xl/`` を足すと二重になる。
        found[sheet.get("name") or ""] = (
            target.lstrip("/") if target.startswith("/") else f"xl/{target}")
    return found


def _dir(part: str) -> str:
    return part.rsplit("/", 1)[0] if "/" in part else ""


def _rels_of(part: str) -> str:
    return f"{_dir(part)}/_rels/{Path(part).name}.rels"


def _relative(base: str, part: str) -> str:
    """``xl/worksheets`` から ``xl/drawings/drawing1.xml`` へ → ``../drawings/…``。"""
    here = [p for p in base.split("/") if p]
    there = part.split("/")
    while here and there[:1] == here[:1]:
        here.pop(0)
        there.pop(0)
    return "../" * len(here) + "/".join(there)


def _with_links(body: bytes | None, added: list[tuple[str, str, str, bool]]) -> bytes:
    rows = "".join(
        f'<Relationship Id="{i}" Type="{t}" Target="{g}"'
        + (' TargetMode="External"' if outside else "") + "/>"
        for i, t, g, outside in added)
    if body is None:
        return ('<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns='
                '"http://schemas.openxmlformats.org/package/2006/relationships">'
                f"{rows}</Relationships>").encode("utf-8")
    return body.decode("utf-8").replace(
        "</Relationships>", f"{rows}</Relationships>").encode("utf-8")


if __name__ == "__main__":
    import sys

    target = Path(sys.argv[1] if len(sys.argv) > 1 else "dataset").resolve()
    for made in build(target):
        print(made)
