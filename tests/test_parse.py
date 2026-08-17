"""① パース ―― **意味を判断しない**ことと、**編集済みを守る**ことを確かめる。"""

from __future__ import annotations

import subprocess
from pathlib import Path

import pytest
from openpyxl import Workbook

from arp4 import mdio, parse
from arp4.paths import Paths, Round
from conftest import sources_dir, write


def _book(path: Path, rows: list[list], start: int = 1,
          title: str = "受注テーブル") -> Path:
    book = Workbook()
    sheet = book.active
    sheet.title = title
    for r, row in enumerate(rows, start=start):
        for c, value in enumerate(row, start=1):
            if value is not None:
                sheet.cell(row=r, column=c, value=value)
    book.save(path)
    return path


def _parse(round_: Round, source: Path, base: Path) -> list[mdio.Doc]:
    targets, findings = parse.plan(round_, [source], base)
    assert not [f for f in findings if f.level == "error"]
    return [t.doc for t in targets]


def test_表とテキストは別の塊になる(project: Paths, round_: Round) -> None:
    """区切りは**提示上の都合**だが、番地は必ず併記される。"""
    path = _book(sources_dir(project) / "a.xlsx",
                 [["表題"], [None], [None], ["論理名", "物理名"], ["受注番号", "ORDER_NO"]])
    doc = _parse(round_, path, sources_dir(project))[0]

    kinds = [(c.anchor, c.at) for c in doc.chunks]
    assert ("s1-x1", "A1") in kinds
    assert any(a.startswith("s1-t") and ":" in at for a, at in kinds)
    assert all(chunk.at for chunk in doc.chunks)


def test_幅1の縦結合は全行へ展開する(project: Paths, round_: Round) -> None:
    """**判断ではなく忠実性の回復**（画面上は全行に値が見えている）。"""
    path = sources_dir(project) / "a.xlsx"
    book = Workbook()
    sheet = book.active
    sheet["A1"] = "区分"
    sheet["B1"] = "項目"
    sheet["A2"] = "受注"
    sheet["B2"] = "受注番号"
    sheet["B3"] = "受注日"
    sheet.merge_cells("A2:A3")
    book.save(path)

    doc = _parse(round_, path, sources_dir(project))[0]
    table = [c for c in doc.chunks if c.rows][0]
    assert table.rows[-1][0] == "受注"


def test_横結合は広げない(project: Paths, round_: Round) -> None:
    path = sources_dir(project) / "a.xlsx"
    book = Workbook()
    sheet = book.active
    sheet["A1"] = "見出し"
    sheet["A2"] = "x"
    sheet["B2"] = "y"
    sheet.merge_cells("A1:B1")
    book.save(path)

    doc = _parse(round_, path, sources_dir(project))[0]
    table = [c for c in doc.chunks if c.rows][0]
    assert table.rows[0][1] == ""


def test_非表示シートは読まない(project: Paths, round_: Round) -> None:
    path = sources_dir(project) / "a.xlsx"
    book = Workbook()
    book.active.title = "見える"
    book.active["A1"] = "x"
    hidden = book.create_sheet("隠し")
    hidden["A1"] = "y"
    hidden.sheet_state = "hidden"
    book.save(path)

    docs = _parse(round_, path, sources_dir(project))
    assert len(docs) == 1


def test_シート1枚がファイル1本になる(project: Paths, round_: Round) -> None:
    path = sources_dir(project) / "資料" / "基本設計書.xlsx"
    path.parent.mkdir(parents=True)
    book = Workbook()
    book.active.title = "受注テーブル"
    book.active["A1"] = "x"
    second = book.create_sheet("画面項目")
    second["A1"] = "y"
    book.save(path)

    targets, _ = parse.plan(round_, [path], sources_dir(project))
    names = sorted(t.path.relative_to(round_.parsed).as_posix() for t in targets)
    assert names == ["資料/基本設計書.xlsx/受注テーブル.md",
                     "資料/基本設計書.xlsx/画面項目.md"]


def test_フォルダ構造で一意になる(project: Paths, round_: Round) -> None:
    """3 は表示名のハッシュで ID を作っていたので、別フォルダの同名が衝突した。"""
    for folder in ("A", "B"):
        target = sources_dir(project) / folder / "基本設計書.xlsx"
        target.parent.mkdir(parents=True, exist_ok=True)
        _book(target, [["x"]])

    targets, _ = parse.plan(round_, [sources_dir(project)], sources_dir(project))
    paths = {t.path for t in targets}
    assert len(paths) == 2


def test_読めない形式はP001(project: Paths, round_: Round) -> None:
    """**読めない形式は数えて申告する**（黙って飛ばさない）。

    検体を `.vsdx`（Visio）にしてあるのは、**業務フローが Visio で届く**のが
    実案件で普通だからである ―― `.docx` や `.csv` のように「いつか読む側」に
    回るものを使うと、読めるようになった日にこのテストが**読めることを失敗と
    して報告する。**
    """
    write(sources_dir(project) / "業務フロー.vsdx", "x")
    targets, findings = parse.plan(round_, [sources_dir(project)], sources_dir(project))

    assert not targets
    assert [f.code for f in findings] == ["P001"]
    # **何が読めるかまで言う。**「読めません」で終わる申告は拾い直されない。
    assert ".xlsx" in findings[0].message and ".docx" in findings[0].message


def test_書けない1本があっても残りを書く(project: Paths, round_: Round) -> None:
    """**読めた資料が書けずに消えるのは、読めなかったものが消えるのと同じ。**

    :func:`parse.plan` は「1 冊が壊れていて 29 冊が落ちる」を避けるために例外を
    握っているのに、書き出しのほうは 1 本目で止まっていた ―― 途中まで書いた
    パース結果だけが残り、**どこまで書けたのかは誰にも分からない**。

    ここでは書き出し先に同名のフォルダを置いて塞ぐ（実案件で効くのは
    Windows の 260 文字と権限だが、壊し方は同じ ``OSError`` である）。
    """
    for name in ("a.xlsx", "b.xlsx", "c.xlsx"):
        _book(sources_dir(project) / name, [["論理名", "物理名"], ["受注番号", "ORDER_NO"]])
    targets, _ = parse.plan(round_, [sources_dir(project)], sources_dir(project))
    assert len(targets) == 3

    blocked = targets[1].path
    blocked.mkdir(parents=True)                  # ここだけ書けなくする

    written, findings = parse.write(targets)
    assert len(written) == 2                     # 残りは書けている
    assert [f.code for f in findings] == ["P011"]
    assert findings[0].target == blocked.name
    assert "書き出しに失敗しました" in findings[0].message


def test_長すぎるパスは長さを言う() -> None:
    """**Windows は 260 文字で頭打ちになる。** そのとき出るのは
    ``FileNotFoundError``（「無い」）なので、権限やディスクを疑って時間を溶かす
    ―― 資料の側は 1 つも間違っていない。"""
    assert "260 文字" in parse._write_advice(Path("C:/" + "あ" * 300 + ".md"))
    assert "260 文字" not in parse._write_advice(Path("C:/短い.md"))


def test_機械が置いたものを資料として数えない(project: Paths) -> None:
    """**客先からもらったフォルダには `.git` が付いてくる。**

    1 つずつ「読めない形式です」（`P001`）と言うと、**本当に読めなかった 1 冊が
    その山に埋もれる**。判定は元の置き場からの**相対**で行う ―― 絶対パスで
    見ると、プロジェクトが `.venv` の下にあるだけで資料が 1 冊も見えなくなる。
    """
    write(sources_dir(project) / ".git" / "HEAD", "ref: refs/heads/main")
    write(sources_dir(project) / "node_modules" / "left-pad" / "index.js", "x")
    write(sources_dir(project) / "src" / "__pycache__" / "a.cpython-311.pyc", "x")
    write(sources_dir(project) / "資料" / "~$受注一覧.xlsx", "x")
    write(sources_dir(project) / "資料" / "受注一覧.txt", "x")

    assert [p.name for p in parse.expand([sources_dir(project)])] == ["受注一覧.txt"]


def test_隠しフォルダの下にある置き場でも資料は見える(tmp_path: Path) -> None:
    """判定は相対で行う ―― **置き場そのものが `.tmp` の下にあることはある。**"""
    sources = tmp_path / ".cache" / "sources"
    write(sources / "資料" / "a.txt", "x")

    assert [p.name for p in parse.expand([sources])] == ["a.txt"]


def test_自分の親を指すリンクで同じ資料を何度も拾わない(
        project: Paths) -> None:
    """**`rglob` は接合点を素直に辿る。**

    自分の親を指すリンクが 1 つあると、同じ `a.txt` が何本も出てくる ――
    実測では 64 本になった（止まったのは Windows のパス長 260 文字で切れた
    からで、リンクを見分けたからではない）。同じ資料が何度もパースされ、
    書き出し先が重なって `P002` の山になる。

    **リンクそのものは飛ばさない** ―― 別のところを指すリンクの先には資料が
    ある。落とすのは「もう歩いた実体」だけなので、消える資料は無い。
    """
    write(sources_dir(project) / "資料" / "a.txt", "x")
    link = sources_dir(project) / "資料" / "自分自身"
    try:
        link.symlink_to(sources_dir(project), target_is_directory=True)
    except (OSError, NotImplementedError):
        # Windows のシンボリックリンクは管理者か開発者モードが要る。接合点
        # （ジャンクション）は要らないので、そちらで作る。
        made = subprocess.run(["cmd", "/c", "mklink", "/J", str(link),
                               str(sources_dir(project))], capture_output=True)
        if made.returncode != 0 or not link.is_dir():
            pytest.skip("リンクを作れない環境です")

    assert [p.name for p in parse.expand([sources_dir(project)])] == ["a.txt"]


def test_別のところを指すリンクの先は歩く(project: Paths,
                                          tmp_path: Path) -> None:
    """**落とすのは「もう歩いた実体」だけである。** 資料置き場を別の
    ドライブへ逃がして接合点で繋ぐのは実案件で普通にやる。"""
    outside = tmp_path / "外"
    write(outside / "受注一覧.txt", "x")
    link = sources_dir(project) / "資料"
    try:
        link.symlink_to(outside, target_is_directory=True)
    except (OSError, NotImplementedError):
        made = subprocess.run(["cmd", "/c", "mklink", "/J", str(link),
                               str(outside)], capture_output=True)
        if made.returncode != 0 or not link.is_dir():
            pytest.skip("リンクを作れない環境です")

    assert [p.name for p in parse.expand([sources_dir(project)])] == ["受注一覧.txt"]


def test_先の消えたリンクは黙って落とさない(project: Paths,
                                            round_: Round,
                                            tmp_path: Path) -> None:
    """**一覧には出るのに開けない。** ここは長く黙って落としていた。

    `is_file()` も `is_dir()` も偽になるのは、リンクの先が消えているとき・
    クラウド同期が実体をまだ落としてきていないときである ―― どちらも
    **0 冊で正常終了する**ので、資料を渡したつもりの人には「中身が空だった」
    ようにしか見えない。
    """
    outside = tmp_path / "外"
    outside.mkdir()
    link = sources_dir(project) / "資料置き場"
    try:
        link.symlink_to(outside, target_is_directory=True)
    except (OSError, NotImplementedError):
        made = subprocess.run(["cmd", "/c", "mklink", "/J", str(link),
                               str(outside)], capture_output=True)
        if made.returncode != 0 or not link.is_dir():
            pytest.skip("リンクを作れない環境です")
    outside.rmdir()                                # 先だけ消える（リンクは残る）

    _, findings = parse.plan(round_, [sources_dir(project)], sources_dir(project))
    said = [f for f in findings if f.code == "P007"]

    assert [f.target for f in said] == ["資料置き場"]
    assert "先が見つかりません" in said[0].message


def test_そのパスに無いことを0冊で正常終了にしない(project: Paths,
                                                  round_: Round) -> None:
    """**打ち間違いも同じ形で消える。** 読めた資料 0 冊は成功に見える。"""
    _, findings = parse.plan(round_, [sources_dir(project) / "そんざいしない.xlsx"],
                             sources_dir(project))
    said = [f for f in findings if f.code == "P007"]

    assert [f.target for f in said] == ["そんざいしない.xlsx"]
    assert "そのパスにはありません" in said[0].message
    # **リンク切れとは次にやることが違う**ので、同じ文句にしない
    assert "同期" not in said[0].message


def test_一覧できないフォルダは中身ごと消えたと言う(
        project: Paths, round_: Round,
        monkeypatch: pytest.MonkeyPatch) -> None:
    """**ここには長く「既知の穴」とだけ書いてあった。**

    権限の付け替え漏れ・まだ落ちてきていない同期フォルダ・使用中のロックで
    フォルダが一覧できないと、**その下の資料は 1 冊残らず消える** ―― どの
    フォルダが消えたのかを 1 行も言っていなかった。
    """
    write(sources_dir(project) / "資料" / "A" / "受注.xlsx", "x")
    blocked = sources_dir(project) / "資料" / "A"
    original = Path.iterdir

    def refuse(self):
        if self == blocked:
            raise PermissionError(13, "Permission denied")
        return original(self)

    monkeypatch.setattr(Path, "iterdir", refuse)
    _, findings = parse.plan(round_, [sources_dir(project)], sources_dir(project))
    said = [f for f in findings if f.code == "P007"]

    assert [f.target for f in said] == ["A"]
    assert "一覧できません" in said[0].message
    assert "1 冊も読んでいません" in said[0].message


def test_権限で開けない資料を壊れていることにしない(
        project: Paths, round_: Round,
        monkeypatch: pytest.MonkeyPatch) -> None:
    """**資料の側は 1 つも間違っていない。**

    `[Errno 13] Permission denied` だけを出していたので、**資料が壊れている
    のだと思われる** ―― 開いたままの Excel・客先から付いてきた権限・まだ
    落ちてきていない同期ファイルがこの形になる。
    """
    path = _book(sources_dir(project) / "受注.xlsx", [["項目", "値"]])
    original = Path.open

    def refuse(self, *args, **kwargs):
        if self == path:
            raise PermissionError(13, "Permission denied")
        return original(self, *args, **kwargs)

    monkeypatch.setattr(Path, "open", refuse)
    monkeypatch.setattr("openpyxl.load_workbook", lambda *a, **k: (_ for _ in ()).throw(
        PermissionError(13, "Permission denied")))

    _, findings = parse.plan(round_, [path], sources_dir(project))
    said = [f for f in findings if f.code == "P010"]

    assert len(said) == 1
    assert "開く権限がありません" in said[0].message
    assert "資料そのものは壊れていません" in said[0].message


def test_gitに聞くのは1度きり(project: Paths, round_: Round,
                              monkeypatch: pytest.MonkeyPatch) -> None:
    """**性能の問題が網羅性の問題になる。**

    1 ファイルにつき `git status` を 1 回起動していた。1 件 23ms はどうという
    ことのない数だが、**シート 1 枚がファイル 1 本**なので 30 冊 201 シートの
    再実行では 201 プロセス ―― 実測 5.0 秒で、パース本体（1.3 秒）より長い。
    ここが遅くなると「遅いから通さない」が始まって資料が落ちる。
    """
    book = sources_dir(project) / "a.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    for i in range(12):
        workbook.create_sheet(f"シート{i}")["A1"] = "x"
    workbook.save(book)

    # **効くのは再実行のときである**（1 度目は書き出し先がまだ無い）。
    parse.write(parse.plan(round_, [book], sources_dir(project))[0])

    calls = []
    original = parse.subprocess.run
    monkeypatch.setattr(parse.subprocess, "run",
                        lambda *a, **k: (calls.append(a), original(*a, **k))[1])
    targets, _ = parse.plan(round_, [book], sources_dir(project))

    assert len(targets) == 12 and all(t.exists for t in targets)
    assert len(calls) <= 2                       # toplevel と status で 2 回まで


def test_zipの外を指す関係でもディスクを読まない() -> None:
    """**細工した `.rels` でパッケージの外を読ませられないこと。**

    `..` がパッケージの外へ出ても、返るのは zip の中のキーである ―― arp4 は
    zip を**展開せず**、この名前は `ZipFile.read` に渡すだけなので、そもそも
    実体のパスにならない。**確かめたうえで防御は足していない**（足すと
    「zip の外を指す関係」をエラー扱いすることになり、ブックを移動した跡で
    普通にそうなっている資料が読めなくなる）。
    """
    assert parse._resolve("xl/worksheets", "../../../../etc/passwd") == "etc/passwd"
    assert parse._resolve("xl/worksheets", "/xl/media/1.png") == "xl/media/1.png"
    assert parse._resolve("xl/worksheets", "../drawings/d1.xml") \
        == "xl/drawings/d1.xml"


def test_gitが使えないときは編集ありに倒す(tmp_path: Path) -> None:
    """**分からないまま黙って上書きするより、確認が 1 回多いほうがましである。**

    `None` は「git が使えない」で、**空集合とは意味が違う**（空集合は
    「聞けたが 1 件も編集されていない」）。
    """
    outside = tmp_path / "リポジトリの外"
    outside.mkdir()
    assert parse._edited(outside / "x.md", None) is True
    assert parse._edited(outside / "x.md", set()) is False


def test_編集済みは確認が要る(project: Paths, round_: Round,
                             monkeypatch: pytest.MonkeyPatch) -> None:
    """**未編集のものは黙って上書きしてよい。** 全部聞くと確認が読み飛ばされる。"""
    source = _book(sources_dir(project) / "a.xlsx", [["論理名", "物理名"], ["x", "Y"]])
    targets, _ = parse.plan(round_, [source], sources_dir(project))
    parse.write(targets)

    monkeypatch.setattr(parse, "_edited", lambda path, dirty: False)
    again, _ = parse.plan(round_, [source], sources_dir(project))
    assert again[0].exists and not again[0].needs_confirm

    monkeypatch.setattr(parse, "_edited", lambda path, dirty: True)
    edited, _ = parse.plan(round_, [source], sources_dir(project))
    assert edited[0].needs_confirm


def test_コードはASTで骨格だけ取る(project: Paths, round_: Round) -> None:
    """**意図の層は出さない**（整理層が原本を直接読む）。"""
    source = write(sources_dir(project) / "order" / "service.py", '''
class OrderService:
    """受注サービス。"""

    def register(self, order: "Order") -> "OrderNo":
        if order.duplicated:
            raise DuplicateOrder("重複")
        return OrderNo("x")


def helper(value: int) -> str:
    return str(value)
''')
    targets, _ = parse.plan(round_, [source], sources_dir(project))
    doc = targets[0].doc

    assert targets[0].path.relative_to(round_.parsed).as_posix() == \
        "order/service.py.md"
    klass = doc.chunks[0]
    assert klass.heading == "クラス: OrderService"
    assert klass.at.startswith("order/service.py#L")
    # 1 行目は**クラスそのもの**（注釈と継承がここに出る）
    assert klass.rows[1][:4] == ["OrderService", "クラス", "", "class OrderService"]
    assert klass.rows[2][:2] == ["register", "メソッド"]
    assert klass.rows[2][5] == "DuplicateOrder"
    assert doc.chunks[1].heading == "モジュール関数"


def test_モジュール直下の定数を出す(project: Paths, round_: Round) -> None:
    """**「決まっていること」がここにしか無い。**

    クラスと関数と取り込みしか見ていなかったので、`__version__` も拡張子も
    予約名も 1 行も出ていなかった ―― 自身のソースでは `__init__.py` が
    **丸ごと空になって黙って消えた**（`parsed/` が 23 本中 22 本）。
    """
    source = write(sources_dir(project) / "arp4" / "__init__.py", f'''
"""ai-ready-pipeline 4。"""

__version__ = "4.0.0.dev0"
EXT: str = ".yml"
_MAGIC = "{"x" * 300}"
''')
    targets, findings = parse.plan(round_, [source], sources_dir(project))
    chunk = targets[0].doc.chunks[0]

    assert not findings                      # 空ではないので P009 は出ない
    assert (chunk.anchor, chunk.heading) == ("v1", "定数")
    assert chunk.at.startswith("arp4/__init__.py#L")
    assert [row[0] for row in chunk.rows[1:]] == ["__version__", "EXT", "_MAGIC"]
    assert chunk.rows[1][:4] == ["__version__", "定数", "",
                                 "__version__ = '4.0.0.dev0'"]
    # 私物（先頭 `_`）も出す ―― どれが仕様かを決めるのは整理層である。
    assert chunk.rows[3][3].endswith("…（以下略・原本を見ること）")


def test_1行も出なかったら言う(project: Paths, round_: Round) -> None:
    """**数えて申告するのはブックだけではない。**

    Excel は `P009` で言っていたのに、コード・DDL・Markdown は黙って何も返さず、
    `sources/` 23 本に対して `parsed/` が 22 本でも**差の 1 本がどれかを言うものが
    無かった**。読めなかったものほど静かに消えるのが、いちばん避けたい壊れ方である。

    **`P001`（読めない形式）とは違う。** あちらは拡張子で対象外と分かるので既に
    申告があった ―― 黙っていたのは**読める形式なのに何も出なかった**ほうである。
    """
    source = write(sources_dir(project) / "arp4" / "説明.py", '"""覚え書きだけ。"""\n')
    targets, findings = parse.plan(round_, [source], sources_dir(project))

    assert targets == []
    assert [f.code for f in findings] == ["P009"]
    assert findings[0].level == "warn"
    assert "1 行も残りません" in findings[0].message


def test_文字コードが当たらないソースは文字コードの話をする(
        project: Paths, round_: Round) -> None:
    """**UTF-8 でも cp932 でも読めないソースがある**（EUC-JP・UTF-16）。

    構文エラーと同じ入口（`P010`）から出るが、**次にやることが違う** ――
    あちらは原本を直接読めばよいが、こちらは開くところから当たらない。
    どちらも「zip として開けません」ではない。
    """
    source = sources_dir(project) / "order" / "旧帳票.py"
    source.parent.mkdir(parents=True, exist_ok=True)
    source.write_bytes("# 受注一覧の出力\n".encode("euc-jp"))

    _, findings = parse.plan(round_, [source], sources_dir(project))
    said = [f for f in findings if f.code == "P010"]

    assert len(said) == 1
    assert "UTF-8 でも cp932 でも読めません" in said[0].message
    assert "zip" not in said[0].message


def test_1本も出なかったブックは黙って消さない(project: Paths,
                                              round_: Round) -> None:
    """**`sources/` に置いた 1 冊が `parsed/` から丸ごと消える。**

    値も図もコメントも無いシートを出さない決まりは変えない（でないと
    **作業用の白紙がぜんぶパース結果になる**）が、それをブック 1 冊に掛けた
    結果は黙れない ―― 置いたブックが 1 行も残らないのは、置かなかったのと
    同じに見える。
    """
    book = Workbook()
    book.remove(book.active)
    book.create_sheet("作業用")
    book.create_sheet("メモ")
    path = sources_dir(project) / "白紙.xlsx"
    book.save(path)

    targets, findings = parse.plan(round_, [path], sources_dir(project))
    said = [f for f in findings if f.code == "P009"]

    assert not targets
    assert len(said) == 1
    assert "シート 2 枚" in said[0].message
    assert "1 本も出ませんでした" in said[0].message


def test_中身のあるブックでP009を出さない(project: Paths, round_: Round) -> None:
    """**何でも申告すると、本当に読めていないものが山に埋もれる。**"""
    path = _book(sources_dir(project) / "普通.xlsx", [["項目", "値"], ["受注番号", "x"]])
    _, findings = parse.plan(round_, [path], sources_dir(project))
    assert not [f for f in findings if f.code == "P009"]


def test_読めなかった描画パートは図0個と同じ顔をさせない(
        project: Paths, round_: Round) -> None:
    """**飲んだ例外は「図が無い」と見分けが付かない。**

    描画パートの XML が壊れていると `_shapes` は静かに空を返す ―― 図の
    描いてあるシートが「図の無いシート」として出てきて、`未読取` を宣言する
    先（`g1`）すら無くなる。`arp4 render` なら絵にはできるので、**そこへ
    案内できるかどうかが分かれ目**である。
    """
    import shutil, zipfile

    path = _book(sources_dir(project) / "業務フロー.xlsx", [["受注の流れ"]])
    original = path.with_suffix(".orig.xlsx")
    shutil.move(path, original)
    with zipfile.ZipFile(original) as source, zipfile.ZipFile(path, "w") as target:
        for entry in source.infolist():
            if entry.filename == "xl/worksheets/_rels/sheet1.xml.rels":
                continue
            target.writestr(entry, source.read(entry.filename))
        target.writestr("xl/worksheets/_rels/sheet1.xml.rels", _RELS)
        target.writestr("xl/drawings/drawing1.xml", "<xdr:wsDr><壊れ")
    original.unlink()

    targets, _ = parse.plan(round_, [path], sources_dir(project))
    doc = targets[0].doc

    assert "XML として読めませんでした" in "\n".join(doc.notes)
    assert "図が無いのではありません" in "\n".join(doc.notes)
    # **宣言する先は出す。** 出さないと、読めていないものほど静かに消える。
    assert "s1-g1" in [c.anchor for c in doc.chunks]


def test_関係を辿れないブックは図が無いことにしない(
        project: Paths, round_: Round) -> None:
    """**ブックまるごと図が消える形。** `xl/_rels/workbook.xml.rels` が
    壊れていると、図形もグラフも SmartArt も 1 つも見えなくなる。"""
    import shutil, zipfile

    path = _with_drawing(_book(sources_dir(project) / "構成図.xlsx", [["構成"]]))
    original = path.with_suffix(".orig.xlsx")
    shutil.move(path, original)
    with zipfile.ZipFile(original) as source, zipfile.ZipFile(path, "w") as target:
        for entry in source.infolist():
            target.writestr(entry, source.read(entry.filename))
        # 描画からもう 1 段（SmartArt・グラフ）を辿る関係だけが壊れている形。
        # openpyxl は救出モードなら描画を 1 つも読まないので**ブックは開く**
        # ―― 落ちるのはこちらの walk だけである。
        target.writestr("xl/drawings/_rels/drawing1.xml.rels",
                        "<Relationships><壊れ".encode("utf-8"))
    original.unlink()

    _, findings = parse.plan(round_, [path], sources_dir(project))
    said = [f for f in findings if f.code == "P008"]

    assert len(said) == 1
    assert "1 つも読めていません" in said[0].message
    assert "資料に無いのではありません" in said[0].message


def test_マクロが入っていることは黙らない(project: Paths, round_: Round) -> None:
    """**申告の規律が掛かっていなかった唯一の入口。**

    `.xlsm` を弾くのをやめたときに書いたのは「マクロが付いているだけで中身は
    同じ」だったが、**それは表の話でしかない** ―― 採番規則・入力チェック・
    帳票の組み立てがマクロにしか無いことがあり、そのときシートに書いてあるのは
    「ボタンを押す」だけである。数式のキャッシュ無しも貼り付け画像も
    「読めていない」と言っているのに、マクロだけは**あることすら伝えて
    いなかった**（読み手にはマクロの無いブックと同じに見える）。

    **検体（`tests/dataset/*.yml`）には置けない。** `vbaProject.bin` は Excel が
    作る OLE 複合ドキュメントで、写せば**バイナリをリポジトリに置く**ことになり、
    でっち上げれば Excel が開かない ―― どちらも検体の決めごとに反する。
    """
    import shutil, zipfile

    path = _book(sources_dir(project) / "受注登録.xlsm", [["項目", "チェック"]])
    original = path.with_suffix(".orig.xlsx")
    shutil.move(path, original)
    with zipfile.ZipFile(original) as source, zipfile.ZipFile(path, "w") as target:
        for entry in source.infolist():
            target.writestr(entry, source.read(entry.filename))
        # 中身は OLE 複合ドキュメント（zip の中にあるが zip では開けない）。
        # **見るのはパートがあるかどうかだけ**なので、先頭の 8 バイトで足りる。
        target.writestr("xl/vbaProject.bin", b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1")
    original.unlink()

    _, findings = parse.plan(round_, [path], sources_dir(project))
    said = [f for f in findings if f.code == "P006"]

    assert len(said) == 1
    assert "マクロ（VBA）が入っています" in said[0].message
    assert "VBE" in said[0].message                 # 次にやることを書く
    # **中身は取らない。** 取れると言えば、読まれないまま読んだことになる。
    assert "中身は取っていません" in said[0].message


def test_マクロの無いブックでP006を出さない(project: Paths, round_: Round) -> None:
    """**何でも申告すると、本当に読めていないものが山に埋もれる。**"""
    path = _book(sources_dir(project) / "普通.xlsx", [["項目", "値"]])
    _, findings = parse.plan(round_, [path], sources_dir(project))
    assert not [f for f in findings if f.code == "P006"]


def test_間の空行と空列は残し両端は広げない() -> None:
    """**番地から表の中の位置を割り出せる**ことが、区切りを許している前提である。

    詰めると ``at`` の範囲と表の形が合わなくなり、``B3:D5`` の 2 列目が
    C 列か D 列かを読み手が決められない ―― 番地を併記していれば読み直せる、
    という約束がそこだけ成り立たなくなる。
    """
    cells = {(3, 2): "a", (3, 4): "b", (5, 2): "c"}
    frame = parse._frame(cells, list(cells))

    assert frame.at == "B3:D5"
    assert parse._grid(cells, frame) == [["a", "", "b"], ["", "", ""], ["c", "", ""]]
    assert frame.addresses == [("B3", "a"), ("D3", "b"), ("B5", "c")]


def test_行頭の字下げは残し行末の空白は落とす(project: Paths, round_: Round) -> None:
    """字下げは**画面に見えている階層**である（項目定義書の「項目名」列の慣習）。"""
    path = _book(sources_dir(project) / "a.xlsx",
                 [["項目名", "型"], ["受注ヘッダ", ""], ["　　受注番号　", "文字列"],
                  ["  子  ", "数値"]])
    doc = _parse(round_, path, sources_dir(project))[0]

    table = [c for c in doc.chunks if c.rows][0]
    assert [row[0] for row in table.rows] == ["項目名", "受注ヘッダ", "　　受注番号", "  子"]
    assert "| 　　受注番号 |" in mdio.dump(doc)


def test_コメントだけのシートも1本出す(project: Paths, round_: Round) -> None:
    """**値で足切りするとシートごと消える。** 空欄への指摘がいちばん多い。"""
    from openpyxl.comments import Comment

    path = sources_dir(project) / "a.xlsx"
    book = Workbook()
    book.active.title = "確認事項"
    book.active["C3"].comment = Comment("桁数が未定。次回までに確定", "設計者")
    book.save(path)

    doc = _parse(round_, path, sources_dir(project))[0]
    memo = [c for c in doc.chunks if c.anchor == "s1-m1"]
    assert memo and memo[0].cells == [("C3", "設計者: 桁数が未定。次回までに確定")]


def test_リンクの行き先はtargetとlocationの両方を見る() -> None:
    """**別ブックの特定シート**は両方に分かれて入る（片方だけ見ると半分になる）。"""
    from openpyxl.worksheet.hyperlink import Hyperlink

    class _Cell:
        def __init__(self, link):
            self.hyperlink = link

    assert parse._link(_Cell(None)) == ""
    assert parse._link(_Cell(Hyperlink(ref="A1", location="受注!A1"))) == "#受注!A1"
    assert parse._link(_Cell(Hyperlink(ref="A1", target="外部.xlsx"))) == "外部.xlsx"
    assert parse._link(_Cell(Hyperlink(ref="A1", target="外部.xlsx",
                                       location="Sheet1!B2"))) == "外部.xlsx#Sheet1!B2"


def test_高さ0の行と幅0の列も画面には出ていない(project: Paths, round_: Round) -> None:
    """**隠すフラグは立っていないのに、画面からは消えている。**

    Excel は行の高さ 0・列の幅 0 を普通に受け取り、``hidden`` は立てない ――
    ツールが書き出した設計書と、境目を誤って詰めたブックで起きる。フラグだけを
    見ていたぶん**画面に見えている値として出していた**（申告も出ないので、
    混ざっていることすら言えない）。

    **潰れたものを数え分けるのは行だけ**である。実物の Excel で確かめたところ、
    「再表示」で幅 0 の列は既定幅に戻るのに高さ 0 の行は戻らない ―― 別の案内が
    要るのは行だけで、列にまで付けると**戻る操作を「戻りません」と言う**ことに
    なる。

    ここは :func:`parse._hidden` を直に見る ―― openpyxl は 0 を偽と見なして
    書けないので、ブックから作ろうとすると XML を手で書く話になる。
    """

    class _Dimension:
        def __init__(self, hidden=False, height=None, width=None,
                     first=None, last=None):
            self.hidden, self.height, self.width = hidden, height, width
            self.min, self.max = first, last

    class _Sheet:
        row_dimensions = {3: _Dimension(height=0), 4: _Dimension(hidden=True),
                          5: _Dimension(height=13.5)}
        column_dimensions = {"C": _Dimension(width=0, first=3, last=3),
                             "E": _Dimension(hidden=True, first=5, last=5),
                             "G": _Dimension(width=8.43, first=7, last=7)}

    invisible = parse._hidden(_Sheet())
    assert invisible.rows == {3, 4} and invisible.crushed_rows == {3}
    assert invisible.columns == {3, 5}          # 幅 0 の列も画面には出ていない


def test_紙にしか出ないものは表に混ぜない(project: Paths, round_: Round) -> None:
    """**フッタは表ではない。** 値と同じ表に混ぜると、どの行の値かが決まらない。"""
    path = sources_dir(project) / "a.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "機能一覧"
    sheet["A1"] = "機能ID"
    sheet["B1"] = "機能名"
    sheet["A2"] = "F-01"
    sheet["B2"] = "受注登録"
    sheet.oddFooter.left.text = "TLG-OMS-BD-014"
    sheet.oddFooter.right.text = "&P / &N"
    sheet.print_title_rows = "1:1"
    book.save(path)

    doc = _parse(round_, path, sources_dir(project))[0]
    setup = [c for c in doc.chunks if c.anchor == "s1-p1"]
    assert setup and dict(setup[0].cells) == {
        "フッタ左": "TLG-OMS-BD-014", "フッタ右": "&P / &N",
        "印刷タイトル行": "$1:$1"}
    assert "&P" in "\n".join(doc.notes)               # 刷るまで決まらない値は埋めない
    assert [c.anchor for c in doc.chunks] == ["s1-t1", "s1-p1"]


def test_フッタしか無いシートは1本出さない(project: Paths, round_: Round) -> None:
    """**ページ設定は全シートまとめて掛ける。** 数に入れると作業用の白紙まで出る。"""
    path = sources_dir(project) / "a.xlsx"
    book = Workbook()
    book.active.title = "作業用"
    book.active.oddFooter.center.text = "社外秘"
    book.save(path)

    assert _parse(round_, path, sources_dir(project)) == []


def test_Windowsの予約名を避ける() -> None:
    assert parse.safe_name("CON") == "_CON"
    assert parse.safe_name("a/b:c") == "a_b_c"
    assert parse.safe_name("末尾の点.") == "末尾の点"


# ── 図形 ────────────────────────────────────────────────────────
_DRAWING = """<?xml version="1.0" encoding="UTF-8"?>
<xdr:wsDr
  xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
  xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:twoCellAnchor>
    <xdr:sp><xdr:txBody><a:p><a:r><a:t>\u53d7\u6ce8\u5165\u529b</a:t></a:r></a:p></xdr:txBody></xdr:sp>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
  <xdr:twoCellAnchor>
    <xdr:grpSp>
      <xdr:sp><xdr:txBody>
        <a:p><a:r><a:t>\u4e0e\u4fe1</a:t></a:r><a:r><a:t>\u5224\u5b9a</a:t></a:r></a:p>
        <a:p><a:r><a:t>NG \u306f\u5dee\u623b\u3057</a:t></a:r></a:p>
      </xdr:txBody></xdr:sp>
      <xdr:sp><xdr:txBody><a:p/></xdr:txBody></xdr:sp>
    </xdr:grpSp>
    <xdr:clientData/>
  </xdr:twoCellAnchor>
  <xdr:oneCellAnchor>
    <xdr:cxnSp/>
    <xdr:clientData/>
  </xdr:oneCellAnchor>
</xdr:wsDr>
"""

_DRAWING_LINKED = """<?xml version="1.0" encoding="UTF-8"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:twoCellAnchor>
    <xdr:sp><xdr:nvSpPr><xdr:cNvPr id="2" name="A"/></xdr:nvSpPr>
      <xdr:txBody><a:p><a:r><a:t>受注入力</a:t></a:r></a:p></xdr:txBody></xdr:sp>
  </xdr:twoCellAnchor>
  <xdr:twoCellAnchor>
    <xdr:sp><xdr:nvSpPr><xdr:cNvPr id="3" name="B"/></xdr:nvSpPr>
      <xdr:txBody><a:p><a:r><a:t>与信判定</a:t></a:r></a:p></xdr:txBody></xdr:sp>
  </xdr:twoCellAnchor>
  <xdr:twoCellAnchor>
    <xdr:cxnSp><xdr:nvCxnSpPr><xdr:cNvCxnSpPr>
        <a:stCxn id="2" idx="3"/><a:endCxn id="3" idx="1"/>
      </xdr:cNvCxnSpPr></xdr:nvCxnSpPr>
      <xdr:spPr><a:ln><a:tailEnd type="triangle"/></a:ln></xdr:spPr></xdr:cxnSp>
  </xdr:twoCellAnchor>
  <xdr:twoCellAnchor>
    <xdr:cxnSp><xdr:nvCxnSpPr><xdr:cNvCxnSpPr>
        <a:stCxn id="3" idx="3"/><a:endCxn id="2" idx="1"/>
      </xdr:cNvCxnSpPr></xdr:nvCxnSpPr>
      <xdr:spPr><a:ln><a:headEnd type="triangle"/></a:ln></xdr:spPr></xdr:cxnSp>
  </xdr:twoCellAnchor>
</xdr:wsDr>
"""

_RELS = """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Target="../drawings/drawing1.xml"
    Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"/>
</Relationships>
"""


def _with_drawing(path: Path) -> Path:
    """openpyxl は図形を書けないので、保存した xlsx へ描画パートを差し込む。"""
    import shutil, zipfile

    original = path.with_suffix(".orig.xlsx")
    shutil.move(path, original)
    with zipfile.ZipFile(original) as source, zipfile.ZipFile(path, "w") as target:
        for entry in source.infolist():
            if entry.filename == "xl/worksheets/_rels/sheet1.xml.rels":
                continue
            target.writestr(entry, source.read(entry.filename))
        target.writestr("xl/worksheets/_rels/sheet1.xml.rels", _RELS)
        target.writestr("xl/drawings/drawing1.xml", _DRAWING)
    original.unlink()
    return path


def test_図形はテキストだけ取り出す(project: Paths, round_: Round) -> None:
    """**線の接続は取れないが、箱の中の語は取れる。** 語だけでも整理層は仕事ができる。"""
    path = _with_drawing(_book(sources_dir(project) / "flow.xlsx", [["業務フロー"]]))
    doc = _parse(round_, path, sources_dir(project))[0]

    shapes = [c for c in doc.chunks if c.anchor == "s1-g1"]
    assert len(shapes) == 1
    values = [text for _, text in shapes[0].cells]
    assert values == ["受注入力", "与信判定\nNG は差戻し"]   # 段落は改行、ラン内は連結


def test_図形の申告は取れなかったぶんを言う(project: Paths, round_: Round) -> None:
    """**黙って空を返さない**（`未読取` にするかどうかの判断材料になる）。

    数えるのは**図形そのもの**であってアンカーではない。グループ化された
    業務フローは箱 10 個でアンカー 1 個になるので、アンカーを数えていた頃は
    「図形が 1 個あり、10 個からテキストを取り出しました」と自分で矛盾していた。
    """
    path = _with_drawing(_book(sources_dir(project) / "flow.xlsx", [["業務フロー"]]))
    doc = _parse(round_, path, sources_dir(project))[0]

    note = "\n".join(doc.notes)
    # sp 3 個（うち 1 個はテキストなし・グループの中も 1 個ずつ数える）＋ 接続子 1 本
    assert "図形 3 個・接続子 1 本" in note
    assert "2 個からテキスト" in note                 # 矢印 1 本ぶんは取れていない
    assert "s1-g1" in note and "未読取" in note


def test_テキストも接続も無ければ読めていないと言う() -> None:
    drawing = parse._shapes(parse.Part("drawing1.xml", b"""<?xml version="1.0"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing">
  <xdr:absoluteAnchor><xdr:cxnSp/></xdr:absoluteAnchor>
</xdr:wsDr>"""), parse.Drawing())
    assert (drawing.shapes, drawing.connectors, drawing.total) == (0, 1, 1)
    assert (drawing.labels, drawing.links) == ([], [])
    assert "テキストも接続も取れていません" in parse._shape_note(1, drawing)


def test_両端が図形に付いていない線は取れないと数える() -> None:
    """**座標から当てない。** 線を目分量で置いた図は「取れなかった」と申告する。"""
    drawing = parse._shapes(parse.Part("drawing1.xml", """<?xml version="1.0"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:twoCellAnchor>
    <xdr:sp><xdr:nvSpPr><xdr:cNvPr id="1"/></xdr:nvSpPr>
      <xdr:txBody><a:p><a:r><a:t>受注入力</a:t></a:r></a:p></xdr:txBody></xdr:sp>
  </xdr:twoCellAnchor>
  <xdr:twoCellAnchor><xdr:cxnSp><xdr:nvCxnSpPr><xdr:cNvCxnSpPr/></xdr:nvCxnSpPr>
    <xdr:spPr><a:ln><a:tailEnd type="triangle"/></a:ln></xdr:spPr></xdr:cxnSp>
  </xdr:twoCellAnchor>
</xdr:wsDr>""".encode("utf-8")), parse.Drawing())

    assert drawing.labels == ["受注入力"]
    assert drawing.links == []
    assert (drawing.loose, drawing.unnamed) == (1, 0)
    # **絵にしても読めない側**である（相手が無名なだけの線とは案内が違う）
    note = parse._shape_note(1, drawing)
    assert "接続子 1 本はどこにも繋がっていません" in note
    assert "絵にしても決まりません" in note


# ── 接続の線種（凡例が意味を描き分ける手がかり） ────────────────
def _connected(*lines: str) -> parse.Drawing:
    """箱 3 つと、``a:ln`` だけを差し替えた接続子を並べた描画パート。"""
    boxes = "".join(f"""\
  <xdr:twoCellAnchor><xdr:sp><xdr:nvSpPr><xdr:cNvPr id="{i}" name="s{i}"/></xdr:nvSpPr>
    <xdr:txBody><a:p><a:r><a:t>箱{i}</a:t></a:r></a:p></xdr:txBody></xdr:sp>
  </xdr:twoCellAnchor>
""" for i in (2, 3, 4))
    joins = "".join(f"""\
  <xdr:twoCellAnchor><xdr:cxnSp><xdr:nvCxnSpPr><xdr:cNvCxnSpPr>
      <a:stCxn id="2" idx="0"/><a:endCxn id="{3 + i}" idx="0"/>
    </xdr:cNvCxnSpPr></xdr:nvCxnSpPr>{line}</xdr:cxnSp></xdr:twoCellAnchor>
""" for i, line in enumerate(lines))
    return parse._shapes(parse.Part("drawing1.xml", f"""<?xml version="1.0"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
{boxes}{joins}</xdr:wsDr>""".encode("utf-8")), parse.Drawing())


def test_接続子の名前は判断せずそのまま出す() -> None:
    """**線に付いた言葉の唯一の在り処が名前である。**

    図形の名前（``図 3``）は自動採番なので取らないが（:func:`parse._shape_alt`）、
    線は代替テキストもテキスト枠も持たない ―― 実測（kotonoha r001）の体制図では
    ``委託`` ``点検依頼`` が名前にだけあり、**「どれが委託か」は他のどこにも
    書かれていなかった**（線種は 7 本とも実線で、凡例が言う破線は資料に無い）。

    **自動名を機械が落とさない。** ``コネクタ 10`` が混ざるのは整理層が見て
    捨てればよいノイズだが、機械が判定して落とすと**落とした判断が誰にも見えなく
    なる** ―― しかも判定は言語ごとの綴りに依るので、資料の言語が変わるだけで
    人の付けた名前を捨てる。
    """
    named = ('<xdr:nvCxnSpPr><xdr:cNvPr id="9" name="{name}"/><xdr:cNvCxnSpPr>'
             '<a:stCxn id="2" idx="0"/><a:endCxn id="{to}" idx="0"/>'
             "</xdr:cNvCxnSpPr></xdr:nvCxnSpPr>"
             "<xdr:spPr><a:ln><a:tailEnd type=\"triangle\"/></a:ln></xdr:spPr>")
    body = f"""<?xml version="1.0"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:sp><xdr:nvSpPr><xdr:cNvPr id="2" name="s2"/></xdr:nvSpPr>
    <xdr:txBody><a:p><a:r><a:t>運用</a:t></a:r></a:p></xdr:txBody></xdr:sp>
  <xdr:sp><xdr:nvSpPr><xdr:cNvPr id="3" name="s3"/></xdr:nvSpPr>
    <xdr:txBody><a:p><a:r><a:t>委託先</a:t></a:r></a:p></xdr:txBody></xdr:sp>
  <xdr:sp><xdr:nvSpPr><xdr:cNvPr id="4" name="s4"/></xdr:nvSpPr>
    <xdr:txBody><a:p><a:r><a:t>経理部</a:t></a:r></a:p></xdr:txBody></xdr:sp>
  <xdr:cxnSp>{named.format(name="委託", to=3)}</xdr:cxnSp>
  <xdr:cxnSp>{named.format(name="コネクタ 10", to=4)}</xdr:cxnSp>
</xdr:wsDr>"""

    drawing = parse._shapes(parse.Part("drawing1.xml", body.encode("utf-8")),
                            parse.Drawing())

    assert [link.name for link in drawing.links] == ["委託", "コネクタ 10"]
    rows = parse._link_rows(drawing.links)
    assert rows[0] == ["元", "向き", "先", "名前", "線種"]
    assert rows[1] == ["運用", "→", "委託先", "委託", "実線"]
    # 自動名も**そのまま**並べる（捨てるのは整理層の仕事）。
    assert rows[2] == ["運用", "→", "経理部", "コネクタ 10", "実線"]


def test_名前の無い接続子は名前の列を出さない() -> None:
    """**空の列を足さない。** 名前を持たない図で表を横に伸ばしても何も伝わらない。"""
    plain = '<xdr:spPr><a:ln><a:tailEnd type="triangle"/></a:ln></xdr:spPr>'

    rows = parse._link_rows(_connected(plain, plain).links)

    assert rows[0] == ["元", "向き", "先", "線種"]


def test_接続の線種を転記する() -> None:
    """**体制図の凡例は線種で意味を描き分ける**（「実線＝指揮命令 / 破線＝委託」）。

    線種を出していなかったあいだ、整理層は接続 1 本ずつの意味を決められなかった
    ―― 実測（kotonoha r001）で体制図がこれで**通し実行の唯一の「未読取」1 件**に
    なった。読むのは ``a:prstDash`` に書いてあることの転記で、**意味は付けない。**
    """
    drawing = _connected(
        '<xdr:spPr><a:ln w="12700"><a:solidFill><a:srgbClr val="404040"/></a:solidFill>'
        '<a:tailEnd type="triangle"/></a:ln></xdr:spPr>',
        '<xdr:spPr><a:ln w="19050"><a:solidFill><a:srgbClr val="C00000"/></a:solidFill>'
        '<a:prstDash val="sysDash"/><a:tailEnd type="triangle"/></a:ln></xdr:spPr>')

    assert [(l.dash, l.color, l.width) for l in drawing.links] == [
        ("実線", "#404040", "1pt"),        # `a:prstDash` が無いのは実線（既定）
        ("破線", "#C00000", "1.5pt")]      # `sysDash` も破線（訳すだけ）
    # 色と太さは**違いがあるので**列になる。
    assert parse._link_rows(drawing.links)[0] == ["元", "向き", "先",
                                                  "線種", "線色", "太さ"]


def test_線が全部同じでも線種は出す() -> None:
    """**「7 本とも実線でした」自体が答え**である。

    凡例が破線での描き分けを謳っていても図に破線が 1 本も無いなら、整理層が
    決められることは何も無い（資料の側に情報が無い）。列が無いと、それが
    「線種の無い図」なのか「線種を読んでいないパーサ」なのか区別できない。
    色と太さは逆で、全行同じ値の列は表を横に伸ばすだけなので畳む。
    """
    same = ('<xdr:spPr><a:ln w="12700"><a:solidFill><a:srgbClr val="404040"/>'
            '</a:solidFill><a:tailEnd type="triangle"/></a:ln></xdr:spPr>')

    rows = parse._link_rows(_connected(same, same).links)

    assert rows[0] == ["元", "向き", "先", "線種"]
    assert [row[-1] for row in rows[1:]] == ["実線", "実線"]


def test_テーマ由来の線種は断定しない() -> None:
    """**解決していないものを埋めない。** ``xdr:style`` の ``a:lnRef`` は
    ``theme1.xml`` の ``lnStyleLst`` を引くので、ここまでは辿っていない ――
    「実線」と埋めると、破線で描き分けてある図が全部同じに見えるうえ**申告が嘘**
    になる。全部空なら列ごと出さない。
    """
    themed = ('<xdr:spPr><a:ln><a:tailEnd type="triangle"/></a:ln></xdr:spPr>'
              '<xdr:style><a:lnRef idx="2"/></xdr:style>')

    drawing = _connected(themed, themed)

    assert [l.dash for l in drawing.links] == ["", ""]
    assert parse._link_rows(drawing.links)[0] == ["元", "向き", "先"]


def test_壊れた描画パートでパースを止めない() -> None:
    drawing = parse._shapes(parse.Part("drawing1.xml", b"<not xml"), parse.Drawing())
    assert (drawing.shapes, drawing.labels, drawing.links) == (0, [], [])
    assert parse._diagram(b"<not xml", parse.Drawing()).diagram_boxes == 0


def test_グラフとSmartArtと埋め込みを別々に数える() -> None:
    """**どれも ``xdr:sp`` ではない。** 数えないと、そのシートは図形 0 個になる。

    何であるかは ``a:graphicData`` の uri に書いてあるので、読むのは転記である
    ―― 分けて数えるのは、**次にやることが 3 つとも違う**からである（グラフは
    元データのシートを読む、SmartArt は箱の文字が取れる、埋め込みは元ファイルを
    sources/ に足す）。
    """
    drawing = parse._shapes(parse.Part("drawing1.xml", """<?xml version="1.0"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:twoCellAnchor><xdr:graphicFrame><a:graphic><a:graphicData
    uri="http://schemas.openxmlformats.org/drawingml/2006/chart"/>
    </a:graphic></xdr:graphicFrame></xdr:twoCellAnchor>
  <xdr:twoCellAnchor><xdr:graphicFrame><a:graphic><a:graphicData
    uri="http://schemas.openxmlformats.org/drawingml/2006/diagram"/>
    </a:graphic></xdr:graphicFrame></xdr:twoCellAnchor>
  <xdr:twoCellAnchor><xdr:graphicFrame><a:graphic><a:graphicData
    uri="http://schemas.openxmlformats.org/presentationml/2006/ole"/>
    </a:graphic></xdr:graphicFrame></xdr:twoCellAnchor>
</xdr:wsDr>""".encode("utf-8")), parse.Drawing())

    assert (drawing.charts, drawing.diagrams, drawing.objects) == (1, 1, 1)
    assert drawing.total == 3                    # 図形 0 個でも「無い」ではない
    assert drawing.summary == "SmartArt 1 個・グラフ 1 個・埋め込みオブジェクト 1 個"


def test_SmartArtの見た目用の複製を二重に数えない() -> None:
    """**箱・繋ぎの点・見た目の複製が同じ一覧に並ぶ。** 全部拾うと語が 2 回出る。"""
    drawing = parse._diagram("""<?xml version="1.0"?>
<dgm:dataModel xmlns:dgm="http://schemas.openxmlformats.org/drawingml/2006/diagram"
               xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <dgm:ptLst>
    <dgm:pt modelId="0" type="doc"/>
    <dgm:pt modelId="1"><dgm:t><a:p><a:r><a:t>受注</a:t></a:r></a:p></dgm:t></dgm:pt>
    <dgm:pt modelId="2" type="sibTrans"><dgm:t><a:p/></dgm:t></dgm:pt>
    <dgm:pt modelId="3" type="pres"><dgm:t><a:p><a:r><a:t>受注</a:t></a:r></a:p></dgm:t></dgm:pt>
  </dgm:ptLst>
</dgm:dataModel>""".encode("utf-8"), parse.Drawing())

    assert drawing.diagram_boxes == 1
    assert drawing.labels == ["受注"]


# ── 貼り付け画像（機械は読まないが、整理層は読む） ──────────────
def _book_with_pictures(path: Path, count: int, title: str = "画面") -> Path:
    """絵柄の違う画像を ``count`` 枚貼ったブック。**実体を持たせる。**"""
    import io

    import picture
    from openpyxl.drawing.image import Image

    book = Workbook()
    sheet = book.active
    sheet.title = title
    sheet["A1"] = "受注入力"
    for index, kind in enumerate(("画面", "帳票", "ロゴ")[:count]):
        shot = Image(io.BytesIO(picture.draw(kind, 120, 80)))
        shot.anchor = f"C{index * 8 + 3}"
        sheet.add_image(shot)
    book.save(path)
    return path


def test_貼り付け画像は実体ごと出して名前をパース結果に書く(
        project: Paths, round_: Round) -> None:
    """**機械が読めないことと、誰にも読まれないことは別である。**

    貼り付け画像に描かれた業務フロー・現行画面は、枚数だけ数えていたころ
    **ブックの中に入ったまま**だった ―― 整理層は開けば読めるのに、開ける場所へ
    出していなかった。出すのはバイト列の転記なので、意味の判断はしていない。
    """
    source = _book_with_pictures(sources_dir(project) / "受注.xlsx", 2)
    targets, findings = parse.plan(round_, [source], sources_dir(project))
    written, said = parse.write(targets)
    assert not [f for f in findings + said if f.level == "error"]

    # 実体は parsed/ の隣（images/ の同じ枝）に、シート名から付けた名前で出る
    where = round_.images / "受注.xlsx"
    assert sorted(p.name for p in where.iterdir()) == ["画面-p1.png", "画面-p2.png"]
    assert (where / "画面-p1.png").read_bytes()[:8] == b"\x89PNG\r\n\x1a\n"

    # **名前はパース結果の中にある。** 出典として指せないと、読み取った内容を
    # 整理結果へ書くときに根拠が残らない。
    body = (round_.parsed / "受注.xlsx" / "画面.md").read_text(encoding="utf-8")
    assert "<!-- a:s1-i1 at=画像 2 枚 -->" in body
    assert "![画面-p1.png](../../images/受注.xlsx/画面-p1.png)" in body
    assert (round_.parsed / "受注.xlsx" / "画面.md").parent.joinpath(
        "../../images/受注.xlsx/画面-p1.png").resolve().is_file()

    # **sources.yml にも名前で残す。** 撮り直したときに、消えた画像とまだ出して
    # いない画像が区別できなくなる。
    parse.record(round_, targets, written)
    prints = round_.prints.read_text(encoding="utf-8")
    assert "受注.xlsx/画面-p1.png" in prints


def test_同じ実体を貼った2枚を別々の画像として出さない() -> None:
    """**1 枚の画像を 2 か所に貼ったシート**は ``xdr:pic`` が 2 個・実体は 1 つ。

    枚数ぶん書き出すと同じバイト列が別の名前で並び、読む側は**違う画像だと
    思って両方開く** ―― 読む予算をそれだけで使う。
    """
    drawing = parse.Drawing(pictures=2, media=[("xl/media/image1.png", "現行画面"),
                                               ("xl/media/image1.png", "現行画面")])
    shots, chunk, _ = parse._pictures(
        3, "画面", Path("受注.xlsx"), drawing, {"xl/media/image1.png": b"\x89PNG"})

    assert [one.name for one in shots] == ["画面-p1.png"]
    assert chunk is not None and chunk.at == "画像 1 枚"


# ── 画像の中の文字（Windows OCR） ────────────────────────────────
def _reads(monkeypatch: pytest.MonkeyPatch, *lines: str) -> None:
    """engine の代わりに、決まった行を返す（`conftest` の差し替えを上書きする）。"""
    from arp4 import ocr

    monkeypatch.setattr(ocr, "_unavailable", lambda: "")
    monkeypatch.setattr(
        ocr, "_run",
        lambda bodies: ("", [ocr.Reading(lines=lines, language="ja")
                             for _ in bodies]))


def test_画像の中の文字は別のアンカーに出る(
        project: Paths, round_: Round, monkeypatch: pytest.MonkeyPatch) -> None:
    """**`i1` に混ぜない。** バイト列の写し（資料）と機械の読みは出自が違う ――
    同じ塊に入れると、整理層は誤読を「資料にそう書いてある」と読む。
    """
    _reads(monkeypatch, "受注番号 ORDER-001", "得意先コード")
    source = _book_with_pictures(sources_dir(project) / "受注.xlsx", 1)
    doc = _parse(round_, source, sources_dir(project))[0]

    chunk = next(c for c in doc.chunks if c.anchor == "s1-o1")
    assert chunk.at == "画像 1 枚"
    assert "読み違えが混ざります" in chunk.heading      # 出典として指されたときに見える
    assert "`画面-p1.png`" in chunk.text                # どの画像から出た字か
    assert "    受注番号 ORDER-001" in chunk.text       # 原文のまま（表に組み直さない）
    # 申告にも枚数で出る（残りが「絵である」ことはこれで言える）
    note = "".join(doc.notes)
    assert "うち 1 枚からは Windows OCR が文字を読み出しました" in note


def test_字の出なかった画像にも空でない塊を出す(
        project: Paths, round_: Round) -> None:
    """**空の `o1` は「画像に文字が無かった」に見える。**

    それは「資料に無い」と「機械が読めていない」の取り違えそのものなので、
    engine が動いたことと字が無かったことを本文に書く（`conftest` の差し替えが
    ちょうどこの形＝「動いたが字は無い」である）。
    """
    source = _book_with_pictures(sources_dir(project) / "受注.xlsx", 1)
    doc = _parse(round_, source, sources_dir(project))[0]

    chunk = next(c for c in doc.chunks if c.anchor == "s1-o1")
    assert "文字は見つかりませんでした" in chunk.text
    assert "Windows OCR では文字を 1 つも読めませんでした" in "".join(doc.notes)


def test_読みにいっていないことと字が無いことを混ぜない(
        project: Paths, round_: Round) -> None:
    """`--no-ocr` は「誰も見ていない」であって「絵だった」ではない。"""
    source = _book_with_pictures(sources_dir(project) / "受注.xlsx", 1)
    targets, _ = parse.plan(round_, [source], sources_dir(project), use_ocr=False)
    doc = targets[0].doc

    chunk = next(c for c in doc.chunks if c.anchor == "s1-o1")
    assert "読みにいっていません" in chunk.text
    assert "文字は見つかりませんでした" not in chunk.text
    assert "読みにいっていません（`--no-ocr`）" in "".join(doc.notes)


def test_engineの無い環境は走らせた人にも1度だけ言う(
        project: Paths, round_: Round, monkeypatch: pytest.MonkeyPatch) -> None:
    """**足りないのは言語パック 1 つ**ということが、入れられる人の目に触れる。

    パース結果の中だけに書いても、それを読むのは整理層である ―― しかも資料
    30 冊ぶん同じ行を並べても、分かることは 1 つも増えない。
    """
    from arp4 import ocr

    monkeypatch.setattr(ocr, "_unavailable", lambda: "言語パックがありません")
    where = sources_dir(project)
    _book_with_pictures(where / "受注.xlsx", 2)
    _book_with_pictures(where / "出荷.xlsx", 2, title="一覧")
    targets, findings = parse.plan(round_, [where], where)

    assert [f.code for f in findings if f.code == "P016"] == ["P016"]
    assert "言語パックがありません" in next(
        f for f in findings if f.code == "P016").message
    chunk = next(c for c in targets[0].doc.chunks if c.anchor == "s1-o1")
    assert "読めませんでした（言語パックがありません）" in chunk.text


def test_画像の無い資料でOCRの申告を出さない(
        project: Paths, round_: Round, monkeypatch: pytest.MonkeyPatch) -> None:
    """**関係の無い資料に環境の話を出さない。** 出すと申告そのものが読み飛ばされる。"""
    from arp4 import ocr

    monkeypatch.setattr(ocr, "_unavailable", lambda: "言語パックがありません")
    path = _book(sources_dir(project) / "a.xlsx", [["論理名", "物理名"]])
    _, findings = parse.plan(round_, [path], sources_dir(project))

    assert not [f for f in findings if f.code == "P016"]


def test_実体を取り出せない画像は数えて申告する() -> None:
    """リンク画像（``r:link``）はブックの中に実体を持たない。

    指しているのは**資料を作った人の手元のパス**なので、こちらからは辿れない
    ―― 黙ると、貼ってある枚数と出した枚数の差が誰にも説明されないまま残る。
    """
    note = parse._picture_note(2, parse.Drawing(pictures=3, picture_alts=1,
                                                media=[("xl/media/image1.png", "")]))
    assert "実体は `s2-i1` に出してあります" in note
    assert "うち 2 枚は実体を取り出せませんでした" in note


_DRAWING_MUTE = """<?xml version="1.0" encoding="UTF-8"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
  <xdr:twoCellAnchor><xdr:sp><xdr:txBody><a:p/></xdr:txBody></xdr:sp></xdr:twoCellAnchor>
  <xdr:twoCellAnchor><xdr:cxnSp/></xdr:twoCellAnchor>
</xdr:wsDr>
"""


def test_テキストが無い図形でもアンカーは出す(project: Paths, round_: Round) -> None:
    """**読めていないシートほど静かに消える**のを止める。

    アンカーが無いとそのシートは `freeze` の未整理一覧（G001）に上がらず、
    `未読取` を宣言しようにも G004「アンカーがありません」で弾かれる ―― つまり
    「機械が読めていない」と申告する先が無くなる。
    """
    import shutil, zipfile

    path = _book(sources_dir(project) / "flow.xlsx", [["業務フロー"]])
    original = path.with_suffix(".orig.xlsx")
    shutil.move(path, original)
    with zipfile.ZipFile(original) as source, zipfile.ZipFile(path, "w") as target:
        for entry in source.infolist():
            if entry.filename != "xl/worksheets/_rels/sheet1.xml.rels":
                target.writestr(entry, source.read(entry.filename))
        target.writestr("xl/worksheets/_rels/sheet1.xml.rels", _RELS)
        target.writestr("xl/drawings/drawing1.xml", _DRAWING_MUTE)
    original.unlink()

    doc = _parse(round_, path, sources_dir(project))[0]
    shapes = [c for c in doc.chunks if c.anchor == "s1-g1"]
    assert len(shapes) == 1
    assert shapes[0].at == "図形 1 個・接続子 1 本"
    assert "テキストの入った図形はありません" in shapes[0].cells[0][1]
    assert "render" in shapes[0].cells[0][1]

    # 書き出して読み戻しても見える（アンカーとして数えられる）
    written = mdio.write(round_.parsed / "flow.md", doc)
    assert "s1-g1" in {a.id for a in mdio.read(written).anchors}


def test_図形の接続は別のアンカーに出す(project: Paths, round_: Round) -> None:
    """**接続は転記であって推測ではない。** 接続子は繋ぐ相手の id を持っている。

    箱の名前（何があるか）と線（どう繋がるか）は別の事実なので、片方だけを出典に
    できるようにアンカーを分ける。
    """
    import shutil, zipfile

    path = _book(sources_dir(project) / "flow.xlsx", [["業務フロー"]])
    original = path.with_suffix(".orig.xlsx")
    shutil.move(path, original)
    with zipfile.ZipFile(original) as source, zipfile.ZipFile(path, "w") as target:
        for entry in source.infolist():
            if entry.filename != "xl/worksheets/_rels/sheet1.xml.rels":
                target.writestr(entry, source.read(entry.filename))
        target.writestr("xl/worksheets/_rels/sheet1.xml.rels", _RELS)
        target.writestr("xl/drawings/drawing1.xml", _DRAWING_LINKED)
    original.unlink()

    doc = _parse(round_, path, sources_dir(project))[0]
    links = [c for c in doc.chunks if c.anchor == "s1-c1"]
    assert len(links) == 1
    assert links[0].at == "接続 2 本"
    # 線種は取れたら必ず出す（色と太さは違いがあるときだけ ―― ここは 2 本とも
    # `a:ln` に色も太さも無いので出ない）。
    assert links[0].rows[0] == ["元", "向き", "先", "線種"]
    # 矢羽根が終点側 → そのまま。始点側 → **向きを起こし直す**（逆に書かない）
    assert links[0].rows[1] == ["受注入力", "→", "与信判定", "実線"]
    assert links[0].rows[2] == ["受注入力", "→", "与信判定", "実線"]
    assert "接続 2 本" in "\n".join(doc.notes)


# ── すかすかな塊・結合の暴走・アンカーの偽造 ────────────────────
def test_すかすかでも小さい枠は表のまま(project: Paths, round_: Round) -> None:
    """**3 行 3 列に値が 3 つは書きかけの表**であって工程表ではない。

    すかすか判定を密度だけで掛けると、値の少ない小さな表がぜんぶ箇条書きに
    なる ―― 大きさと密度の両方を見る理由である。
    """
    path = sources_dir(project) / "small.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "小さい表"
    sheet["A1"], sheet["C3"], sheet["E5"] = "区分", "値", "備考"
    book.save(path)

    doc = _parse(round_, path, sources_dir(project))[0]
    assert [c.anchor for c in doc.chunks] == ["s1-t1"]
    assert "番地付きの箇条書き" not in "\n".join(doc.notes)


def test_すかすかな塊は境目で切り替わる(project: Paths, round_: Round) -> None:
    """判定は ``面積 ≧ 100`` かつ ``面積 > セル数 × 8``。**両方を確かめる。**"""
    def build(rows: int, columns: int, every: int) -> mdio.Doc:
        path = sources_dir(project) / f"grid{rows}x{columns}x{every}.xlsx"
        book = Workbook()
        sheet = book.active
        sheet.title = "格子"
        for row in range(1, rows + 1):
            for column in range(1, columns + 1, every):
                sheet.cell(row=row, column=column, value="■")
        book.save(path)
        return _parse(round_, path, sources_dir(project))[0]

    # 12x12=144 の枠に 1 列おき（72 セル）→ 144 < 72*8 なので表のまま
    assert [c.anchor for c in build(12, 12, 2).chunks] == ["s1-t1"]
    # 同じ枠に 12 列おき（12 セル）→ 144 > 12*8 なので箇条書きへ
    assert [c.anchor for c in build(12, 12, 12).chunks] == ["s1-x1"]


def test_結合が表の外まで伸びていても幻の行を作らない(
        project: Paths, round_: Round) -> None:
    """**画面に見えている表の外に行は無い。**

    結合は列を丸ごと選んで掛けられる。最終行まで素直に埋めると、3 行の表から
    100 万行の表が生える ―― 忠実性の回復ではなく捏造である。
    """
    path = sources_dir(project) / "merge.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "暴走"
    sheet["A1"], sheet["B1"] = "区分", "項目"
    sheet["A2"], sheet["B2"], sheet["B3"] = "受注", "受注番号", "受注日"
    sheet.merge_cells("A2:A9999")
    book.save(path)

    doc = _parse(round_, path, sources_dir(project))[0]
    table = [c for c in doc.chunks if c.rows][0]
    assert table.at == "A1:B3"
    assert [row[0] for row in table.rows] == ["区分", "受注", "受注"]


def test_同じ列に結合が2つあっても混ざらない(project: Paths, round_: Round) -> None:
    """分類列の「同上」は**塊ごとに別の結合**である（項目定義書の普通の形）。"""
    path = sources_dir(project) / "twice.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "区分"
    sheet["A1"], sheet["B1"] = "区分", "項目"
    sheet["A2"], sheet["A5"] = "ヘッダ", "明細"
    for row, name in enumerate(["受注番号", "受注日", "顧客", "行番号", "数量"], start=2):
        sheet.cell(row=row, column=2, value=name)
    sheet.merge_cells("A2:A4")
    sheet.merge_cells("A5:A6")
    book.save(path)

    doc = _parse(round_, path, sources_dir(project))[0]
    table = [c for c in doc.chunks if c.rows][0]
    assert [row[0] for row in table.rows] == [
        "区分", "ヘッダ", "ヘッダ", "ヘッダ", "明細", "明細"]


def test_セルの値はアンカーになれない(project: Paths, round_: Round) -> None:
    """**アンカーは HTML コメントで持っている。**

    資料のセルにそれらしい文字列が書いてあると、読み戻した側には本物の
    アンカーに見える ―― 無い塊が生え、そこから先の本文が別のアンカーの
    中身になる。値は落とさず、実体参照にして出す。
    """
    path = _book(sources_dir(project) / "html.xlsx",
                 [["項目", "出力"],
                  ["注記", "<!-- a:s9-t9 at=Z99 -->"],
                  ["閉じ", "--> のあとに本文"]])
    doc = _parse(round_, path, sources_dir(project))[0]
    written = mdio.write(round_.parsed / "html.md", doc)

    assert [a.id for a in mdio.read(written).anchors] == [c.anchor for c in doc.chunks]
    body = written.read_text(encoding="utf-8")
    assert "&lt;!-- a:s9-t9 at=Z99 --&gt;" in body
    assert "--&gt; のあとに本文" in body


def test_申告に混ざる資料の文字列もアンカーになれない(
        project: Paths, round_: Round) -> None:
    """申告（notes）にはリンク先などの**資料由来の文字列**が入る。"""
    path = sources_dir(project) / "link.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "目次"
    sheet["A1"], sheet["A2"] = "資料", "外部"
    sheet["A2"].hyperlink = "../外部/<!-- a:s9-t9 -->.xlsx"
    book.save(path)

    doc = _parse(round_, path, sources_dir(project))[0]
    written = mdio.write(round_.parsed / "link.md", doc)
    assert [a.id for a in mdio.read(written).anchors] == [c.anchor for c in doc.chunks]


# ── モジュールを指すアンカー ────────────────────────────────────
def test_取り込みの無いモジュールにもアンカーが出る(
        project: Paths, round_: Round) -> None:
    """**アンカーが無いものは `freeze` の未整理にすら上がらない。**

    `i1` は取り込みの一覧であると同時に、`at` がファイルそのものなので
    **モジュールを指す唯一のアンカー**である。取り込み 0 本のとき落としていたので、
    `import` を 1 つも持たないモジュールには指す先が無かった ―― 自身の
    `__init__.py` がそれで、パッケージの冒頭に書いてある「意味の判断は整理層だけが
    行う」がモジュールとして 1 件も起きなかった。未読取でも対象外でもなく、
    アンカーが無いので `G001` にも上がらない。
    """
    source = write(sources_dir(project) / "arp4" / "__init__.py",
                   '"""ai-ready-pipeline 4。"""\n\n__version__ = "4.0.0.dev0"\n')
    targets, _ = parse.plan(round_, [source], sources_dir(project))
    chunks = {c.anchor: c for c in targets[0].doc.chunks}

    assert set(chunks) == {"v1", "i1"}
    assert chunks["i1"].at == "arp4/__init__.py"      # ファイルそのものを指す
    assert len(chunks["i1"].rows) == 1                # 見出しだけ（取り込み 0 本）


def test_中身の無いファイルにはアンカーを立てない(
        project: Paths, round_: Round) -> None:
    """**書くことが無いものに宣言を迫らない。**

    通すために対象外宣言を量産させると、宣言そのものが読まれなくなる。
    """
    source = write(sources_dir(project) / "arp4" / "空.py", "")
    targets, _ = parse.plan(round_, [source], sources_dir(project))

    assert targets == []


# ── 例外の欄 ────────────────────────────────────────────────────
def _rows(project: Paths, round_: Round, body: str) -> dict[str, list[str]]:
    source = write(sources_dir(project) / "m.py", body)
    targets, _ = parse.plan(round_, [source], sources_dir(project))
    chunk = [c for c in targets[0].doc.chunks if c.heading == "モジュール関数"][0]
    return {row[0]: row for row in chunk.rows[1:]}


def test_raiseに渡す補助関数は戻り値の型で出す(
        project: Paths, round_: Round) -> None:
    """**どちらにも正解が無い欄**を出さない。

    `raise _broken(...)` の `_broken` は private なヘルパの名前であって例外の型では
    ないので、整理層は書き写せず（`raises: _broken` は嘘）、落とすと `G018` が
    「欄が埋まっているのに落ちている」と鳴る。原本には `-> YamlError` と
    **書いてある**ので、辿るのは転記であって推測ではない。
    """
    rows = _rows(project, round_, '''
class YamlError(Exception):
    pass


def _broken(exc: Exception) -> YamlError:
    return YamlError(str(exc))


def marked(text: str) -> int:
    try:
        return int(text)
    except ValueError as exc:
        raise _broken(exc) from exc
''')
    assert rows["marked"][5] == "YamlError"
    assert rows["_broken"][5] == ""           # 返すだけで投げていない


def test_注釈の無い補助関数は名前のまま出す(
        project: Paths, round_: Round) -> None:
    """**読めなかったものは読めなかったと出す。** 当てにいかない。"""
    rows = _rows(project, round_, '''
def _broken(exc):
    return RuntimeError(str(exc))


def marked(text: str) -> int:
    raise _broken(text)
''')
    assert rows["marked"][5] == "_broken"


def test_変数をraiseするときは辿らない(project: Paths, round_: Round) -> None:
    """呼び出しの形のときだけ辿る ―― 名前がたまたま関数と同じでも別物である。"""
    rows = _rows(project, round_, '''
def saved() -> int:
    return 1


def go() -> None:
    saved = ValueError("x")
    raise saved
''')
    assert rows["go"][5] == "saved"


# ── 原本の指紋 ──────────────────────────────────────────────────
def _record(project: Paths, round_: Round, source: Path) -> None:
    targets, _ = parse.plan(round_, [source], sources_dir(project))
    written, _ = parse.write(targets)
    parse.record(round_, targets, written)


def test_撮った原本の指紋を残す(project: Paths, round_: Round) -> None:
    """**上流のずれを言えるようにする。** `freeze.verify` と向きが対称である。"""
    source = write(sources_dir(project) / "a.py", "X = 1\n")
    _record(project, round_, source)

    assert parse.drifted(round_) == []
    body = round_.prints.read_text(encoding="utf-8")
    assert "資料/a.py" in body                     # 根からの相対で持つ
    assert "a.py.md" in body                          # 撮り直す先まで言う


def test_撮ったあとで原本が変わったらG019(project: Paths, round_: Round) -> None:
    """**「原本を変えるな」ではない。** 撮った版といまの版が違う、とだけ言う。"""
    source = write(sources_dir(project) / "a.py", "X = 1\n")
    _record(project, round_, source)
    write(source, "X = 2\n")

    found = parse.drifted(round_)
    assert [f.code for f in found] == ["G019"]
    assert found[0].level == "warn"                   # build は落ちない
    assert "a.py.md" in found[0].message              # どれが古いのかを言う


def test_原本が消えたらG019(project: Paths, round_: Round) -> None:
    source = write(sources_dir(project) / "a.py", "X = 1\n")
    _record(project, round_, source)
    source.unlink()

    found = parse.drifted(round_)
    assert [f.code for f in found] == ["G019"]
    assert "見つかりません" in found[0].message


def test_指紋の無いラウンドは黙る(round_: Round) -> None:
    """指紋を残す前に撮ったラウンドを、あとから壊れていることにしない。"""
    assert parse.drifted(round_) == []


def test_一部だけ撮り直しても前の指紋を消さない(
        project: Paths, round_: Round) -> None:
    """`arp4 parse` は資料の一部だけを撮り直す使い方をする。"""
    first = write(sources_dir(project) / "a.py", "X = 1\n")
    second = write(sources_dir(project) / "b.py", "Y = 1\n")
    _record(project, round_, first)
    _record(project, round_, second)

    assert parse.drifted(round_) == []
    body = round_.prints.read_text(encoding="utf-8")
    assert "資料/a.py" in body and "資料/b.py" in body


def test_上書きを見送った原本の指紋は更新しない(
        project: Paths, round_: Round) -> None:
    """**守ったつもりの編集が、いちばん黙って腐る形にならないように。**

    ディスクの上には古いパース結果が残っているのに指紋だけ新しくすると、
    ずれが消えたように見える。
    """
    source = write(sources_dir(project) / "a.py", "X = 1\n")
    _record(project, round_, source)
    write(source, "X = 2\n")

    targets, _ = parse.plan(round_, [source], sources_dir(project))
    parse.record(round_, targets, [])         # 1 本も書けなかった（＝見送った）

    assert [f.code for f in parse.drifted(round_)] == ["G019"]


# ── --exclude（0-6） ────────────────────────────────────────────
def test_excludeで外した資料は読まず件数を言う(project: Paths,
                                              round_: Round) -> None:
    """``tests/`` を丸ごと渡すと**期待値（正解）を資料として拾う**罠がある。

    除外は黙って行わない ―― 何件飛ばしたかを P014 で必ず言う。
    """
    root = sources_dir(project)
    write(root / "a.py", "X = 1\n")
    write(root / "正解" / "b.md", "# 期待値\n")

    targets, findings = parse.plan(round_, [root], root, exclude=["正解/*"])

    assert [t.path.name for t in targets] == ["a.py.md"]
    said = [f for f in findings if f.code == "P014"]
    assert len(said) == 1 and "1 ファイル" in said[0].message


def test_当たらないexcludeは打ち間違いとして言う(project: Paths,
                                                round_: Round) -> None:
    root = sources_dir(project)
    write(root / "a.py", "X = 1\n")

    targets, findings = parse.plan(round_, [root], root, exclude=["seikai/*"])

    assert len(targets) == 1                     # 何も外れていない
    said = [f for f in findings if f.code == "P014"]
    assert said and "当たりませんでした" in said[0].message


def test_excludeは名前だけでも当たる(project: Paths, round_: Round) -> None:
    root = sources_dir(project)
    write(root / "a.py", "X = 1\n")
    write(root / "深い" / "場所" / "b.golden.md", "# 期待値\n")

    targets, _ = parse.plan(round_, [root], root, exclude=["*.golden.md"])

    assert [t.path.name for t in targets] == ["a.py.md"]


# ── argparse の骨格（Phase 1-5） ────────────────────────────────
def test_argparseのコマンドを骨格として出す(project: Paths,
                                            round_: Round) -> None:
    """add_parser / add_argument の宣言は AST に書いてある ―― 転記だけで取れる。"""
    source = write(sources_dir(project) / "cli.py", '''\
import argparse


def main():
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers()
    parse = sub.add_parser("parse", help="resources to parsed")
    parse.add_argument("--round", help="round name")
    build = sub.add_parser("build", help="organized to spec")
''')
    targets, _ = parse.plan(round_, [source], sources_dir(project))
    doc = targets[0].doc
    chunk = [c for c in doc.chunks if c.anchor == "p1"][0]

    assert chunk.heading == "コマンド（argparse）"
    kinds = {(row[0], row[1]) for row in chunk.rows[1:]}
    assert ("'parse'", "コマンド") in kinds       # 定数は引用符ごと（原文のまま）
    assert ("'build'", "コマンド") in kinds
    assert ("'--round'", "引数") in kinds


def test_名前が変数のadd_parserも落とさない(project: Paths,
                                            round_: Round) -> None:
    """補助関数で包んだ CLI（arp4 自身の形）では実引数が変数になる ――
    書いてあるとおりの式を出す（黙って落とすと argparse 不使用に見える）。"""
    source = write(sources_dir(project) / "cli.py", '''\
import argparse


def add(sub, name, help_text):
    return sub.add_parser(name, help=help_text)
''')
    targets, _ = parse.plan(round_, [source], sources_dir(project))
    chunk = [c for c in targets[0].doc.chunks if c.anchor == "p1"][0]

    assert ["name", "コマンド", "help_text", "sub", "5"] in chunk.rows


def test_補助関数で包んだコマンドは1ホップ辿って出す(project: Paths,
                                                     round_: Round) -> None:
    """仮引数がそのまま add_parser へ渡る補助関数（arp4 自身の ``add()`` の形）は、
    その呼び出しがコマンドの宣言である ―― ``_factories`` と同じ 1 ホップの転記。"""
    source = write(sources_dir(project) / "cli.py", '''\
import argparse


def main():
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers()

    def add(name, help_text, strict=False):
        return sub.add_parser(name, help=help_text)

    add("init", "make skeleton")
    add("parse", "sources to parsed", strict=True)
''')
    targets, _ = parse.plan(round_, [source], sources_dir(project))
    chunk = [c for c in targets[0].doc.chunks if c.anchor == "p1"][0]
    commands = {(row[0], row[2]) for row in chunk.rows[1:]
                if row[1] == "コマンド"}

    assert ("'init'", "make skeleton") in commands
    assert ("'parse'", "sources to parsed") in commands


def test_add_parserが無ければ骨格の塊を出さない(project: Paths,
                                                round_: Round) -> None:
    source = write(sources_dir(project) / "tool.py", '''\
import argparse

parser = argparse.ArgumentParser()
parser.add_argument("path")
''')
    targets, _ = parse.plan(round_, [source], sources_dir(project))

    assert not [c for c in targets[0].doc.chunks if c.anchor == "p1"]


def test_no_ocrでもスキャンしたページの絵は出す(project: Paths, round_: Round,
                                                pdf_reader) -> None:
    """**`--no-ocr` が落とすのは機械の読みだけである。**

    実体を渡さないことではない ―― 整理層は絵を開いて読めるので、そこまで
    止めると**読める資料を読めなくする**（Excel の貼り付け画像を `--no-ocr`
    でも `images/` へ出しているのと同じ規律）。読みにいっていないことは
    `o1` に必ず書くので、「読めなかった」と「読まなかった」は混ざらない。
    """
    import paper

    path = sources_dir(project) / "受入確認書.pdf"
    paper.build(path, {"ページ": [{"スキャン": True}]})
    targets, findings = parse.plan(round_, [path], sources_dir(project),
                                   use_ocr=False)

    assert len(targets) == 1
    anchors = {c.anchor: c for c in targets[0].doc.chunks}
    assert set(anchors) == {"p1-i1", "p1-o1"}
    assert targets[0].images                       # 実体が出ている
    assert "読みにいっていません" in anchors["p1-o1"].text
    assert [f.code for f in findings if f.code == "P017"] == ["P017"]
