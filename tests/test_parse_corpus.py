"""コーパス（:mod:`corpus`）を丸ごとパースして、**実物にある書かれ方**を確かめる。

ここのテストは 1 本ずつブックを組み立てない。``arp4 parse sources/`` と同じく
**20 冊を 1 度に**通し、その結果に対して各テストが 1 つの観点だけを見る
―― 実運用で起きるのは「1 冊が壊れていて 19 冊が落ちる」であって、
「1 冊だけを完璧にパースする」ではない。

規律はどれも同じところに戻ってくる ―― **読めなかったものを黙らない**。
空欄に見えるものが「資料に無い」のか「機械が読めていない」のかで、
次のラウンドで拾い直すかどうかが正反対になる。
"""

from __future__ import annotations

import time
from pathlib import Path

import pytest

import corpus
from arp4 import mdio, parse
from arp4.paths import Paths, Round


@pytest.fixture(scope="module")
def corpus_source(tmp_path_factory: pytest.TempPathFactory) -> Path:
    directory = tmp_path_factory.mktemp("corpus") / "sources"
    directory.mkdir()
    corpus.build(directory)
    return directory


@pytest.fixture(scope="module")
def parsed(corpus_source: Path, tmp_path_factory: pytest.TempPathFactory):
    """コーパス全体のパース結果 ``{パース結果の相対パス: Doc}`` と findings。"""

    class _Round:
        parsed = tmp_path_factory.mktemp("parsed")
        images = tmp_path_factory.mktemp("images")

    targets, findings = parse.plan(_Round(), [corpus_source], corpus_source)
    docs = {t.path.relative_to(_Round.parsed).as_posix(): t.doc for t in targets}
    return docs, findings


def _doc(parsed, name: str):
    docs, _ = parsed
    assert name in docs, f"{name} がありません（あるのは {sorted(docs)}）"
    return docs[name]


def _table(doc, anchor: str | None = None) -> list[list[str]]:
    tables = [c for c in doc.chunks if c.rows
              and (anchor is None or c.anchor == anchor)]
    assert tables, f"表がありません（{[c.anchor for c in doc.chunks]}）"
    return tables[0].rows


def _notes(doc) -> str:
    return "\n".join(doc.notes)


# ── 全体 ────────────────────────────────────────────────────────
def test_壊れた1冊で残りが落ちない(parsed) -> None:
    """**30 冊のうち 1 冊は必ず開けない。** そこで止まると誰も残りを見ない。"""
    docs, findings = parsed
    assert [f.code for f in findings if f.level == "error"] == ["P010"]
    assert len(docs) >= 28                       # 19 冊ぶんのシートは読めている


def test_同名別フォルダは別のファイルになる(parsed) -> None:
    docs, _ = parsed
    assert "資料/A/基本設計書.xlsx/受注テーブル.md" in docs
    assert "資料/B/基本設計書.xlsx/受注テーブル.md" in docs


def test_マクロ付きも読む(parsed) -> None:
    """**中身は .xlsx と同じ。** 読めるものを「読めない」と言わない。"""
    doc = _doc(parsed, "資料/A/処理仕様書.xlsm/受注登録.md")
    assert ["2", "在庫引当", "引当可能数が不足するときは受注を保留にする"] in _table(doc)


def test_旧形式は何をすれば読めるかまで言う(tmp_path: Path) -> None:
    """「読めません」で終わると、その資料は誰にも拾い直されない。"""

    class _Round:
        parsed = tmp_path / "parsed"
        images = tmp_path / "images"

    (tmp_path / "旧.xls").write_bytes(b"legacy")
    _, findings = parse.plan(_Round(), [tmp_path], tmp_path)
    assert [f.code for f in findings] == ["P001"]
    assert "xlsx として保存し直して" in findings[0].message


# ── 値の忠実性 ──────────────────────────────────────────────────
def test_画面に見えている表記で出す(parsed) -> None:
    """``str()`` に任せると、資料に書いていない表記になる。

    ``2026-08-02 00:00:00`` の ``00:00:00`` は資料のどこにも無く、
    ``0.153`` は画面上 ``15.3%`` である ―― **整理層は別の値を読む。**
    """
    rows = {row[0]: row[1] for row in _table(_doc(parsed, "資料/D/型の見本.xlsx/型.md"))}
    assert rows["納期"] == "2026-08-02"          # 時刻 0 は落とす
    assert rows["締切"] == "2026-08-02 17:30:00"  # 時刻があるなら残す
    assert rows["受付開始"] == "09:30"
    assert rows["対象外"] == "FALSE"
    assert rows["達成率"] == "15.3%"             # 書式の小数点以下に合わせる
    assert rows["消費税率"] == "10%"
    # **通貨・桁区切りは触らない。** 数は同じなので、直し始めると線が引けなくなる。
    assert rows["単価"] == "1200"
    assert rows["想定件数"] == "1200000"


def test_やっかいな値でも中身を落とさない(parsed) -> None:
    """改行・パイプ・コロン・波括弧はどれも実物の設計書に出てくる。"""
    doc = _doc(parsed, "資料/D/やっかいな値.xlsx/備考.md")
    rows = {row[0]: row[1] for row in _table(doc)}
    assert rows["採番"] == "受注番号は\n年度 2 桁 + 連番 8 桁"
    assert rows["区分"] == "A|B|C のいずれか"
    assert rows["メッセージ"] == "{0}を入力してください。"

    # Markdown にしたときも表が壊れない（読み戻してアンカーが見える）。
    body = mdio.dump(doc)
    assert "受注番号は<br>年度 2 桁 + 連番 8 桁" in body
    assert r"A\|B\|C のいずれか" in body


def test_縦結合は全行へ展開し横結合は広げない(parsed) -> None:
    """縦は「同上」の慣習（＝画面上は全行に見えている）、横はそうではない。"""
    # 受注テーブルはブックの 2 枚目なので、アンカーは s2-*（1 枚目は表紙）。
    table = _table(_doc(parsed, "資料/A/基本設計書.xlsx/受注テーブル.md"), "s2-t1")
    区分 = [row[0] for row in table]
    assert 区分[1:] == ["ヘッダ", "ヘッダ", "ヘッダ", "明細", "明細"]

    screen = _table(_doc(parsed, "資料/A/基本設計書.xlsx/画面項目.md"))
    assert screen[0][0] == "受注入力画面" and screen[0][1] == ""


# ── 読めなかったものの申告 ──────────────────────────────────────
def test_計算結果の無い数式を空欄と言わない(parsed) -> None:
    """**表の上では空欄と見分けが付かない。** 番地を並べて宣言先を作る。"""
    doc = _doc(parsed, "資料/A/集計表.xlsx/件数集計.md")
    assert ["合計", ""] in _table(doc)            # 値は入れない（推測しない）
    assert "計算結果が保存されていない" in _notes(doc)
    gap = [c for c in doc.chunks if c.anchor == "s1-f1"]
    assert gap and [ref for ref, _ in gap[0].cells] == ["B4", "B5"]


def test_数式しかないシートを丸ごと落とさない(parsed) -> None:
    """非空セルが 0 になるので、黙っているとシートごと消える。"""
    doc = _doc(parsed, "資料/A/集計表.xlsx/派生値.md")
    assert [c.anchor for c in doc.chunks if c.anchor.endswith("-f1")] == ["s2-f1"]


def test_番地を並べきれないときは省いたと言う(tmp_path: Path) -> None:
    """**黙って切らない。** 20 個で切れているのに黙ると「20 個だけ」と読める。"""
    from openpyxl import Workbook

    class _Round:
        parsed = tmp_path / "parsed"
        images = tmp_path / "images"

    book = Workbook()
    for row in range(1, 26):
        book.active.cell(row=row, column=1, value=f"項目{row}")
        book.active.cell(row=row, column=2, value=f"=A{row}&\"x\"")
    book.save(tmp_path / "多い.xlsx")

    targets, _ = parse.plan(_Round(), [tmp_path], tmp_path)
    gap = [c for c in targets[0].doc.chunks if c.anchor == "s1-f1"][0]
    assert len(gap.cells) == 21                  # 20 個＋「ほか」の 1 行
    assert gap.cells[-1] == ("…", "ほか 5 個（番地はここに並べきれません）")
    assert "ほか 5 個" in "\n".join(targets[0].doc.notes)


def test_非表示の行と列は読むが混ざっていると言う(parsed) -> None:
    """**落とすのは意味の判断。** 折りたたまれただけの生きた仕様のことがある。"""
    doc = _doc(parsed, "資料/C/旧システム調査.xlsx/現行機能.md")
    table = _table(doc)
    assert any("F003" in row for row in table)   # 隠れている行も読む
    assert any("ORD_ENT" in row for row in table)
    assert "非表示の行 1 行・列 1 列" in _notes(doc)


def test_貼り付け画像は図形と呼ばない(parsed) -> None:
    """撮り直しても中身は読めない ―― 図形と同じ案内をすると空振りさせる。"""
    doc = _doc(parsed, "資料/C/画面レイアウト集.xlsx/受注入力.md")
    assert "画像 2 枚" in _notes(doc)
    shapes = [c for c in doc.chunks if c.anchor == "s1-g1"]
    assert shapes and shapes[0].at == "埋め込みオブジェクト 1 個・画像 2 枚"


def test_グループ図形は箱の数で数える(parsed) -> None:
    """アンカーを数えていた頃は「図形 1 個から 3 個のテキスト」と矛盾していた。"""
    doc = _doc(parsed, "資料/A/基本設計書.xlsx/業務フロー.md")
    assert "図形 3 個・接続子 3 本" in _notes(doc)
    labels = [c for c in doc.chunks if c.anchor == "s4-g1"][0].cells
    assert [text for _, text in labels] == ["受注登録", "与信判定\nNG は差戻し", "出荷指示"]


def test_接続は転記し目分量の線は取れないと言う(parsed) -> None:
    """矢羽根が始点側にある線は**向きを起こし直す**（逆に書かない）。"""
    doc = _doc(parsed, "資料/A/基本設計書.xlsx/業務フロー.md")
    links = [c for c in doc.chunks if c.anchor == "s4-c1"][0]
    assert links.rows[1:] == [["受注登録", "→", "与信判定\nNG は差戻し", "実線"],
                              ["与信判定\nNG は差戻し", "→", "出荷指示", "実線"]]
    assert "接続子 1 本はどこにも繋がっていません" in _notes(doc)


def test_エラー値を値と言わない(parsed) -> None:
    """``#REF!`` は**表の上では値のように見える**が、その欄の仕様は取れていない。

    数式の未計算（`s1-f1`）と分けるのは、**次にやることが正反対**だからである
    ―― あちらは開いて保存し直せば値が入るが、こちらは開き直しても直らない。
    """
    doc = _doc(parsed, "資料/C/移行計画.xlsx/課題一覧.md")
    assert ["2", "受注履歴の移送", "#REF!", "基盤"] in _table(doc)  # 画面どおり残す
    errors = [c for c in doc.chunks if c.anchor == "s1-e1"]
    assert errors and errors[0].cells == [("C3", "#REF!"), ("C4", "#N/A")]
    assert "開き直しても直りません" in _notes(doc)


def test_セルのコメントを落とさない(parsed) -> None:
    """**コメントは表に出てこない。** 黙ると、決定の理由が丸ごと消える。

    値の無いセルに付いたものも取る ―― 値で足切りすると、いちばん多い
    「空欄への指摘」だけが消える。
    """
    doc = _doc(parsed, "資料/A/項目定義書.xlsx/項目.md")
    memo = [c for c in doc.chunks if c.anchor == "s1-m1"]
    assert memo and memo[0].cells == [
        ("A5", "設計者: 採番は 2026 年度から 10 桁へ。旧 8 桁は移行対象"),
        ("E6", "レビュア: 区分は入力規則から選ぶ")]      # E6 の値は空
    assert "表には出てこない補足" in _notes(doc)


def test_字下げは階層なので落とさない(parsed) -> None:
    """項目定義書の「項目名」列は**字下げで親子を表す**（日本の慣習）。

    まとめて ``strip()`` していたぶん、親と子が同じ深さの項目として並んでいた。
    """
    doc = _doc(parsed, "資料/A/項目定義書.xlsx/項目.md")
    項目名 = [row[0] for row in _table(doc)]
    assert 項目名 == ["項目名", "受注ヘッダ", "　　受注番号", "　　受注日",
                     "明細", "　　行番号"]
    assert "| 　　受注番号 |" in mdio.dump(doc)          # 書き出しても残る


def test_行末の空白は落とす(parsed) -> None:
    """行頭と違い、**行末の空白は画面に出ていない**（残すと diff がうるさい）。"""
    rows = {row[0]: row[1] for row in _table(_doc(parsed, "資料/D/やっかいな値.xlsx/備考.md"))}
    assert rows["子項目"] == "　　受注番号"


def test_リンク先は表に出ないので別に出す(parsed) -> None:
    """目次シートは**表示文字列だけ見ると語の一覧**だが、実体はリンクの束である。

    リンク先を落とすと、**まだ集めていない資料があること自体が分からない。**
    """
    doc = _doc(parsed, "資料/A/目次.xlsx/目次.md")
    links = [c for c in doc.chunks if c.anchor == "s1-l1"]
    assert links and links[0].cells == [
        ("B2", "#受注テーブル!A1"),                       # ブックの中（location）
        ("B3", "../外部/外部インタフェース仕様書.xlsx"),   # 別ブック（target）
        ("B4", "https://example.invalid/standards"),
        ("B5", "mailto:sekkei@example.invalid")]
    assert "このブックの外" in _notes(doc)
    assert "../外部/外部インタフェース仕様書.xlsx" in _notes(doc)


def test_スレッドコメントは本文と返信を出す(parsed) -> None:
    """**いまの Excel のコメントは旧形式ではない。**

    openpyxl が返すのは古い Excel 向けの**なりすまし**で、本文は 200 字の
    Microsoft の但し書きの末尾に埋もれ、記入者は `tc={GUID}` になる ――
    レビュー指摘は資料でいちばん新しい情報であることが多いのに、そこが
    いちばん読めない形で出ていた。返信・記入者・解決済みはそこにしか無い。
    """
    doc = _doc(parsed, "資料/A/レビュー記録.xlsx/指摘.md")
    memo = [c for c in doc.chunks if c.anchor == "s1-m1"][0]
    assert memo.cells == [
        ("A2", "とりまとめ: この表は 2026-07 版"),          # 旧形式は旧形式のまま
        ("C3", "設計者（2026-07-01）: 採番は 2026 年度から 10 桁へ。旧 8 桁は移行対象"),
        ("C3", "返信 レビュア（2026-07-02）: 移行対象の洗い出しは 8/20 まで"),
        ("C4", "[解決済み] レビュア（2026-06-20）: 桁あふれの検討結果は別紙のとおり"),
        # **名簿から消えた記入者**（退職者・外部レビュア）でもコメントは落とさない
        ("D5", "記入者不明（2026-07-05）: ここは仕様が固まっていない")]


def test_なりすましのコメントを二重に数えない(parsed) -> None:
    """**同じ指摘が旧形式でも書かれている。** 素直に出すと件数が倍に見える。"""
    doc = _doc(parsed, "資料/A/レビュー記録.xlsx/指摘.md")
    body = mdio.dump(doc)
    assert "[Threaded comment]" not in body
    assert "go.microsoft.com" not in body
    assert body.count("採番は 2026 年度から 10 桁へ") == 1
    assert "コメント（メモ）が 5 件" in _notes(doc)    # 旧 1 件＋スレッド 4 件


def test_スレッドのパートが無くても本文は出す(parsed) -> None:
    """**パートを落として配られた資料**（抜粋・別ツールで再保存）がある。

    なりすましだけが残るので、但し書きを剥がして本文を出すしかない ――
    記入者も返信も取れないが、**但し書き 200 字より本文のほうが要る。**
    """
    doc = _doc(parsed, "資料/A/レビュー記録.xlsx/旧メモ.md")
    memo = [c for c in doc.chunks if c.anchor == "s2-m1"][0]
    assert memo.cells == [("B2", "10 桁化は次期対応")]


def test_取り消し線は画面に見えているので申告する(parsed) -> None:
    """**表の上では生きた行と見分けが付かない。**

    ここには長く「取り消し線＝廃止と読むのは整理層の仕事」と書いてあったが、
    **整理層に届いていたのは `受注一括取込` という文字だけ**で、画面に見えて
    いる `~~受注一括取込~~` ではなかった ―― 判断を任せた先が、判断の材料を
    受け取っていない。値は落とさず、番地を並べて宣言先を作る。
    """
    doc = _doc(parsed, "資料/C/廃止一覧.xlsx/機能一覧.md")
    assert ["F002", "受注一括取込", "2026-04 廃止"] in _table(doc)   # 値は残す
    struck = [c for c in doc.chunks if c.anchor == "s1-d1"][0]
    assert struck.cells == [("A3", "F002"), ("B3", "受注一括取込"),
                            ("C3", "2026-04 廃止"), ("B5", "受注照会")]
    assert "機械は判断していません" in _notes(doc)


def test_太字や色は取らない(parsed) -> None:
    """**強調は値を偽らない。** 取り消し線だけが「消してある」と見せている。

    ここを開けると罫線・塗り・フォントまで同じ理屈で入ってきて、どこまでが
    「資料に書いてあること」なのかを機械が決めることになる。
    """
    doc = _doc(parsed, "資料/C/廃止一覧.xlsx/機能一覧.md")
    struck = [c for c in doc.chunks if c.anchor == "s1-d1"][0]
    assert "B2" not in {ref for ref, _ in struck.cells}   # 太字＋赤字のセル
    assert "太字" not in mdio.dump(doc) and "色" not in mdio.dump(doc)


def test_代替テキストは人が書いた文字なので取る(parsed) -> None:
    """**貼り付け画像の中身は取れない**が、何の画像かは分かることがある。

    名前（`Picture 1`）は Excel が自動で振るので取らない ―― 並べても行が
    増えるだけで何も伝わらない。`descr` は人が書いたときにしか入らない。
    """
    doc = _doc(parsed, "資料/C/画面レイアウト集.xlsx/受注入力.md")
    alts = [c for c in doc.chunks if c.anchor == "s1-a1"][0]
    assert alts.cells == [("画像", "受注入力画面のイメージ（明細は 20 行まで）"),
                          ("埋め込みオブジェクト", "項目定義（受注入力）.docx")]
    assert "Picture 1" not in mdio.dump(doc)              # 自動で振られた名前
    assert "画像 2 枚" in _notes(doc)                     # 枚数の申告は変わらない


def test_外部ブック参照はまだ手元に無い資料の一覧になる(parsed) -> None:
    """**表からいっそう見えない。** 参照先の値はキャッシュされている。

    画面に出ているのはただの数で、それが別ブック由来だとは誰も気付かない
    ―― その資料が集まっていなければ、数の根拠は次のラウンドでも確かめられない。
    """
    _, findings = parsed
    external = {f.target: f.message for f in findings if f.code == "P004"}
    assert "外部参照.xlsx" in external
    assert "../../外部/単価表.xlsx" in external["外部参照.xlsx"]
    assert "シート: 単価・改定履歴" in external["外部参照.xlsx"]   # キャッシュ済み
    assert all(f.level == "warn" for f in findings if f.code == "P004")


def test_入力規則の選択肢は取らない(parsed) -> None:
    """**クリックするまで画面に出ないものは「見えている表記」ではない。**

    区分の候補（通常・返品・見積）は入力規則にしか無い。取れば整理層は楽になるが、
    そこを取り始めると条件付き書式・数式の分岐まで境目が無くなる ―― 資料に
    書いてあることの転記ではなく、**Excel の設定を仕様として読み直す**作業になる。
    取らないと決めたので、**取れていないとも言わない**（申告は画面に見えている
    ものを取り落としたときだけ出す）。
    """
    doc = _doc(parsed, "資料/A/項目定義書.xlsx/項目.md")
    assert "通常" in {row[4] for row in _table(doc)}      # 入力済みの値は取れている
    assert "返品" not in mdio.dump(doc)                   # 候補は出さない
    assert "入力規則" not in _notes(doc)


# ── シートの扱い ────────────────────────────────────────────────
def test_書き出し先の衝突で1枚も消さない(parsed) -> None:
    """記号を落とすと別のシートが同じファイル名になる。**先に出たほうは動かさない。**"""
    docs, findings = parsed
    assert "資料/D/様式集.xlsx/受注.md" in docs
    assert "資料/D/様式集.xlsx/受注~2.md" in docs
    assert docs["資料/D/様式集.xlsx/受注.md"].chunks[0].cells[0][1] == "受注（旧）"
    assert docs["資料/D/様式集.xlsx/受注~2.md"].chunks[0].cells[0][1] == "受注（新）"
    assert [f.code for f in findings if f.code == "P002"] == ["P002"]


def test_空のシートと非表示シートは出さない(parsed) -> None:
    docs, _ = parsed
    assert "資料/D/様式集.xlsx/白紙.md" not in docs
    assert "資料/D/様式集.xlsx/作業用.md" not in docs
    assert "資料/D/様式集.xlsx/_CON.md" in docs   # 予約名は避けるが落とさない


def test_1x1はテキストとして番地付きで出す(parsed) -> None:
    doc = _doc(parsed, "資料/D/様式集.xlsx/注記.md")
    assert doc.chunks[0].cells == [("C5", "本書は 2026 年度版である")]
    assert doc.chunks[0].at == "C5"


def test_非表示シートは読まないがあったことは言う(parsed) -> None:
    """**行・列と扱いが違うぶん、消え方も違う。**

    行・列の非表示は読んで申告するのに、シートは読まない ―― そこまでは決めごと
    だが、**黙って落とすところまで同じにはしない**。「旧版」を隠しただけの
    ブックは実案件にごく普通にあり、シートが 1 枚も出てこなければ誰も疑わない。
    """
    _, findings = parsed
    hidden = {f.target: f.message for f in findings if f.code == "P003"}
    assert "移行計画.xlsx" in hidden and "様式集.xlsx" in hidden
    assert "旧版_v1.0" in hidden["移行計画.xlsx"]
    assert all(f.level == "warn" for f in findings if f.code == "P003")


def test_グラフだけのシートが消えない(parsed) -> None:
    """**グラフは ``xdr:sp`` ではない。** 数えていなかった頃はファイルが出なかった。

    セルが 0 個・図形が 0 個になるので、シートが存在したことすら整理層に
    伝わらない ―― `未読取` を宣言する先も無い、いちばん静かな消え方である。
    """
    doc = _doc(parsed, "資料/B/実績報告.xlsx/推移グラフ.md")
    assert "グラフ 1 個" in _notes(doc)
    assert [c.anchor for c in doc.chunks] == ["s2-k1", "s2-g1"]


def test_グラフの中身が取れたら取れていないと言わない(parsed) -> None:
    """**申告が本文と食い違うのがいちばん困る。**

    `k1` にタイトルも参照範囲も出ているシートで「中身は取れていません」と
    言うと、読み手はどちらを信じればいいのか分からなくなる ―― 申告が
    信用されなくなるのは、申告しないのと同じくらい悪い。
    """
    doc = _doc(parsed, "資料/B/実績報告.xlsx/推移グラフ.md")
    assert "中身は取れていません" not in _notes(doc)
    assert "arp4 render" not in _notes(doc)              # 絵にしても何も増えない
    assert "セルの値は 1 つもありません" in _notes(doc)


def test_グラフはどのシートを読めばいいかまで言う(parsed) -> None:
    """**「別のシートの値です」だけでは誰も探しに行かない。**

    系列が指している範囲は `c:f` に `'元データ'!$B$2:$B$3` と**そのまま
    書いてある** ―― 読むのは座標からの復元ではなく転記である。
    """
    doc = _doc(parsed, "資料/B/実績報告.xlsx/推移グラフ.md")
    chart = [c for c in doc.chunks if c.anchor == "s2-k1"][0]
    assert chart.rows[0] == ["グラフ", "系列", "分類", "値"]
    assert [row[0] for row in chart.rows[1:]] == ["月別の受注・出荷件数"] * 2
    assert [row[3] for row in chart.rows[1:]] == ["'元データ'!$B$2:$B$3",
                                                  "'元データ'!$C$2:$C$3"]
    assert "参照先は「元データ」シート" in _notes(doc)


def test_グラフシートも1本出す(parsed) -> None:
    """**セルを持たないシート**もシートである（アンカーの番号は並び順で振る）。

    セルが 1 つも無いので、グラフを読まないと**書くことが何も無い** ――
    タイトルと参照範囲だけが、次にどこを読めばいいかを伝える。
    """
    doc = _doc(parsed, "資料/B/実績報告.xlsx/売上グラフ.md")
    assert "グラフシート" in _notes(doc)
    assert [c.anchor for c in doc.chunks] == ["s3-g1", "s3-k1"]   # 3 枚目
    chart = [c for c in doc.chunks if c.anchor == "s3-k1"][0]
    assert chart.rows[1][0] == "受注件数の推移"
    assert chart.rows[1][3] == "'元データ'!$B$2:$B$3"


def test_SmartArtの箱から文字を取る(parsed) -> None:
    """**SmartArt は図形を 1 つも持たない。** 文字は別のパートにしかない。

    近年の設計書の体制図・業務フローはこれで描かれていることが多く、
    ``xdr:sp`` だけを見ていた頃は「図形 0 個」でシートごと消えていた。
    """
    doc = _doc(parsed, "資料/B/体制図.xlsx/体制.md")
    assert "SmartArt 1 個（箱 3 個）" in _notes(doc)
    labels = [c for c in doc.chunks if c.anchor == "s1-g1"][0].cells
    # 見た目用の複製（pres）と繋ぎの点（parTrans/sibTrans）は数えない
    assert [text for _, text in labels] == ["PM", "業務チーム\n受注・請求", "基盤チーム"]
    assert "箱どうしの繋がり（親子・順序）は取れていません" in _notes(doc)


def test_空列で隔てた表は地続きにならない(parsed) -> None:
    """**番地から表の中の位置が割り出せる**ようにする（区切りの前提そのもの）。

    空の行・列を詰めていた頃は ``at=B3:F8`` と書いてある表の列が 4 本しかなく、
    3 本目が D 列か E 列かを読み手が決められなかった。おまけに横に並んだ別々の
    表が地続きになり、``| 画面ID | 画面名 | ボタン | 動作 |`` と**同じ行に無い
    ものが同じ行に**並んでいた。
    """
    doc = _doc(parsed, "資料/D/一覧表.xlsx/画面一覧.md")
    table = [c for c in doc.chunks if c.rows][0]
    assert table.at == "B3:F8"
    assert len(table.rows) == 6                          # 3〜8 行
    assert all(len(row) == 5 for row in table.rows)      # B〜F 列
    assert [row[2] for row in table.rows] == [""] * 6    # D 列は空のまま残る
    # 番地の算術が成り立つ ―― 上から 5 行目は 7 行目、左から 4 列目は E 列
    assert table.rows[4][:2] == ["帳票ID", "帳票名"]
    assert table.rows[0][3] == "ボタン"
    # 2 行空ければ別の塊のまま（詰めないこととは別の話）
    assert [c.at for c in doc.chunks if c.cells] == ["B11"]


def test_すかすかな塊は表にしない(parsed) -> None:
    """**工程表を格子で並べると、中身の何十倍にもなる。**

    斜めに並んだ ``■`` は空白 1 マスずつで繋がって 1 つの塊になるので、
    62 行 × 121 列の枠に 362 セルという表が出る ―― 塊の中の空きを詰めない
    約束は、たかだか空行 1 本ぶんの膨らみを想定していた。トークンを減らす
    ための区切り（規律の 4）が、逆にいちばん大きなファイルを作っていた。

    **番地付きの箇条書きなら何も落ちない** ―― どのセルにも番地が付くので、
    格子から位置を数えるより精確である（区切りの前提そのもの）。
    """
    doc = _doc(parsed, "資料/E/工程表.xlsx/工程.md")
    assert [c.anchor for c in doc.chunks] == ["s1-x1"]        # 表にはしない
    chunk = doc.chunks[0]
    assert chunk.at == "A1:DQ62" and len(chunk.cells) == 362
    assert ("A1", "開発工程表（2026 年度）") in chunk.cells
    assert ("B3", "■") in chunk.cells                        # 番地は残っている
    assert "表ではなく番地付きの箇条書き" in _notes(doc)
    assert len(mdio.dump(doc)) < 12000                        # 表なら 24,026 字


def test_詰まった表はすかすか判定に巻き込まれない(parsed) -> None:
    """**空欄の多い表がぜんぶ箇条書きになっては困る。** 大きさと密度を両方見る。"""
    doc = _doc(parsed, "資料/D/一覧表.xlsx/画面一覧.md")
    assert [c.anchor for c in doc.chunks] == ["s1-t1", "s1-x1"]   # 表のまま
    項目 = _doc(parsed, "資料/A/項目定義書.xlsx/項目.md")
    assert any(c.rows for c in 項目.chunks)                       # 空欄だらけでも表


def test_列ごとの縦結合で幻の行を作らない(parsed) -> None:
    """**画面に見えている表は 3 行しかない。**

    結合は列を丸ごと選んで掛けられる。素直に最終行まで展開すると 2000 行の
    表が生えるが、それは忠実性の回復ではなく捏造である ―― 表の外に行は無い。
    """
    doc = _doc(parsed, "資料/E/工程表.xlsx/体制.md")
    table = _table(doc)
    assert table == [["区分", "担当"], ["業務", "受注チーム"], ["業務", "請求チーム"]]
    assert [c.at for c in doc.chunks if c.rows] == ["A1:B3"]


def test_セルの値がアンカーを偽造できない(parsed) -> None:
    """**アンカーは HTML コメントで持っている。**

    資料のセルに `<!-- a:s9-t9 -->` と書いてあると、書き出した Markdown を
    読み戻した側には**本物のアンカーに見える** ―― 表の途中に無い塊が生え、
    そこから先の本文が別のアンカーの中身になる。整理結果が偽のアンカーを
    出典にしても `freeze` は通ってしまう。HTML の画面仕様書には普通に出てくる。
    """
    doc = _doc(parsed, "資料/D/やっかいな値.xlsx/備考.md")
    rows = {row[0]: row[1] for row in _table(doc)}
    assert rows["埋め込み例"] == "<!-- a:s9-t9 at=Z99 --> を出力する"  # 値は落とさない
    body = mdio.dump(doc)
    assert "&lt;!-- a:s9-t9 at=Z99 --&gt;" in body               # 実体参照で出す
    assert body.count("<!-- a:") == len(doc.chunks)              # 本物だけ


def test_偽造したアンカーは読み戻しても増えない(tmp_path: Path, parsed) -> None:
    """書いて読み戻すところまでやらないと、壊れているかどうかは分からない。"""
    doc = _doc(parsed, "資料/D/やっかいな値.xlsx/備考.md")
    path = mdio.write(tmp_path / "備考.md", doc)
    assert [a.id for a in mdio.read(path).anchors] == [c.anchor for c in doc.chunks]


def test_サロゲートペアを落とさない(parsed) -> None:
    """JIS 第 3・4 水準（`𠮷`）は人名・地名に出てくる。"""
    rows = {row[0]: row[1] for row
            in _table(_doc(parsed, "資料/D/やっかいな値.xlsx/備考.md"))}
    assert rows["担当者"] == "𠮷田・﨑山"


# ── 性能 ────────────────────────────────────────────────────────
def test_使用範囲が膨らんでも実セルしか回らない(corpus_source: Path) -> None:
    """書式だけのセルが 1 個遠くにあるだけで、``iter_rows()`` は 400 万回回る。

    削除済みデータの残骸で使用範囲が膨らんだブックは実案件でごく普通に出てくる
    ―― 30 冊のパースが分単位になり、「遅いから通さない」が始まる。
    """
    from openpyxl import load_workbook

    book = load_workbook(corpus_source / "資料/C/旧システム調査.xlsx", data_only=True)
    sheet = book["現行機能"]
    assert sheet.max_row >= 20000                # 使用範囲は確かに膨らんでいる
    assert len(parse._live(sheet)) < 100         # 回るのは実セルだけ

    start = time.perf_counter()
    parse._cells(sheet)
    assert time.perf_counter() - start < 2.0     # 直す前は 7.5 秒だった
    book.close()
