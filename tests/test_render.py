"""画像化 ―― **切り方**と、**黙って落とさない**ことを確かめる。

Excel を起動するテストは置かない（CI に Office は無い）。撮影は COM の差し替えで
経路だけを見て、判断のいるところ（どこで切るか・どこまで広がっているか）は
純粋な関数として確かめる。
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from arp4 import render
from arp4.paths import Paths
from conftest import sources_dir


def _sheet(rows: int = 10, cols: int = 5, *, row_height: float = 15.0,
           col_width: float = 8.43):
    book = Workbook()
    sheet = book.active
    for row in range(1, rows + 1):
        sheet.row_dimensions[row].height = row_height
        for column in range(1, cols + 1):
            sheet.cell(row=row, column=column, value="x")
    for column in range(1, cols + 1):
        sheet.column_dimensions[get_column_letter(column)].width = col_width
    return sheet


# ── 切り方 ──────────────────────────────────────────────────────
def test_収まるシートは1枚のまま() -> None:
    tiles = render.plan_worksheet(_sheet(rows=10))
    assert len(tiles) == 1
    assert tiles[0].rows == (1, 10)


def test_縦に長いシートは分割される() -> None:
    """行の高さの合計が上限を超えたら切る。**行数ではなくピクセルで決める。**"""
    tall = _sheet(rows=200)                      # 200 行 x 20 px = 4000 px
    tiles = render.plan_worksheet(tall, max_px=1400)

    assert len(tiles) > 1
    assert all(tile.height_px <= 1400 for tile in tiles)
    assert tiles[0].rows[0] == 1
    assert tiles[-1].rows[1] == 200


def test_分割は数行ぶん重ねる() -> None:
    """境界で切れた行を拾い直せるように。**重複はレビューで捨てられる。**"""
    tiles = render.plan_worksheet(_sheet(rows=200), max_px=1400, overlap=2)
    for previous, following in zip(tiles, tiles[1:]):
        assert following.rows[0] < previous.rows[1]        # 重なっている
        assert following.rows[0] > previous.rows[0]        # かつ必ず前へ進む


def test_横は広いときだけ切る() -> None:
    """1 行が 2 枚に分かれると対応が取れないので、横は最後の手段。"""
    wide = _sheet(rows=5, cols=25)                          # 25 列 x 64 px
    assert len(render.plan_worksheet(wide, wide_px=2200)) == 1

    tiles = render.plan_worksheet(wide, wide_px=800)
    assert len({tile.cols for tile in tiles}) > 1


def test_結合セルを割らないように境界を戻す() -> None:
    """日本の設計書は結合セルで見出しを作る。**跨いで切ると見出しだけが残る。**"""
    sheet = _sheet(rows=200)
    sheet.merge_cells(start_row=68, start_column=1, end_row=72, end_column=3)

    tiles = render.plan_worksheet(sheet, max_px=1400, overlap=0)
    boundaries = {tile.rows[0] for tile in tiles[1:]}
    assert not any(68 < boundary <= 72 for boundary in boundaries)


def test_図形を割らないように境界を戻す() -> None:
    """**結合セルより損が大きい ―― 撮る理由そのものが図形だからである。**

    切れても絵は 2 枚とも「写ってはいる」ので、**撮れなかったことにならない**
    （実測: `画面レイアウト` の貼り付け画像は左半分と右半分に分かれていた）。
    """
    sheet = _sheet(rows=5, cols=60)                     # 60 列 x 64 px = 3840 px
    boxes = [(1, 30, 5, 40)]                            # AD:AN に置いた 1 個

    tiles = render.plan_worksheet(sheet, (5, 40), boxes=boxes,
                                  wide_px=2200, overlap=0)
    assert len(tiles) > 1                               # 横に割れてはいる
    assert any(tile.cols[0] <= 30 and 40 <= tile.cols[1] for tile in tiles)


def test_予算より大きい図形は諦めて切る() -> None:
    """**進まないより切れたほうがよい**（結合セルと同じ扱い）。"""
    sheet = _sheet(rows=5, cols=60)
    tiles = render.plan_worksheet(sheet, (5, 58), boxes=[(1, 2, 5, 58)],
                                  wide_px=2200)

    assert len(tiles) > 1
    assert tiles[-1].cols[1] >= 58


def test_割れない結合でも進む() -> None:
    """予算より長い結合があっても止まらない。**進まないより切れたほうがよい。**"""
    sheet = _sheet(rows=200)
    sheet.merge_cells(start_row=2, start_column=1, end_row=199, end_column=2)

    tiles = render.plan_worksheet(sheet, max_px=1400)
    assert len(tiles) > 1
    assert tiles[-1].rows[1] == 200


def test_列の宣言はminからmaxまで効く() -> None:
    """openpyxl は ``<col min=1 max=12>`` を**先頭の列にしか登録しない**。

    そこで既定幅を使うと、細い列が並ぶ方眼シートの幅を何倍にも見積もり、
    要らない横分割が出る。
    """
    book = Workbook()
    sheet = book.active
    for column in range(1, 41):
        sheet.cell(row=1, column=column, value="x")
    narrow = sheet.column_dimensions["A"]
    narrow.width = 1.8
    narrow.min, narrow.max = 1, 40

    widths = render._col_px(sheet, 40)
    assert max(widths) < 20                                 # 1.8 文字ぶんのまま
    assert len(render.plan_worksheet(sheet, wide_px=2200)) == 1


def test_行も列も無ければ撮らない() -> None:
    book = Workbook()
    assert render.plan_worksheet(book.active) == []


# ── 広がり（図形はセルの外にいる） ──────────────────────────────
def test_図形しか無いシートも広がりを持つ(tmp_path: Path) -> None:
    """**セルが空でも図形はそこにある。** 画像化したいのは大抵このシートである。"""
    book = Workbook()
    sheet = book.active
    sheet["A1"] = "業務フロー"
    path = tmp_path / "flow.xlsx"
    book.save(path)
    _add_drawing(path, to_row=40, to_col=30)

    extent = render.drawing_extent(path)
    assert extent[sheet.title] == (41, 31)

    from openpyxl import load_workbook
    reopened = load_workbook(path)
    tiles = render.plan_worksheet(reopened[sheet.title],
                                  extent[sheet.title])
    assert tiles[-1].rows[1] >= 41
    assert tiles[0].cols[1] >= 31


def test_図形は1個ずつ位置を持つ(tmp_path: Path) -> None:
    """**右下だけでは切り口を決められない。** どこからどこまでが 1 個かが要る。"""
    book = Workbook()
    book.active["A1"] = "業務フロー"
    path = tmp_path / "flow.xlsx"
    book.save(path)
    _add_drawing(path, to_row=40, to_col=30)

    boxes = render.drawing_boxes(path)[book.active.title]
    assert boxes == [(2, 2, 41, 31)]
    assert render.extent_of(boxes) == (41, 31)          # 広がりはここから導く


def test_描画が読めなくても止まらない(tmp_path: Path) -> None:
    """壊れた資料で**パイプラインごと落とさない**（広がりが分からないだけ）。"""
    broken = tmp_path / "broken.xlsx"
    broken.write_bytes(b"not a zip")
    assert render.drawing_extent(broken) == {}
    assert render.drawing_boxes(broken) == {}


# ── 実測（現場の 1 冊を丸ごと写した検体） ──────────────────────
@pytest.fixture(scope="module")
def dataset_source(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """検体を実ファイルにする（**1 度だけ**組む）。"""
    import dataset

    directory = tmp_path_factory.mktemp("render") / "sources"
    directory.mkdir()
    dataset.build(directory)
    return directory


def test_貼り付け画像は1枚の絵に収まる(dataset_source: Path) -> None:
    """**合成した検体では出ない。** 図形が切れ目をまたぐには、セルより図形の
    ほうが右まで届いていて、なおかつ横に割れるだけ広い 1 枚が要る。

    実測（`画面レイアウト`）: セルは AB 列までなのに図形は AO 列まで届くので
    横に割れ、切れ目が注記の箱と貼り付け画像 2 枚を縦に貫いていた ―― 左側が
    1 枚目・右側が 2 枚目に写るので、**どちらの絵にも画像が丸ごとは出ない。**
    """
    book = dataset_source / "資料/K/受注管理システム基本設計書（第3.2版）.xlsx"
    boxes = render.drawing_boxes(book)["画面レイアウト"]
    tiles = render.plan_sheet(book, "画面レイアウト")

    assert len(boxes) == 3                              # 注記 1 + 貼り付け画像 2
    assert len(tiles) > 1, "横に割れていないと検体にならない"
    for box in boxes:
        assert any(tile.rows[0] <= box[0] and box[2] <= tile.rows[1]
                   and tile.cols[0] <= box[1] and box[3] <= tile.cols[1]
                   for tile in tiles), f"{box} を切り口が貫いている"


def test_開けない資料があっても残りは計画する(tmp_path: Path) -> None:
    """**1 冊の失敗で全部を落とさない。**

    撮影側は最初からそうしてあり `parse` も飛ばして続けるのに、**計画側にだけ
    守りが無かった** ―― zip として壊れた 1 冊を渡すと `BadZipFile` が素通しに
    なり、**残りの 13 冊が 1 枚も撮れない**（実測: `test_data` 14 冊）。
    """
    from arp4 import cli

    good = tmp_path / "無事.xlsx"
    book = Workbook()
    book.active.title = "業務フロー"
    book.active["A1"] = "受注業務フロー"
    book.save(good)
    broken = tmp_path / "添付_破損.xlsx"
    broken.write_bytes("PK\x03\x04 添付が途中で切れたファイル".encode())

    plans, _, unopenable = cli._render_plan([broken, good], _args(all=True))

    assert [where for where, _, _ in plans] == [good]    # 続きは撮れる
    assert [where for where, _ in unopenable] == [broken]
    assert "BadZipFile" in unopenable[0][1]              # **理由まで残す**


# ── 撮影（COM は差し替える） ────────────────────────────────────
class _FakeExcel:
    """撮った範囲を記録するだけの Excel。**書き出しは呼び出し側が差し替える。**"""

    def __init__(self, exported: list[str], unopenable: str = "") -> None:
        self.exported = exported
        self.titles: list[str] = []
        self.unopenable = unopenable            # 名前に含まれたら開けないブック
        self.opened: list[str] = []
        self.Visible = False
        self.DisplayAlerts = True
        self.PrintCommunication = True
        self.Workbooks = self
        self.quit = False

    # Workbooks.Open → ブック / ブック.Worksheets(名前) → シート
    def Open(self, path: str, ReadOnly: bool = False) -> "_FakeExcel":
        if self.unopenable and self.unopenable in path:
            raise RuntimeError("Workbooks クラスの Open メソッドが失敗しました。")
        self.opened.append(path)
        return self

    def Worksheets(self, name: str) -> "_FakeExcel":
        return self

    @property
    def PageSetup(self) -> "_FakeSetup":
        return _FakeSetup(self)

    def ExportAsFixedFormat(self, kind: int, path: str) -> None:
        self.exported.append(_FakeSetup.print_area)
        self.titles.append(_FakeSetup.title_rows)

    def Close(self, SaveChanges: bool = True) -> None:
        pass

    def Quit(self) -> None:
        self.quit = True


#: ヘッダ・フッタの置き場（Excel の ``PageSetup`` と同じ名前）。
_BANNERS = ("LeftHeader", "CenterHeader", "RightHeader",
            "LeftFooter", "CenterFooter", "RightFooter")


class _FakeSetup:
    """``PageSetup`` の代わり。**実物の癖を 1 つだけ写してある。**

    ``PrintCommunication`` を落としている間、Excel は**ヘッダ・フッタへの代入
    だけを黙って捨てる**（印刷範囲・見出し行・余白・ズームは通る）。写して
    おかないと、消去をまとめの中へ戻す変更がテストを通ってしまう ―― 絵に紙の
    装飾が焼き込まれるのは**実物で撮った人にだけ見える**壊れ方である。
    """

    print_area = ""
    title_rows = ""
    #: 撮る前に入っていたヘッダ・フッタ。テストが仕込み、消えたかを見る。
    banners: dict[str, str] = {}

    def __init__(self, app: object = None) -> None:
        object.__setattr__(self, "app", app)

    def __setattr__(self, name: str, value: object) -> None:
        if name == "PrintArea":
            _FakeSetup.print_area = str(value)
        if name == "PrintTitleRows":
            _FakeSetup.title_rows = str(value)
        if name in _BANNERS:
            batching = getattr(self.app, "PrintCommunication", True) is False
            if not batching:                    # まとめの中の代入は捨てられる
                _FakeSetup.banners[name] = str(value)


class _FakeCom:
    def CoUninitialize(self) -> None:
        pass


def test_区画ごとに印刷範囲を切り替えて撮る(tmp_path: Path, monkeypatch) -> None:
    exported: list[str] = []
    excel = _FakeExcel(exported)
    monkeypatch.setattr(render, "_rasterize", lambda pdf, target: _Pixel())

    tiles = [render.Tile((1, 40), (1, 5), 400, 800),
             render.Tile((39, 70), (1, 5), 400, 600)]
    shots, blank = render.render_sheet(
        tmp_path / "book.xlsx", "業務フロー", tmp_path / "out", tiles=tiles,
        acquire=lambda: (_FakeCom(), excel))

    assert exported == ["A1:E40", "A39:E70"]
    assert [s.path.name for s in shots] == ["業務フロー-1.png", "業務フロー-2.png"]
    assert not blank
    assert excel.quit                                       # **必ず終了させる**


def test_紙の装飾は絵に焼き込まない(tmp_path: Path, monkeypatch) -> None:
    """**ヘッダ・フッタを消してから撮る。** 消し損ねると撮る対象そのものを潰す。

    余白を 0 にしてあるので、ヘッダは 1 行目と同じ高さに来る ―― 落ちていると
    紙の装飾が**表の見出しへ重ね刷り**になり、雑音どころか読めなくなる。文書
    番号・版・機密区分はパースが `s?-p1` で別に申告しているので、絵に要るものは
    1 つも無い。

    消去を ``PrintCommunication`` のまとめの**外**でやっているかを見ている
    （中でやると Excel が黙って捨てる ―― :class:`_FakeSetup`）。長く
    まとめの中に書いてあり、**`ヘッダ左` と `ヘッダ中` を持つ検体が 1 つも
    無かったので**、右上と下端に出るぶんだけが無害に見えていた。
    """
    excel = _FakeExcel([])
    _FakeSetup.banners = {name: "社外秘（取扱注意）" for name in _BANNERS}
    monkeypatch.setattr(render, "_rasterize", lambda pdf, target: _Pixel())

    render.render_sheet(
        tmp_path / "book.xlsx", "項目編集仕様", tmp_path / "out",
        tiles=[render.Tile((1, 17), (1, 16), 900, 500)],
        acquire=lambda: (_FakeCom(), excel))

    assert set(_FakeSetup.banners.values()) == {""}, _FakeSetup.banners


def test_白紙は書かずに数える(tmp_path: Path, monkeypatch) -> None:
    """白紙は「資料が空」ではなく「**広がりの見積もりが行き過ぎた**」の合図。"""
    monkeypatch.setattr(render, "_rasterize", lambda pdf, target: None)

    tiles = [render.Tile((1, 40), (1, 5), 400, 800)]
    shots, blank = render.render_sheet(
        tmp_path / "book.xlsx", "業務フロー", tmp_path / "out", tiles=tiles,
        acquire=lambda: (_FakeCom(), _FakeExcel([])))

    assert not shots
    assert blank == tiles
    assert not list((tmp_path / "out").glob("*.png"))


def test_開けないブックがあっても残りは撮る(tmp_path: Path, monkeypatch) -> None:
    """**1 冊の失敗で全部を落とさない。** それまでに撮れた絵まで無かったことになる。"""
    monkeypatch.setattr(render, "_rasterize", lambda pdf, target: _Pixel())
    excel = _FakeExcel([], unopenable="壊れ")

    jobs = [
        render.Job(tmp_path / "壊れ.xlsx", "業務フロー", tmp_path / "out",
                   (render.Tile((1, 10), (1, 5), 400, 200),)),
        render.Job(tmp_path / "無事.xlsx", "業務フロー", tmp_path / "out",
                   (render.Tile((1, 10), (1, 5), 400, 200),)),
    ]
    results = render.render_all(jobs, acquire=lambda: (_FakeCom(), excel))

    assert results[0].error                                 # 理由が残っている
    assert "Open" in results[0].error or "開け" in results[0].error
    assert not results[0].shots
    assert len(results[1].shots) == 1                       # 続きは撮れている
    assert excel.quit


def test_同じブックは1回しか開かない(tmp_path: Path, monkeypatch) -> None:
    """起動と読み込みが所要のほとんどを占めるので、シートごとに開き直さない。"""
    monkeypatch.setattr(render, "_rasterize", lambda pdf, target: _Pixel())
    excel = _FakeExcel([])

    book = tmp_path / "一冊.xlsx"
    tile = (render.Tile((1, 10), (1, 5), 400, 200),)
    jobs = [render.Job(book, name, tmp_path / "out", tile)
            for name in ("業務フロー", "ER図", "画面遷移図")]
    results = render.render_all(jobs, acquire=lambda: (_FakeCom(), excel))

    assert len(excel.opened) == 1
    assert all(len(r.shots) == 1 for r in results)
    assert {r.shots[0].path.name for r in results} == {
        "業務フロー.png", "ER図.png", "画面遷移図.png"}


def test_見出し行は2枚目以降にだけ載せる(tmp_path: Path, monkeypatch) -> None:
    """1 枚目には元から入っている。**重ねると同じ行が 2 度出る。**"""
    monkeypatch.setattr(render, "_rasterize", lambda pdf, target: _Pixel())
    excel = _FakeExcel([])

    job = render.Job(tmp_path / "一覧.xlsx", "項目一覧", tmp_path / "out",
                     (render.Tile((1, 50), (1, 8), 900, 1000),
                      render.Tile((49, 100), (1, 8), 900, 1000)),
                     title_rows=3)
    render.render_all([job], acquire=lambda: (_FakeCom(), excel))

    assert excel.titles == ["", "$1:$3"]


class _Pixel:
    """1 枚ぶんの絵の代わり（``save`` だけ持てばよい）。"""

    def save(self, path: Path) -> None:
        Path(path).write_bytes(
            b"\x89PNG\r\n\x1a\n" + b"\x00" * 4 + b"IHDR"
            + (7).to_bytes(4, "big") + (5).to_bytes(4, "big"))


# ── 出来た画像 ──────────────────────────────────────────────────
def test_PNGの大きさを読む(tmp_path: Path) -> None:
    path = tmp_path / "a.png"
    _Pixel().save(path)
    assert render.png_size(path) == (7, 5)


def test_PNGでなければ0を返す(tmp_path: Path) -> None:
    path = tmp_path / "a.png"
    path.write_bytes(b"garbage")
    assert render.png_size(path) == (0, 0)
    assert render.png_size(tmp_path / "無い.png") == (0, 0)


# ── 道具 ────────────────────────────────────────────────────────
def _add_drawing(path: Path, *, to_row: int, to_col: int) -> None:
    """xlsx に図形を 1 個ねじ込む（openpyxl は書けないので zip を直接いじる）。"""
    import shutil
    import zipfile

    xdr = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    main = "http://schemas.openxmlformats.org/drawingml/2006/main"
    drawing = (
        f'<?xml version="1.0"?><xdr:wsDr xmlns:xdr="{xdr}" xmlns:a="{main}">'
        '<xdr:twoCellAnchor>'
        '<xdr:from><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>'
        '<xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
        f'<xdr:to><xdr:col>{to_col}</xdr:col><xdr:colOff>0</xdr:colOff>'
        f'<xdr:row>{to_row}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>'
        '<xdr:sp><xdr:txBody><a:p><a:r><a:t>受注登録</a:t></a:r></a:p>'
        '</xdr:txBody></xdr:sp></xdr:twoCellAnchor></xdr:wsDr>')
    rels = ('<?xml version="1.0"?><Relationships xmlns="http://schemas.openxml'
            'formats.org/package/2006/relationships"><Relationship Id="rId9" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            'relationships/drawing" Target="../drawings/drawing1.xml"/>'
            '</Relationships>')

    work = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path) as source, zipfile.ZipFile(work, "w") as out:
        for item in source.infolist():
            if item.filename == "xl/worksheets/_rels/sheet1.xml.rels":
                continue
            out.writestr(item, source.read(item.filename))
        out.writestr("xl/drawings/drawing1.xml", drawing)
        out.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)
    shutil.move(work, path)


# ── 絵を本文から辿れるようにする ────────────────────────────────
def test_撮った絵をパース結果のアンカーへ貼る(tmp_path: Path, monkeypatch) -> None:
    """**撮っただけでは使われない。** 整理層が読むのは `parsed/**.md` である。"""
    monkeypatch.setattr(render, "_rasterize", lambda pdf, target: _Pixel())
    parsed = tmp_path / ".arp" / "rounds" / "r001" / "parsed" / "受注.xlsx" / "業務フロー.md"
    parsed.parent.mkdir(parents=True)
    parsed.write_text("# 受注.xlsx / 業務フロー\n\n"
                      "## 図形（テキストのみ）  <!-- a:s2-g1 at=図形 19 個 -->\n\n"
                      "- `図形1` 受注登録\n", encoding="utf-8")

    job = render.Job(tmp_path / "受注.xlsx", "業務フロー",
                     tmp_path / ".arp" / "rounds" / "r001" / "images" / "受注.xlsx",
                     (render.Tile((1, 40), (1, 5), 400, 800),),
                     parsed=parsed, anchor="s2-g1")
    result = render.render_all([job], acquire=lambda: (_FakeCom(), _FakeExcel([])))[0]

    assert result.attached == parsed
    body = parsed.read_text(encoding="utf-8")
    assert "![業務フロー（A1:E40）](../../images/受注.xlsx/業務フロー.png)" in body
    assert "- `図形1` 受注登録" in body                 # 本文は消さない


def test_撮り直しても絵は増えない(tmp_path: Path, monkeypatch) -> None:
    """**貼り直しは差し替え。** 何度も回すたびに同じ絵が積まれては読めない。"""
    monkeypatch.setattr(render, "_rasterize", lambda pdf, target: _Pixel())
    parsed = tmp_path / "業務フロー.md"
    parsed.write_text("# a\n\n## 図形  <!-- a:s1-g1 -->\n\n- `図形1` 受注登録\n",
                      encoding="utf-8")
    job = render.Job(tmp_path / "a.xlsx", "業務フロー", tmp_path / "out",
                     (render.Tile((1, 40), (1, 5), 400, 800),),
                     parsed=parsed, anchor="s1-g1")

    for _ in range(3):
        render.render_all([job], acquire=lambda: (_FakeCom(), _FakeExcel([])))
    assert parsed.read_text(encoding="utf-8").count("![") == 1


def test_貼り先が無くても撮影は成功する(tmp_path: Path, monkeypatch) -> None:
    """まだ parse していないだけ。**絵が撮れたことまで失敗にしない。**"""
    monkeypatch.setattr(render, "_rasterize", lambda pdf, target: _Pixel())
    job = render.Job(tmp_path / "a.xlsx", "業務フロー", tmp_path / "out",
                     (render.Tile((1, 40), (1, 5), 400, 800),),
                     parsed=tmp_path / "無い.md", anchor="s1-g1")
    result = render.render_all([job], acquire=lambda: (_FakeCom(), _FakeExcel([])))[0]

    assert result.shots and not result.error
    assert result.attached is None


# ── 範囲を名指しして撮り直す（拡大の入口） ──────────────────────
def test_名指しした範囲だけを撮る() -> None:
    """**人が範囲を決めたら機械は広げない**（図形の広がりも右の逃げも足さない）。"""
    sheet = _sheet(rows=60, cols=20)
    tiles = render.plan_range(sheet, "C10:H30")

    assert len(tiles) == 1
    assert tiles[0].rows == (10, 30)
    assert tiles[0].cols == (3, 8)
    assert tiles[0].range == "C10:H30"


def test_名指しの範囲でも長ければ割る() -> None:
    """拡大したいのに 1 枚へ詰めたら元の木阿弥である。"""
    tiles = render.plan_range(_sheet(rows=300, cols=10), "B10:E290", max_px=1400)

    assert len(tiles) > 1
    assert tiles[0].rows[0] == 10                       # **1 行目から数え直さない**
    assert tiles[-1].rows[1] == 290
    assert all(t.cols == (2, 5) for t in tiles)
    assert all(t.height_px <= 1400 for t in tiles)


def test_名指しの範囲でも図形は割らない() -> None:
    """寄るのは**図形を読むため**である。割った拍子に切れたら来た意味が無い。"""
    tiles = render.plan_range(_sheet(rows=5, cols=60), "A1:BH5",
                              boxes=[(1, 30, 5, 40)], wide_px=2200, overlap=0)
    assert any(tile.cols[0] <= 30 and 40 <= tile.cols[1] for tile in tiles)


def test_名指しの範囲でも結合セルは割らない() -> None:
    sheet = _sheet(rows=300, cols=10)
    sheet.merge_cells(start_row=78, start_column=2, end_row=82, end_column=5)

    tiles = render.plan_range(sheet, "B10:E290", max_px=1400, overlap=0)
    boundaries = {t.rows[0] for t in tiles[1:]}
    assert not any(78 < boundary <= 82 for boundary in boundaries)


def test_範囲として読めなければ言う() -> None:
    with pytest.raises(ValueError, match="範囲として読めません"):
        render.plan_range(_sheet(), "ここからそこまで")
    with pytest.raises(ValueError, match="行と列の両方"):
        render.plan_range(_sheet(), "B:H")


def test_拡大図は全体図に足される(tmp_path: Path, monkeypatch) -> None:
    """**拡大図は全体図の代わりではない。** 差し替えると根拠が減る。"""
    monkeypatch.setattr(render, "_rasterize", lambda pdf, target: _Pixel())
    parsed = tmp_path / "業務フロー.md"
    parsed.write_text("# a\n\n## 図形  <!-- a:s1-g1 at=図形 19 個 -->\n\n"
                      "- `図形1` 受注登録\n", encoding="utf-8")
    common = dict(xlsx=tmp_path / "a.xlsx", sheet="業務フロー",
                  out_dir=tmp_path / "out", parsed=parsed, anchor="s1-g1")

    render.render_all([render.Job(tiles=(render.Tile((1, 40), (1, 8), 400, 800),),
                                  **common)],
                      acquire=lambda: (_FakeCom(), _FakeExcel([])))
    render.render_all([render.Job(tiles=(render.Tile((5, 25), (2, 6), 300, 400),),
                                  stem="業務フロー-B5_F25", **common)],
                      acquire=lambda: (_FakeCom(), _FakeExcel([])))

    body = parsed.read_text(encoding="utf-8")
    assert body.count("![") == 2                        # 全体図と拡大図が並ぶ
    assert "業務フロー.png" in body and "業務フロー-B5_F25.png" in body
    assert "- `図形1` 受注登録" in body


# ── --pending（宣言そのものが撮り直しの依頼になる） ─────────────
def _pending_project(tmp_path: Path, *, reason_kind: str = "未読取"):
    """`未読取` を 1 件持つ最小のプロジェクトを組む。"""
    from openpyxl import Workbook
    from arp4 import parse, paths as paths_module
    from conftest import organized, write

    project = paths_module.create(tmp_path)
    round_ = project.round("r001")
    book = sources_dir(project) / "資料" / "設計書.xlsx"
    book.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "業務フロー"
    workbook.active["B2"] = "2. 業務フロー"
    workbook.save(book)

    targets, _ = parse.plan(round_, [sources_dir(project)], sources_dir(project))
    parse.write(targets)
    organized(round_, "資料/設計書.xlsx/業務フロー.yml",
              "out_of_scope:\n"
              f"  - {{ anchor: s1-x1, kind: {reason_kind}, reason: 図で描かれている }}\n")
    return project, round_


def test_未読取の宣言からシートを引き当てる(tmp_path: Path) -> None:
    """宣言は表題のセルに付く。**撮るのはシート全体**（番地を撮っても何も写らない）。"""
    from arp4 import cli

    project, round_ = _pending_project(tmp_path)
    args = _args(pending=True)
    plans, skipped, note, _ = cli._pending_plan(round_, sources_dir(project), args)

    assert skipped == 0 and "1 件" in note
    (book, sheet, tiles), = plans
    assert book.name == "設計書.xlsx" and sheet == "業務フロー"
    assert tiles and tiles[0].rows[0] == 1               # 宣言の番地 B2 ではない


def test_対象外は撮り直さない(tmp_path: Path) -> None:
    """`対象外` は「資料に仕様が無い」。撮っても読むものは無い。"""
    from arp4 import cli

    project, round_ = _pending_project(tmp_path, reason_kind="対象外")
    plans, _, note, _ = cli._pending_plan(round_, sources_dir(project),
                                          _args(pending=True))
    assert not plans and not note


def test_元シートに戻れなければ数える(tmp_path: Path) -> None:
    """**黙って飛ばさない。** 資料が動いた／消えたことは分かるようにする。"""
    from arp4 import cli

    project, round_ = _pending_project(tmp_path)
    (sources_dir(project) / "資料" / "設計書.xlsx").unlink()

    plans, skipped, _, _ = cli._pending_plan(round_, sources_dir(project),
                                             _args(pending=True))
    assert not plans and skipped == 1


def test_一巡目に読めた資料が壊れていても止まらない(tmp_path: Path) -> None:
    """**`--pending` は前のラウンドの宣言から辿る。** その間に資料は差し替わる。"""
    from arp4 import cli

    project, round_ = _pending_project(tmp_path)
    book = sources_dir(project) / "資料" / "設計書.xlsx"
    book.write_bytes("PK\x03\x04 添付が途中で切れたファイル".encode())

    plans, _, _, unopenable = cli._pending_plan(round_, sources_dir(project),
                                                _args(pending=True))
    assert not plans
    assert [path for path, _ in unopenable] == [book]   # 名指しで残る


def _args(**overrides):
    import argparse

    values = dict(pending=False, sheet=None, range=None, all=False,
                  target_px=render.TARGET_PX, wide_px=render.WIDE_PX)
    values.update(overrides)
    return argparse.Namespace(**values)
